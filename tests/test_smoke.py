import json
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch
import unittest

from openpyxl import Workbook, load_workbook
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QApplication, QDialog, QHeaderView, QScrollArea, QSizePolicy

from src.models.validation_result import ValidationIssue, ValidationResult
from src.pipeline import process_file
from src.ui.desktop_app import ValidationDesktopApp, get_qt_app
from src.validators.engine import ValidationEngine


class TestSmoke(unittest.TestCase):
    def test_text_file_is_loaded_and_validated(self) -> None:
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "users.txt"
            file_path.write_text("user_id;name\n1;Dana\n2;Noam\n", encoding="utf-8")

            result = process_file(file_path, required_columns=["user_id", "name"])

            self.assertEqual(result.summary.total_rows, 2)
            self.assertTrue(result.summary.is_valid)

    def test_excel_file_missing_required_value_is_reported(self) -> None:
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "users.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["user_id", "name"])
            sheet.append([1, "Dana"])
            sheet.append([2, None])
            workbook.save(file_path)

            result = process_file(file_path, required_columns=["user_id", "name"])

            self.assertEqual(result.summary.total_rows, 2)
            self.assertFalse(result.summary.is_valid)
            self.assertEqual(len(result.issues), 1)
            self.assertEqual(result.issues[0].column_name, "name")

    def test_validation_engine_detects_missing_columns(self) -> None:
        engine = ValidationEngine(required_columns=["user_id", "name", "email"])

        result = engine.validate([
            {"user_id": 1, "name": "Dana"},
        ])

        self.assertFalse(result.summary.is_valid)
        self.assertEqual(result.issues[0].column_name, "email")

    def test_excel_report_is_created_with_summary_and_issues(self) -> None:
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "users.txt"
            output_dir = Path(temp_dir) / "output"
            input_path.write_text(
                "user_id;name;email\n1;Dana;dana@example.com\n2;Noam;\n",
                encoding="utf-8",
            )

            result = process_file(
                input_path,
                required_columns=["user_id", "name", "email"],
                output_dir=output_dir,
            )

            self.assertIsNotNone(result.report_path)
            self.assertTrue(result.report_path.exists())

            workbook = load_workbook(result.report_path)
            self.assertIn("סיכום", workbook.sheetnames)
            self.assertIn("שגיאות", workbook.sheetnames)
            self.assertNotIn("נתונים", workbook.sheetnames)
            self.assertEqual(workbook["סיכום"]["A2"].value, "שורות שנבדקו")
            self.assertEqual(workbook["סיכום"]["B2"].value, 2)
            self.assertEqual(workbook["סיכום"]["A5"].value, "הקובץ תקין")
            self.assertEqual(workbook["סיכום"]["B5"].value, False)
            self.assertEqual(workbook["סיכום"]["A9"].value, "תאריך הפקה")
            self.assertRegex(str(workbook["סיכום"]["B9"].value), r"\d{4}-\d{2}-\d{2}")
            self.assertEqual(workbook["סיכום"]["A10"].value, "שעת הפקה")
            self.assertRegex(str(workbook["סיכום"]["B10"].value), r"\d{2}:\d{2}:\d{2}")
            self.assertEqual(workbook["שגיאות"]["A1"].value, "מספר שורה")
            self.assertEqual(workbook["שגיאות"]["A2"].value, 2)
            self.assertEqual(workbook["שגיאות"]["B2"].value, "email")
            self.assertEqual(workbook["שגיאות"]["C2"].value, "ערך חובה חסר")

    def test_process_file_accepts_multiple_files_for_slot(self) -> None:
        with TemporaryDirectory() as temp_dir:
            first_path = Path(temp_dir) / "usr02_part1.txt"
            second_path = Path(temp_dir) / "usr02_part2.txt"
            first_path.write_text("BNAME;UFLAG;TRDAT;LTIME\nUSER_A;0;20260101;080000\n", encoding="utf-8")
            second_path.write_text("BNAME;UFLAG;TRDAT;LTIME\nUSER_B;0;20260102;090000\n", encoding="utf-8")

            result = process_file(
                [first_path, second_path],
                required_columns=["BNAME", "UFLAG"],
                source_name_override="USR02",
            )

            self.assertEqual(result.summary.total_rows, 2)
            self.assertTrue(result.summary.is_valid)
            self.assertEqual(sorted(result.source_files), sorted([first_path.name, second_path.name]))

    def test_sap_usr02_export_with_metadata_and_hash_is_parsed(self) -> None:
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "usr02.txt"
            output_dir = Path(temp_dir) / "output"
            input_path.write_text(
                "Table:\t\t\tUSR02\n"
                "Displayed Fields:\t\t\t44\tof\t44\n\n"
                "\tMANDT\tBNAME\tUFLAG\tTRDAT\tLTIME\tPWDSALTEDHASH\n"
                "\t100\tAROMI\t0\t17.11.2025\t11:10:44\t{x-isSHA512, 15000}ABCDEF\n",
                encoding="utf-8",
            )

            result = process_file(
                input_path,
                required_columns=["BNAME", "UFLAG", "TRDAT", "LTIME"],
                output_dir=output_dir,
                source_name_override="USR02",
            )

            self.assertEqual(result.summary.total_rows, 1)
            self.assertIsNotNone(result.report_path)
            self.assertTrue(result.report_path.exists())

    def test_sap_e070_export_with_windows_encoding_is_parsed(self) -> None:
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "e070.txt"
            content = (
                "Table:\t\tE070\n"
                "Displayed Fields:\t\t\t10\tof\t\t10\n\n"
                "\tTRKORR\tTRFUNCTION\tTRSTATUS\tTARSYSTEM\tKORRDEV\tAS4USER\tAS4DATE\tAS4TIME\tSTRKORR\tAS4TEXT\n\n"
                "\tFPDK901838\tW\tR\tFPQ\tCUST\tPICCOLOG\t17.01.2025\t13:42:40\t\tDER “Customizing”\n"
            )
            input_path.write_bytes(content.encode("cp1252"))

            result = process_file(
                input_path,
                required_columns=["TRKORR", "AS4USER", "TRFUNCTION"],
                source_name_override="E070",
            )

            self.assertEqual(result.summary.total_rows, 1)
            self.assertTrue(result.summary.is_valid)

    def test_sap_e070_export_with_legend_row_is_parsed(self) -> None:
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "e070_legend.txt"
            content = (
                "Table:\t\tE070\t\t\t\t\t\t\n"
                "Displayed Fields:\t\t\t10\tof\t\t10\t\tFixed Columns:\t1\tList Width\t1023\n"
                '" K workbench requests, W customized requests, X unclassified tasks, Q customizing tasks "\t\t\t\t\n'
                "\tTRKORR\t\tTRFUNCTION\t\tTRSTATUS\t\tTARSYSTEM\tKORRDEV\tAS4USER\t\tAS4DATE\t\tAS4TIME\tSTRKORR\tAS4TEXT\n"
                "\tFPDK902313\t\tW\t\tR\t\tFPQ\tCUST\tTMISHALI\t\t09/11/2025\t\t10:21:31\t\tTomer update\n"
            )
            input_path.write_bytes(content.encode("cp1252"))

            result = process_file(
                input_path,
                required_columns=["TRKORR", "AS4USER", "TRFUNCTION"],
                source_name_override="E070",
            )

            self.assertEqual(result.summary.total_rows, 1)
            self.assertTrue(result.summary.is_valid)
            self.assertFalse(any(issue.message == "עמודת חובה חסרה" for issue in result.issues))

    def test_desktop_gui_initializes_with_hebrew_labels(self) -> None:
        qt_app = get_qt_app()
        self.assertIsInstance(qt_app, QApplication)

        window = ValidationDesktopApp()
        try:
            self.assertEqual(window.windowTitle(), "כלי להערכת בקרות ITGC בסביבת SAP HANA APP")
            self.assertEqual(window.tabs.count(), 4)
            self.assertEqual(window.tabs.tabText(0), "קליטת קבצים")
            self.assertEqual(window.tabs.tabText(1), "ביצוע ניתוח לביקורת")
            self.assertEqual(window.tabs.tabText(2), "סקירת דוח משתמשים")
            self.assertEqual(window.tabs.tabText(3), "הגדרות מערכת לביקורת")
            self.assertIn("QTabBar::tab:selected", window.tabs.styleSheet())
            self.assertIn("background-color: #6d002f", window.tabs.styleSheet())
            self.assertIn("color: white", window.tabs.styleSheet())
            self.assertIn("בצע ניתוח", window.audit_run_button.text())
            self.assertIn("ייצוא", window.export_log_button.text())
            self.assertIn("מסך בדיקת קלטי SAP HANA DB", ValidationDesktopApp.format_rtl_text(window.header_label.text()))
            self.assertIn("כלי להערכת בקרות ITGC", ValidationDesktopApp.format_rtl_text(window.app_title_label.text()))
            self.assertIs(window.header_label.parentWidget(), window.intake_tab)
            self.assertIs(window.hint_label.parentWidget(), window.intake_tab)
            self.assertIs(window.actions_row.parentWidget(), window.intake_tab)
            self.assertIn("מקורות קלט לבדיקת SAP HANA APP", window.slots_group.title())
            self.assertTrue(window.slots_group.alignment() & Qt.AlignRight)
            self.assertEqual(window.header_label.layoutDirection(), Qt.RightToLeft)
            self.assertEqual(window.hint_label.layoutDirection(), Qt.RightToLeft)
            self.assertEqual(window.actions_row.layoutDirection(), Qt.RightToLeft)
            self.assertEqual(window.actions_row.sizePolicy().horizontalPolicy(), QSizePolicy.Expanding)
            self.assertTrue(window.header_label.alignment() & Qt.AlignRight)
            self.assertTrue(window.hint_label.alignment() & Qt.AlignRight)
            self.assertFalse(window.required_columns_group.isVisible())
            self.assertFalse(window.summary_group.isVisible())
            self.assertFalse(window.results_group.isVisible())
            self.assertIn("לחיצה כפולה", window.run_log_table.toolTip())
            self.assertEqual(window.run_log_table.columnCount(), 10)
            self.assertEqual(window.run_log_table.horizontalHeaderItem(1).text(), "קבוצת דוחות")
            self.assertEqual(window.run_log_table.horizontalHeaderItem(3).text(), "תאריך הפקה")
            self.assertEqual(window.run_log_table.horizontalHeaderItem(4).text(), "רשומות שנקלטו")
            self.assertEqual(window.run_log_table.horizontalHeaderItem(7).text(), "תיאור שגיאה")
            self.assertEqual(window.run_log_table.horizontalHeaderItem(8).text(), "תאריך בדיקה")
            self.assertEqual(window.run_log_table.horizontalHeaderItem(9).text(), "שעת בדיקה")
            self.assertIn("USR02", window.slot_widgets)
            self.assertIn("extraction_date_edit", window.slot_widgets["USR02"])
            self.assertIn("extraction_date_label", window.slot_widgets["USR02"])
            self.assertTrue(window.slot_widgets["USR02"]["extraction_date_label"].alignment() & Qt.AlignRight)
            self.assertEqual(window.slot_widgets["USR02"]["path_label"].layoutDirection(), Qt.RightToLeft)
            self.assertEqual(window.SLOT_DEFINITIONS["ADR6_USR21"]["label"], "ADR6 / USER_ADDR")
            self.assertIn("USER_ADDR", window.SLOT_DEFINITIONS["ADR6_USR21"]["description"])
            self.assertIs(window.run_log_group.parentWidget(), window.intake_tab)
            self.assertIs(window.user_preview_group.parentWidget(), window.review_tab)
            self.assertEqual(ValidationDesktopApp.format_rtl_text(window.user_preview_group.title()), "רשימת משתמשים שנטענו")
            self.assertGreaterEqual(window.user_preview_table.columnCount(), 24)
            self.assertEqual(window.user_preview_table.verticalScrollBarPolicy(), Qt.ScrollBarAlwaysOn)
            self.assertEqual(window.user_preview_table.horizontalHeaderItem(0).text(), "CLIENT")
            self.assertEqual(window.user_preview_table.horizontalHeaderItem(1).text(), "משתמש")
            self.assertEqual(window.user_preview_table.horizontalHeaderItem(8).text(), "מספר כתובת")
            self.assertEqual(window.user_preview_table.horizontalHeaderItem(9).text(), "מספר פרסונה")
            self.assertEqual(window.user_preview_table.horizontalHeaderItem(12).text(), "סיסמה ראשונית")
            self.assertEqual(window.user_preview_table.horizontalHeaderItem(13).text(), "תאריך שינוי סיסמה")
            self.assertEqual(window.user_preview_table.horizontalHeaderItem(14).text(), "תאריך הגדרת סיסמה")
            self.assertIn("טרם נבחר קובץ", window.slot_widgets["USR02"]["path_label"].text())
            window.slot_widgets["USR02"]["selected_paths"] = ["C:/temp/usr02_100.txt"]
            window._update_slot_path_label("USR02")
            self.assertEqual(window.slot_widgets["USR02"]["path_label"].layoutDirection(), Qt.LeftToRight)
            self.assertTrue(window.slot_widgets["USR02"]["path_label"].alignment() & Qt.AlignLeft)
            self.assertIn("AGR_USERS", window.slot_widgets)
            self.assertIn("RSPARAM", window.slot_widgets)
            self.assertIn("טבלאות משתמשים", window.category_run_buttons)
            self.assertIn("טבלאות משתמשים", window.category_sections)
            self.assertEqual(window.category_run_buttons["טבלאות משתמשים"].text(), "הרץ בדיקה")
            self.assertNotEqual(window.category_run_buttons["טבלאות משתמשים"].styleSheet(), "")
            users_layout = window.category_sections["טבלאות משתמשים"].layout()
            self.assertGreaterEqual(users_layout.columnStretch(1), 1)
            self.assertGreaterEqual(users_layout.columnStretch(2), 2)
        finally:
            window.close()

    def test_user_preview_table_supports_column_configuration_and_interactive_resize(self) -> None:
        get_qt_app()
        window = ValidationDesktopApp()
        try:
            self.assertEqual(window.user_preview_columns_button.text(), "הוסף / מחק עמודות")
            self.assertFalse(isinstance(window.centralWidget(), QScrollArea))
            self.assertEqual(window.user_preview_table.horizontalHeader().sectionResizeMode(0), QHeaderView.Interactive)
            self.assertEqual(window.user_preview_table.horizontalHeader().sectionResizeMode(1), QHeaderView.Interactive)
            self.assertGreater(len(window.USER_PREVIEW_COLUMN_DEFINITIONS), window.user_preview_table.columnCount() - 1)
            defined_fields = {column["field"] for column in window.USER_PREVIEW_COLUMN_DEFINITIONS}
            self.assertIn("PWDINITIAL", defined_fields)
            self.assertIn("PWDCHGDATE", defined_fields)
            self.assertIn("PWDSETDATE", defined_fields)
            self.assertIn("GLTGV", defined_fields)
            self.assertIn("GLTGB", defined_fields)
            self.assertIn("USTYP", defined_fields)
            self.assertIn("LOCNT", defined_fields)
            self.assertIn("OCOD1", defined_fields)
            self.assertIn("PASSCODE", defined_fields)
            self.assertIn("PWDSALTEDHASH", defined_fields)
            self.assertIn("SECURITY_POLICY", defined_fields)
            self.assertIn("DEPARTMENT", defined_fields)
        finally:
            window.close()

    def test_user_preview_grid_uses_available_review_tab_space(self) -> None:
        get_qt_app()
        window = ValidationDesktopApp()
        try:
            self.assertEqual(window.user_preview_group.sizePolicy().verticalPolicy(), QSizePolicy.Expanding)
            self.assertEqual(window.user_preview_table.sizePolicy().verticalPolicy(), QSizePolicy.Expanding)
            self.assertGreaterEqual(window.user_preview_table.minimumHeight(), 360)
            self.assertGreater(window.user_preview_table.maximumHeight(), 1000)
        finally:
            window.close()

    def test_user_preview_table_supports_sorting_and_period_filter(self) -> None:
        get_qt_app()
        window = ValidationDesktopApp()
        try:
            self.assertTrue(window.user_preview_table.isSortingEnabled())
            self.assertEqual(window.user_preview_status_filter.count(), 3)
            self.assertEqual(window.audit_period_from_edit.text(), "")
            self.assertEqual(window.audit_period_to_edit.text(), "")
        finally:
            window.close()

    def test_user_preview_filters_users_by_activity_in_selected_period(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            usr02_path = base_dir / "usr02_100.txt"
            usr02_path.write_text(
                "MANDT;BNAME;UFLAG;TRDAT;LTIME\n"
                "100;ACTIVE_OPEN;0;20250115;080000\n"
                "100;ACTIVE_LOCKED;64;20251231;090000\n"
                "100;INACTIVE_OLD;0;20240101;100000\n"
                "100;INACTIVE_EMPTY;0;;110000\n",
                encoding="utf-8",
            )

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window.slot_widgets["USR02"]["selected_paths"] = [str(usr02_path)]

                window.user_preview_status_filter.setCurrentIndex(0)
                window.refresh_user_preview()
                self.assertEqual(window.user_preview_table.rowCount(), 4)

                window.audit_period_from_edit.setText("2025-01-01")
                window.audit_period_to_edit.setText("2025-12-31")
                window.user_preview_status_filter.setCurrentIndex(1)
                window.refresh_user_preview()

                active_users = {
                    window.user_preview_table.item(row, 1).text()
                    for row in range(window.user_preview_table.rowCount())
                }
                self.assertEqual(window.user_preview_table.rowCount(), 2)
                self.assertEqual(active_users, {"ACTIVE_OPEN", "ACTIVE_LOCKED"})

                window.user_preview_status_filter.setCurrentIndex(2)
                window.refresh_user_preview()

                inactive_users = {
                    window.user_preview_table.item(row, 1).text()
                    for row in range(window.user_preview_table.rowCount())
                }
                self.assertEqual(window.user_preview_table.rowCount(), 2)
                self.assertEqual(inactive_users, {"INACTIVE_OLD", "INACTIVE_EMPTY"})
            finally:
                window.close()

    def test_user_preview_sorts_date_columns_chronologically(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            usr02_path = base_dir / "usr02_100.txt"
            usr02_path.write_text(
                "MANDT;BNAME;UFLAG;TRDAT;LTIME;PWDCHGDATE;PWDSETDATE;GLTGV;GLTGB\n"
                "100;USER_C;0;15.01.2024;080000;01.03.2024;01.02.2024;15.02.2024;31.12.2026\n"
                "100;USER_A;0;01.01.2020;090000;05.01.2020;02.01.2020;10.01.2020;01.01.2024\n"
                "100;USER_B;0;09.09.2023;100000;10.10.2023;10.09.2023;11.11.2023;31.12.2025\n",
                encoding="utf-8",
            )

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window.slot_widgets["USR02"]["selected_paths"] = [str(usr02_path)]
                window._apply_user_preview_columns(["BNAME", "GLTGV", "TRDAT", "PWDCHGDATE", "PWDSETDATE", "GLTGB"])
                window.refresh_user_preview()

                gltgv_column = next(
                    index
                    for index in range(window.user_preview_table.columnCount())
                    if window.user_preview_table.horizontalHeaderItem(index).text() == "תקף מתאריך"
                )
                window.user_preview_table.sortItems(gltgv_column, Qt.DescendingOrder)
                sorted_users = [window.user_preview_table.item(row, 0).text() for row in range(window.user_preview_table.rowCount())]
                displayed_dates = [window.user_preview_table.item(row, gltgv_column).text() for row in range(window.user_preview_table.rowCount())]

                self.assertEqual(sorted_users, ["USER_C", "USER_B", "USER_A"])
                self.assertEqual(displayed_dates, ["15.02.2024", "11.11.2023", "10.01.2020"])
            finally:
                window.close()

    def test_user_preview_column_selection_updates_headers_and_persists(self) -> None:
        get_qt_app()
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window._apply_user_preview_columns(["BNAME", "STATUS", "UFLAG", "PWDINITIAL", "PWDCHGDATE", "PWDSETDATE"])

                self.assertEqual(window.user_preview_table.columnCount(), 6)
                self.assertEqual(window.user_preview_table.horizontalHeaderItem(0).text(), "משתמש")
                self.assertEqual(window.user_preview_table.horizontalHeaderItem(1).text(), "סטטוס משתמש")
                self.assertEqual(window.user_preview_table.horizontalHeaderItem(2).text(), "סיסמה ראשונית")
                self.assertEqual(window.user_preview_table.horizontalHeaderItem(3).text(), "תאריך שינוי סיסמה")
                self.assertEqual(window.user_preview_table.horizontalHeaderItem(4).text(), "תאריך הגדרת סיסמה")
                self.assertEqual(window.user_preview_table.horizontalHeaderItem(5).text(), "קוד נעילה")
            finally:
                window.close()

            second_window = ValidationDesktopApp(base_dir=base_dir)
            try:
                self.assertEqual(second_window.user_preview_table.columnCount(), 6)
                self.assertEqual(second_window.user_preview_visible_columns, ["BNAME", "STATUS", "PWDINITIAL", "PWDCHGDATE", "PWDSETDATE", "UFLAG"])
            finally:
                second_window.close()

    def test_cancel_in_user_preview_column_dialog_keeps_existing_columns(self) -> None:
        get_qt_app()
        window = ValidationDesktopApp()
        try:
            original_headers = [
                window.user_preview_table.horizontalHeaderItem(index).text()
                for index in range(window.user_preview_table.columnCount())
            ]
            dialog, selection_table = window._create_user_preview_columns_dialog()
            selection_table.item(0, 2).setCheckState(Qt.Unchecked)

            with patch.object(window, "_create_user_preview_columns_dialog", return_value=(dialog, selection_table)), patch.object(
                dialog,
                "exec",
                return_value=QDialog.Rejected,
            ):
                window.show_user_preview_column_dialog()

            current_headers = [
                window.user_preview_table.horizontalHeaderItem(index).text()
                for index in range(window.user_preview_table.columnCount())
            ]
            self.assertEqual(current_headers, original_headers)
        finally:
            window.close()

    def test_user_preview_reviewer_fields_persist_in_json_by_mandt_and_bname(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            usr02_path = base_dir / "usr02_100.txt"
            usr02_path.write_text(
                "MANDT;BNAME;UFLAG;TRDAT;LTIME\n"
                "100;USER_A;0;20260101;080000\n",
                encoding="utf-8",
            )

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window.slot_widgets["USR02"]["selected_paths"] = [str(usr02_path)]
                window._apply_user_preview_columns(["MANDT", "BNAME", "REVIEW_STATUS", "REVIEW_NOTES"])
                window.refresh_user_preview()

                self.assertEqual(window.user_preview_table.rowCount(), 1)
                reviewer_combo = window.user_preview_table.cellWidget(0, 2)
                self.assertIsNotNone(reviewer_combo)
                self.assertEqual(reviewer_combo.currentText(), "טרם נבדק")

                reviewer_combo.setCurrentText("נבדק")
                notes_item = window.user_preview_table.item(0, 3)
                notes_item.setText("נסקר ואושר")
                window.refresh_user_preview()

                reviewer_combo = window.user_preview_table.cellWidget(0, 2)
                self.assertEqual(reviewer_combo.currentText(), "נבדק")
                self.assertEqual(window.user_preview_table.item(0, 3).text(), "נסקר ואושר")

                state_path = base_dir / "data" / "output" / "user_preview_reviewer_state.json"
                self.assertTrue(state_path.exists())
                state_data = json.loads(state_path.read_text(encoding="utf-8"))
                self.assertIn("100|USER_A", state_data)
                self.assertEqual(state_data["100|USER_A"]["REVIEW_STATUS"], "נבדק")
                self.assertEqual(state_data["100|USER_A"]["REVIEW_NOTES"], "נסקר ואושר")
            finally:
                window.close()

    def test_export_user_preview_to_excel_creates_file_with_reviewer_columns(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            usr02_path = base_dir / "usr02_100.txt"
            usr02_path.write_text(
                "MANDT;BNAME;UFLAG;TRDAT;LTIME\n"
                "100;USER_A;0;20260101;080000\n",
                encoding="utf-8",
            )

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window.slot_widgets["USR02"]["selected_paths"] = [str(usr02_path)]
                window._apply_user_preview_columns(["MANDT", "BNAME", "REVIEW_STATUS", "REVIEW_NOTES"])
                window.refresh_user_preview()

                reviewer_combo = window.user_preview_table.cellWidget(0, 2)
                reviewer_combo.setCurrentText("לבירור")
                window.user_preview_table.item(0, 3).setText("נדרשת בדיקה נוספת")

                export_path = window.export_user_preview_to_excel()

                self.assertIsNotNone(export_path)
                self.assertTrue(export_path.exists())
                self.assertIn("users_review_", export_path.name)
                workbook = load_workbook(export_path)
                self.assertIn("סקירת משתמשים", workbook.sheetnames)
                self.assertEqual(workbook["סקירת משתמשים"]["A1"].value, "CLIENT")
                self.assertEqual(workbook["סקירת משתמשים"]["C1"].value, "בוצעה סקירה")
                self.assertEqual(workbook["סקירת משתמשים"]["D1"].value, "הערות סוקר")
                self.assertEqual(workbook["סקירת משתמשים"]["C2"].value, "לבירור")
                self.assertEqual(workbook["סקירת משתמשים"]["D2"].value, "נדרשת בדיקה נוספת")
            finally:
                window.close()

    def test_slot_controls_are_visibly_rendered(self) -> None:
        qt_app = get_qt_app()
        window = ValidationDesktopApp()
        try:
            window.show()
            qt_app.processEvents()
            self.assertGreater(window.slot_widgets["USR02"]["button"].height(), 20)
            self.assertGreater(window.slot_widgets["USR02"]["clear_button"].height(), 20)
            self.assertGreater(window.slot_widgets["USR02"]["path_label"].height(), 20)
            self.assertGreater(window.slot_widgets["USR02"]["extraction_date_edit"].height(), 20)
            self.assertGreater(
                window.slots_scroll.widget().minimumSizeHint().height(),
                window.slots_scroll.viewport().height(),
            )
        finally:
            window.close()

    def test_slot_clear_button_removes_loaded_file(self) -> None:
        get_qt_app()
        window = ValidationDesktopApp()
        try:
            with patch("src.ui.desktop_app.QFileDialog.getOpenFileNames", return_value=(["C:/temp/e070_100.txt"], "")):
                window.choose_file("E070")

            self.assertEqual(window.slot_widgets["E070"]["selected_paths"], ["C:/temp/e070_100.txt"])

            window.slot_widgets["E070"]["clear_button"].click()

            self.assertEqual(window.slot_widgets["E070"]["selected_paths"], [])
            self.assertIn("טרם נבחר קובץ", window.slot_widgets["E070"]["path_label"].text())
        finally:
            window.close()

    def test_clear_last_load_button_removes_only_last_loaded_slot(self) -> None:
        get_qt_app()
        window = ValidationDesktopApp()
        try:
            with patch("src.ui.desktop_app.QFileDialog.getOpenFileNames", return_value=(["C:/temp/usr02_100.txt"], "")):
                window.choose_file("USR02")
            with patch("src.ui.desktop_app.QFileDialog.getOpenFileNames", return_value=(["C:/temp/e070_100.txt"], "")):
                window.choose_file("E070")

            window.clear_last_load_button.click()

            self.assertEqual(window.slot_widgets["E070"]["selected_paths"], [])
            self.assertEqual(window.slot_widgets["USR02"]["selected_paths"], ["C:/temp/usr02_100.txt"])
            self.assertEqual(window.selected_slot_key, "USR02")
        finally:
            window.close()

    def test_file_picker_remembers_last_used_folder_between_restarts(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            remembered_dir = base_dir / "remembered"
            remembered_dir.mkdir(parents=True, exist_ok=True)
            selected_multi_file = remembered_dir / "usr02_100.txt"
            selected_multi_file.write_text("BNAME;UFLAG\nUSER_A;0\n", encoding="utf-8")
            selected_single_file = remembered_dir / "rsparam.xlsx"
            Workbook().save(selected_single_file)

            get_qt_app()
            first_window = ValidationDesktopApp(base_dir=base_dir)
            try:
                with patch(
                    "src.ui.desktop_app.QFileDialog.getOpenFileNames",
                    return_value=([str(selected_multi_file)], ""),
                ) as multi_dialog_mock:
                    first_window.choose_file("USR02")

                self.assertEqual(multi_dialog_mock.call_args.args[2], str(base_dir / "data" / "input"))
            finally:
                first_window.close()

            reopened_window = ValidationDesktopApp(base_dir=base_dir)
            try:
                with patch(
                    "src.ui.desktop_app.QFileDialog.getOpenFileName",
                    return_value=(str(selected_single_file), ""),
                ) as single_dialog_mock:
                    reopened_window.choose_file("RSPARAM")

                self.assertEqual(single_dialog_mock.call_args.args[2], str(remembered_dir))
            finally:
                reopened_window.close()

    def test_file_picker_falls_back_to_default_input_when_saved_folder_is_missing(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            saved_state_path = base_dir / "data" / "output" / "file_dialog_state.json"
            saved_state_path.parent.mkdir(parents=True, exist_ok=True)
            saved_state_path.write_text(
                json.dumps({"last_directory": str(base_dir / "missing-folder")}, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            fallback_file = base_dir / "data" / "input" / "rsparam.xlsx"
            fallback_file.parent.mkdir(parents=True, exist_ok=True)
            Workbook().save(fallback_file)

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                with patch(
                    "src.ui.desktop_app.QFileDialog.getOpenFileName",
                    return_value=(str(fallback_file), ""),
                ) as dialog_mock:
                    window.choose_file("RSPARAM")

                self.assertEqual(dialog_mock.call_args.args[2], str(base_dir / "data" / "input"))
            finally:
                window.close()

    def test_export_run_log_to_excel_creates_file(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                result = ValidationResult(
                    rows=[{"BNAME": "USER_A", "__source_file": "usr02.txt"}],
                    issues=[],
                    source_files=["usr02.txt"],
                )
                result.file_row_counts = {"usr02.txt": 1}
                window._append_run_log_entries("USR02", ["C:/temp/usr02.txt"], result)

                export_path = window.export_run_log_to_excel()

                self.assertIsNotNone(export_path)
                self.assertTrue(export_path.exists())
                workbook = load_workbook(export_path)
                self.assertIn("קבצים שנבדקו", workbook.sheetnames)
                self.assertEqual(workbook["קבצים שנבדקו"]["A1"].value, "משבצת")
            finally:
                window.close()

    def test_category_run_button_validates_selected_group(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            usr02_path = base_dir / "usr02_100.txt"
            adr6_path = base_dir / "adr6.txt"
            usr02_path.write_text(
                "BNAME;UFLAG;TRDAT;LTIME\nUSER_A;0;20260101;080000\n",
                encoding="utf-8",
            )
            adr6_path.write_text(
                "ADDRNUMBER;PERSNUMBER;SMTP_ADDR\n1001;2001;user@example.com\n",
                encoding="utf-8",
            )

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window.slot_widgets["USR02"]["selected_paths"] = [str(usr02_path)]
                window.slot_widgets["ADR6_USR21"]["selected_paths"] = [str(adr6_path)]

                with patch("src.ui.desktop_app.QMessageBox.information") as information_mock, patch(
                    "src.ui.desktop_app.QMessageBox.warning"
                ) as warning_mock:
                    window.run_category_validation("טבלאות משתמשים")

                self.assertEqual(window.run_log_table.rowCount(), 2)
                self.assertEqual(window.run_log_table.item(1, 0).text(), "ADR6 / USER_ADDR")
                self.assertEqual(window.tabs.currentIndex(), 0)
                self.assertTrue(window.report_button.isEnabled())
                self.assertTrue(information_mock.called)
                self.assertFalse(warning_mock.called)
            finally:
                window.close()

    def test_category_run_warns_when_required_slot_is_missing(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            adr6_path = base_dir / "adr6_only.txt"
            adr6_path.write_text(
                "ADDRNUMBER;PERSNUMBER;SMTP_ADDR\n1001;2001;user@example.com\n",
                encoding="utf-8",
            )

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window.slot_widgets["ADR6_USR21"]["selected_paths"] = [str(adr6_path)]

                with patch("src.ui.desktop_app.QMessageBox.information") as information_mock, patch(
                    "src.ui.desktop_app.QMessageBox.warning"
                ) as warning_mock:
                    window.run_category_validation("טבלאות משתמשים")

                self.assertEqual(window.run_log_table.rowCount(), 1)
                self.assertTrue(warning_mock.called)
                self.assertTrue(information_mock.called)
            finally:
                window.close()

    def test_run_log_is_recorded_per_file_and_exposes_details(self) -> None:
        window = ValidationDesktopApp()
        try:
            window.slot_widgets["USR02"]["extraction_date_edit"].setText("2026-04-15")
            result = ValidationResult(
                rows=[
                    {"BNAME": "USER_A", "__source_file": "usr02_a.txt"},
                    {"BNAME": "USER_B", "__source_file": "usr02_b.txt"},
                ],
                issues=[
                    ValidationIssue(
                        row_number=1,
                        column_name="BNAME",
                        message="ערך חריג",
                        source_file="usr02_a.txt",
                    )
                ],
                source_files=["usr02_a.txt", "usr02_b.txt"],
            )

            window._append_run_log_entries("USR02", ["C:/temp/usr02_a.txt", "C:/temp/usr02_b.txt"], result)

            self.assertEqual(window.run_log_table.rowCount(), 2)
            self.assertEqual(window.run_log_table.item(0, 0).text(), "USR02")
            self.assertEqual(window.run_log_table.item(0, 1).text(), "טבלאות משתמשים")
            self.assertEqual(window.run_log_table.item(0, 3).text(), "2026-04-15")
            self.assertEqual(window.run_log_table.item(0, 4).text(), "1")
            self.assertEqual(window.run_log_table.item(0, 5).text(), "שגוי")
            self.assertEqual(window.run_log_table.item(1, 5).text(), "תקין")
            self.assertIn("ערך חריג", window.run_log_table.item(0, 7).text())
            self.assertIn("ללא שגיאות", window.run_log_table.item(1, 7).text())
            self.assertRegex(window.run_log_table.item(0, 8).text(), r"\d{4}-\d{2}-\d{2}")
            self.assertRegex(window.run_log_table.item(0, 9).text(), r"\d{2}:\d{2}:\d{2}")
            invalid_details = window._build_log_details(0)
            valid_details = window._build_log_details(1)
            self.assertIn("usr02_a.txt", invalid_details)
            self.assertIn("טבלאות משתמשים", invalid_details)
            self.assertIn("2026-04-15", invalid_details)
            self.assertIn("מספר רשומות שנקלטו: 1", invalid_details)
            self.assertIn("ערך חריג", invalid_details)
            self.assertIn("usr02_b.txt", valid_details)
            self.assertIn("לא נמצאו שגיאות", valid_details)
        finally:
            window.close()

    def test_rtl_formatter_keeps_text_clean_without_control_markers(self) -> None:
        value = ValidationDesktopApp.format_rtl_text("קובץ Excel, users.txt (2026)")

        self.assertEqual(value, "קובץ Excel, users.txt (2026)")
        self.assertNotIn("\u2066", value)
        self.assertNotIn("\u2067", value)
        self.assertNotIn("\u2069", value)

    def test_hana_app_slot_catalog_contains_expected_sources(self) -> None:
        self.assertEqual(ValidationDesktopApp.SLOT_DEFINITIONS["USR02"]["category"], "טבלאות משתמשים")
        self.assertEqual(ValidationDesktopApp.SLOT_DEFINITIONS["AGR_USERS"]["category"], "טבלאות הרשאות כלליות")
        self.assertEqual(ValidationDesktopApp.SLOT_DEFINITIONS["E070"]["category"], "טבלאות שינויים")
        self.assertEqual(ValidationDesktopApp.SLOT_DEFINITIONS["RSPARAM"]["category"], "מדיניות סיסמאות")
        self.assertIn("משתמשים", ValidationDesktopApp.SLOT_DEFINITIONS["USR02"]["description"])

    def test_password_policy_profile_detects_security_gaps(self) -> None:
        rows = [
            {"PROPERTY": "minimal_password_length", "VALUE": "6"},
            {"PROPERTY": "force_first_password_change", "VALUE": "FALSE"},
            {"PROPERTY": "maximum_invalid_connect_attempts", "VALUE": "10"},
        ]

        result = ValidationEngine().validate(rows, source_name="password_policy.csv")
        messages = {issue.message for issue in result.issues}

        self.assertIn("אורך סיסמה מינימלי חייב להיות לפחות 8", messages)
        self.assertIn("חובת החלפת סיסמה ראשונית חייבת להיות פעילה", messages)
        self.assertIn("מספר ניסיונות התחברות שגויים חייב להיות מוגבל", messages)

    def test_audit_policies_profile_requires_active_policy(self) -> None:
        rows = [
            {"AUDIT_POLICY_NAME": "USER_CHANGES", "IS_AUDIT_POLICY_ACTIVE": "FALSE"},
        ]

        result = ValidationEngine().validate(rows, source_name="audit_policies.csv")

        self.assertFalse(result.summary.is_valid)
        self.assertTrue(any(issue.message == "לפחות מדיניות Audit אחת חייבת להיות פעילה" for issue in result.issues))

    def test_users_profile_requires_last_login_field(self) -> None:
        rows = [
            {"USER_NAME": "DANA"},
        ]

        result = ValidationEngine().validate(rows, source_name="users_export.csv")

        self.assertFalse(result.summary.is_valid)
        self.assertTrue(any(issue.message == "נדרשת לפחות אחת מהעמודות: LAST_SUCCESSFUL_CONNECT / LAST_SUCCESSFUL_CONNECT_DATE" for issue in result.issues))

    def test_adr6_usr21_slot_accepts_adr6_only_structure(self) -> None:
        rows = [
            {"ADDRNUMBER": "1001", "PERSNUMBER": "2001", "SMTP_ADDR": "user@example.com"},
        ]

        result = ValidationEngine().validate(rows, source_name="ADR6_USR21")

        self.assertFalse(any(issue.column_name == "BNAME" and issue.message == "עמודת חובה חסרה" for issue in result.issues))

    def test_adr6_usr21_slot_accepts_addr_users_structure(self) -> None:
        rows = [
            {
                "MANDT": "100",
                "BNAME": "USER_A",
                "NAME_FIRST": "Dana",
                "NAME_LAST": "Levi",
                "NAME_TEXTC": "Dana Levi",
                "COMPANY": "Ayalon",
            },
        ]

        result = ValidationEngine().validate(rows, source_name="ADR6_USR21")

        self.assertFalse(any("אינו תואם למבנה המצופה עבור המשבצת ADR6 / ADDR_USERS" in issue.message for issue in result.issues))

    def test_user_preview_table_merges_usr02_adr6_and_addr_users(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            usr02_path = base_dir / "usr02_100.txt"
            addr_users_path = base_dir / "addr_users.txt"
            adr6_path = base_dir / "adr6.txt"

            usr02_path.write_text(
                "MANDT;BNAME;UFLAG;TRDAT;LTIME;PWDINITIAL;PWDCHGDATE;PWDSETDATE;GLTGV;GLTGB;USTYP;LOCNT;OCOD1;PASSCODE;PWDSALTEDHASH;SECURITY_POLICY\n"
                "100;USER_A;0;20260101;080000;1;20250101;20241231;20240101;20261231;A;2;SECRET;HASH160;SALTEDHASHVALUE;STRICT_POLICY\n",
                encoding="utf-8",
            )
            addr_users_path.write_text(
                "MANDT;BNAME;NAME_FIRST;NAME_LAST;NAME_TEXTC;COMPANY;ADDRNUMBER;PERSNUMBER;DEPARTMENT\n"
                "100;USER_A;Dana;Levi;Dana Levi;Ayalon;1001;2001;Finance\n",
                encoding="utf-8",
            )
            adr6_path.write_text(
                "ADDRNUMBER;PERSNUMBER;SMTP_ADDR\n"
                "1001;2001;user@example.com\n",
                encoding="utf-8",
            )

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window.slot_widgets["USR02"]["selected_paths"] = [str(usr02_path)]
                window.slot_widgets["ADR6_USR21"]["selected_paths"] = [str(addr_users_path), str(adr6_path)]
                window._apply_user_preview_columns([
                    "BNAME", "DEPARTMENT", "GLTGV", "GLTGB", "USTYP", "LOCNT", "OCOD1", "PASSCODE", "PWDSALTEDHASH", "SECURITY_POLICY"
                ])

                window.refresh_user_preview()

                self.assertEqual(ValidationDesktopApp.format_rtl_text(window.user_preview_group.title()), "רשימת משתמשים שנטענו")
                self.assertIs(window.user_preview_group.parentWidget(), window.review_tab)
                self.assertEqual(window.user_preview_table.rowCount(), 1)
                self.assertEqual(window.user_preview_table.item(0, 0).text(), "USER_A")
                self.assertEqual(window.user_preview_table.item(0, 1).text(), "Finance")
                self.assertEqual(window.user_preview_table.item(0, 2).text(), "20240101")
                self.assertEqual(window.user_preview_table.item(0, 3).text(), "20261231")
                self.assertEqual(window.user_preview_table.item(0, 4).text(), "A")
                self.assertEqual(window.user_preview_table.item(0, 5).text(), "2")
                self.assertEqual(window.user_preview_table.item(0, 6).text(), "SECRET")
                self.assertEqual(window.user_preview_table.item(0, 7).text(), "HASH160")
                self.assertEqual(window.user_preview_table.item(0, 8).text(), "SALTEDHASHVALUE")
                self.assertEqual(window.user_preview_table.item(0, 9).text(), "STRICT_POLICY")
            finally:
                window.close()

    def test_unmatched_addr_users_rows_are_excluded_when_usr02_present(self) -> None:
        with TemporaryDirectory() as temp_dir:
            base_dir = Path(temp_dir)
            usr02_path = base_dir / "usr02_100.txt"
            addr_users_path = base_dir / "addr_users.txt"

            usr02_path.write_text(
                "MANDT;BNAME;UFLAG;TRDAT;LTIME\n"
                "600;UZIZ;0;20260101;080000\n",
                encoding="utf-8",
            )
            addr_users_path.write_text(
                "MANDT;BNAME;NAME_FIRST;NAME_LAST;NAME_TEXTC;COMPANY\n"
                "600;UZIZ;Uzi;Ziv;Uzi Ziv;AYALON\n"
                "600;BASISADMIN;;;Basis Admin;AYALON\n",
                encoding="utf-8",
            )

            get_qt_app()
            window = ValidationDesktopApp(base_dir=base_dir)
            try:
                window.slot_widgets["USR02"]["selected_paths"] = [str(usr02_path)]
                window.slot_widgets["ADR6_USR21"]["selected_paths"] = [str(addr_users_path)]

                window.refresh_user_preview()

                self.assertEqual(window.user_preview_table.rowCount(), 1)
                self.assertEqual(window.user_preview_table.item(0, 1).text(), "UZIZ")
                self.assertEqual(window.user_preview_table.item(0, 7).text(), "פעיל")
            finally:
                window.close()

    def test_agr_1251_allows_empty_high_value_in_normal_sap_rows(self) -> None:
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "agr_1251_100.txt"
            input_path.write_text(
                "MANDT;AGR_NAME;COUNTER;OBJECT;AUTH;VARIANT;FIELD;LOW;HIGH\n"
                "100;/AIF/ARC_CREATE;000001;S_TCODE;T_XX93001900;;TCD;SARA;\n"
                "100;/AIF/ARC_CREATE;000002;S_ADMI_FCD;T_XX93001900;;S_ADMI_FCD;;\n",
                encoding="utf-8",
            )

            result = process_file(
                input_path,
                required_columns=["AGR_NAME", "OBJECT", "FIELD", "LOW", "HIGH"],
                source_name_override="AGR_1251",
            )

            range_issues = [issue for issue in result.issues if issue.column_name in {"LOW", "HIGH"}]
            self.assertEqual(range_issues, [])

    def test_stms_accepts_formal_sap_headers(self) -> None:
        rows = [
            {
                "Number": "1",
                "Date": "17.01.25",
                "Time": "13:42:40",
                "Request": "FPDK901838",
                "Clt": "400",
                "Owner": "PICCOLOG",
                "User": "PICCOLOG",
                "Project": "PICCOLOG",
                "Short Text": "DER_Customizing_Finetuning_16012025",
                "RC": "0",
            }
        ]

        result = ValidationEngine().validate(rows, source_name="STMS")

        self.assertFalse(any(issue.message == "עמודת חובה חסרה" for issue in result.issues))
        self.assertFalse(any("אינו תואם למבנה המצופה עבור המשבצת STMS" in issue.message for issue in result.issues))

    def test_stms_text_export_with_intro_rows_is_parsed(self) -> None:
        with TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "stms.txt"
            input_path.write_text(
                "Entries for FPP:\t\t\t155\t\t\t\t\t23.11.2025\t07:52:18\n\n"
                "\t\tTime Interval\t\t01.01.25 00:00:00\t\t\t\t  to\t23.11.25 24:00:00\n\n"
                "\tNumber\t\tDate\t\t\tTime\tRequest\t\t\tClt\tOwner\tUser\tProject\tShort Text\t\t\tRC\n\n"
                "\t     1\t\t17.01.25\t\t\t13:42:40\tFPDK901838\t\t\t400\tPICCOLOG\tPICCOLOG\t\tDER_Customizing_Finetuning_16012025\t0\n",
                encoding="utf-8",
            )

            result = process_file(
                input_path,
                required_columns=["TRKORR"],
                source_name_override="STMS",
            )

            self.assertEqual(result.summary.total_rows, 1)
            self.assertTrue(result.summary.is_valid)
            self.assertFalse(any(issue.message == "עמודת חובה חסרה" for issue in result.issues))
            self.assertFalse(any("אינו תואם למבנה המצופה עבור המשבצת STMS" in issue.message for issue in result.issues))

    def test_stms_blank_rc_value_is_not_treated_as_missing_status(self) -> None:
        rows = [
            {
                "Request": "FPDK901838",
                "RC": "",
                "Owner": "PICCOLOG",
            }
        ]

        result = ValidationEngine(required_columns=["TRKORR", "STATUS"]).validate(rows, source_name="STMS")

        self.assertFalse(any(issue.column_name == "STATUS" for issue in result.issues))

    def test_failed_slot_validation_is_logged_in_run_log(self) -> None:
        get_qt_app()
        window = ValidationDesktopApp()
        try:
            with patch("src.ui.desktop_app.process_file", side_effect=RuntimeError("boom")):
                summary = window._run_slot_validation("E070", ["C:/temp/e070.txt"], show_feedback=False)

            self.assertEqual(summary["status"], "error")
            self.assertEqual(window.run_log_table.rowCount(), 1)
            self.assertEqual(window.run_log_table.item(0, 0).text(), "E070")
            self.assertEqual(window.run_log_table.item(0, 1).text(), "טבלאות שינויים")
            self.assertEqual(window.run_log_table.item(0, 4).text(), "0")
            self.assertEqual(window.run_log_table.item(0, 5).text(), "שגיאה")
            self.assertIn("boom", window.run_log_table.item(0, 7).text())
            self.assertIn("boom", window._build_log_details(0))
        finally:
            window.close()

    def test_usr02_slot_blocks_wrong_rsparam_structure(self) -> None:
        rows = [
            {"PARAMETER": "login/min_password_lng", "VALUE": "8"},
        ]

        result = ValidationEngine().validate(rows, source_name="USR02")

        self.assertFalse(result.summary.is_valid)
        self.assertTrue(any("אינו תואם למבנה המצופה עבור המשבצת USR02" in issue.message for issue in result.issues))

    def test_rsparam_excel_with_title_row_is_detected(self) -> None:
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "rsparam.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Parameter export"])
            sheet.append([None, None])
            sheet.append(["NAME", "CURRENT VALUE"])
            sheet.append(["login/min_password_lng", "8"])
            workbook.save(file_path)

            result = process_file(
                file_path,
                required_columns=["PARAMETER", "VALUE"],
                source_name_override="RSPARAM",
            )

            self.assertFalse(any(issue.message == "עמודת חובה חסרה" for issue in result.issues))
            self.assertFalse(any("אינו תואם למבנה המצופה עבור המשבצת RSPARAM" in issue.message for issue in result.issues))

    def test_rz10_key_value_text_is_detected(self) -> None:
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "rz10.txt"
            file_path.write_text(
                "login/min_password_lng = 8\nlogin/password_lock_for_system_user = 1\n",
                encoding="utf-8",
            )

            result = process_file(
                file_path,
                required_columns=["PARAMETER", "VALUE"],
                source_name_override="TPFET",
            )

            self.assertFalse(any(issue.message == "עמודת חובה חסרה" for issue in result.issues))
            self.assertFalse(any("אינו תואם למבנה המצופה עבור המשבצת TPFET" in issue.message for issue in result.issues))

    def test_rsparam_csv_with_user_defined_and_system_default_headers_is_detected(self) -> None:
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "rsparam.csv"
            file_path.write_text(
                "Parameter Name,User-Defined Value,System Default Value,System Default Value(Unsubstituted Form),Comment\n"
                "login/min_password_lng,,8,8,Minimum password length\n",
                encoding="utf-8",
            )

            result = process_file(
                file_path,
                required_columns=["PARAMETER", "VALUE"],
                source_name_override="RSPARAM",
            )

            self.assertFalse(any(issue.message == "עמודת חובה חסרה" for issue in result.issues))
            self.assertFalse(any("אינו תואם למבנה המצופה עבור המשבצת RSPARAM" in issue.message for issue in result.issues))

    def test_rsparam_excel_with_cover_sheet_is_detected(self) -> None:
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "rsparam_cover.xlsx"
            workbook = Workbook()
            cover = workbook.active
            cover.title = "Cover"
            cover.append(["Export summary"])
            sheet = workbook.create_sheet("RSPARAM")
            sheet.append(["Parameter Name", "Parameter Value"])
            sheet.append(["login/min_password_lng", "8"])
            workbook.save(file_path)

            result = process_file(
                file_path,
                required_columns=["PARAMETER", "VALUE"],
                source_name_override="RSPARAM",
            )

            self.assertFalse(any(issue.message == "עמודת חובה חסרה" for issue in result.issues))
            self.assertFalse(any("אינו תואם למבנה המצופה עבור המשבצת RSPARAM" in issue.message for issue in result.issues))

    def test_rsparam_blank_value_is_ignored_for_irrelevant_parameter(self) -> None:
        rows = [
            {"PARAMETER": "dbs/db2/generic", "VALUE": ""},
        ]

        result = ValidationEngine(required_columns=["PARAMETER", "VALUE"]).validate(rows, source_name="RSPARAM")

        self.assertFalse(any(issue.column_name == "VALUE" and issue.message == "ערך חובה חסר" for issue in result.issues))

    def test_rsparam_blank_value_is_reported_for_itgc_relevant_parameter(self) -> None:
        rows = [
            {"PARAMETER": "login/min_password_lng", "VALUE": ""},
        ]

        result = ValidationEngine(required_columns=["PARAMETER", "VALUE"]).validate(rows, source_name="RSPARAM")

        self.assertTrue(any(issue.column_name == "VALUE" and issue.message == "ערך חובה חסר" for issue in result.issues))


if __name__ == "__main__":
    unittest.main()
