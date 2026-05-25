"""Generate Excel template for filling in audit control metadata.

Run once:
    python generate_controls_template.py

Output:
    data/output/controls_metadata_template.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

CONTROLS = [
    ("MC7-25_AYALON_44", "MC - ניהול שינויים", "גבוה", "STMS - Import מורשים בלבד",
     "Import לסביבת ייצור יתבצע רק על ידי משתמשים מורשים."),
    ("MA1-1&MA7-17_AYALON_2", "MA - ניהול גישה", "גבוה", "השלמת סקירת משתמשים",
     "אחת לשנה ישלף דוח כל המשתמשים הפעילים במערכת ויישלח לאחראי ה-IT לסקירה. הסקירה תכלול בחינה אם המשתמש פעיל/מחוק/נעול, תאריך התחברות אחרון, סוג המשתמש. אם סקירת המשתמשים טרם הושלמה - נוצר ממצא."),
    ("MA1-1_AYALON_5", "MA - ניהול גישה", "גבוה", "משתמשי מערכת",
     "DDIC וSAP* הינם משתמשים עם הרשאות הגבוהות ביותר במערכת. יש לוודא שפרמטר login/no_automatic_user_sapstar = 1 להשבתת משתמש SAP* האוטומטי."),
    ("MA2-2_AYALON_6", "MA - ניהול גישה", "גבוה", "מדיניות סיסמאות",
     "מדיניות הסיסמאות במערכת הSAP הינה על פי הbest practice: 1. אורך סיסמה מינימלי 8 תווים. 2. נעילה לאחר לכל היותר 6 ניסיונות כושלים. 3. ביטול נעילה אוטומטי מבוטל. 4. תפוגת סיסמה לכל היותר 90 ימים. 5. היסטוריית סיסמאות לפחות 5 ערכים. 6. מורכבות סיסמה (ספרות, אותיות, תווים מיוחדים). 7. Maximum idle time <=1800 שניות (30 דקות)."),
    ("MA3-3_AYALON_14", "MA - ניהול גישה", "גבוה", "פרופילים חזקים",
     "הקצאת פרופילי-על למשתמש (כגון SAP_ALL, SAP_NEW, S_A.SYSTEM) מעניקה הרשאות מערכת רחבות ודורשת בקרה הדוקה. נבדק על פי UST04 ו-USH04."),
    ("MA1-1_AYALON_10", "MA - ניהול גישה", "גבוה", "הרשאות ניהול משתמשים",
     "משתמשים בעלי הרשאות לניהול משתמשים (S_TCODE/SU01/SU10, S_USER_*) זוהו לפי אובייקטי הרשאה ב-AGR_1251."),
    ("MA1-1_AYALON_11", "MA - ניהול גישה", "גבוה", "הרשאות ניהול הרשאות",
     "משתמשים בעלי הרשאות לניהול הרשאות (S_TCODE/PFCG, S_DEVELOP/ACGR, S_USER_ADM) זוהו לפי אובייקטי הרשאה ב-AGR_1251."),
    ("MA1-1_AYALON_12", "MA - ניהול גישה", "גבוה", "הרשאות לתוכנית RSCDOK99",
     "משתמשים בעלי הרשאה להרצת תוכנית RSCDOK99 (S_PROGRAM עם P_GROUP=RSCDOK99 ו-P_ACTION=SUB) זוהו לפי אובייקטי הרשאה ב-AGR_1251."),
    ("MA1-1_AYALON_16", "MA - ניהול גישה", "גבוה", "הרשאות לניהול נתונים",
     "משתמשים בעלי הרשאות לניהול נתונים (S_TCODE/SE16/SM30/SM31/SE16N/SE17/SM38/SE37, S_TABU_DIS, S_TABU_NAM, S_TABU_CLI, S_DATASET) זוהו לפי אובייקטי הרשאה ב-AGR_1251."),
    ("MA1-1_AYALON_43", "MA - ניהול גישה", "גבוה", "הרשאה להעברת שינויים",
     "משתמשים בעלי הרשאה להעברת שינויים (S_TCODE/STMS/SCC4, S_TABU_DIS/DICBERCLS=SS, S_TRANSPORT, S_CTS_ADMI) זוהו לפי אובייקטי הרשאה ב-AGR_1251."),
    ("MA1-1_AYALON_45", "MA - ניהול גישה", "גבוה", "הרשאות לשימוש ב-DEBUG",
     "משתמשים בעלי הרשאות DEBUG (S_TCODE/SE38/SA38/SE80/ST05, S_DEVELOP/OBJTYPE=DEBUG, S_DEVELOP/ACTVT, S_PROGRAM/P_ACTION=SUB, S_PROGRAM/P_GROUP=*, S_ADMI_FCD/PADM) זוהו לפי אובייקטי הרשאה ב-AGR_1251."),
    ("MA1-1_AYALON_67", "MA - ניהול גישה", "גבוה", "הרשאה לעידכון ג'ובים",
     "משתמשים בעלי הרשאות ניהול ג'ובים (S_TCODE/SM36 וכד', S_BTCH_ADM, S_BTCH_JOB/DELE/RELE/PROT, S_BTCH_NAM/*, S_BTCH_MONI/DELE/RELE) זוהו לפי אובייקטי הרשאה ב-AGR_1251."),
    ("MA5.1-13_AYALON_24", "MA - ניהול גישה", "גבוה", "משתמשים חדשים",
     "זיהוי משתמשים שנוצרו במהלך תקופת הביקורת על פי תאריך הוקמת המשתמש (GLTGV) בUSR02."),
    ("MA5.3-13_AYALON_25", "MA - ניהול גישה", "בינוני", "משתמשים מנויידים",
     "זיהוי משתמשים שפרופיל ההרשאות שלהם השתנה במהלך תקופת הביקורת על פי שינוי ב-PROFS בטבלת USH04."),
    ("MA7-17_AYALON_30", "MA - ניהול גישה", "בינוני", "סקירת הרשאות משתמשים",
     "סקירת הרשאות תקופתית של כלל המשתמשים הפעילים במערכת על ידי גורמים עסקיים, לווידוא נחיצות ותקינות ההרשאות."),
    ("MC5-23_AYALON_48", "MC - ניהול שינויים", "גבוה", "הפרדת תפקידים - מפתחים בסביבת ייצור",
     "זיהוי מפתחים המוגדרים כמשתמשים פעילים בסביבת הייצור (Segregation of Duties)."),
]

HEADERS = [
    "מזהה בקרה",
    "קטגוריה",
    "רמת סיכון",
    "סוג בדיקה",
    "תיאור הבקרה (קיים)",
    "תהליך (למילוי)",
    "תיאור הסיכון (למילוי)",
    "צעדי טסט (אופציונלי - אם ריק יילקח טקסט גנרי)",
    "תיעוד נדרש (אופציונלי)",
    "הערות נוספות (אופציונלי)",
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "מטא-דאטה בקרות"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="right", vertical="top", wrap_text=True)
    existing_fill = PatternFill("solid", fgColor="E7E6E6")
    fillable_fill = PatternFill("solid", fgColor="FFF2CC")
    side = Side(style="thin", color="BFBFBF")
    border = Border(left=side, right=side, top=side, bottom=side)

    for col, header in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    for row_idx, (cid, cat, risk, ctype, desc) in enumerate(CONTROLS, start=2):
        values = [cid, cat, risk, ctype, desc, "", "", "", "", ""]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.alignment = cell_align
            c.border = border
            c.fill = existing_fill if col <= 5 else fillable_fill

    widths = [24, 20, 12, 28, 60, 40, 60, 50, 30, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for row_idx in range(2, len(CONTROLS) + 2):
        ws.row_dimensions[row_idx].height = 90
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("הוראות")
    ws2.sheet_view.rightToLeft = True
    instructions = [
        ("שדה", "הסבר"),
        ("תהליך", "התהליך העסקי / IT שאליו שייכת הבקרה (למשל: ניהול גישה למערכת, ניהול שינויים, אבטחת מידע)"),
        ("תיאור הסיכון", "תיאור הסיכון העסקי/בקרתי שהבקרה נועדה למתן (מה יכול להשתבש אם הבקרה לא קיימת)"),
        ("צעדי טסט", "אם ריק - יילקח טקסט גנרי. אפשר לדרוס במלל מפורט פר-בקרה"),
        ("תיעוד נדרש", "אילו טבלאות/קבצים נדרשים. אם ריק - יילקח אוטומטית מהפרופיל"),
        ("הערות נוספות", "כל מידע משלים שיופיע בנייר העבודה"),
    ]
    for r, (a, b) in enumerate(instructions, start=1):
        ca = ws2.cell(row=r, column=1, value=a)
        cb = ws2.cell(row=r, column=2, value=b)
        if r == 1:
            for c in (ca, cb):
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = border
        else:
            for c in (ca, cb):
                c.alignment = cell_align
                c.border = border
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 80

    output = Path(__file__).parent / "data" / "output" / "controls_metadata_template.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"Created: {output.resolve()}")


if __name__ == "__main__":
    main()
