"""Controls Catalog loader — reads/writes data/knowledge_base/controls_catalog.json.

The catalog stores per-control metadata (description, process, risk_description,
test_steps_override, notes, in_scope, analysis_type) and is the single source of
truth for all fields editable by the user.

At application startup, call ``load_and_apply_catalog`` to:
  1. Load controls_catalog.json from the knowledge_base directory.
  2. Merge the user-editable fields into AUDIT_CONTROL_DEFINITIONS in-place
     (so the rest of the application reads them transparently).

The legacy CSV loader (controls_metadata_loader) is still tried AFTER this so
that any hand-edited CSV can still override individual fields.

Excel import/export (Phase 2) is scaffolded but not yet wired into the UI.
"""
from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

CATALOG_FILENAME = "controls_catalog.json"

# Fields stored in the catalog that are merged into AUDIT_CONTROL_DEFINITIONS.
_MERGEABLE_FIELDS = (
    "description",
    "process",
    "risk_description",
    "test_steps_override",
    "notes",
)

# Fields that identify and categorise a control (read-only from catalog perspective).
_IDENTITY_FIELDS = (
    "control_id",
    "category",
    "sub_category",
    "risk_level",
    "check_type",
    "analysis_type",
    "in_scope",
)


def _catalog_path(knowledge_base_dir: Path) -> Path:
    return Path(knowledge_base_dir) / CATALOG_FILENAME


def load_catalog(knowledge_base_dir: Path) -> list[dict[str, Any]]:
    """Load controls_catalog.json.  Returns a list of control dicts.

    Returns an empty list when the file does not exist or cannot be parsed.
    """
    path = _catalog_path(knowledge_base_dir)
    if not path.exists():
        logger.debug("controls_catalog.json not found at %s", path)
        return []
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        controls = data.get("controls", [])
        logger.info("Loaded controls catalog with %d entries from %s", len(controls), path)
        return controls
    except Exception as exc:
        logger.exception("Failed to load controls catalog %s: %s", path, exc)
        return []


def save_catalog(controls: list[dict[str, Any]], knowledge_base_dir: Path) -> None:
    """Save the controls list back to controls_catalog.json."""
    path = _catalog_path(knowledge_base_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "_schema_version": 1,
        "_description": "רשימת בקרות לניתוח — ניתן לערוך ישירות, לייבא מ-Excel, או לייצא ל-Excel. נשמר ב-Git.",
        "controls": controls,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("Saved controls catalog (%d entries) to %s", len(controls), path)


def apply_catalog_to_definitions(
    controls: list[dict[str, Any]],
    definitions: dict[str, dict[str, str]],
) -> None:
    """Merge user-editable catalog fields into AUDIT_CONTROL_DEFINITIONS in-place.

    Only non-empty values in the catalog overwrite existing definition entries.
    ``in_scope`` and ``analysis_type`` are stored as extra keys so the rest of the
    application can read them via ``AUDIT_CONTROL_DEFINITIONS[control_id]['in_scope']``.
    """
    for entry in controls:
        control_id = str(entry.get("control_id", "")).strip()
        if not control_id:
            continue
        target = definitions.get(control_id)
        if target is None:
            logger.debug("Catalog references unknown control_id %s — skipping", control_id)
            continue

        # Merge text fields
        for field in _MERGEABLE_FIELDS:
            value = str(entry.get(field, "") or "").strip()
            if value:
                target[field] = value

        # Persist analysis_type and in_scope (not already in definitions schema)
        analysis_type = str(entry.get("analysis_type", "") or "").strip()
        if analysis_type:
            target["analysis_type"] = analysis_type

        # in_scope stored as string "true"/"false" for JSON-compat in dict[str, str]
        in_scope = entry.get("in_scope")
        if in_scope is not None:
            target["in_scope"] = "true" if in_scope else "false"


def load_and_apply_catalog(
    knowledge_base_dir: Path,
    definitions: dict[str, dict[str, str]],
) -> list[dict[str, Any]]:
    """Convenience: load catalog and merge into definitions.  Returns the raw list."""
    controls = load_catalog(knowledge_base_dir)
    if controls:
        apply_catalog_to_definitions(controls, definitions)
    return controls


def get_control_in_scope(definitions: dict[str, dict[str, str]], control_id: str) -> bool:
    """Return True when the control is in scope (default: True when not set)."""
    entry = definitions.get(control_id, {})
    return str(entry.get("in_scope", "true")).lower() != "false"


def get_analysis_type(definitions: dict[str, dict[str, str]], control_id: str) -> str:
    """Return the analysis_type string for a control (empty string when not set)."""
    return str(definitions.get(control_id, {}).get("analysis_type", "")).strip()


# ---------------------------------------------------------------------------
# Excel export (Phase 2 scaffold — not yet wired to UI)
# ---------------------------------------------------------------------------

_EXCEL_HEADERS = [
    "מזהה בקרה",
    "קטגוריה",
    "תת-קטגוריה",
    "רמת סיכון",
    "סוג בדיקה",
    "סוג ניתוח",
    "בסקופ (TRUE/FALSE)",
    "תיאור הבקרה",
    "תהליך",
    "תיאור הסיכון",
    "צעדי טסט (override)",
    "הערות",
]

_EXCEL_FIELD_MAP = [
    ("control_id",          "מזהה בקרה"),
    ("category",            "קטגוריה"),
    ("sub_category",        "תת-קטגוריה"),
    ("risk_level",          "רמת סיכון"),
    ("check_type",          "סוג בדיקה"),
    ("analysis_type",       "סוג ניתוח"),
    ("in_scope",            "בסקופ (TRUE/FALSE)"),
    ("description",         "תיאור הבקרה"),
    ("process",             "תהליך"),
    ("risk_description",    "תיאור הסיכון"),
    ("test_steps_override", "צעדי טסט (override)"),
    ("notes",               "הערות"),
]


def export_catalog_to_excel(controls: list[dict[str, Any]], output_path: Path) -> Path:
    """Export the catalog to an Excel file for editing outside the tool.

    Requires openpyxl (already a project dependency).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "רשימת בקרות לניתוח"
    ws.sheet_view.rightToLeft = True

    _HDR_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
    _HDR_FONT = Font(bold=True, color="FFFFFFFF", size=11)

    # Header row
    for col_idx, (_, header) in enumerate(_EXCEL_FIELD_MAP, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for row_idx, entry in enumerate(controls, start=2):
        for col_idx, (field, _) in enumerate(_EXCEL_FIELD_MAP, start=1):
            raw = entry.get(field, "")
            if field == "in_scope":
                value = "TRUE" if raw else "FALSE"
            else:
                value = str(raw or "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    # Column widths
    widths = [20, 22, 30, 12, 28, 20, 18, 55, 35, 55, 30, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Exported catalog to %s", output_path)
    return output_path


def import_catalog_from_excel(xlsx_path: Path) -> list[dict[str, Any]]:
    """Import controls catalog from a previously-exported Excel file.

    Returns a list of control dicts ready for ``save_catalog`` / ``apply_catalog_to_definitions``.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel import") from exc

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Map header labels → column index
    header_row = [str(cell or "").strip() for cell in rows[0]]
    col_map: dict[str, int] = {label: idx for idx, label in enumerate(header_row)}

    controls: list[dict[str, Any]] = []
    for row in rows[1:]:
        entry: dict[str, Any] = {}
        for field, header in _EXCEL_FIELD_MAP:
            idx = col_map.get(header)
            if idx is None:
                continue
            raw = row[idx] if idx < len(row) else None
            value = str(raw or "").strip()
            if field == "in_scope":
                entry[field] = value.upper() not in {"FALSE", "0", "לא", "NO", ""}
            else:
                entry[field] = value
        if entry.get("control_id"):
            controls.append(entry)

    logger.info("Imported %d controls from %s", len(controls), xlsx_path)
    return controls
