import re
from itertools import islice
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook


class ExcelFileReader:
    def read(self, file_path: Path) -> list[dict[str, Any]]:
        return [row for batch in self.read_in_batches(file_path) for row in batch]

    def read_in_batches(self, file_path: Path, chunk_size: int = 20000) -> Iterator[list[dict[str, Any]]]:
        workbook = load_workbook(file_path, data_only=True, read_only=True)
        try:
            best_sheet = None
            best_header_index = 0
            best_score = -1

            for sheet in workbook.worksheets:
                preview_rows = list(islice(sheet.iter_rows(values_only=True), 50))
                if not preview_rows:
                    continue

                header_index, score = self._find_header_row_index(preview_rows)
                if score > best_score:
                    best_sheet = sheet
                    best_header_index = header_index
                    best_score = score

            if best_sheet is None:
                return

            rows_iter = best_sheet.iter_rows(values_only=True)
            headers: list[str] | None = None
            header_positions: list[int] = []
            current_batch: list[dict[str, Any]] = []

            for row_index, row in enumerate(rows_iter):
                if row_index < best_header_index:
                    continue
                if row_index == best_header_index:
                    header_positions = [index for index, value in enumerate(row) if value not in (None, "")]
                    headers = [str(row[index]).strip() for index in header_positions]
                    continue

                if headers is None:
                    continue
                if not any(value not in (None, "") for value in row):
                    continue

                item = {
                    headers[position]: row[index] if index < len(row) else None
                    for position, index in enumerate(header_positions)
                }
                current_batch.append(item)
                if len(current_batch) >= chunk_size:
                    yield current_batch
                    current_batch = []

            if current_batch:
                yield current_batch
        finally:
            workbook.close()

    @staticmethod
    def _find_header_row_index(rows: list[tuple[Any, ...]]) -> tuple[int, int]:
        best_index = 0
        best_score = -1

        for index, row in enumerate(rows[:50]):
            cells = [str(value).strip() for value in row if value not in (None, "")]
            if len(cells) < 2:
                continue

            identifier_like = [cell for cell in cells if re.fullmatch(r"[A-Za-z_][A-Za-z0-9_/@ .-]*", cell)]
            score = len(identifier_like)
            if score >= 2 and score > best_score:
                best_index = index
                best_score = score

        return best_index, best_score
