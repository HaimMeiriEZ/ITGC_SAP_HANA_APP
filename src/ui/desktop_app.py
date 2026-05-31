import fnmatch
import json
import os
import re
import subprocess
import sys
import copy

from typing import Any, Sequence
from datetime import datetime, date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from PySide6.QtCore import QCoreApplication, QDate, QEvent, QObject, QThread, Qt, Signal, Slot
from PySide6.QtGui import QBrush, QColor, QFont, QPixmap
from PySide6.QtWidgets import (
    QAbstractItemView,
    QCheckBox,
    QDoubleSpinBox,
    QStyledItemDelegate,
    QSizePolicy,
    QTabWidget,
    QApplication,
    QFileDialog,
    QComboBox,
    QDateEdit,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QDialog,
    QDialogButtonBox,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QPlainTextEdit,
    QTextEdit,
    QProgressBar,
    QProgressDialog,
    QRadioButton,
    QVBoxLayout,
    QWidget,
    QHeaderView,
)

from src.config import AppConfig, CONTROL_GROUPS, CONTROL_LABELS, SLOT_DEFAULT_CONTROLS
from src.models.validation_result import ValidationIssue
from src.pipeline import process_file
from src.persistence.ui_state_repository import IpeEvidenceRepository, UiStateRepository
from src.persistence.controls_metadata_loader import (
    apply_metadata_to_definitions,
    load_controls_metadata_csv,
)
from src.readers.excel_reader import ExcelFileReader
from src.readers.text_reader import TextFileReader
from src.reporting.excel_report import ExcelReportWriter
from src.reporting.working_paper_report import write_control_working_paper
from src.services.ai_service import OllamaClient, RECOMMENDED_MODELS
from src.services.audit_service import (
    build_audit_detail_row,
    build_audit_detail_values,
    build_audit_summary_values,
    sorted_audit_summary_rows,
    sync_user_review_completion_finding,
    upsert_audit_control_data,
)
from src.services.user_preview_service import (
    build_user_preview_rows,
    filter_user_preview_rows,
    format_user_preview_value_for_display,
    get_user_preview_sort_value,
    parse_user_preview_date,
)
from src.services.user_review_service import (
    build_user_review_incomplete_reason,
    default_reviewer_values,
    has_review_note,
    is_user_review_complete,
    normalize_review_field,
    normalize_reviewer_status,
    reviewer_state_key,
)
from src.validators.spec_rules import (
    AUDIT_CONTROL_DEFINITIONS,
    AUTH_MGMT_PERMISSION_CRITERIA,
    DATA_MGMT_PERMISSION_CRITERIA,
    DEBUG_PERMISSION_CRITERIA,
    JOB_MGMT_PERMISSION_CRITERIA,
    RSCDOK99_PERMISSION_CRITERIA,
    SAP_APP_RSPARAM_RULES,
    TRANSPORT_PERMISSION_CRITERIA,
    USER_MGMT_PERMISSION_CRITERIA,
    get_audit_control_definition,
    get_column_aliases,
    get_profile_audit_controls,
)


def get_qt_app() -> QApplication:
    if "unittest" in sys.modules and "QT_QPA_PLATFORM" not in os.environ:
        os.environ["QT_QPA_PLATFORM"] = "offscreen"

    app = QApplication.instance()
    if not isinstance(app, QApplication):
        app = QApplication(sys.argv)
        app.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        app.setFont(QFont("Segoe UI", 9))
    return app


class _RightAlignDelegate(QStyledItemDelegate):
    """Forces physical-right alignment on all cells, regardless of RTL layout direction."""

    def initStyleOption(self, option: Any, index: Any) -> None:  # type: ignore[override]
        super().initStyleOption(option, index)
        option.displayAlignment = (
            Qt.AlignmentFlag.AlignAbsolute
            | Qt.AlignmentFlag.AlignRight
            | Qt.AlignmentFlag.AlignVCenter
        )


class _IpeMappingCellDelegate(QStyledItemDelegate):
    """IPE mapping slot cells: filled accent color when checked, plain when unchecked. No checkbox drawn."""

    _CHECKED_COLOR = QColor("#6d002f")

    def paint(self, painter: Any, option: Any, index: Any) -> None:
        bg = index.data(Qt.ItemDataRole.BackgroundRole)
        painter.save()
        if isinstance(bg, QBrush) and bg.style() != Qt.BrushStyle.NoBrush:
            painter.fillRect(option.rect, bg)
        else:
            painter.fillRect(option.rect, option.palette.base().color())
        painter.restore()

    def editorEvent(self, event: Any, model: Any, option: Any, index: Any) -> bool:  # type: ignore[override]
        if (
            event.type() == QEvent.Type.MouseButtonRelease
            and event.button() == Qt.MouseButton.LeftButton
        ):
            bg = index.data(Qt.ItemDataRole.BackgroundRole)
            is_checked = isinstance(bg, QBrush) and bg.style() != Qt.BrushStyle.NoBrush
            new_state = Qt.CheckState.Unchecked if is_checked else Qt.CheckState.Checked
            model.setData(index, new_state, Qt.ItemDataRole.CheckStateRole)
            return True
        return bool(super().editorEvent(event, model, option, index))


class SortableTableWidgetItem(QTableWidgetItem):
    SORT_ROLE = Qt.ItemDataRole.UserRole + 2

    def __lt__(self, other: object) -> bool:
        if isinstance(other, QTableWidgetItem):
            self_sort_value = self.data(self.SORT_ROLE)
            other_sort_value = other.data(self.SORT_ROLE)
            if self_sort_value is not None or other_sort_value is not None:
                left = "" if self_sort_value is None else str(self_sort_value)
                right = "" if other_sort_value is None else str(other_sort_value)
                return left < right
            return super().__lt__(other)
        return NotImplemented


class SlotValidationWorker(QObject):
    succeeded = Signal(str, list, object)
    failed = Signal(str, list, str)
    finished = Signal()

    def __init__(
        self,
        slot_key: str,
        file_paths: list[str],
        input_files_dict: dict[str, list[str | Path]],
        required_columns: list[str],
        output_dir: Path,
        authorized_users: list[str],
        strong_profiles: list[str] | None = None,
    ) -> None:
        super().__init__()
        self.slot_key = slot_key
        self.file_paths = file_paths
        self.input_files_dict = input_files_dict
        self.required_columns = required_columns
        self.output_dir = output_dir
        self.authorized_users = authorized_users
        self.strong_profiles = strong_profiles

    @Slot()
    def run(self) -> None:
        try:
            result = process_file(
                input_files=self.input_files_dict,
                required_columns=self.required_columns,
                output_dir=self.output_dir,
                source_name_override=self.slot_key,
                authorized_users=self.authorized_users,
                strong_profiles=self.strong_profiles,
            )
            self.succeeded.emit(self.slot_key, list(self.file_paths), result)
        except Exception as error:
            self.failed.emit(self.slot_key, list(self.file_paths), str(error))
        finally:
            self.finished.emit()


class BatchNarrationWorker(QObject):
    """Background worker that generates AI narrations for a list of user findings.

    Emits ``progress(done, total, bname)`` per processed row and ``finished(processed, skipped, failed)``
    at the end.  Designed to be moved to a ``QThread`` and connected via signals.
    """

    progress = Signal(int, int, str)
    finished = Signal(int, int, int)

    def __init__(
        self,
        items: list[dict[str, Any]],
        ai_settings: dict[str, Any],
        work_environment: str,
    ) -> None:
        super().__init__()
        self._items = items
        self._ai_settings = ai_settings
        self._work_environment = work_environment
        self._cancel_requested = False

    @Slot()
    def request_cancel(self) -> None:
        self._cancel_requested = True

    @Slot()
    def run(self) -> None:
        processed = 0
        skipped = 0
        failed = 0
        total = len(self._items)
        try:
            from src.services.findings_narrator import narrate_user_finding
            client = OllamaClient(self._ai_settings)
        except Exception:
            self.finished.emit(0, 0, total)
            return

        for idx, item in enumerate(self._items, start=1):
            if self._cancel_requested:
                break
            row = item.get("row", {})
            raw_findings = str(item.get("raw_findings", "") or "")
            bname = str(row.get("BNAME", ""))
            self.progress.emit(idx, total, bname)
            if not raw_findings.strip():
                skipped += 1
                continue
            try:
                result = narrate_user_finding(row, raw_findings, client, self._work_environment)
                if result and result != raw_findings:
                    processed += 1
                else:
                    skipped += 1
            except Exception:
                failed += 1
        self.finished.emit(processed, skipped, failed)


class _ImportReviewConfirmDialog(QDialog):
    """Confirmation dialog displayed before applying an imported review Excel file.

    Shows import statistics and any integrity warnings (missing records with
    reviewed data, notes that will be erased).  The user chooses how to proceed.
    """

    MODE_ALL = "all"              # apply every value from the file, including empty notes
    MODE_PRESERVE_NOTES = "preserve_notes"  # keep existing notes when the imported value is empty

    def __init__(
        self,
        parent: QWidget | None,
        total_in_file: int,
        missing_with_data: list[str],   # BNAME values absent from the file but with reviewed data
        notes_cleared: list[tuple[str, list[str]]],  # (BNAME, [formal field names]) for notes that will be erased
    ) -> None:
        super().__init__(
            parent,
            Qt.WindowType.Dialog
            | Qt.WindowType.CustomizeWindowHint
            | Qt.WindowType.WindowTitleHint,
        )
        self.setWindowTitle("אישור ייבוא סקירת משתמשים")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.setWindowModality(Qt.WindowModality.WindowModal)
        self.setMinimumWidth(500)

        self._mode = self.MODE_ALL
        has_issues = bool(missing_with_data or notes_cleared)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 14, 16, 14)

        # ── Summary ────────────────────────────────────────────────────────
        summary_label = QLabel(f"נמצאו <b>{total_in_file}</b> רשומות בקובץ לעדכון.")
        summary_label.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        summary_label.setAlignment(Qt.AlignmentFlag.AlignRight)
        layout.addWidget(summary_label)

        # ── Warning section (shown only when issues exist) ─────────────────
        if has_issues:
            warn_widget = QWidget()
            warn_widget.setStyleSheet(
                "background-color:#FFF3CD; border:1px solid #FFC107; border-radius:4px;"
            )
            warn_layout = QVBoxLayout(warn_widget)
            warn_layout.setContentsMargins(10, 8, 10, 8)
            warn_layout.setSpacing(6)

            if missing_with_data:
                max_shown = 8
                display_names = missing_with_data[:max_shown]
                extra = len(missing_with_data) - max_shown
                miss_header = QLabel(
                    f"<b>⚠ {len(missing_with_data)} רשומות עם נתונים קיימים אינן מופיעות בקובץ (לא ישתנו):</b>"
                )
                miss_header.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
                miss_header.setAlignment(Qt.AlignmentFlag.AlignRight)
                miss_header.setWordWrap(True)
                miss_header.setStyleSheet("color:#856404;")
                names_text = ", ".join(display_names)
                if extra > 0:
                    names_text += f" ועוד {extra} נוספים..."
                miss_names = QLabel(names_text)
                miss_names.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
                miss_names.setAlignment(Qt.AlignmentFlag.AlignRight)
                miss_names.setWordWrap(True)
                miss_names.setStyleSheet("color:#856404;")
                warn_layout.addWidget(miss_header)
                warn_layout.addWidget(miss_names)

            if notes_cleared:
                cleared_lbl = QLabel(
                    f"<b>⚠ {len(notes_cleared)} רשומות שבהן הערות שאוכלסו יימחקו אם תבחר 'כל השינויים'.</b>"
                )
                cleared_lbl.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
                cleared_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)
                cleared_lbl.setWordWrap(True)
                cleared_lbl.setStyleSheet("color:#856404;")
                warn_layout.addWidget(cleared_lbl)

            layout.addWidget(warn_widget)

            # ── Radio options ────────────────────────────────────────────────
            options_lbl = QLabel("<b>כיצד ברצונך להמשיך?</b>")
            options_lbl.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
            options_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)
            layout.addWidget(options_lbl)

            self._radio_all = QRadioButton("המשך עם כל השינויים (הערות ריקות בקובץ יימחקו)")
            self._radio_all.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
            self._radio_all.setChecked(True)
            self._radio_preserve = QRadioButton("המשך ושמור הערות קיימות (הערות לא יימחקו אם הקובץ ריק בשדה זה)")
            self._radio_preserve.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
            layout.addWidget(self._radio_all)
            layout.addWidget(self._radio_preserve)
        else:
            self._radio_all = None
            self._radio_preserve = None

        # ── Buttons ──────────────────────────────────────────────────────────
        btn_box = QDialogButtonBox()
        btn_box.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        confirm_btn = btn_box.addButton("אשר ייבוא", QDialogButtonBox.ButtonRole.AcceptRole)
        confirm_btn.setDefault(True)
        btn_box.addButton("ביטול", QDialogButtonBox.ButtonRole.RejectRole)
        btn_box.accepted.connect(self._on_confirm)
        btn_box.rejected.connect(self.reject)
        layout.addWidget(btn_box)

        self.adjustSize()

    def _on_confirm(self) -> None:
        if self._radio_preserve is not None and self._radio_preserve.isChecked():
            self._mode = self.MODE_PRESERVE_NOTES
        else:
            self._mode = self.MODE_ALL
        self.accept()

    @property
    def selected_mode(self) -> str:
        return self._mode


class _LoadingProgressDialog(QDialog):
    """Non-closable modal progress dialog shown while a slot validation worker is running."""

    def __init__(self, parent: QWidget | None, slot_label: str) -> None:
        super().__init__(
            parent,
            Qt.WindowType.Dialog
            | Qt.WindowType.CustomizeWindowHint
            | Qt.WindowType.WindowTitleHint,
        )
        self.setWindowTitle("טעינת נתונים")
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.setWindowModality(Qt.WindowModality.ApplicationModal)
        self.resize(360, 130)
        self.setMinimumWidth(300)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(14)

        label_text = f"טוען נתוני משבצת: {slot_label}\nאנא המתן לסיום הטעינה..."
        msg_label = QLabel(label_text)
        msg_label.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        msg_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        msg_label.setWordWrap(True)
        msg_label.setStyleSheet("font-size: 13px; color: #16325c;")
        layout.addWidget(msg_label)

        progress_bar = QProgressBar()
        progress_bar.setRange(0, 0)  # indeterminate / animated
        progress_bar.setTextVisible(False)
        progress_bar.setMinimumHeight(16)
        layout.addWidget(progress_bar)

    def closeEvent(self, event: Any) -> None:  # type: ignore[override]
        """Prevent the user from manually closing the dialog."""
        event.ignore()


class _FilterableHeaderView(QHeaderView):
    """QHeaderView subclass that renders a per-section filter (▼) button and emits filterRequested(int)."""

    filterRequested = Signal(int)  # emits logicalIndex

    _ICON_WIDTH = 18

    def __init__(self, orientation: Qt.Orientation, parent: QWidget | None = None) -> None:
        super().__init__(orientation, parent)
        self._active_filter_sections: set[int] = set()

    def set_active_filter_sections(self, sections: set[int]) -> None:
        self._active_filter_sections = set(sections)
        self.viewport().update()

    def paintSection(self, painter: Any, rect: Any, logicalIndex: int) -> None:  # type: ignore[override]
        painter.save()
        super().paintSection(painter, rect, logicalIndex)
        painter.restore()

        is_active = logicalIndex in self._active_filter_sections
        icon_rect_type = rect.__class__
        from PySide6.QtCore import QRect  # noqa: PLC0415
        icon_rect = QRect(rect.left(), rect.top() + 2, self._ICON_WIDTH, rect.height() - 4)

        painter.save()
        if is_active:
            from PySide6.QtGui import QBrush as _Brush  # noqa: PLC0415
            painter.setBrush(_Brush(QColor("#c8dff8")))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(icon_rect, 3, 3)
            painter.setPen(QColor("#1a5da8"))
        else:
            painter.setPen(QColor("#8899aa"))

        f = painter.font()
        f.setPointSize(7)
        painter.setFont(f)
        painter.drawText(icon_rect, Qt.AlignmentFlag.AlignCenter, "▼")
        painter.restore()

    def mousePressEvent(self, event: Any) -> None:  # type: ignore[override]
        if event.button() == Qt.MouseButton.LeftButton:
            pos = event.position().toPoint() if hasattr(event, "position") else event.pos()
            logical_index = self.logicalIndexAt(pos)
            if logical_index >= 0:
                section_left = self.sectionViewportPosition(logical_index)
                if section_left <= pos.x() < section_left + self._ICON_WIDTH:
                    self.filterRequested.emit(logical_index)
                    return
        super().mousePressEvent(event)


class ValidationDesktopApp(QMainWindow):
    USER_PREVIEW_SLOTS = {"USR02", "ADR6_USR21"}

    MULTI_FILE_SLOTS = {
        "USR02",
        "ADR6_USR21",
        "AGR_USERS",
        "AGR_1251",
        "AGR_1252",
        "AGR_DEFINE",
        "UST04",
        "USH04",
        "E070",
        "STMS",
    }

    USER_PREVIEW_COLUMN_DEFINITIONS = [
        {"field": "MANDT", "formal": "CLIENT", "technical": "MANDT", "source": "USR02", "default": True, "width": 90},
        {"field": "WORK_ENVIRONMENT", "formal": "סביבת עבודה", "technical": "WORK_ENVIRONMENT", "source": "הגדרות מערכת", "default": True, "width": 210},
        {"field": "BNAME", "formal": "משתמש", "technical": "BNAME", "source": "USR02", "default": True, "width": 130},
        {"field": "NAME_FIRST", "formal": "שם פרטי", "technical": "NAME_FIRST", "source": "USER_ADDR", "default": True, "width": 120},
        {"field": "NAME_LAST", "formal": "שם משפחה", "technical": "NAME_LAST", "source": "USER_ADDR", "default": True, "width": 120},
        {"field": "NAME_TEXTC", "formal": "שם מלא", "technical": "NAME_TEXTC", "source": "USER_ADDR", "default": True, "width": 180},
        {"field": "COMPANY", "formal": "חברה", "technical": "COMPANY", "source": "USER_ADDR", "default": True, "width": 150},
        {"field": "SMTP_ADDR", "formal": 'דוא"ל', "technical": "SMTP_ADDR", "source": "ADR6", "default": True, "width": 200},
        {"field": "STATUS", "formal": "סטטוס משתמש", "technical": "STATUS", "source": "USR02", "default": True, "width": 120},
        {"field": "ADDRNUMBER", "formal": "מספר כתובת", "technical": "ADDRNUMBER", "source": "USER_ADDR / ADR6", "default": True, "width": 120},
        {"field": "PERSNUMBER", "formal": "מספר פרסונה", "technical": "PERSNUMBER", "source": "USER_ADDR / ADR6", "default": True, "width": 120},
        {"field": "TRDAT", "formal": "תאריך כניסה אחרון", "technical": "TRDAT", "source": "USR02", "default": True, "width": 140},
        {"field": "LTIME", "formal": "שעת כניסה אחרונה", "technical": "LTIME", "source": "USR02", "default": True, "width": 140},
        {"field": "PWDINITIAL", "formal": "סיסמה ראשונית", "technical": "PWDINITIAL", "source": "USR02", "default": True, "width": 120},
        {"field": "PWDCHGDATE", "formal": "תאריך שינוי סיסמה", "technical": "PWDCHGDATE", "source": "USR02", "default": True, "width": 145},
        {"field": "PWDSETDATE", "formal": "תאריך הגדרת סיסמה", "technical": "PWDSETDATE", "source": "USR02", "default": True, "width": 145},
        {"field": "DEPARTMENT", "formal": "מחלקה", "technical": "DEPARTMENT", "source": "USER_ADDR", "default": True, "width": 150},
        {"field": "GLTGV", "formal": "תקף מתאריך", "technical": "GLTGV", "source": "USR02", "default": True, "width": 120},
        {"field": "GLTGB", "formal": "תקף עד תאריך", "technical": "GLTGB", "source": "USR02", "default": True, "width": 120},
        {"field": "USTYP", "formal": "סוג משתמש", "technical": "USTYP", "source": "USR02", "default": True, "width": 110},
        {"field": "LOCNT", "formal": "מספר ניסיונות כניסה כושלים", "technical": "LOCNT", "source": "USR02", "default": True, "width": 125},
        {"field": "OCOD1", "formal": "סיסמה", "technical": "OCOD1", "source": "USR02", "default": True, "width": 130},
        {"field": "PASSCODE", "formal": "ערך Hash של סיסמה", "technical": "PASSCODE", "source": "USR02", "default": True, "width": 220},
        {"field": "PWDSALTEDHASH", "formal": "ערך Hash מוצפן של סיסמה", "technical": "PWDSALTEDHASH", "source": "USR02", "default": True, "width": 220},
        {"field": "SECURITY_POLICY", "formal": "מדיניות אבטחה", "technical": "SECURITY_POLICY", "source": "USR02", "default": True, "width": 160},
        {"field": "REVIEW_STATUS", "formal": "בוצעה סקירה", "technical": "REVIEW_STATUS", "source": "סוקר", "default": True, "width": 165},
        {"field": "FINDINGS_DESCRIPTION", "formal": "תיאור ממצאים", "technical": "FINDINGS_DESCRIPTION", "source": "מערכת", "default": True, "width": 280},
        {"field": "FINDINGS_DESCRIPTION_AI", "formal": "נרטיב AI לממצאים", "technical": "FINDINGS_DESCRIPTION_AI", "source": "AI", "default": False, "width": 340},
        {"field": "ANOMALY_SCORE", "formal": "ציון אנומליה", "technical": "ANOMALY_SCORE", "source": "ניתוח סטטיסטי", "default": False, "width": 110},
        {"field": "ANOMALY_CODES", "formal": "קודי אנומליה", "technical": "ANOMALY_CODES", "source": "ניתוח סטטיסטי", "default": False, "width": 200},
        {"field": "TECH_REVIEW_NOTES", "formal": "הערות סוקר גורם טכני", "technical": "TECH_REVIEW_NOTES", "source": "סוקר טכני", "default": True, "width": 240},
        {"field": "BUS_REVIEW_NOTES", "formal": "הערות סוקר גורם מהכספים", "technical": "BUS_REVIEW_NOTES", "source": "סוקר מהכספים", "default": True, "width": 240},
        {"field": "LAST_IMPORT_DATE", "formal": "עודכן לאחרונה", "technical": "LAST_IMPORT_DATE", "source": "ייבוא סקירה", "default": True, "width": 145},
        {"field": "UFLAG", "formal": "קוד נעילה", "technical": "UFLAG", "source": "USR02", "default": False, "width": 100},
    ]
    DEFAULT_USER_PREVIEW_COLUMNS = [
        column["field"]
        for column in USER_PREVIEW_COLUMN_DEFINITIONS
        if bool(column.get("default"))
    ]
    CURRENT_USER_PREVIEW_SETTINGS_VERSION = 10
    USER_PREVIEW_SETTINGS_MIGRATIONS = {
        2: ["PWDINITIAL", "PWDCHGDATE", "PWDSETDATE"],
        3: ["DEPARTMENT", "GLTGV", "GLTGB", "USTYP", "LOCNT", "OCOD1", "PASSCODE", "PWDSALTEDHASH", "SECURITY_POLICY"],
        4: ["REVIEW_STATUS", "REVIEW_NOTES"],
        5: ["FINDINGS_DESCRIPTION"],
        6: ["TECH_REVIEW_NOTES", "BUS_REVIEW_NOTES"],
        7: ["WORK_ENVIRONMENT"],
        8: ["LAST_IMPORT_DATE"],
        9: ["FINDINGS_DESCRIPTION_AI"],
        10: ["ANOMALY_SCORE", "ANOMALY_CODES"],
    }
    USER_PREVIEW_FILTER_OPTIONS = [
        ("all", "כלל האוכלוסייה"),
        ("active", "פעילים בתקופה הנבדקת"),
        ("inactive", "לא פעילים בתקופה הנבדקת"),
    ]
    REVIEW_STATUS_OPTIONS = ["טרם נבדק", "נבדק - תקין", "נבדק - לא תקין"]
    DEFAULT_REVIEW_STATUS = "טרם נבדק"
    REVIEWED_STATUSES = {"נבדק - תקין", "נבדק - לא תקין"}
    REVIEW_COMPLETION_CONTROL_ID = "MA1-1&MA7-17_AYALON_2"
    USER_TYPE_RULES = {
        "Dialog": ["A"],
        "System": ["B"],
        "Communication": ["C"],
        "Service": ["S"],
        "Reference": ["L"],
    }
    USER_PREVIEW_DATE_FIELDS = {"TRDAT", "PWDCHGDATE", "PWDSETDATE", "GLTGV", "GLTGB"}
    EXPORT_REVIEW_FIELDS = [
        "MANDT", "WORK_ENVIRONMENT", "BNAME", "NAME_TEXTC", "SMTP_ADDR", "STATUS", "USTYP",
        "GLTGV", "GLTGB", "TRDAT", "PWDSETDATE", "PWDCHGDATE",
        "FINDINGS_DESCRIPTION", "REVIEW_STATUS", "TECH_REVIEW_NOTES", "BUS_REVIEW_NOTES",
    ]
    WORK_ENVIRONMENT_OPTIONS = [
        ("FPD", "FPD - DEV - סביבת מפתחים"),
        ("PFT", "PFT - PRE PROD - סביבת קדם ייצור"),
        ("FPP", "FPP - PROD - סביבת ייצור"),
        ("FPQ", "FPQ - QA - סביבת בדיקות"),
        ("FDR", "FDR - DR - סביבת התאוששות מאסון"),
    ]

    SLOT_DEFINITIONS = {
        "USR02": {
            "domain": "MA - ניהול גישה",
            "sub_category": "1.1 - Joiners / Movers / Leavers",
            "description": "משתמשים - מקור חובה לבדיקות גישה, סטטוס ותאריכי התחברות.",
            "expected_file": "usr02_100.txt",
            "required": True,
        },
        "ADR6_USR21": {
            "label": "ADR6 / USER_ADDR",
            "domain": "MA - ניהול גישה",
            "sub_category": "1.1 - Joiners / Movers / Leavers",
            "description": "ניתן להזין קובצי ADR6 או USER_ADDR או את שניהם יחד לצורך העשרת נתוני המשתמשים מתוך USR02.",
            "expected_file": "adr6.txt או user_addr.txt",
            "required": False,
        },
        "AGR_USERS": {
            "domain": "MA - ניהול גישה",
            "sub_category": "1.2 - סקר הרשאות תקופתי",
            "description": "רולים-משתמשים - מיפוי המשתמשים לרולים במערכת.",
            "expected_file": "agr_users_100.txt",
            "required": True,
        },
        "AGR_1251": {
            "domain": "MA - ניהול גישה",
            "sub_category": "1.2 - סקר הרשאות תקופתי",
            "description": "רולים-אובייקטי הרשאה - זיהוי אובייקטי הרשאות רגישים.",
            "expected_file": "agr_1251_100.txt",
            "required": True,
        },
        "UST04": {
            "domain": "MA - ניהול גישה",
            "sub_category": "1.2 - סקר הרשאות תקופתי",
            "description": "פרופילים-משתמשים - שיוך פרופילים ישיר למשתמשים.",
            "expected_file": "ust04.txt",
            "required": True,
        },
        "USH04": {
            "domain": "MA - ניהול גישה",
            "sub_category": "1.2 - סקר הרשאות תקופתי",
            "description": "היסטוריית שיוך פרופילים למשתמשים - זיהוי שינויים בפרופילים ופרופילים חזקים.",
            "expected_file": "ush04.txt",
            "required": False,
        },
        "AGR_1252": {
            "domain": "MA - ניהול גישה",
            "sub_category": "1.2 - סקר הרשאות תקופתי",
            "description": "רולים-טרנזקציות - זיהוי גישות עסקיות וטרנזקציות.",
            "expected_file": "agr_1252_100.txt",
            "required": False,
        },
        "AGR_DEFINE": {
            "domain": "MA - ניהול גישה",
            "sub_category": "1.2 - סקר הרשאות תקופתי",
            "description": "רולים מורחב - מידע כללי על הגדרת הרול.",
            "expected_file": "agr_define.txt",
            "required": False,
        },
        "RSPARAM": {
            "domain": "MA - ניהול גישה",
            "sub_category": "1.3 - משתמשי-על ומדיניות סיסמאות",
            "description": "פרמטרים סיסטמאיים - פרמטרי אבטחה והקשחת מערכת.",
            "expected_file": "rsparam.xlsx",
            "required": True,
        },
        "TPFET": {
            "label": "TPFET / RZ10",
            "domain": "MA - ניהול גישה",
            "sub_category": "1.3 - משתמשי-על ומדיניות סיסמאות",
            "description": "פרמטרים סיסטמאיים נוספים, כולל פרופילי login כגון RZ10.",
            "expected_file": "rz10.txt",
            "required": False,
        },
        "E070": {
            "domain": "MC - ניהול שינויים",
            "sub_category": "2.1 - תיעוד ובקשות שינוי",
            "description": "רשימת שינויים - נתוני transport requests ושינויים בסביבה.",
            "expected_file": "e070_100.txt",
            "required": True,
        },
        "T000": {
            "domain": "MC - ניהול שינויים",
            "sub_category": "2.3 - הפרדת סביבות DEV/QA/PRD",
            "description": "לוג פעילות שינוי SCC4 - בקרות שינוי ברמת client.",
            "expected_file": "t000.txt",
            "required": False,
        },
        "STMS": {
            "domain": "MC - ניהול שינויים",
            "sub_category": "2.4 - Import לסביבת ייצור",
            "description": "רשימת שינויים שהועברה דרך SCC4 או STMS.",
            "expected_file": "stms.txt",
            "required": False,
        },
    }

    DOMAIN_DEFINITIONS: dict[str, Any] = {
        "MA - ניהול גישה": {
            "in_development": False,
            "sub_categories": [
                {
                    "key": "1.1 - Joiners / Movers / Leavers",
                    "description": "תהליכי הצטרפות, שינוי תפקיד ועזיבה — ווידוא שקיים אישור מנהל ושהגישה הוסרה בזמן.",
                },
                {
                    "key": "1.2 - סקר הרשאות תקופתי",
                    "description": "ביצוע סקר הרשאות תקופתי על ידי גורמים עסקיים (Business Owners) לווידוא נחיצות ההרשאות.",
                },
                {
                    "key": "1.3 - משתמשי-על ומדיניות סיסמאות",
                    "description": "ניהול משתמשי-על (BASIS, SAP*, DDIC) ובדיקת פרמטרי מדיניות סיסמאות במערכת.",
                },
            ],
        },
        "MC - ניהול שינויים": {
            "in_development": False,
            "sub_categories": [
                {
                    "key": "2.1 - תיעוד ובקשות שינוי",
                    "description": "תיעוד בקשת השינוי ואישור עסקי לפני הפיתוח.",
                },
                {
                    "key": "2.3 - הפרדת סביבות DEV/QA/PRD",
                    "description": "הפרדה בין סביבת הפיתוח (DEV), הבחינה (QA) והייצור (PRD).",
                },
                {
                    "key": "2.4 - Import לסביבת ייצור",
                    "description": "בחינה של תהליך ה-Import לסביבת הייצור ומי מורשה לבצע אותו (בדרך כלל צוות ה-Basis).",
                },
            ],
        },
        "MO - ניהול תפעולי": {
            "in_development": True,
            "sub_categories": [
                {
                    "key": "3.1 - ניטור תהליכי רקע (Batch Jobs)",
                    "description": "בודקים מה קורה אם ג'וב קריטי נכשל והאם יש בקרה על שינוי לוחות זמנים של ג'ובים.",
                },
                {
                    "key": "3.2 - גיבויים ושחזור (Backups)",
                    "description": "ווידוא שהגיבויים מבוצעים כסדרם ושיש שחזור תקופתי מוצלח (Restoration Test).",
                },
                {
                    "key": "3.3 - ניהול תקלות (Incidents)",
                    "description": "תהליך הטיפול בתקלות מערכת ותיעודן.",
                },
            ],
        },
    }

    SETTINGS_SECTION_DEPENDENCIES = {
        "user_review_period": set(),
        "authorized_stms_users": set(),
        "super_users": set(),
        "generic_users": set(),
        "critical_roles": set(),
        "critical_privileges": set(),
        "password_policy_defaults": set(),
        "file_mappings": set(),
        "inactive_days_threshold": set(),
    }

    def __init__(self, base_dir: Path | None = None) -> None:
        super().__init__()
        self.config = AppConfig.default(base_dir or Path.cwd())
        self.ui_state_repository = UiStateRepository(self.config.output_dir, self.config.input_dir)
        self.ipe_repository = IpeEvidenceRepository(
            self.config.output_dir, base_dir or Path.cwd()
        )
        self.ipe_evidence_data: dict[str, list[dict[str, Any]]] = {}
        self.config.input_dir.mkdir(parents=True, exist_ok=True)
        self.config.output_dir.mkdir(parents=True, exist_ok=True)

        # Merge CSV-supplied metadata into AUDIT_CONTROL_DEFINITIONS at startup
        try:
            csv_metadata = load_controls_metadata_csv(self.config.output_dir)
            if csv_metadata:
                apply_metadata_to_definitions(csv_metadata, AUDIT_CONTROL_DEFINITIONS)
        except Exception:  # pragma: no cover - never block startup
            pass
        self.report_path: Path | None = None
        self.log_export_path: Path | None = None
        self.slot_widgets: dict[str, dict[str, Any]] = {}
        self.category_run_buttons: dict[str, QPushButton] = {}
        self.category_sections: dict[str, QGroupBox] = {}
        self.selected_slot_key: str | None = None
        self.load_history: list[str] = []
        self.summary_labels: dict[str, QLabel] = {}
        self.run_log_records: list[dict[str, Any]] = []
        self.audit_summary_records: dict[str, dict[str, Any]] = {}
        self.audit_details_by_control: dict[str, list[dict[str, Any]]] = {}
        # Raw slot rows captured per control_id for working-paper export
        self.control_to_slot_rows: dict[str, list[dict[str, Any]]] = {}
        self.control_to_slot_key: dict[str, str] = {}
        # Population count per slot_key (rows ingested) for IPE working-paper sheet
        self.slot_to_row_count: dict[str, int] = {}
        # Per-slot, per-file row counts. Keyed slot_key -> {file_stem_upper: count}
        self._slot_file_row_counts: dict[str, dict[str, int]] = {}
        self.permissions_summary_records: dict[str, dict[str, Any]] = {}
        self.permissions_users_by_control: dict[str, list[dict[str, Any]]] = {}
        # Per-source-profile staging for MA3-3 strong profiles control.
        # Structure: {detected_profile (UST04/USH04): {client_name: {user_name: set(profile_names)}}}
        self._strong_profile_data: dict[str, dict[str, dict[str, set[str]]]] = {}
        self.agr_1251_cached_rows: list[dict[str, Any]] = []
        self.agr_users_cached_rows: list[dict[str, Any]] = []
        self.agr_users_population_by_mandt: dict[str, int] = {}
        self.user_mgmt_summary_records: dict[str, dict[str, Any]] = {}
        self.user_mgmt_users_by_control: dict[str, list[dict[str, Any]]] = {}
        self.auth_mgmt_summary_records: dict[str, dict[str, Any]] = {}
        self.auth_mgmt_users_by_control: dict[str, list[dict[str, Any]]] = {}
        self.rscdok99_summary_records: dict[str, dict[str, Any]] = {}
        self.rscdok99_users_by_control: dict[str, list[dict[str, Any]]] = {}
        self.data_mgmt_summary_records: dict[str, dict[str, Any]] = {}
        self.data_mgmt_users_by_control: dict[str, list[dict[str, Any]]] = {}
        self.transport_summary_records: dict[str, dict[str, Any]] = {}
        self.transport_users_by_control: dict[str, list[dict[str, Any]]] = {}
        self.debug_summary_records: dict[str, dict[str, Any]] = {}
        self.debug_users_by_control: dict[str, list[dict[str, Any]]] = {}
        self.job_mgmt_summary_records: dict[str, dict[str, Any]] = {}
        self.job_mgmt_users_by_control: dict[str, list[dict[str, Any]]] = {}
        self.audit_findings_export_path: Path | None = None
        self.validation_thread: QThread | None = None
        self.validation_worker: SlotValidationWorker | None = None
        self.batch_narration_thread: QThread | None = None
        self.batch_narration_worker: BatchNarrationWorker | None = None
        self.batch_narration_progress_dialog: QProgressDialog | None = None
        self._allow_user_preview_persistence = base_dir is not None or "unittest" not in sys.modules
        self.last_file_dialog_directory = self._load_last_file_dialog_directory()
        self._refreshing_user_preview = False
        self.user_preview_export_path: Path | None = None
        self.user_reviewer_state = self._load_user_reviewer_state()
        self.user_preview_visible_columns = self._load_user_preview_column_selection()
        self.user_preview_column_filters: dict[str, set[str]] = self._load_user_preview_column_filters()
        self._loading_dialog: QDialog | None = None
        self.system_settings_widgets: dict[str, Any] = {}
        self.system_settings_sections: dict[str, QGroupBox] = {}
        self.system_settings_unavailable_labels: dict[str, QLabel] = {}
        self.system_settings_file_mapping_order: list[str] = []

        self._slot_control_mapping: dict[str, list[str]] = dict(SLOT_DEFAULT_CONTROLS)
        self._configure_window()
        self._build_ui()
        self._populate_all_slot_thumbnails()
        self._load_system_settings_into_form(
            self._current_system_settings(),
            load_review_period=self._system_settings_path().exists(),
        )
        _saved_mapping = self._current_system_settings().get("slot_ipe_control_mapping")
        self._slot_control_mapping = (
            dict(_saved_mapping) if isinstance(_saved_mapping, dict) else dict(SLOT_DEFAULT_CONTROLS)
        )
        self._apply_system_settings_availability()

    @staticmethod
    def format_rtl_text(text: object) -> str:
        raw_text = "" if text is None else str(text)
        return re.sub(r"[\u2066\u2067\u2068\u2069\u200e\u200f]", "", raw_text)

    @staticmethod
    def format_ui_rtl_text(text: object) -> str:
        normalized_text = ValidationDesktopApp.format_rtl_text(text).strip()
        if normalized_text and re.search(r"[\u0590-\u05FF]", normalized_text):
            return f"\u202B{normalized_text}\u202C"
        return normalized_text

    def _configure_window(self) -> None:
        self.setWindowTitle(self.format_rtl_text("כלי להערכת בקרות ITGC בסביבת SAP HANA APP"))
        self.setMinimumSize(1180, 760)
        self.resize(1280, 860)
        self.setLayoutDirection(Qt.LayoutDirection.RightToLeft)

    def _log_to_console(self, message: str, level: str = "info") -> None:
        """Append a timestamped message to the activity console and yield the event loop."""
        _colors: dict[str, str] = {
            "info":  "#90c090",
            "warn":  "#e0c860",
            "error": "#e07070",
            "title": "#60b0e0",
            "dim":   "#556677",
        }
        color = _colors.get(level, "#cccccc")
        ts = datetime.now().strftime("%H:%M:%S")
        escaped = (
            message
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
        html = (
            f'<p dir="rtl" style="margin:0;">'
            f'<span style="color:#445566;">[{ts}]</span>'
            f'&nbsp;<span style="color:{color};">{escaped}</span>'
            f'</p>'
        )
        if hasattr(self, "console_output"):
            self.console_output.append(html)
            self.console_output.ensureCursorVisible()
        QApplication.processEvents()

    def _enter_intake_stage_b(self) -> None:
        """Show the run-log table (Stage B of progressive intake UI). Console is always visible."""
        self.run_log_group.show()

    def _build_ui(self) -> None:
        central_widget = QWidget()
        central_widget.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.setCentralWidget(central_widget)

        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(16, 16, 16, 16)
        main_layout.setSpacing(12)

        _title_container = QWidget()
        _title_container.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        _title_row = QHBoxLayout(_title_container)
        _title_row.setContentsMargins(0, 0, 0, 0)
        _title_row.setSpacing(12)

        # ── Company logo (left-aligned) ────────────────────────────────────
        _logo_path = Path(__file__).parent / "assets" / "ayalon_logo.png"
        if _logo_path.exists():
            _logo_pixmap = QPixmap(str(_logo_path))
            if not _logo_pixmap.isNull():
                _logo_label = QLabel()
                _logo_label.setPixmap(
                    _logo_pixmap.scaledToHeight(36, Qt.TransformationMode.SmoothTransformation)
                )
                _logo_label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                _logo_label.setContentsMargins(0, 0, 0, 0)
                _title_row.addWidget(_logo_label)

        _title_row.addStretch(1)
        self.app_title_label = QLabel("כלי להערכת בקרות ITGC בסביבת SAP HANA APP")
        self.app_title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #16325c;")
        _title_row.addWidget(self.app_title_label)
        main_layout.addWidget(_title_container)

        _header_container = QWidget()
        _header_container.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        _header_row = QHBoxLayout(_header_container)
        _header_row.setContentsMargins(0, 0, 0, 0)
        _header_row.setSpacing(0)
        self.header_label = QLabel("מסך בדיקת קלטי SAP HANA APP")
        self.header_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #16325c;")
        _header_row.addStretch(1)
        _header_row.addWidget(self.header_label)

        self.hint_label = QTextEdit()
        self.hint_label.setReadOnly(True)
        self.hint_label.setHtml(
            '<p dir="rtl" style="color: #4f5d73; margin: 0; padding: 0;">'
            "בחר קבצים לפי המשבצת המתאימה. כוכבית מציינת משבצת חובה. חובה לציין את תאריך ההפקה של הקבצים. ניתן להריץ בדיקה נפרדת לכל קבוצת קבצים בלי להמתין לטעינת כל הדוחות."
            "</p>"
        )
        self.hint_label.setFixedHeight(38)
        self.hint_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.hint_label.setStyleSheet(
            "background: transparent; border: none; padding: 0;"
        )

        self.work_environment_row = QWidget()
        self.work_environment_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        work_environment_layout = QHBoxLayout(self.work_environment_row)
        work_environment_layout.setContentsMargins(0, 0, 0, 0)
        work_environment_layout.setSpacing(8)
        work_environment_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self.work_environment_label = QLabel(self.format_ui_rtl_text("סביבת עבודה:"))
        self.work_environment_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.work_environment_combo = QComboBox()
        self.work_environment_combo.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.work_environment_combo.setMinimumWidth(260)
        for env_code, env_label in self.WORK_ENVIRONMENT_OPTIONS:
            self.work_environment_combo.addItem(self.format_rtl_text(env_label), env_code)
        self.work_environment_combo.currentIndexChanged.connect(self._persist_work_environment_selection)

        work_environment_layout.addWidget(self.work_environment_label)
        work_environment_layout.addWidget(self.work_environment_combo)
        work_environment_layout.addStretch(1)

        self.actions_row = QWidget()
        self.actions_row.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.actions_row.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        buttons_layout = QHBoxLayout(self.actions_row)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(8)
        buttons_layout.addStretch(1)

        self.clear_button = QPushButton(self.format_ui_rtl_text("נקה מסך"))
        self.clear_button.clicked.connect(self.clear_results)
        buttons_layout.addWidget(self.clear_button)

        self.export_log_button = QPushButton(self.format_ui_rtl_text("ייצוא רישום לאקסל"))
        self.export_log_button.clicked.connect(lambda: self.export_run_log_to_excel(open_after_export=True))
        buttons_layout.addWidget(self.export_log_button)

        self.output_button = QPushButton(self.format_ui_rtl_text("פתח תיקיית פלט"))
        self.output_button.clicked.connect(self.open_output_folder)
        buttons_layout.addWidget(self.output_button)

        self.report_button = QPushButton(self.format_ui_rtl_text("פתח דוח אקסל"))
        self.report_button.setEnabled(False)
        self.report_button.clicked.connect(self.open_report)
        buttons_layout.addWidget(self.report_button)

        self.tabs = QTabWidget()
        self.tabs.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.tabs.setDocumentMode(True)
        self.tabs.setTabPosition(QTabWidget.TabPosition.North)
        self.tabs.setMovable(False)
        self.tabs.setStyleSheet(
            """
            QTabBar::tab {
                background-color: #e9eef7;
                color: #16325c;
                border: 1px solid #b7c4d8;
                border-bottom: none;
                padding: 6px 12px;
                margin-left: 2px;
                min-width: 150px;
                font-weight: bold;
            }
            QTabBar::tab:selected {
                background-color: #6d002f;
                color: white;
            }
            QTabWidget::pane {
                border: 1px solid #c7cfda;
                top: -1px;
                background: #f5f7fb;
            }
            """
        )

        self.intake_tab = QWidget()
        self.intake_tab.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.intake_layout = QVBoxLayout(self.intake_tab)
        self.intake_layout.setContentsMargins(8, 8, 8, 8)
        self.intake_layout.setSpacing(10)
        self.intake_layout.addWidget(_header_container)
        self.intake_layout.addWidget(self.hint_label)
        self.intake_layout.addWidget(self.work_environment_row)
        self.intake_layout.addWidget(self.actions_row)

        self.analysis_tab = QWidget()
        self.analysis_tab.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.analysis_layout = QVBoxLayout(self.analysis_tab)
        self.analysis_layout.setContentsMargins(12, 12, 12, 12)
        self.analysis_layout.setSpacing(10)
        self.analysis_hint_label = QLabel(
            self.format_ui_rtl_text("לאחר טעינת הקבצים ניתן לבצע ניתוח לביקורת ולסקור כאן את הממצאים המרכזיים.")
        )
        self.analysis_hint_label.setWordWrap(True)
        self.analysis_hint_label.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.analysis_hint_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        self.analysis_layout.addWidget(self.analysis_hint_label)
        self.audit_run_button = QPushButton(self.format_ui_rtl_text("בצע ניתוח לביקורת עבור המשבצת שנבחרה"))
        self.audit_run_button.clicked.connect(self.run_validation)
        self.analysis_layout.addWidget(self.audit_run_button, 0, Qt.AlignmentFlag.AlignRight)

        self.analysis_progress_container = QWidget()
        self.analysis_progress_container.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        analysis_progress_layout = QHBoxLayout(self.analysis_progress_container)
        analysis_progress_layout.setContentsMargins(0, 0, 0, 0)
        analysis_progress_layout.setSpacing(8)
        analysis_progress_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self.analysis_progress_label = QLabel(self.format_ui_rtl_text("מעבד קובץ..."))
        self.analysis_progress_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.analysis_progress_label.setStyleSheet("color: #16325c; font-weight: bold;")

        self.analysis_progress_bar = QProgressBar()
        self.analysis_progress_bar.setMinimumWidth(220)
        self.analysis_progress_bar.setMaximumWidth(320)
        self.analysis_progress_bar.setTextVisible(False)
        self.analysis_progress_bar.setRange(0, 0)

        analysis_progress_layout.addWidget(self.analysis_progress_label)
        analysis_progress_layout.addWidget(self.analysis_progress_bar)
        analysis_progress_layout.addStretch(1)
        self.analysis_progress_container.hide()
        self.analysis_layout.addWidget(self.analysis_progress_container, 0, Qt.AlignmentFlag.AlignRight)

        self.audit_export_button = QPushButton(self.format_ui_rtl_text("ייצוא ממצאים לאקסל"))
        self.audit_export_button.clicked.connect(lambda: self.export_audit_findings_to_excel(open_after_export=True))
        self.analysis_layout.addWidget(self.audit_export_button, 0, Qt.AlignmentFlag.AlignRight)

        self.audit_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי ביקורת כללי - ריכוז"))
        self.audit_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        audit_summary_layout = QVBoxLayout(self.audit_summary_group)
        audit_summary_layout.setContentsMargins(8, 14, 8, 8)
        self.audit_summary_table = QTableWidget(0, 11)
        self.audit_summary_table.setItemDelegate(_RightAlignDelegate(self.audit_summary_table))
        self.audit_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("מזהה בקרה"),
            self.format_rtl_text("סוג בדיקה"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("רשומות תקינות"),
            self.format_rtl_text("רשומות עם ממצא"),
            self.format_rtl_text("סהכ רשומות"),
            self.format_rtl_text("נייר עבודה"),
            self.format_rtl_text("סביבת עבודה"),
            self.format_rtl_text("קובץ מקור"),
            self.format_rtl_text("תאריך הפקה"),
            self.format_rtl_text("תיאור בדיקה"),
        ])
        _audit_summary_hdr = self.audit_summary_table.horizontalHeader()
        _audit_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _audit_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _audit_summary_hdr.setStretchLastSection(False)
        self.audit_summary_table.setColumnWidth(1, 200)  # סוג בדיקה
        self.audit_summary_table.setColumnWidth(6, 90)   # נייר עבודה
        self.audit_summary_table.setColumnWidth(10, 220) # תיאור בדיקה
        self.audit_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.audit_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.audit_summary_table.setAlternatingRowColors(True)
        self.audit_summary_table.setMinimumHeight(220)
        self.audit_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.audit_summary_table.itemSelectionChanged.connect(self._refresh_selected_audit_detail)
        audit_summary_layout.addWidget(self.audit_summary_table)
        self.analysis_layout.addWidget(self.audit_summary_group)

        self.audit_detail_group = QGroupBox(self.format_ui_rtl_text("פירוט ממצאי ביקורת"))
        self.audit_detail_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        audit_detail_layout = QVBoxLayout(self.audit_detail_group)
        audit_detail_layout.setContentsMargins(8, 14, 8, 8)
        self.audit_detail_table = QTableWidget(0, 13)
        self.audit_detail_table.setItemDelegate(_RightAlignDelegate(self.audit_detail_table))
        self.audit_detail_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קובץ מקור"),
            self.format_rtl_text("תאריך הפקה"),
            self.format_rtl_text("סביבת עבודה"),
            self.format_rtl_text("קטגוריה"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("תיאור"),
            self.format_rtl_text("סוג בדיקה"),
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
            self.format_rtl_text("ערך בפועל"),
            self.format_rtl_text("ערך מצופה"),
            self.format_rtl_text("סטטוס"),
            self.format_rtl_text("תיאור מלא"),
        ])
        _audit_detail_hdr = self.audit_detail_table.horizontalHeader()
        _audit_detail_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _audit_detail_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _audit_detail_hdr.setStretchLastSection(True)  # תיאור מלא
        self.audit_detail_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.audit_detail_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.audit_detail_table.setAlternatingRowColors(True)
        self.audit_detail_table.setMinimumHeight(220)
        self.audit_detail_table.setToolTip(self.format_ui_rtl_text("לחיצה כפולה על שורה תציג פירוט מלא של הממצא"))
        self.audit_detail_table.cellDoubleClicked.connect(self.show_audit_detail_dialog)
        audit_detail_layout.addWidget(self.audit_detail_table)
        self.analysis_layout.addWidget(self.audit_detail_group)

        self.review_tab = QWidget()
        self.review_tab.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.review_layout = QVBoxLayout(self.review_tab)
        self.review_layout.setContentsMargins(6, 6, 6, 6)
        self.review_layout.setSpacing(6)

        self.permissions_review_tab = QWidget()
        self.permissions_review_tab.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.permissions_review_layout = QVBoxLayout(self.permissions_review_tab)
        self.permissions_review_layout.setContentsMargins(12, 12, 12, 12)
        self.permissions_review_layout.setSpacing(10)

        self.permissions_hint_label = QLabel(
            self.format_ui_rtl_text("סקירת הרשאות מרכזת ממצאים רוחביים על הרשאות משתמשים במערכת.")
        )
        self.permissions_hint_label.setWordWrap(True)
        self.permissions_hint_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        self.permissions_review_layout.addWidget(self.permissions_hint_label)

        self.permissions_inner_tabs = QTabWidget()
        self.permissions_inner_tabs.setLayoutDirection(Qt.LayoutDirection.RightToLeft)

        strong_profiles_page = QWidget()
        strong_profiles_page_layout = QVBoxLayout(strong_profiles_page)
        strong_profiles_page_layout.setContentsMargins(0, 0, 0, 0)
        strong_profiles_page_layout.setSpacing(8)
        self._build_strong_profiles_permissions_section(strong_profiles_page_layout)
        self.permissions_inner_tabs.addTab(strong_profiles_page, self.format_rtl_text("פרופילים למשתמשים חזקים"))

        placeholder_titles: list[str] = []

        user_mgmt_page = QWidget()
        user_mgmt_page_layout = QVBoxLayout(user_mgmt_page)
        user_mgmt_page_layout.setContentsMargins(0, 0, 0, 0)
        user_mgmt_page_layout.setSpacing(8)
        self._build_user_mgmt_permissions_section(user_mgmt_page_layout)
        self.permissions_inner_tabs.addTab(user_mgmt_page, self.format_rtl_text("הרשאות ניהול משתמשים"))

        auth_mgmt_page = QWidget()
        auth_mgmt_page_layout = QVBoxLayout(auth_mgmt_page)
        auth_mgmt_page_layout.setContentsMargins(0, 0, 0, 0)
        auth_mgmt_page_layout.setSpacing(8)
        self._build_auth_mgmt_permissions_section(auth_mgmt_page_layout)
        self.permissions_inner_tabs.addTab(auth_mgmt_page, self.format_rtl_text("הרשאות ניהול הרשאות"))

        rscdok99_page = QWidget()
        rscdok99_page_layout = QVBoxLayout(rscdok99_page)
        rscdok99_page_layout.setContentsMargins(0, 0, 0, 0)
        rscdok99_page_layout.setSpacing(8)
        self._build_rscdok99_permissions_section(rscdok99_page_layout)
        self.permissions_inner_tabs.addTab(rscdok99_page, self.format_rtl_text("הרשאות לתוכנית RSCDOK99"))

        data_mgmt_page = QWidget()
        data_mgmt_page_layout = QVBoxLayout(data_mgmt_page)
        data_mgmt_page_layout.setContentsMargins(0, 0, 0, 0)
        data_mgmt_page_layout.setSpacing(8)
        self._build_data_mgmt_permissions_section(data_mgmt_page_layout)
        self.permissions_inner_tabs.addTab(data_mgmt_page, self.format_rtl_text("הרשאות לניהול נתונים"))

        transport_page = QWidget()
        transport_page_layout = QVBoxLayout(transport_page)
        transport_page_layout.setContentsMargins(0, 0, 0, 0)
        transport_page_layout.setSpacing(8)
        self._build_transport_permissions_section(transport_page_layout)
        self.permissions_inner_tabs.addTab(transport_page, self.format_rtl_text("הרשאה להעברת שינויים"))

        debug_page = QWidget()
        debug_page_layout = QVBoxLayout(debug_page)
        debug_page_layout.setContentsMargins(0, 0, 0, 0)
        debug_page_layout.setSpacing(8)
        self._build_debug_permissions_section(debug_page_layout)
        self.permissions_inner_tabs.addTab(debug_page, self.format_rtl_text("הרשאות לשימוש ב DEBUG"))

        job_mgmt_page = QWidget()
        job_mgmt_page_layout = QVBoxLayout(job_mgmt_page)
        job_mgmt_page_layout.setContentsMargins(0, 0, 0, 0)
        job_mgmt_page_layout.setSpacing(8)
        self._build_job_mgmt_permissions_section(job_mgmt_page_layout)
        self.permissions_inner_tabs.addTab(job_mgmt_page, self.format_rtl_text("הרשאה לעידכון ג'ובים"))

        for placeholder_title in placeholder_titles:
            placeholder_page = QWidget()
            placeholder_layout = QVBoxLayout(placeholder_page)
            placeholder_layout.setContentsMargins(0, 0, 0, 0)
            placeholder_layout.setSpacing(8)
            placeholder_label = QLabel(
                self.format_ui_rtl_text("חוצץ זה ישופעל בשלב הבא של סקירת ההרשאות.")
            )
            placeholder_label.setWordWrap(True)
            placeholder_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
            placeholder_layout.addWidget(placeholder_label)
            placeholder_layout.addStretch(1)
            self.permissions_inner_tabs.addTab(placeholder_page, self.format_rtl_text(placeholder_title))

        self.permissions_review_layout.addWidget(self.permissions_inner_tabs)

        self.settings_tab = QWidget()
        self.settings_tab.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.settings_tab_layout = QVBoxLayout(self.settings_tab)
        self.settings_tab_layout.setContentsMargins(12, 12, 12, 12)
        self.settings_tab_layout.setSpacing(0)

        self.settings_scroll = QScrollArea()
        self.settings_scroll.setWidgetResizable(True)
        self.settings_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.settings_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.settings_scroll.setMinimumHeight(520)
        self.settings_scroll.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.settings_content = QWidget()
        self.settings_content.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.settings_content.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.settings_layout = QVBoxLayout(self.settings_content)
        self.settings_layout.setContentsMargins(0, 0, 0, 0)
        self.settings_layout.setSpacing(10)
        self.settings_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        self.settings_intro_label = QLabel(
            self.format_ui_rtl_text("בטאב זה ניתן לנהל את הגדרות הביקורת והעמודות הנדרשות לכל משבצת.")
        )
        # השתמש ב-QLabel עם rich text HTML RTL ו-align right
        self.settings_intro_label.setText(
            '<div dir="rtl" align="right">בטאב זה ניתן לנהל את הגדרות הביקורת והעמודות הנדרשות לכל משבצת.</div>'
        )
        self.settings_intro_label.setWordWrap(True)
        self.settings_intro_label.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.settings_intro_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        self.settings_intro_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.settings_intro_label.setMinimumHeight(40)
        self.settings_intro_label.setMaximumWidth(16777215)
        self.settings_layout.addWidget(self.settings_intro_label)
        self._build_system_settings_sections()

        self.settings_scroll.setWidget(self.settings_content)
        self.settings_scroll.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.settings_tab_layout.addWidget(self.settings_scroll)

        self.tabs.addTab(self.intake_tab, self.format_rtl_text("קליטת קבצים"))
        self.tabs.addTab(self.settings_tab, self.format_rtl_text("הגדרות מערכת לביקורת"))
        self.tabs.addTab(self.review_tab, self.format_rtl_text("סקירת דוח משתמשים"))
        self.tabs.addTab(self.permissions_review_tab, self.format_rtl_text("סקירת הרשאות"))
        self.tabs.addTab(self.analysis_tab, self.format_rtl_text("ביצוע ניתוח לביקורת"))
        main_layout.addWidget(self.tabs)

        # ── Activity console ──────────────────────────────────────────────
        self._console_container = QWidget()
        self._console_container.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        _console_vlayout = QVBoxLayout(self._console_container)
        _console_vlayout.setContentsMargins(0, 2, 0, 0)
        _console_vlayout.setSpacing(2)

        _console_header = QWidget()
        _console_header_layout = QHBoxLayout(_console_header)
        _console_header_layout.setContentsMargins(0, 0, 0, 0)
        _console_header_layout.setSpacing(6)
        _console_title_lbl = QLabel("◈ לוג פעולות")
        _console_title_lbl.setStyleSheet(
            "color: #667788; font-size: 10px; font-family: Consolas, monospace;"
        )
        self.console_clear_btn = QPushButton("נקה")
        self.console_clear_btn.setFixedSize(46, 18)
        self.console_clear_btn.setStyleSheet(
            "QPushButton { font-size: 9px; padding: 0 4px; color: #889; background: #252535;"
            " border: 1px solid #444; border-radius: 2px; }"
            "QPushButton:hover { background: #353545; color: #aab; }"
        )
        self.console_clear_btn.clicked.connect(lambda: self.console_output.clear())
        _console_header_layout.addStretch(1)
        _console_header_layout.addWidget(_console_title_lbl)
        _console_header_layout.addWidget(self.console_clear_btn)
        _console_vlayout.addWidget(_console_header)

        self.console_output = QTextEdit()
        self.console_output.setReadOnly(True)
        self.console_output.setFixedHeight(105)
        self.console_output.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.console_output.setStyleSheet(
            "QTextEdit {"
            "  background-color: #181826;"
            "  color: #90c090;"
            "  font-family: Consolas, 'Courier New', monospace;"
            "  font-size: 11px;"
            "  border: 1px solid #2c3c4c;"
            "  padding: 4px 6px;"
            "}"
        )
        _console_vlayout.addWidget(self.console_output)

        self.slots_group = QGroupBox(self.format_ui_rtl_text("מקורות קלט לבדיקת SAP HANA APP"))
        self.slots_group.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        self.slots_group.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        slots_group_layout = QVBoxLayout(self.slots_group)
        slots_group_layout.setContentsMargins(8, 18, 8, 8)

        self.slots_scroll = QScrollArea()
        self.slots_scroll.setWidgetResizable(True)
        self.slots_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.slots_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.slots_scroll.setMinimumHeight(280)
        self.slots_scroll.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        slots_container = QWidget()
        slots_container.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        slots_layout = QGridLayout(slots_container)
        slots_layout.setContentsMargins(12, 12, 12, 12)
        slots_layout.setHorizontalSpacing(12)
        slots_layout.setVerticalSpacing(10)
        slots_layout.setColumnStretch(0, 0)
        slots_layout.setColumnStretch(1, 1)
        slots_layout.setColumnStretch(2, 2)
        slots_layout.setColumnStretch(3, 0)
        slots_layout.setColumnMinimumWidth(0, 140)
        slots_layout.setColumnMinimumWidth(3, 120)
        slots_layout.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignRight)

        current_row = 0
        for domain in self._ordered_categories():
            palette = self._category_palette(domain)
            in_development = bool(self.DOMAIN_DEFINITIONS.get(domain, {}).get("in_development", False))

            domain_section = QGroupBox(self.format_ui_rtl_text(domain))
            domain_section.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            domain_section.setStyleSheet(
                f"""
                QGroupBox {{
                    font-weight: bold;
                    border: 2px solid {palette['border']};
                    border-radius: 10px;
                    margin-top: 12px;
                    padding-top: 14px;
                    background-color: #ffffff;
                }}
                QGroupBox::title {{
                    subcontrol-origin: margin;
                    subcontrol-position: top left;
                    padding: 4px 12px;
                    background-color: {palette['header']};
                    color: white;
                    border-radius: 6px;
                    font-weight: bold;
                }}
                """
            )
            self.category_sections[domain] = domain_section
            domain_section.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
            domain_layout = QVBoxLayout(domain_section)
            domain_layout.setContentsMargins(12, 18, 12, 12)
            domain_layout.setSpacing(10)

            domain_button = QPushButton("הרץ בדיקות תחום")
            domain_button.setMinimumHeight(34)
            domain_button.setToolTip(self.format_rtl_text(f"הרצת בדיקות עבור תחום {domain}"))
            domain_button.setStyleSheet(
                f"background-color: {palette['button']}; border: 2px solid {palette['border']}; color: white; font-weight: bold;"
            )
            domain_button.clicked.connect(
                lambda _checked=False, d=domain: self.run_domain_validation(d)
            )
            self.category_run_buttons[domain] = domain_button

            sub_cat_style = f"""
                QGroupBox {{
                    font-weight: bold;
                    border: 1px solid {palette['border']};
                    border-radius: 7px;
                    margin-top: 8px;
                    padding-top: 10px;
                    background-color: #f9fafb;
                }}
                QGroupBox::title {{
                    subcontrol-origin: margin;
                    subcontrol-position: top left;
                    padding: 2px 8px;
                    background-color: {palette['header']};
                    color: white;
                    border-radius: 4px;
                    font-weight: bold;
                    font-size: 11px;
                    opacity: 0.85;
                }}
            """

            for sub_cat_info in self.DOMAIN_DEFINITIONS.get(domain, {}).get("sub_categories", []):
                sub_cat_key = str(sub_cat_info["key"])
                sub_cat_desc = str(sub_cat_info["description"])

                sub_group = QGroupBox(self.format_ui_rtl_text(sub_cat_key))
                sub_group.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                sub_group.setStyleSheet(sub_cat_style)
                sub_group.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
                sub_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

                if in_development:
                    sub_group_layout = QVBoxLayout(sub_group)
                    sub_group_layout.setContentsMargins(10, 14, 10, 10)
                    sub_group_layout.setSpacing(6)

                    desc_label = QLabel(self.format_ui_rtl_text(sub_cat_desc))
                    desc_label.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
                    desc_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
                    desc_label.setWordWrap(True)
                    desc_label.setStyleSheet("color: #4f5d73;")

                    dev_label = QLabel(self.format_ui_rtl_text("⚙ בפיתוח — אין בדיקות אוטומטיות זמינות עדיין"))
                    dev_label.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
                    dev_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    dev_label.setStyleSheet(
                        "color: #6b7280; font-style: italic; background: #f3f4f6;"
                        " border: 1px dashed #d1d5db; border-radius: 4px; padding: 4px 8px;"
                    )

                    sub_group_layout.addWidget(desc_label)
                    sub_group_layout.addWidget(dev_label)
                else:
                    category_layout = QGridLayout(sub_group)
                    category_layout.setContentsMargins(12, 16, 12, 10)
                    category_layout.setHorizontalSpacing(12)
                    category_layout.setVerticalSpacing(10)
                    category_layout.setColumnStretch(0, 0)
                    category_layout.setColumnStretch(1, 1)
                    category_layout.setColumnStretch(2, 2)
                    category_layout.setColumnStretch(3, 0)
                    category_layout.setColumnMinimumWidth(0, 140)
                    category_layout.setColumnMinimumWidth(3, 120)
                    category_layout.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignRight)

                    # Sub-category description row
                    sub_desc_label = QLabel(self.format_ui_rtl_text(sub_cat_desc))
                    sub_desc_label.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
                    sub_desc_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    sub_desc_label.setWordWrap(True)
                    sub_desc_label.setStyleSheet("color: #4f5d73; padding: 2px 0;")
                    category_layout.addWidget(sub_desc_label, 0, 0, 1, 4)

                    section_row = 1
                    for slot_key, metadata in self.SLOT_DEFINITIONS.items():
                        if metadata.get("sub_category") != sub_cat_key:
                            continue

                        display_name = metadata.get("label", slot_key)
                        slot_title = QLabel(self.format_ui_rtl_text(f"{display_name}{' *' if metadata['required'] else ''}"))
                        slot_title.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
                        slot_title.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
                        slot_title.setStyleSheet("font-weight: bold;")
                        slot_title.setMinimumWidth(110)

                        description = QLabel(self.format_ui_rtl_text(metadata["description"]))
                        description.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
                        description.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
                        description.setWordWrap(True)
                        description.setMinimumHeight(34)
                        description.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

                        sample = QLabel(self.format_ui_rtl_text(f"קובץ צפוי: {metadata['expected_file']}"))
                        sample.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
                        sample.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
                        sample.setWordWrap(True)
                        sample.setStyleSheet("color: #5b6573;")
                        sample.setMinimumWidth(120)
                        sample.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

                        status_label = QLabel(self.format_ui_rtl_text("טרם נבחר קובץ"))
                        status_label.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
                        status_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        status_label.setWordWrap(True)
                        status_label.setMinimumHeight(32)
                        status_label.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
                        status_label.setStyleSheet("padding: 6px; background: #ffffff; border: 1px solid #cfd6e4;")

                        extraction_date_label = QLabel(self.format_ui_rtl_text("תאריך הפקה:"))
                        extraction_date_label.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
                        extraction_date_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        extraction_date_label.setStyleSheet("color: #5b6573;")
                        extraction_date_edit = QLineEdit(self._default_extraction_date())
                        extraction_date_edit.setAlignment(Qt.AlignmentFlag.AlignRight)
                        extraction_date_edit.setPlaceholderText("YYYY-MM-DD")
                        extraction_date_edit.setMinimumHeight(32)
                        extraction_date_edit.setMaximumWidth(170)

                        extraction_date_row = QWidget()
                        extraction_date_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
                        extraction_date_layout = QHBoxLayout(extraction_date_row)
                        extraction_date_layout.setContentsMargins(0, 0, 0, 0)
                        extraction_date_layout.setSpacing(6)
                        extraction_date_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        extraction_date_layout.addWidget(extraction_date_label, 0, Qt.AlignmentFlag.AlignRight)
                        extraction_date_layout.addWidget(extraction_date_edit, 0, Qt.AlignmentFlag.AlignRight)
                        extraction_date_layout.addStretch(1)

                        select_button = QPushButton("בחירת קבצים" if slot_key in self.MULTI_FILE_SLOTS else "בחירת קובץ")
                        select_button.setMinimumHeight(34)
                        select_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
                        select_button.clicked.connect(lambda _checked=False, sk=slot_key: self.choose_file(sk))

                        clear_slot_button = QPushButton("נקה")
                        clear_slot_button.setMinimumHeight(34)
                        clear_slot_button.setMinimumWidth(74)
                        clear_slot_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
                        clear_slot_button.clicked.connect(lambda _checked=False, sk=slot_key: self.clear_slot_selection(sk))

                        slot_buttons = QWidget()
                        slot_buttons.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
                        slot_buttons_layout = QHBoxLayout(slot_buttons)
                        slot_buttons_layout.setContentsMargins(0, 0, 0, 0)
                        slot_buttons_layout.setSpacing(6)
                        slot_buttons_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        slot_buttons_layout.addWidget(select_button)
                        slot_buttons_layout.addWidget(clear_slot_button)

                        # IPE evidence button
                        ipe_button = QPushButton("📎 ראיה (IPE)")
                        ipe_button.setToolTip("צרף תמונת מסך כראיית שליפה אותנטית (IPE)")
                        ipe_button.setMinimumHeight(34)
                        ipe_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
                        ipe_button.setStyleSheet(
                            "QPushButton { background-color: #e8f4fd; border: 1px solid #90caf9; color: #0d47a1; }"
                            "QPushButton:hover { background-color: #bbdefb; }"
                        )
                        ipe_button.clicked.connect(lambda _checked=False, sk=slot_key: self._add_ipe_evidence(sk))

                        slot_buttons_layout.addWidget(ipe_button)

                        # Thumbnail strip (hidden until images are attached)
                        thumb_scroll = QScrollArea()
                        thumb_scroll.setWidgetResizable(True)
                        thumb_scroll.setFixedHeight(90)
                        thumb_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
                        thumb_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
                        thumb_scroll.setVisible(False)
                        thumb_scroll.setStyleSheet("QScrollArea { border: 1px solid #90caf9; background: #f5f9ff; }")

                        thumb_container = QWidget()
                        thumb_container.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
                        thumb_layout = QHBoxLayout(thumb_container)
                        thumb_layout.setContentsMargins(4, 4, 4, 4)
                        thumb_layout.setSpacing(6)
                        thumb_layout.addStretch(1)
                        thumb_scroll.setWidget(thumb_container)

                        category_layout.setRowMinimumHeight(section_row, 42)
                        category_layout.addWidget(slot_title, section_row, 3)
                        category_layout.addWidget(description, section_row, 2)
                        category_layout.addWidget(sample, section_row, 1)
                        category_layout.addWidget(slot_buttons, section_row, 0)
                        section_row += 1
                        category_layout.setRowMinimumHeight(section_row, 36)
                        category_layout.addWidget(status_label, section_row, 0, 1, 4)
                        section_row += 1
                        category_layout.setRowMinimumHeight(section_row, 34)
                        category_layout.addWidget(extraction_date_row, section_row, 0, 1, 4)
                        section_row += 1
                        # thumbnail strip row (hidden by default)
                        category_layout.addWidget(thumb_scroll, section_row, 0, 1, 4)
                        section_row += 1

                        self.slot_widgets[slot_key] = {
                            "path_label": status_label,
                            "button": select_button,
                            "clear_button": clear_slot_button,
                            "ipe_button": ipe_button,
                            "thumb_scroll": thumb_scroll,
                            "thumb_layout": thumb_layout,
                            "metadata": metadata,
                            "selected_paths": [],
                            "extraction_date_edit": extraction_date_edit,
                            "extraction_date_label": extraction_date_label,
                        }

                domain_layout.addWidget(sub_group)

            domain_layout.addWidget(domain_button, 0, Qt.AlignmentFlag.AlignRight)
            domain_section.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            slots_layout.addWidget(domain_section, current_row, 0, 1, 4)
            current_row += 1

        bottom_spacer = QLabel("")
        bottom_spacer.setMinimumHeight(120)
        slots_layout.addWidget(bottom_spacer, current_row, 0, 1, 4)
        current_row += 1
        slots_layout.setRowStretch(current_row, 1)
        slots_layout.setRowMinimumHeight(current_row, 20)
        self.slots_scroll.setWidget(slots_container)
        slots_group_layout.addWidget(self.slots_scroll)

        self.slots_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.intake_layout.addWidget(self.slots_group, 1)


        self.user_preview_group = QGroupBox(self.format_ui_rtl_text("רשימת משתמשים שנטענו"))
        self.user_preview_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.user_preview_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.user_preview_group.setMinimumHeight(460)
        user_preview_layout = QVBoxLayout(self.user_preview_group)
        user_preview_layout.setContentsMargins(8, 12, 8, 8)
        user_preview_layout.setSpacing(4)
        user_preview_layout.setAlignment(Qt.AlignmentFlag.AlignTop)

        # --- Create progress group first ---
        self.user_review_progress_group = QGroupBox(self.format_ui_rtl_text("סיכום התקדמות סקירה"))
        self.user_review_progress_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.user_review_progress_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        user_review_progress_layout = QVBoxLayout(self.user_review_progress_group)
        user_review_progress_layout.setContentsMargins(8, 8, 8, 8)
        user_review_progress_layout.setSpacing(6)

        user_review_counts_row = QWidget()
        user_review_counts_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        user_review_counts_layout = QHBoxLayout(user_review_counts_row)
        user_review_counts_layout.setContentsMargins(0, 0, 0, 0)
        user_review_counts_layout.setSpacing(14)
        user_review_counts_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self.user_review_total_label = QLabel(self.format_ui_rtl_text("סה\"כ משתמשים בדוח: 0"))
        self.user_review_total_label.setStyleSheet("font-weight: bold;")
        self.user_review_reviewed_label = QLabel(self.format_ui_rtl_text("משתמשים שנבדקו: 0"))
        self.user_review_reviewed_label.setStyleSheet("font-weight: bold; color: #2e7d32;")
        self.user_review_unreviewed_label = QLabel(self.format_ui_rtl_text("משתמשים שטרם נבדקו: 0"))
        self.user_review_unreviewed_label.setStyleSheet("font-weight: bold; color: #1565c0;")

        user_review_counts_layout.addWidget(self.user_review_total_label, 0, Qt.AlignmentFlag.AlignRight)
        user_review_counts_layout.addWidget(self.user_review_reviewed_label, 0, Qt.AlignmentFlag.AlignRight)
        user_review_counts_layout.addWidget(self.user_review_unreviewed_label, 0, Qt.AlignmentFlag.AlignRight)
        user_review_counts_layout.addStretch(1)
        user_review_progress_layout.addWidget(user_review_counts_row)

        self.user_review_progress_bar = QProgressBar()
        self.user_review_progress_bar.setMinimum(0)
        self.user_review_progress_bar.setMaximum(100)
        self.user_review_progress_bar.setValue(0)
        self.user_review_progress_bar.setTextVisible(True)
        self.user_review_progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.user_review_progress_bar.setFormat("0%")
        self.user_review_progress_bar.setStyleSheet(
            "QProgressBar {"
            "border: 1px solid #b0bec5; border-radius: 4px;"
            "background-color: #f5f7fa; text-align: center;"
            "font-weight: bold; color: #0d47a1;"
            "}"
            "QProgressBar::chunk {"
            "background-color: #42a5f5; border-radius: 3px;"
            "}"
        )
        user_review_progress_layout.addWidget(self.user_review_progress_bar)

        self.user_review_progress_percent_label = QLabel(self.format_ui_rtl_text("התקדמות השלמת סקירה: 0%"))
        self.user_review_progress_percent_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.user_review_progress_percent_label.setStyleSheet("font-weight: bold; color: #0d47a1;")
        user_review_progress_layout.addWidget(self.user_review_progress_percent_label)

        # --- Now create actions row ---
        self.user_preview_actions_row = QWidget()
        self.user_preview_actions_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.user_preview_actions_row.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.user_preview_actions_row.setMaximumHeight(40)
        user_preview_actions_layout = QHBoxLayout(self.user_preview_actions_row)
        user_preview_actions_layout.setContentsMargins(0, 0, 0, 0)
        user_preview_actions_layout.setSpacing(8)
        user_preview_actions_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self.user_preview_export_button = QPushButton("ייצוא סקירה לאקסל")
        self.user_preview_export_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.user_preview_export_button.clicked.connect(lambda: self.export_user_preview_to_excel(open_after_export=True))
        user_preview_actions_layout.addWidget(self.user_preview_export_button, 0, Qt.AlignmentFlag.AlignRight)

        self.user_preview_import_button = QPushButton("ייבוא סקירה מאקסל")
        self.user_preview_import_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.user_preview_import_button.clicked.connect(self.import_user_review_from_excel)
        user_preview_actions_layout.addWidget(self.user_preview_import_button, 0, Qt.AlignmentFlag.AlignRight)

        self.user_preview_ai_narrate_button = QPushButton("הפק נרטיב AI לכל הממצאים")
        self.user_preview_ai_narrate_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.user_preview_ai_narrate_button.setToolTip(
            "מפיק תיאור ממצאים מנוסח על-ידי מודל AI מקומי (Ollama) לכל המשתמשים עם ממצאים. דורש שה-AI מופעל בהגדרות המערכת."
        )
        self.user_preview_ai_narrate_button.clicked.connect(self.generate_ai_narration_for_all_findings)
        user_preview_actions_layout.addWidget(self.user_preview_ai_narrate_button, 0, Qt.AlignmentFlag.AlignRight)

        self.user_preview_send_business_button = QPushButton("שליחת הדוח לגורם מהכספים")
        self.user_preview_send_business_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.user_preview_send_business_button.clicked.connect(self.draft_user_review_email_to_business)
        user_preview_actions_layout.addWidget(self.user_preview_send_business_button, 0, Qt.AlignmentFlag.AlignRight)

        self.user_preview_send_technical_button = QPushButton("שליחת הדוח לגורם טכני")
        self.user_preview_send_technical_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.user_preview_send_technical_button.clicked.connect(self.draft_user_review_email_to_technical)
        user_preview_actions_layout.addWidget(self.user_preview_send_technical_button, 0, Qt.AlignmentFlag.AlignRight)

        self.user_preview_columns_button = QPushButton("הוסף / מחק עמודות")
        self.user_preview_columns_button.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.user_preview_columns_button.clicked.connect(self.show_user_preview_column_dialog)
        user_preview_actions_layout.addWidget(self.user_preview_columns_button, 0, Qt.AlignmentFlag.AlignRight)
        user_preview_actions_layout.addStretch(1)

        # Add the progress group first, then the actions row
        user_preview_layout.addWidget(self.user_review_progress_group, 0, Qt.AlignmentFlag.AlignTop)
        user_preview_layout.addWidget(self.user_preview_actions_row, 0, Qt.AlignmentFlag.AlignTop)

        self.user_preview_filter_row = QWidget()
        self.user_preview_filter_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.user_preview_filter_row.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        user_preview_filter_layout = QHBoxLayout(self.user_preview_filter_row)
        user_preview_filter_layout.setContentsMargins(0, 0, 0, 0)
        user_preview_filter_layout.setSpacing(8)
        user_preview_filter_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        user_preview_filter_layout.addStretch(1)

        self.clear_column_filters_button = QPushButton(self.format_ui_rtl_text("נקה סינוני עמודות"))
        self.clear_column_filters_button.setToolTip(self.format_rtl_text("נקה את כל סינוני העמודות הפעילים"))
        self.clear_column_filters_button.clicked.connect(self._clear_all_user_preview_column_filters)
        user_preview_filter_layout.addWidget(self.clear_column_filters_button, 0, Qt.AlignmentFlag.AlignLeft)

        user_preview_layout.addWidget(self.user_preview_filter_row, 0, Qt.AlignmentFlag.AlignTop)

        self.user_preview_hint = QLabel(
            '<p align="right" style="margin:0">הטבלה מציגה את משתמשי USR02 עם העשרת נתונים מקובצי USER_ADDR ו-ADR6 בלבד.</p>'
        )
        self.user_preview_hint.setWordWrap(True)
        self.user_preview_hint.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        self.user_preview_hint.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        self.user_preview_hint.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.user_preview_hint.setMaximumHeight(44)
        user_preview_layout.addWidget(self.user_preview_hint, 0, Qt.AlignmentFlag.AlignTop)

        self.user_preview_table = QTableWidget(0, 0)
        self.user_preview_table.setItemDelegate(_RightAlignDelegate(self.user_preview_table))
        self.user_preview_table.setEditTriggers(
            QAbstractItemView.EditTrigger.DoubleClicked | QAbstractItemView.EditTrigger.EditKeyPressed | QAbstractItemView.EditTrigger.SelectedClicked
        )
        self.user_preview_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.user_preview_table.setAlternatingRowColors(True)
        self.user_preview_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.user_preview_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.user_preview_table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.user_preview_table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.user_preview_table.setMinimumHeight(420)
        self.user_preview_table.setMaximumHeight(16777215)
        self._configure_user_preview_table()
        self.user_preview_table.itemChanged.connect(self._handle_user_preview_item_changed)
        user_preview_layout.addWidget(self.user_preview_table, 1)
        self.review_layout.addWidget(self.user_preview_group, 1)

        self.run_log_group = QGroupBox(self.format_ui_rtl_text("לוג קבצים שנקלטו"))
        self.run_log_group.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        run_log_layout = QVBoxLayout(self.run_log_group)
        run_log_layout.setContentsMargins(12, 18, 12, 12)
        self.run_log_table = QTableWidget(0, 10)
        self.run_log_table.setItemDelegate(_RightAlignDelegate(self.run_log_table))
        self.run_log_table.setHorizontalHeaderLabels(["משבצת", "קבוצת דוחות", "קובץ", "תאריך הפקה", "רשומות שנקלטו", "סטטוס", "מספר שגיאות", "תיאור שגיאה", "תאריך בדיקה", "שעת בדיקה"])
        self.run_log_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        self.run_log_table.verticalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        _run_log_hdr = self.run_log_table.horizontalHeader()
        _run_log_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _run_log_hdr.setStretchLastSection(False)
        self.run_log_table.setColumnWidth(1, 150)
        self.run_log_table.setColumnWidth(2, 180)
        self.run_log_table.setColumnWidth(7, 220)  # תיאור שגיאה
        self.run_log_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.run_log_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.run_log_table.setAlternatingRowColors(True)
        self.run_log_table.setWordWrap(True)
        self.run_log_table.setTextElideMode(Qt.TextElideMode.ElideMiddle)
        self.run_log_table.setMinimumHeight(160)
        self.run_log_table.setToolTip("לחיצה כפולה על שורה תפתח פירוט מלא עבור הקובץ")
        self.run_log_table.cellDoubleClicked.connect(self.show_log_details)
        run_log_layout.addWidget(self.run_log_table)
        self.run_log_group.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.intake_layout.addWidget(self.run_log_group)
        self.run_log_group.hide()

        self.required_columns_group = QGroupBox(self.format_ui_rtl_text("עמודות חובה לבדיקה"))
        self.required_columns_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        required_layout = QHBoxLayout(self.required_columns_group)
        self.required_columns_edit = QLineEdit("")
        self.required_columns_edit.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.required_columns_edit.setPlaceholderText("יוזן אוטומטית לפי המשבצת שנבחרה")
        required_layout.addWidget(self.required_columns_edit)
        self.required_columns_group.hide()

        self.summary_group = QGroupBox(self.format_ui_rtl_text("סיכום בדיקה"))
        self.summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        summary_layout = QGridLayout(self.summary_group)
        summary_layout.setContentsMargins(12, 18, 12, 12)
        summary_layout.setHorizontalSpacing(10)
        summary_layout.setVerticalSpacing(8)
        summary_items = [
            ("שורות שנבדקו", "total", "0"),
            ("שורות תקינות", "valid", "0"),
            ("שורות שגויות", "invalid", "0"),
            ("סטטוס", "status", "ממתין להרצה"),
        ]
        for column, (title, key, default_value) in enumerate(summary_items):
            title_label = QLabel(title)
            title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            title_label.setStyleSheet("font-weight: bold; qproperty-alignment: 'AlignCenter|AlignVCenter';")
            value_label = QLabel(default_value)
            value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            value_label.setStyleSheet("font-size: 13px; padding: 2px; qproperty-alignment: 'AlignCenter|AlignVCenter';")
            summary_layout.addWidget(title_label, 0, column)
            summary_layout.addWidget(value_label, 1, column)
            self.summary_labels[key] = value_label
        self.summary_group.hide()
        self.intake_layout.insertWidget(4, self.summary_group)

        self.results_group = QGroupBox(self.format_ui_rtl_text("שגיאות קליטה"))
        self.results_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        results_layout = QVBoxLayout(self.results_group)
        results_layout.setContentsMargins(12, 18, 12, 12)
        self.issues_table = QTableWidget(0, 3)
        self.issues_table.setItemDelegate(_RightAlignDelegate(self.issues_table))
        self.issues_table.setHorizontalHeaderLabels(["מספר שורה", "שם עמודה", "הודעת שגיאה"])
        self.issues_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        self.issues_table.verticalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        _issues_hdr = self.issues_table.horizontalHeader()
        _issues_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _issues_hdr.setStretchLastSection(True)  # הודעת שגיאה
        self.issues_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.issues_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.issues_table.setAlternatingRowColors(True)
        results_layout.addWidget(self.issues_table)
        self.issues_table.setMinimumHeight(180)
        self.results_group.hide()
        self.intake_layout.addWidget(self.results_group)
        self.intake_layout.addWidget(self._console_container)

        central_widget.setStyleSheet(
            """
            QWidget {
                background-color: #f5f7fb;
                font-family: 'Segoe UI';
                font-size: 9.5pt;
            }
            QLabel {
                qproperty-alignment: 'AlignRight|AlignVCenter';
            }
            QGroupBox {
                font-weight: bold;
                border: 1px solid #c7cfda;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 12px;
                background-color: #f9fbfe;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 10px;
                background-color: #f5f7fb;
            }
            QPushButton {
                background-color: #e9eef7;
                border: 1px solid #b7c4d8;
                border-radius: 6px;
                padding: 4px 10px;
                font-weight: bold;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #dbe7f8;
            }
            QLineEdit {
                background-color: white;
                border: 1px solid #cfd6e4;
                padding: 4px;
            }
            QTableWidget {
                background-color: white;
                border: 1px solid #cfd6e4;
                gridline-color: #d7deea;
            }
            QTableWidget::item {
                padding-right: 4px;
            }
            QTableWidget::item:selected,
            QTableView::item:selected {
                background-color: #1f4e79;
                color: white;
            }
            QTableWidget::item:selected:!active,
            QTableView::item:selected:!active {
                background-color: #1f4e79;
                color: white;
            }
            QComboBox {
                background-color: #eef3fc;
                border: 1px solid #b7c4d8;
                border-radius: 4px;
                padding: 3px 6px;
            }
            QComboBox:hover {
                background-color: #dbe7f8;
                border-color: #7a9ec8;
            }
            QComboBox QAbstractItemView {
                background-color: #f5f8ff;
                border: 1px solid #b7c4d8;
                selection-background-color: #dbe7f8;
            }
            """
        )
        self._log_to_console("מוכן לקליטה...", "dim")

    def _ordered_categories(self) -> list[str]:
        return list(self.DOMAIN_DEFINITIONS.keys())

    def _ordered_sub_categories(self, domain: str) -> list[str]:
        return [
            sub["key"]
            for sub in self.DOMAIN_DEFINITIONS.get(domain, {}).get("sub_categories", [])
        ]

    def _category_palette(self, category: str) -> dict[str, str]:
        if category.startswith("MA"):
            return {"header": "#16325c", "button": "#16325c", "border": "#16325c"}
        if category.startswith("MC"):
            return {"header": "#1b5e20", "button": "#1b5e20", "border": "#1b5e20"}
        # MO — in development, shown in gray
        return {"header": "#6b7280", "button": "#6b7280", "border": "#9ca3af"}

    @staticmethod
    def _default_extraction_date() -> str:
        return datetime.now().strftime("%Y-%m-%d")


    def _build_system_settings_sections(self) -> None:
        # --- Outer wrapper for flush right alignment ---
        buttons_row = QWidget()
        buttons_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        buttons_row.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        buttons_layout = QHBoxLayout(buttons_row)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(8)
        buttons_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        self.settings_save_btn = QPushButton(self.format_ui_rtl_text("שמור הגדרות"))
        self.settings_save_btn.clicked.connect(self._save_system_settings)
        self.settings_save_btn.setStyleSheet("text-align: left; padding-right: 18px;")

        self.settings_reset_btn = QPushButton(self.format_ui_rtl_text("טען ברירות מחדל"))
        self.settings_reset_btn.clicked.connect(self._reset_system_settings_form)
        self.settings_reset_btn.setStyleSheet("text-align: left; padding-right: 18px;")

        buttons_layout.addWidget(self.settings_save_btn)
        buttons_layout.addWidget(self.settings_reset_btn)
        # Ensure the row itself is aligned right in the parent layout
        self.settings_layout.addWidget(buttons_row, alignment=Qt.AlignmentFlag.AlignRight)

        review_group, review_layout, review_unavailable_label = self._build_settings_group(
            "טווח תקופת הביקורת",
        )

        start_widget = QDateEdit()
        start_widget.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        start_widget.setDisplayFormat("yyyy-MM-dd")
        start_widget.setCalendarPopup(True)
        end_widget = QDateEdit()
        end_widget.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        end_widget.setDisplayFormat("yyyy-MM-dd")
        end_widget.setCalendarPopup(True)

        self.system_settings_widgets["user_review_period.start_date"] = start_widget
        self.system_settings_widgets["user_review_period.end_date"] = end_widget

        date_row = QWidget()
        date_row.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        date_row_layout = QHBoxLayout(date_row)
        date_row_layout.setContentsMargins(0, 0, 0, 0)
        date_row_layout.setSpacing(8)
        # LTR layout: stretch pushes content to the right.
        # Items added last are rightmost. For RTL label convention (label to the right of field):
        # Visual (RTL read): מתאריך: [start] | עד תאריך: [end]
        date_row_layout.addStretch(1)
        date_row_layout.addWidget(end_widget)
        date_row_layout.addWidget(QLabel(self.format_ui_rtl_text("עד תאריך:")))
        date_row_layout.addWidget(start_widget)
        date_row_layout.addWidget(QLabel(self.format_ui_rtl_text("מתאריך:")))
        review_layout.addWidget(date_row)

        self.settings_layout.addWidget(review_group)
        self.system_settings_sections["user_review_period"] = review_group
        self.system_settings_unavailable_labels["user_review_period"] = review_unavailable_label

        authorized_stms_group, authorized_stms_table, authorized_stms_unavailable_label = self._build_super_users_section()
        self.settings_layout.addWidget(authorized_stms_group)
        self.system_settings_widgets["authorized_stms_users"] = authorized_stms_table
        self.system_settings_widgets["super_users"] = authorized_stms_table
        self.system_settings_sections["authorized_stms_users"] = authorized_stms_group
        self.system_settings_sections["super_users"] = authorized_stms_group
        self.system_settings_unavailable_labels["authorized_stms_users"] = authorized_stms_unavailable_label
        self.system_settings_unavailable_labels["super_users"] = authorized_stms_unavailable_label

        dev_group, dev_table, dev_unavailable_label = self._build_developer_list_section()
        self.settings_layout.addWidget(dev_group)
        self.system_settings_widgets["authorized_developers"] = dev_table
        self.system_settings_sections["authorized_developers"] = dev_group
        self.system_settings_unavailable_labels["authorized_developers"] = dev_unavailable_label

        generic_users_group = self._add_settings_text_list_section(
            "generic_users",
            "משתמשים גנריים",
            "רשימה מופרדת שורות",
            read_only=True,
        )
        self.system_settings_sections["generic_users"] = generic_users_group

        self._add_settings_text_list_section(
            "critical_roles",
            "פרופילים משתמשיי על",
            "רשימה מופרדת שורות",
            read_only=True,
        )

        password_group, password_layout, password_unavailable_label = self._build_settings_group(
            "ברירות מחדל למדיניות סיסמה",
            "ערכים לוגיים לבקרות סיסמה ב-APP (כאשר נתוני המקור זמינים).",
        )
        password_grid = QGridLayout()
        password_fields = [
            ("minimal_password_length", "אורך סיסמה מינימלי"),
            ("maximum_invalid_connect_attempts", "login/fails_to_session_end - מקסימום ניסיונות כושלים"),
            ("max_password_age_days", "תוקף סיסמה (ימים)"),
            ("password_max_idle_initial", "מספר הימים לתוקף סיסמה ראשונית"),
            ("password_change_for_SSO", "חובת שינוי סיסמה ראשונית SSO"),
            ("login/fails_to_user_lock", "מקסימום ניסיונות כושלים נעילת משתמש"),
            ("password_history_size", "היסטוריית סיסמאות"),
            ("min_password_digits", "ספרות מינימליות"),
            ("min_password_letters", "אותיות מינימליות"),
            ("min_password_lowercase", "אותיות קטנות מינימליות"),
            ("min_password_uppercase", "אותיות גדולות מינימליות"),
            ("min_password_specials", "תווים מיוחדים מינימליים"),
            ("failed_user_auto_unlock", "שיחרור אוטומטי של משתמש נעול"),
            ("rdisp/gui_auto_logout", "התנתקות אוטומטית GUI")
        ]
        for index, (field_name, label_text) in enumerate(password_fields):
            row = index // 2
            col = (index % 2) * 2
            label = QLabel(self.format_ui_rtl_text(label_text))
            widget: object
            if field_name in {"password_change_for_SSO", "failed_user_auto_unlock"}:
                widget = QComboBox()
                widget.addItems(["TRUE", "FALSE"])
            else:
                widget = QLineEdit()
                if isinstance(widget, QLineEdit):
                    widget.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
                    widget.setMaxLength(6)
            if hasattr(widget, "setMaximumWidth"):
                widget.setMaximumWidth(120)
            self.system_settings_widgets[f"password_policy_defaults.{field_name}"] = widget
            password_grid.addWidget(label, row, col)
            password_grid.addWidget(widget, row, col + 1)
        password_layout.addLayout(password_grid)
        self.settings_layout.addWidget(password_group)
        self.system_settings_sections["password_policy_defaults"] = password_group
        self.system_settings_unavailable_labels["password_policy_defaults"] = password_unavailable_label

        threshold_group, threshold_layout, threshold_unavailable_label = self._build_settings_group(
            "הגדרות נוספות",
            "סף חוסר פעילות משמש לבניית ממצאים אוטומטיים בסקירת משתמשים.",
        )
        threshold_form = QFormLayout()
        threshold_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        threshold_form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)
        threshold_widget = QLineEdit()
        self.system_settings_widgets["inactive_days_threshold"] = threshold_widget
        threshold_form.addRow("סף חוסר פעילות (ימים)", threshold_widget)
        threshold_layout.addLayout(threshold_form)
        self.settings_layout.addWidget(threshold_group)
        self.system_settings_sections["inactive_days_threshold"] = threshold_group
        self.system_settings_unavailable_labels["inactive_days_threshold"] = threshold_unavailable_label

        email_group, email_layout, email_unavailable_label = self._build_settings_group(
            "הגדרות תפוצת מייל",
            "הגדר כתובות מייל של גורם מהכספים וגורם טכני עבור יצירת טיוטות שליחת דוח סקירת משתמשים.",
        )
        email_form = QFormLayout()
        email_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        email_form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)

        business_email_widget = QLineEdit()
        business_email_widget.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        business_email_widget.setPlaceholderText("business.reviewer@company.com")
        technical_email_widget = QLineEdit()
        technical_email_widget.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        technical_email_widget.setPlaceholderText("technical.reviewer@company.com")

        self.system_settings_widgets["business_reviewer_email"] = business_email_widget
        self.system_settings_widgets["technical_reviewer_email"] = technical_email_widget

        email_form.addRow("כתובת מייל גורם מהכספים", business_email_widget)
        email_form.addRow("כתובת מייל גורם טכני", technical_email_widget)
        email_layout.addLayout(email_form)
        self.settings_layout.addWidget(email_group)
        self.system_settings_sections["business_reviewer_email"] = email_group
        self.system_settings_sections["technical_reviewer_email"] = email_group
        self.system_settings_unavailable_labels["business_reviewer_email"] = email_unavailable_label
        self.system_settings_unavailable_labels["technical_reviewer_email"] = email_unavailable_label

        # ---- IPE Evidence Mapping Table ----
        ipe_group = QGroupBox(self.format_ui_rtl_text("מיפוי ראיות IPE לבקרות"))
        ipe_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        ipe_group_layout = QVBoxLayout(ipe_group)
        ipe_group_layout.setContentsMargins(8, 14, 8, 8)

        ipe_desc = QLabel(self.format_ui_rtl_text(
            "הגדר אילו בקרות ישויכו אוטומטית לכל ראיית IPE שתתווסף לסלוט."
        ))
        ipe_desc.setWordWrap(True)
        ipe_desc.setAlignment(Qt.AlignmentFlag.AlignRight)
        ipe_group_layout.addWidget(ipe_desc)

        ipe_reset_btn = QPushButton(self.format_ui_rtl_text("אפס לברירות מחדל"))
        ipe_reset_btn.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        ipe_reset_btn.clicked.connect(self._reset_ipe_mapping_to_defaults)
        ipe_group_layout.addWidget(ipe_reset_btn, alignment=Qt.AlignmentFlag.AlignLeft)

        self._ipe_mapping_control_ids: list[str] = list(CONTROL_LABELS.keys())
        self._ipe_mapping_slot_keys: list[str] = list(self.SLOT_DEFINITIONS.keys())
        _ipe_col_offset = 2  # columns 0 and 1 are fixed: control ID + short name

        self.ipe_mapping_table = QTableWidget(
            len(self._ipe_mapping_control_ids),
            _ipe_col_offset + len(self._ipe_mapping_slot_keys),
        )
        self.ipe_mapping_table.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.ipe_mapping_table.setAlternatingRowColors(True)
        self.ipe_mapping_table.setMinimumHeight(440)
        self.ipe_mapping_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.ipe_mapping_table.verticalHeader().setVisible(False)
        self.ipe_mapping_table.verticalHeader().setDefaultSectionSize(26)
        _ipe_hdr = self.ipe_mapping_table.horizontalHeader()
        _ipe_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)

        # Column headers: two fixed labels then slot/table names
        for col_idx, header_text in enumerate(["מזהה בקרה", "שם קצר"] + self._ipe_mapping_slot_keys):
            self.ipe_mapping_table.setHorizontalHeaderItem(col_idx, QTableWidgetItem(header_text))
        self.ipe_mapping_table.setColumnWidth(0, 195)
        self.ipe_mapping_table.setColumnWidth(1, 165)
        for _sc in range(len(self._ipe_mapping_slot_keys)):
            self.ipe_mapping_table.setColumnWidth(_ipe_col_offset + _sc, 82)

        # Custom delegate for slot columns: colored fill = checked, plain = unchecked
        _ipe_delegate = _IpeMappingCellDelegate(self.ipe_mapping_table)
        for _sc in range(len(self._ipe_mapping_slot_keys)):
            self.ipe_mapping_table.setItemDelegateForColumn(_ipe_col_offset + _sc, _ipe_delegate)

        # One row per control: col-0 = full ID, col-1 = short Hebrew name, col-2+ = slot checkboxes
        self.ipe_mapping_table.blockSignals(True)
        for row_idx, ctrl_id in enumerate(self._ipe_mapping_control_ids):
            id_item = QTableWidgetItem(ctrl_id)
            id_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            self.ipe_mapping_table.setItem(row_idx, 0, id_item)

            raw_label = CONTROL_LABELS.get(ctrl_id, ctrl_id)
            hebrew_part = raw_label.split(" — ", 1)[1] if " — " in raw_label else raw_label
            words = hebrew_part.split()
            short_name = " ".join(words[:5]) + ("..." if len(words) > 5 else "")
            name_item = QTableWidgetItem(short_name)
            name_item.setFlags(Qt.ItemFlag.ItemIsEnabled)
            name_item.setToolTip(hebrew_part)
            self.ipe_mapping_table.setItem(row_idx, 1, name_item)

            for slot_col, slot_key in enumerate(self._ipe_mapping_slot_keys):
                default_controls = set(SLOT_DEFAULT_CONTROLS.get(slot_key, []))
                is_checked = ctrl_id in default_controls
                cb_item = QTableWidgetItem()
                cb_item.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
                cb_item.setCheckState(Qt.CheckState.Checked if is_checked else Qt.CheckState.Unchecked)
                if is_checked:
                    cb_item.setBackground(QBrush(QColor("#6d002f")))
                self.ipe_mapping_table.setItem(row_idx, _ipe_col_offset + slot_col, cb_item)
        self.ipe_mapping_table.blockSignals(False)
        self.ipe_mapping_table.itemChanged.connect(self._on_ipe_mapping_changed)

        ipe_group_layout.addWidget(self.ipe_mapping_table)
        self.settings_layout.addWidget(ipe_group)

        # ---- AI Settings Section ----
        self._build_ai_settings_section()

    def _build_ai_settings_section(self) -> None:
        """Build the 'הגדרות AI' settings group and register its widgets."""
        ai_group, ai_layout, ai_unavailable_label = self._build_settings_group(
            "הגדרות AI — Ollama LLM מקומי",
            "כל עיבוד ה-AI מתבצע באופן מקומי בלבד. אין שליחת נתונים לרשת חיצונית.",
        )

        # Enable / disable toggle
        ai_enable_cb = self._build_checkbox(self.format_ui_rtl_text("הפעל יכולות AI (Ollama נדרש)"))
        self.system_settings_widgets["ai_settings.enabled"] = ai_enable_cb
        ai_layout.addWidget(ai_enable_cb)

        # Ollama host
        host_form = QFormLayout()
        host_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        host_form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)

        host_edit = QLineEdit()
        host_edit.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        host_edit.setPlaceholderText("http://localhost:11434")
        self.system_settings_widgets["ai_settings.ollama_host"] = host_edit
        host_form.addRow(self.format_ui_rtl_text("כתובת שרת Ollama"), host_edit)
        ai_layout.addLayout(host_form)

        # Model combo
        model_row = QWidget()
        model_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        model_row_layout = QHBoxLayout(model_row)
        model_row_layout.setContentsMargins(0, 0, 0, 0)
        model_row_layout.setSpacing(8)

        model_label = QLabel(self.format_ui_rtl_text("מודל:"))
        model_combo = QComboBox()
        model_combo.setEditable(True)
        model_combo.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        for m in RECOMMENDED_MODELS:
            model_combo.addItem(m)
        model_combo.setMinimumWidth(220)
        self.system_settings_widgets["ai_settings.model"] = model_combo

        model_row_layout.addStretch(1)
        model_row_layout.addWidget(model_combo)
        model_row_layout.addWidget(model_label)
        ai_layout.addWidget(model_row)

        # Temperature
        temp_form = QFormLayout()
        temp_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        temp_form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)

        from PySide6.QtWidgets import QDoubleSpinBox
        temp_spin = QDoubleSpinBox()
        temp_spin.setRange(0.0, 1.0)
        temp_spin.setSingleStep(0.1)
        temp_spin.setDecimals(1)
        temp_spin.setValue(0.3)
        temp_spin.setMaximumWidth(80)
        temp_spin.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        self.system_settings_widgets["ai_settings.temperature"] = temp_spin
        temp_form.addRow(self.format_ui_rtl_text("טמפרטורה (יצירתיות 0.0–1.0)"), temp_spin)
        ai_layout.addLayout(temp_form)

        # Timeout
        timeout_form = QFormLayout()
        timeout_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        timeout_form.setFieldGrowthPolicy(QFormLayout.FieldGrowthPolicy.ExpandingFieldsGrow)

        timeout_edit = QLineEdit()
        timeout_edit.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
        timeout_edit.setMaximumWidth(80)
        timeout_edit.setPlaceholderText("60")
        self.system_settings_widgets["ai_settings.timeout_seconds"] = timeout_edit
        timeout_form.addRow(self.format_ui_rtl_text("Timeout (שניות)"), timeout_edit)
        ai_layout.addLayout(timeout_form)

        # Feature toggles
        feat_narration_cb = self._build_checkbox(self.format_ui_rtl_text("נרטיב AI לממצאי בקרות (Feature 2.1)"))
        feat_compensating_cb = self._build_checkbox(self.format_ui_rtl_text("המלצות בקרות מפצות (Feature 2.2)"))
        feat_anomaly_cb = self._build_checkbox(self.format_ui_rtl_text("זיהוי אנומליות סטטיסטיות במשתמשים (Feature 2.4)"))
        self.system_settings_widgets["ai_settings.features.findings_narration"] = feat_narration_cb
        self.system_settings_widgets["ai_settings.features.compensating_controls"] = feat_compensating_cb
        self.system_settings_widgets["ai_settings.features.anomaly_detection"] = feat_anomaly_cb
        ai_layout.addWidget(feat_narration_cb)
        ai_layout.addWidget(feat_compensating_cb)
        ai_layout.addWidget(feat_anomaly_cb)

        # Connection test button + status label
        test_row = QWidget()
        test_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        test_row_layout = QHBoxLayout(test_row)
        test_row_layout.setContentsMargins(0, 0, 0, 0)
        test_row_layout.setSpacing(8)

        test_btn = QPushButton(self.format_ui_rtl_text("בדוק חיבור ל-Ollama"))
        test_btn.setMaximumWidth(200)
        self._ai_connection_status_label = QLabel("")
        self._ai_connection_status_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        def _on_test_connection() -> None:
            ai_cfg = self._collect_ai_settings_from_widgets()
            client = OllamaClient(ai_cfg)
            if client.is_available():
                local_models = client.list_local_models()
                models_str = ", ".join(local_models[:5]) if local_models else "(לא נמצאו מודלים)"
                self._ai_connection_status_label.setText(
                    self.format_ui_rtl_text(f"✓ חיבור תקין | מודלים: {models_str}")
                )
                self._ai_connection_status_label.setStyleSheet("color: #2e7d32; font-weight: bold;")
            else:
                self._ai_connection_status_label.setText(
                    self.format_ui_rtl_text("✗ לא ניתן להתחבר לשרת Ollama")
                )
                self._ai_connection_status_label.setStyleSheet("color: #c62828; font-weight: bold;")

        test_btn.clicked.connect(_on_test_connection)
        test_row_layout.addStretch(1)
        test_row_layout.addWidget(self._ai_connection_status_label)
        test_row_layout.addWidget(test_btn)
        ai_layout.addWidget(test_row)

        self.settings_layout.addWidget(ai_group)
        self.system_settings_sections["ai_settings"] = ai_group
        self.system_settings_unavailable_labels["ai_settings"] = ai_unavailable_label

    def _build_checkbox(self, label_text: str) -> Any:
        from PySide6.QtWidgets import QCheckBox
        cb = QCheckBox(label_text)
        cb.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        return cb

    def _collect_ai_settings_from_widgets(self) -> dict[str, Any]:
        """Read current widget values for ai_settings and return as dict."""
        ai = {}
        enabled_cb = self.system_settings_widgets.get("ai_settings.enabled")
        from PySide6.QtWidgets import QCheckBox, QDoubleSpinBox
        if isinstance(enabled_cb, QCheckBox):
            ai["enabled"] = enabled_cb.isChecked()
        host_w = self.system_settings_widgets.get("ai_settings.ollama_host")
        if isinstance(host_w, QLineEdit):
            ai["ollama_host"] = host_w.text().strip() or "http://localhost:11434"
        model_w = self.system_settings_widgets.get("ai_settings.model")
        if isinstance(model_w, QComboBox):
            ai["model"] = model_w.currentText().strip()
        temp_w = self.system_settings_widgets.get("ai_settings.temperature")
        if isinstance(temp_w, QDoubleSpinBox):
            ai["temperature"] = temp_w.value()
        timeout_w = self.system_settings_widgets.get("ai_settings.timeout_seconds")
        if isinstance(timeout_w, QLineEdit):
            ai["timeout_seconds"] = self._safe_int(timeout_w.text(), 60)
        feat = {}
        for feat_key in ("findings_narration", "compensating_controls", "anomaly_detection"):
            cb = self.system_settings_widgets.get(f"ai_settings.features.{feat_key}")
            if isinstance(cb, QCheckBox):
                feat[feat_key] = cb.isChecked()
        if feat:
            ai["features"] = feat
        return ai

    def _build_settings_group(self, title: str, description: str | None = None) -> tuple[QGroupBox, QVBoxLayout, QLabel]:
        group = QGroupBox(self.format_ui_rtl_text(title))
        group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        layout = QVBoxLayout(group)
        layout.setContentsMargins(10, 14, 10, 10)
        layout.setSpacing(8)
        if description:
            # Use HTML for robust RTL right alignment
            html = f'<div dir="rtl" align="right">{self.format_ui_rtl_text(description)}</div>'
            description_label = QLabel(html)
            description_label.setWordWrap(True)
            description_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
            layout.addWidget(description_label)

        unavailable_label = QLabel(self.format_ui_rtl_text("הגדרה זו לא זמינה ללא קובץ מקור רלוונטי"))
        unavailable_label.setWordWrap(True)
        unavailable_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        unavailable_label.setStyleSheet("color: gray; font-style: italic;")
        unavailable_label.setVisible(False)
        layout.addWidget(unavailable_label)
        return group, layout, unavailable_label

    def _build_super_users_section(self) -> tuple[QGroupBox, QTableWidget, QLabel]:
        group, layout, unavailable_label = self._build_settings_group(
            "מורשי STMS / ניהול שינויים",
            "רשימה לבנה של משתמשים מורשים להעברת טרנספורטים לייצור. יש להזין CLIENT ו-BNAME.",
        )
        table = QTableWidget(0, 2)
        table.setItemDelegate(_RightAlignDelegate(table))
        table.setHorizontalHeaderLabels(["CLIENT", "BNAME מורשה STMS"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        table.setMinimumHeight(140)
        layout.addWidget(table)

        control_row = QWidget()
        control_layout = QHBoxLayout(control_row)
        control_layout.setContentsMargins(0, 0, 0, 0)
        control_layout.setSpacing(8)
        control_layout.addStretch(1)

        add_button = QPushButton(self.format_ui_rtl_text("הוסף שורה"))
        remove_button = QPushButton(self.format_ui_rtl_text("הסר שורה"))
        add_button.clicked.connect(lambda: self._append_super_user_row(table))
        remove_button.clicked.connect(lambda: self._remove_selected_super_user_row(table))
        control_layout.addWidget(remove_button)
        control_layout.addWidget(add_button)
        layout.addWidget(control_row)

        return group, table, unavailable_label

    def _build_developer_list_section(self) -> tuple[QGroupBox, QTableWidget, QLabel]:
        group, layout, unavailable_label = self._build_settings_group(
            "רשימת מפתחים - הפרדת תפקידים",
            "רשימת משתמשים המוגדרים כמפתחים בארגון. הבקרה תזהה אם משתמשים אלו קיימים ופעילים בסביבת הייצור.",
        )
        table = QTableWidget(0, 2)
        table.setItemDelegate(_RightAlignDelegate(table))
        table.setHorizontalHeaderLabels(["CLIENT", "BNAME מפתח"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        table.setMinimumHeight(140)
        layout.addWidget(table)

        control_row = QWidget()
        control_layout = QHBoxLayout(control_row)
        control_layout.setContentsMargins(0, 0, 0, 0)
        control_layout.setSpacing(8)
        control_layout.addStretch(1)

        add_button = QPushButton(self.format_ui_rtl_text("הוסף שורה"))
        remove_button = QPushButton(self.format_ui_rtl_text("הסר שורה"))
        add_button.clicked.connect(lambda: self._append_super_user_row(table))
        remove_button.clicked.connect(lambda: self._remove_selected_super_user_row(table))
        control_layout.addWidget(remove_button)
        control_layout.addWidget(add_button)
        layout.addWidget(control_row)

        return group, table, unavailable_label

    def _append_super_user_row(self, table: QTableWidget) -> None:
        row = table.rowCount()
        table.insertRow(row)
        table.setItem(row, 0, QTableWidgetItem(""))
        table.setItem(row, 1, QTableWidgetItem(""))

    def _remove_selected_super_user_row(self, table: QTableWidget) -> None:
        selected_rows = sorted((index.row() for index in table.selectionModel().selectedRows()), reverse=True)
        for row_index in selected_rows:
            table.removeRow(row_index)

    def _add_settings_text_list_section(
        self,
        key: str,
        title: str,
        description: str,
        read_only: bool = False,
    ) -> QGroupBox:
        group, group_layout, unavailable_label = self._build_settings_group(title, description)
        editor = QPlainTextEdit()
        editor.setMinimumHeight(90)
        if read_only:
            editor.setReadOnly(True)
            editor.setToolTip(self.format_ui_rtl_text("שדה זה מוצג לקריאה בלבד ואינו ניתן לשינוי"))
        self.system_settings_widgets[key] = editor
        group_layout.addWidget(editor)
        self.settings_layout.addWidget(group)
        self.system_settings_sections[key] = group
        self.system_settings_unavailable_labels[key] = unavailable_label
        return group

    @staticmethod
    def _safe_int(value: object, fallback: int) -> int:
        try:
            return int(str(value).strip())
        except Exception:
            return fallback

    def _system_settings_path(self) -> Path:
        return self.ui_state_repository.system_settings_path()

    def _default_system_settings(self) -> dict[str, Any]:
        return {
            "work_environment": "FPP",
            "business_reviewer_email": "",
            "technical_reviewer_email": "",
            "generic_users": ["SAP", "DDIC", "TMSADM", "SAPCPIC"],
            "authorized_stms_users": [],
            "authorized_developers": [],
            "super_users": [],
            "critical_roles": ["SAP_ALL", "SAP_NEW"],
            "critical_privileges": ["S_TABU_DIS", "S_USER_GRP", "S_USER_AGR"],
            "password_policy_defaults": {
                "minimal_password_length": 8,
                "maximum_invalid_connect_attempts": 6,
                "password_expire_warning_time": 14,
                "max_password_age_days": 90,
                "initial_password_change_max_days": 2,
                "password_change_for_SSO": "TRUE",
                "password_max_idle_initial": 0,
                "login/fails_to_user_lock": 0,
                "password_history_size": 5,
                "min_password_digits": 0,
                "min_password_letters": 0,
                "min_password_lowercase": 0,
                "min_password_uppercase": 0,
                "min_password_specials": 0,
                "failed_user_auto_unlock": "FALSE",
                "fails_to_session_end": 0,
                "rdisp/gui_auto_logout": 0,
            },
            "inactive_days_threshold": 90,
            "user_review_period": {
                "start_date": datetime.now().replace(month=1, day=1).strftime("%Y-%m-%d"),
                "end_date": datetime.now().replace(month=12, day=31).strftime("%Y-%m-%d"),
            },
            "file_mappings": {
                slot_key: str(metadata.get("expected_file", ""))
                for slot_key, metadata in self.SLOT_DEFINITIONS.items()
            },
            "slot_ipe_control_mapping": {k: list(v) for k, v in SLOT_DEFAULT_CONTROLS.items()},
            "ai_settings": {
                "enabled": False,
                "ollama_host": "http://localhost:11434",
                "model": "aya-expanse:8b",
                "temperature": 0.3,
                "timeout_seconds": 60,
                "features": {
                    "findings_narration": True,
                    "compensating_controls": True,
                    "anomaly_detection": True,
                },
            },
        }

    def _current_system_settings(self) -> dict[str, Any]:
        return self.ui_state_repository.load_system_settings(self._default_system_settings())

    def _sync_review_filters_from_settings(self, settings: dict[str, Any]) -> None:
        period_cfg = settings.get("user_review_period", {}) if isinstance(settings, dict) else {}
        start_text = str(period_cfg.get("start_date", "")).strip()
        end_text = str(period_cfg.get("end_date", "")).strip()
        if hasattr(self, "audit_period_from_edit") and start_text:
            self.audit_period_from_edit.setText(start_text)
        if hasattr(self, "audit_period_to_edit") and end_text:
            self.audit_period_to_edit.setText(end_text)

    def _load_system_settings_into_form(self, settings: dict[str, Any], load_review_period: bool = True) -> None:
        settings = settings or self._default_system_settings()

        selected_environment = self._normalize_work_environment_code(settings.get("work_environment", "FPP"))
        if hasattr(self, "work_environment_combo") and isinstance(self.work_environment_combo, QComboBox):
            for index in range(self.work_environment_combo.count()):
                if str(self.work_environment_combo.itemData(index) or "").strip().upper() == selected_environment:
                    self.work_environment_combo.setCurrentIndex(index)
                    break

        def _fill_lines(key: str) -> None:
            editor = self.system_settings_widgets.get(key)
            values = settings.get(key, [])
            if isinstance(editor, QPlainTextEdit):
                editor.setPlainText("\n".join(str(item).strip() for item in values if str(item).strip()))

        _fill_lines("generic_users")
        _fill_lines("critical_roles")
        _fill_lines("critical_privileges")

        authorized_table = self.system_settings_widgets.get("authorized_stms_users")
        authorized_users = settings.get("authorized_stms_users", []) if isinstance(settings, dict) else []
        if not authorized_users and isinstance(settings, dict):
            authorized_users = settings.get("super_users", [])
        if isinstance(authorized_table, QTableWidget):
            authorized_table.setRowCount(0)
            if isinstance(authorized_users, list):
                for authorized_user in authorized_users:
                    if isinstance(authorized_user, dict):
                        mandt = str(authorized_user.get("MANDT", "")).strip()
                        bname = str(authorized_user.get("BNAME", "")).strip()
                        if mandt or bname:
                            row = authorized_table.rowCount()
                            authorized_table.insertRow(row)
                            authorized_table.setItem(row, 0, QTableWidgetItem(mandt))
                            authorized_table.setItem(row, 1, QTableWidgetItem(bname))

        dev_table = self.system_settings_widgets.get("authorized_developers")
        dev_list = settings.get("authorized_developers", [])
        if isinstance(dev_table, QTableWidget):
            dev_table.setRowCount(0)
            if isinstance(dev_list, list):
                for dev_entry in dev_list:
                    if isinstance(dev_entry, dict):
                        mandt = str(dev_entry.get("MANDT", "")).strip()
                        bname = str(dev_entry.get("BNAME", "")).strip()
                        row = dev_table.rowCount()
                        dev_table.insertRow(row)
                        dev_table.setItem(row, 0, QTableWidgetItem(mandt))
                        dev_table.setItem(row, 1, QTableWidgetItem(bname))

        if load_review_period:
            period_cfg = settings.get("user_review_period", {}) if isinstance(settings, dict) else {}
            start_text = str(period_cfg.get("start_date", self._default_extraction_date())).strip()
            end_text = str(period_cfg.get("end_date", self._default_extraction_date())).strip()

            start_widget = self.system_settings_widgets.get("user_review_period.start_date")
            end_widget = self.system_settings_widgets.get("user_review_period.end_date")
            if isinstance(start_widget, QDateEdit):
                date_value = QDate.fromString(start_text, "yyyy-MM-dd")
                start_widget.setDate(date_value if date_value.isValid() else QDate.currentDate())
            if isinstance(end_widget, QDateEdit):
                date_value = QDate.fromString(end_text, "yyyy-MM-dd")
                end_widget.setDate(date_value if date_value.isValid() else QDate.currentDate())

        for mapping_key in self.system_settings_file_mapping_order:
            widget = self.system_settings_widgets.get(f"file_mappings.{mapping_key}")
            file_mappings = settings.get("file_mappings", {})
            if isinstance(widget, QLineEdit) and isinstance(file_mappings, dict):
                widget.setText(str(file_mappings.get(mapping_key, "")))

        threshold_widget = self.system_settings_widgets.get("inactive_days_threshold")
        if isinstance(threshold_widget, QLineEdit):
            threshold_widget.setText(str(settings.get("inactive_days_threshold", 90)))

        business_email_widget = self.system_settings_widgets.get("business_reviewer_email")
        if isinstance(business_email_widget, QLineEdit):
            business_email_widget.setText(str(settings.get("business_reviewer_email", "")).strip())
        technical_email_widget = self.system_settings_widgets.get("technical_reviewer_email")
        if isinstance(technical_email_widget, QLineEdit):
            technical_email_widget.setText(str(settings.get("technical_reviewer_email", "")).strip())

        password_defaults = settings.get("password_policy_defaults", {}) if isinstance(settings, dict) else {}
        if isinstance(password_defaults, dict):
            for key, value in password_defaults.items():
                widget = self.system_settings_widgets.get(f"password_policy_defaults.{key}")
                if isinstance(widget, QComboBox):
                    widget.setCurrentText(str(value))
                elif isinstance(widget, QLineEdit):
                    widget.setText(str(value))

        # Load IPE mapping into table
        if hasattr(self, "ipe_mapping_table"):
            ipe_mapping = settings.get("slot_ipe_control_mapping") if isinstance(settings, dict) else None
            if not isinstance(ipe_mapping, dict):
                ipe_mapping = {k: list(v) for k, v in SLOT_DEFAULT_CONTROLS.items()}
            _offset = 2
            self.ipe_mapping_table.blockSignals(True)
            for row_idx, ctrl_id in enumerate(self._ipe_mapping_control_ids):
                for slot_col, slot_key in enumerate(self._ipe_mapping_slot_keys):
                    checked_controls = set(ipe_mapping.get(slot_key, []))
                    is_checked = ctrl_id in checked_controls
                    item = self.ipe_mapping_table.item(row_idx, _offset + slot_col)
                    if item is not None:
                        item.setCheckState(Qt.CheckState.Checked if is_checked else Qt.CheckState.Unchecked)
                        item.setBackground(QBrush(QColor("#6d002f")) if is_checked else QBrush())
            self.ipe_mapping_table.blockSignals(False)

        # Load AI settings
        ai_cfg = settings.get("ai_settings", {}) if isinstance(settings, dict) else {}
        if isinstance(ai_cfg, dict):
            from PySide6.QtWidgets import QCheckBox, QDoubleSpinBox
            enabled_cb = self.system_settings_widgets.get("ai_settings.enabled")
            if isinstance(enabled_cb, QCheckBox):
                enabled_cb.setChecked(bool(ai_cfg.get("enabled", False)))
            host_w = self.system_settings_widgets.get("ai_settings.ollama_host")
            if isinstance(host_w, QLineEdit):
                host_w.setText(str(ai_cfg.get("ollama_host", "http://localhost:11434")))
            model_w = self.system_settings_widgets.get("ai_settings.model")
            if isinstance(model_w, QComboBox):
                model_val = str(ai_cfg.get("model", "aya-expanse:8b"))
                idx = model_w.findText(model_val)
                if idx >= 0:
                    model_w.setCurrentIndex(idx)
                else:
                    model_w.setCurrentText(model_val)
            temp_w = self.system_settings_widgets.get("ai_settings.temperature")
            if isinstance(temp_w, QDoubleSpinBox):
                temp_w.setValue(float(ai_cfg.get("temperature", 0.3)))
            timeout_w = self.system_settings_widgets.get("ai_settings.timeout_seconds")
            if isinstance(timeout_w, QLineEdit):
                timeout_w.setText(str(ai_cfg.get("timeout_seconds", 60)))
            features = ai_cfg.get("features", {})
            if isinstance(features, dict):
                for feat_key in ("findings_narration", "compensating_controls", "anomaly_detection"):
                    cb = self.system_settings_widgets.get(f"ai_settings.features.{feat_key}")
                    if isinstance(cb, QCheckBox):
                        cb.setChecked(bool(features.get(feat_key, True)))

    def _collect_system_settings_from_form(self) -> dict[str, Any]:
        def _lines_from_editor(editor: object) -> list[str]:
            if not isinstance(editor, QPlainTextEdit):
                return []
            return [line.strip() for line in editor.toPlainText().splitlines() if line.strip()]

        settings = copy.deepcopy(self._current_system_settings())
        default_settings = self._default_system_settings()
        settings["work_environment"] = self._current_work_environment_code()

        generic_users_editor = self.system_settings_widgets.get("generic_users")
        if isinstance(generic_users_editor, QPlainTextEdit):
            settings["generic_users"] = _lines_from_editor(generic_users_editor)

        critical_roles_editor = self.system_settings_widgets.get("critical_roles")
        if isinstance(critical_roles_editor, QPlainTextEdit):
            settings["critical_roles"] = _lines_from_editor(critical_roles_editor)

        critical_privileges_editor = self.system_settings_widgets.get("critical_privileges")
        if isinstance(critical_privileges_editor, QPlainTextEdit):
            settings["critical_privileges"] = _lines_from_editor(critical_privileges_editor)

        authorized_table = self.system_settings_widgets.get("authorized_stms_users")
        authorized_users: list[dict[str, str]] = []
        if isinstance(authorized_table, QTableWidget):
            for row_index in range(authorized_table.rowCount()):
                mandt_item = authorized_table.item(row_index, 0)
                bname_item = authorized_table.item(row_index, 1)
                mandt_text = str(mandt_item.text()).strip() if isinstance(mandt_item, QTableWidgetItem) else ""
                bname_text = str(bname_item.text()).strip() if isinstance(bname_item, QTableWidgetItem) else ""
                if mandt_text or bname_text:
                    authorized_users.append({"MANDT": mandt_text, "BNAME": bname_text})
        settings["authorized_stms_users"] = authorized_users
        settings["super_users"] = authorized_users

        dev_table = self.system_settings_widgets.get("authorized_developers")
        dev_list: list[dict[str, str]] = []
        if isinstance(dev_table, QTableWidget):
            for row_index in range(dev_table.rowCount()):
                m_item = dev_table.item(row_index, 0)
                b_item = dev_table.item(row_index, 1)
                m_val = m_item.text().strip() if m_item else ""
                b_val = b_item.text().strip() if b_item else ""
                if m_val or b_val:
                    dev_list.append({"MANDT": m_val, "BNAME": b_val})
        settings["authorized_developers"] = dev_list

        period_start_widget = self.system_settings_widgets.get("user_review_period.start_date")
        period_end_widget = self.system_settings_widgets.get("user_review_period.end_date")
        if isinstance(period_start_widget, QDateEdit) and isinstance(period_end_widget, QDateEdit):
            settings["user_review_period"] = {
                "start_date": period_start_widget.date().toString("yyyy-MM-dd"),
                "end_date": period_end_widget.date().toString("yyyy-MM-dd"),
            }

        file_mappings = {}
        has_mapping_widgets = False
        for mapping_key in self.system_settings_file_mapping_order:
            widget = self.system_settings_widgets.get(f"file_mappings.{mapping_key}")
            if isinstance(widget, QLineEdit):
                has_mapping_widgets = True
                file_mappings[mapping_key] = widget.text().strip()
        if has_mapping_widgets:
            settings["file_mappings"] = file_mappings

        threshold_widget = self.system_settings_widgets.get("inactive_days_threshold")
        if isinstance(threshold_widget, QLineEdit):
            settings["inactive_days_threshold"] = self._safe_int(threshold_widget.text(), 90)

        business_email_widget = self.system_settings_widgets.get("business_reviewer_email")
        if isinstance(business_email_widget, QLineEdit):
            settings["business_reviewer_email"] = business_email_widget.text().strip()
        technical_email_widget = self.system_settings_widgets.get("technical_reviewer_email")
        if isinstance(technical_email_widget, QLineEdit):
            settings["technical_reviewer_email"] = technical_email_widget.text().strip()

        password_defaults = {}
        for field_name in [
            "minimal_password_length",
            "maximum_invalid_connect_attempts",
            "max_password_age_days",
            "password_max_idle_initial",
            "password_change_for_SSO",
            "login/fails_to_user_lock",
            "password_history_size",
            "min_password_digits",
            "min_password_letters",
            "min_password_lowercase",
            "min_password_uppercase",
            "min_password_specials",
            "failed_user_auto_unlock",
            "rdisp/gui_auto_logout",
        ]:
            widget = self.system_settings_widgets.get(f"password_policy_defaults.{field_name}")
            if isinstance(widget, QComboBox):
                password_defaults[field_name] = widget.currentText().strip()
            elif isinstance(widget, QLineEdit):
                default_value = default_settings["password_policy_defaults"].get(field_name, 0)
                password_defaults[field_name] = self._safe_int(widget.text(), int(default_value))
        settings["password_policy_defaults"] = password_defaults

        # Collect IPE mapping from table
        if hasattr(self, "ipe_mapping_table"):
            ipe_mapping: dict[str, list[str]] = {}
            _offset = 2
            for slot_col, slot_key in enumerate(self._ipe_mapping_slot_keys):
                checked: list[str] = []
                for row_idx, ctrl_id in enumerate(self._ipe_mapping_control_ids):
                    item = self.ipe_mapping_table.item(row_idx, _offset + slot_col)
                    if item is not None and item.checkState() == Qt.CheckState.Checked:
                        checked.append(ctrl_id)
                ipe_mapping[slot_key] = checked
            settings["slot_ipe_control_mapping"] = ipe_mapping

        # Collect AI settings
        settings["ai_settings"] = self._collect_ai_settings_from_widgets()

        return settings

    def _normalize_work_environment_code(self, value: object) -> str:
        normalized = str(value or "").strip().upper()
        valid_codes = {code for code, _label in self.WORK_ENVIRONMENT_OPTIONS}
        return normalized if normalized in valid_codes else "FPP"

    def _current_work_environment_code(self) -> str:
        if hasattr(self, "work_environment_combo") and isinstance(self.work_environment_combo, QComboBox):
            selected_data = self.work_environment_combo.currentData()
            return self._normalize_work_environment_code(selected_data)
        settings = self._current_system_settings()
        return self._normalize_work_environment_code(settings.get("work_environment", "FPP"))

    def _current_work_environment_label(self) -> str:
        selected_code = self._current_work_environment_code()
        for env_code, env_label in self.WORK_ENVIRONMENT_OPTIONS:
            if env_code == selected_code:
                return env_label
        return selected_code

    def _persist_work_environment_selection(self, _index: int) -> None:
        try:
            settings = self._current_system_settings()
            settings["work_environment"] = self._current_work_environment_code()
            self.ui_state_repository.save_system_settings(settings)
        except Exception:
            # Avoid interrupting the user flow if persistence fails.
            pass

    def _save_system_settings(self) -> None:
        try:
            settings = self._collect_system_settings_from_form()
            self.ui_state_repository.save_system_settings(settings)
            _saved_mapping = settings.get("slot_ipe_control_mapping")
            self._slot_control_mapping = (
                dict(_saved_mapping) if isinstance(_saved_mapping, dict) else dict(SLOT_DEFAULT_CONTROLS)
            )
            self._sync_review_filters_from_settings(settings)
            self.refresh_user_preview()
            QMessageBox.information(self, "הצלחה", "הגדרות המערכת נשמרו בהצלחה.")
        except Exception as error:
            QMessageBox.critical(self, "שגיאת הגדרות", f"לא ניתן לשמור את הגדרות המערכת.\n\n{error}")

    def _reset_system_settings_form(self) -> None:
        defaults = self._default_system_settings()
        self._load_system_settings_into_form(defaults)
        self._apply_system_settings_availability()

    def _on_ipe_mapping_changed(self, item: QTableWidgetItem) -> None:
        if self.ipe_mapping_table.column(item) < 2:
            return
        self.ipe_mapping_table.blockSignals(True)
        try:
            is_checked = item.checkState() == Qt.CheckState.Checked
            item.setBackground(QBrush(QColor("#6d002f")) if is_checked else QBrush())
        finally:
            self.ipe_mapping_table.blockSignals(False)

    def _reset_ipe_mapping_to_defaults(self) -> None:
        if not hasattr(self, "ipe_mapping_table"):
            return
        _offset = 2
        self.ipe_mapping_table.blockSignals(True)
        for row_idx, ctrl_id in enumerate(self._ipe_mapping_control_ids):
            for slot_col, slot_key in enumerate(self._ipe_mapping_slot_keys):
                is_checked = ctrl_id in set(SLOT_DEFAULT_CONTROLS.get(slot_key, []))
                item = self.ipe_mapping_table.item(row_idx, _offset + slot_col)
                if item is not None:
                    item.setCheckState(Qt.CheckState.Checked if is_checked else Qt.CheckState.Unchecked)
                    item.setBackground(QBrush(QColor("#6d002f")) if is_checked else QBrush())
        self.ipe_mapping_table.blockSignals(False)

    def _available_selected_slots(self) -> set[str]:
        return {
            slot_key
            for slot_key, widget_data in self.slot_widgets.items()
            if list(widget_data.get("selected_paths", []))
        }

    def _apply_system_settings_availability(self) -> None:
        for section_key, section_widget in self.system_settings_sections.items():
            section_widget.setVisible(True)
            section_widget.setEnabled(True)
            if section_key in self.system_settings_unavailable_labels:
                self.system_settings_unavailable_labels[section_key].setVisible(False)
            section_widget.setToolTip("")

    def _build_issue_preview(self, issues: list) -> str:
        if not issues:
            return "ללא שגיאות"

        unique_messages: list[str] = []
        for issue in issues:
            if issue.message not in unique_messages:
                unique_messages.append(issue.message)
            if len(unique_messages) == 2:
                break

        preview = " | ".join(unique_messages)
        return preview if len(preview) <= 90 else f"{preview[:87]}..."

    @staticmethod
    def _is_intake_issue(issue: "ValidationIssue") -> bool:
        """Returns True only for structural/technical intake issues.

        Intake issues (shown in the intake log):
          - Missing required column  (row_number==0, message contains "עמודת חובה חסרה")
          - File structure mismatch  (row_number==0, message contains "אינו תואם למבנה")
          - Missing required column group (row_number==0, message contains "נדרשת לפחות")
          - Missing required value in a data row (row_number>0, message contains "ערך חובה חסר")

        NOT intake issues (audit/analysis findings shown in the analysis tab):
          - RSPARAM / TPFET policy violations
          - "לא נמצא פרמטר נדרש"
          - Any other business-logic row-level check
        """
        msg = issue.message
        if issue.row_number == 0:
            return (
                "עמודת חובה חסרה" in msg
                or "אינו תואם למבנה" in msg
                or "נדרשת לפחות" in msg
            )
        return "ערך חובה חסר" in msg

    @staticmethod
    def _compute_intake_summary(total_rows: int, intake_issues: list["ValidationIssue"]) -> tuple[int, int]:
        """Return (valid_rows, invalid_rows) for intake-only issues."""
        row_level_issues = {issue.row_number for issue in intake_issues if issue.row_number > 0}
        invalid_rows = len(row_level_issues)

        if any(issue.row_number == 0 for issue in intake_issues):
            invalid_rows = total_rows if total_rows else invalid_rows

        valid_rows = max(total_rows - invalid_rows, 0)
        return valid_rows, invalid_rows

    def _get_slot_category(self, slot_key: str) -> str:
        return str(self.SLOT_DEFINITIONS.get(slot_key, {}).get("sub_category", "לא סווג"))

    def _get_slot_display_name(self, slot_key: str) -> str:
        metadata = self.SLOT_DEFINITIONS.get(slot_key, {})
        return str(metadata.get("label", slot_key))

    def _get_slot_extraction_date(self, slot_key: str) -> str:
        widget_data = self.slot_widgets.get(slot_key, {})
        date_edit = widget_data.get("extraction_date_edit")
        if isinstance(date_edit, QLineEdit):
            date_text = date_edit.text().strip()
            return date_text or "לא צוין"
        return "לא צוין"

    @staticmethod
    def _is_a_dialog_user(user_type: object) -> bool:
        normalized_value = "" if user_type is None else str(user_type).strip().upper()
        return normalized_value in ValidationDesktopApp.USER_TYPE_RULES.get("Dialog", [])

    @staticmethod
    def _has_initial_password(password_flag: object) -> bool:
        normalized_value = "" if password_flag is None else str(password_flag).strip().upper()
        return normalized_value in {"X", "1", "TRUE", "YES", "Y"}

    def _is_user_locked(self, uflag: object) -> bool:
        normalized_value = "" if uflag is None else str(uflag).strip()
        if normalized_value in {"", "0", "00"}:
            return False
        try:
            return int(normalized_value) != 0
        except ValueError:
            return True

    def _validity_period_overlaps(
        self,
        gltgv_value: object,
        gltgb_value: object,
        period_start: date | None,
        period_end: date | None,
    ) -> bool:
        if period_start is None or period_end is None:
            return False

        gltgv = self._parse_user_preview_date(gltgv_value)
        gltgb = self._parse_user_preview_date(gltgb_value)
        if gltgv is None and gltgb is None:
            return False

        valid_from = gltgv.date() if gltgv is not None else datetime.min.date()
        valid_to = gltgb.date() if gltgb is not None else datetime.max.date()
        return valid_from <= period_end and valid_to >= period_start

    def _is_user_active_in_period(
        self,
        usr_entry: dict[str, str],
        period_start: date | None,
        period_end: date | None,
    ) -> bool:
        if period_start is None or period_end is None:
            return False

        last_login_date = self._parse_user_preview_date(usr_entry.get("TRDAT", ""))
        if last_login_date is not None and period_start <= last_login_date.date() <= period_end:
            return True

        return self._validity_period_overlaps(
            usr_entry.get("GLTGV", ""),
            usr_entry.get("GLTGB", ""),
            period_start,
            period_end,
        )

    def _is_generic_user(self, bname: object, settings: dict[str, Any]) -> bool:
        if not bname or not isinstance(settings, dict):
            return False
        normalized_name = str(bname).strip().casefold()
        generic_users = settings.get("generic_users", [])
        if not isinstance(generic_users, list):
            return False
        return any(
            fnmatch.fnmatch(normalized_name, str(item).strip().casefold())
            for item in generic_users
        )

    def _is_super_user(self, mandt: object, bname: object, settings: dict[str, Any]) -> bool:
        if not bname or not isinstance(settings, dict):
            return False
        normalized_mandt = str(mandt).strip()
        normalized_bname = str(bname).strip().casefold()
        super_users = settings.get("authorized_stms_users", [])
        if not super_users:
            super_users = settings.get("super_users", [])
        if not isinstance(super_users, list):
            return False
        for row in super_users:
            if not isinstance(row, dict):
                continue
            if str(row.get("MANDT", "")).strip() == normalized_mandt and str(row.get("BNAME", "")).strip().casefold() == normalized_bname:
                return True
        return False

    def _is_developer_user(self, mandt: object, bname: object, settings: dict[str, Any]) -> bool:
        if not bname or not isinstance(settings, dict):
            return False
        normalized_mandt = str(mandt).strip()
        normalized_bname = str(bname).strip().casefold()
        dev_users = settings.get("authorized_developers", [])
        if not isinstance(dev_users, list):
            return False
        for row in dev_users:
            if not isinstance(row, dict):
                continue
            if str(row.get("MANDT", "")).strip() == normalized_mandt and str(row.get("BNAME", "")).strip().casefold() == normalized_bname:
                return True
        return False

    def _build_user_findings_description(self, usr_entry: dict[str, str], extraction_date_text: str) -> str:
        findings: list[str] = []
        settings = self._current_system_settings()
        period_cfg = settings.get("user_review_period", {}) if isinstance(settings, dict) else {}
        start_date = self._parse_user_preview_date(period_cfg.get("start_date", ""))
        end_date = self._parse_user_preview_date(period_cfg.get("end_date", ""))
        period_start = start_date.date() if start_date is not None else None
        period_end = end_date.date() if end_date is not None else None

        is_super_user = self._is_super_user(usr_entry.get("MANDT", ""), usr_entry.get("BNAME", ""), settings)
        is_generic_user = self._is_generic_user(usr_entry.get("BNAME", ""), settings)
        is_locked = self._is_user_locked(usr_entry.get("UFLAG", ""))
        is_active = self._is_user_active_in_period(usr_entry, period_start, period_end)

        work_environment_raw = usr_entry.get("WORK_ENVIRONMENT", "")
        work_environment = work_environment_raw.strip().upper()
        mandt_val = usr_entry.get("MANDT", "")
        bname_val = usr_entry.get("BNAME", "")
        is_developer = self._is_developer_user(mandt_val, bname_val, settings)
        if is_developer and work_environment.startswith("FPP"):
            findings.append("המשתמש הוא מפתח בסביבת ייצור - יש לבחון הרשאות ומומלץ להסיר אותו מהסביבה")

        if (is_super_user or is_generic_user) and not is_locked and is_active:
            if is_super_user and is_generic_user:
                findings.append("משתמש על / גנרי פעיל ולא נעול")
            elif is_super_user:
                findings.append("משתמש על פעיל ולא נעול")
            else:
                findings.append("משתמש גנרי פעיל ולא נעול")

        if self._is_a_dialog_user(usr_entry.get("USTYP", "")):
            inactivity_threshold = self._safe_int(settings.get("inactive_days_threshold", 90), 90)
            password_policy = settings.get("password_policy_defaults", {}) if isinstance(settings, dict) else {}
            max_password_age_days = self._safe_int(
                password_policy.get("max_password_age_days", 90) if isinstance(password_policy, dict) else 90,
                90,
            )
            initial_password_change_max_days = self._safe_int(
                password_policy.get("initial_password_change_max_days", 2) if isinstance(password_policy, dict) else 2,
                2,
            )

            extraction_date = self._parse_user_preview_date(extraction_date_text)
            last_login_date = self._parse_user_preview_date(usr_entry.get("TRDAT", ""))
            password_change_date = self._parse_user_preview_date(usr_entry.get("PWDCHGDATE", ""))
            password_set_date = self._parse_user_preview_date(usr_entry.get("PWDSETDATE", ""))

            if extraction_date is not None and last_login_date is not None:
                inactivity_days = (extraction_date.date() - last_login_date.date()).days
                if inactivity_days > inactivity_threshold:
                    findings.append(f"משתמש לא פעיל מעל {inactivity_threshold} יום")

            if last_login_date is not None and password_change_date is not None:
                password_gap_days = (last_login_date.date() - password_change_date.date()).days
                if password_gap_days > max_password_age_days:
                    findings.append(f"סיסמה לא הוחלפה מעל {max_password_age_days} יום")

            if extraction_date is not None and password_set_date is not None and self._has_initial_password(usr_entry.get("PWDINITIAL", "")):
                initial_password_age_days = (extraction_date.date() - password_set_date.date()).days
                if initial_password_age_days > initial_password_change_max_days:
                    if initial_password_change_max_days == 2:
                        findings.append("סיסמה ראשונית לא הוחלפה תוך 48 שעות")
                    else:
                        findings.append(f"סיסמה ראשונית לא הוחלפה תוך {initial_password_change_max_days} ימים")

            security_policy = str(usr_entry.get("SECURITY_POLICY", "")).strip()
            if security_policy:
                findings.append("עולה חשד שלמשתמש הוגדרה מדיניות סיסמה מיוחדת - ודא בטרנזקציית 'SEC_POLICY' שהמדיניות של משתמש זה אינה כוללת החרגות סיסמה")

        ustyp = str(usr_entry.get("USTYP", "")).strip().upper()
        if ustyp in ("B", "S"):
            if is_generic_user:
                findings.append("משתמש גנרי - יש לוודא המשתמש נעול ולא היה פעיל במהלך השנה ללא אישור מתועד")
            else:
                findings.append("יש לוודא צורך במשתמש ולנעול כשאין צורך, לוודא שהמשתמש הוא אכן משתמש מערכת ולא אנושי")

        return " | ".join(findings)

    @staticmethod
    def _export_sort_key(preview_row: dict[str, str]) -> int:
        review_status = preview_row.get("REVIEW_STATUS", "").strip()
        has_findings = bool(preview_row.get("FINDINGS_DESCRIPTION", "").strip())
        if review_status == "טרם נבדק":
            return 1
        if review_status == "נבדק - לא תקין":
            return 2
        if review_status == "נבדק - תקין" and has_findings:
            return 3
        return 4

    @staticmethod
    def _has_review_note(technical_note: object, business_note: object) -> bool:
        return has_review_note(technical_note, business_note)

    def _is_user_review_complete(
        self,
        review_status: object,
        findings_description: object,
        technical_note: object,
        business_note: object,
    ) -> bool:
        return is_user_review_complete(
            review_status,
            findings_description,
            technical_note,
            business_note,
            self.REVIEWED_STATUSES,
            self.REVIEW_STATUS_OPTIONS,
            self.DEFAULT_REVIEW_STATUS,
        )

    def _load_all_user_preview_rows(self) -> list[dict[str, str]]:
        usr02_rows = self._load_preview_rows("USR02")
        combined_rows = self._load_preview_rows("ADR6_USR21")
        return self._build_user_preview_rows(usr02_rows, combined_rows)

    def _get_user_review_completion_snapshot(self) -> tuple[list[dict[str, str]], int, list[dict[str, str]]]:
        preview_rows = self._load_all_user_preview_rows()
        incomplete_rows: list[dict[str, str]] = []
        reviewed_rows = 0
        for preview_row in preview_rows:
            if self._is_user_review_complete(
                preview_row.get("REVIEW_STATUS", ""),
                preview_row.get("FINDINGS_DESCRIPTION", ""),
                preview_row.get("TECH_REVIEW_NOTES", ""),
                preview_row.get("BUS_REVIEW_NOTES", ""),
            ):
                reviewed_rows += 1
            else:
                incomplete_rows.append(preview_row)
        return preview_rows, reviewed_rows, incomplete_rows

    def _build_user_review_incomplete_reason(self, preview_row: dict[str, str]) -> str:
        return build_user_review_incomplete_reason(
            preview_row.get("REVIEW_STATUS", ""),
            preview_row.get("FINDINGS_DESCRIPTION", ""),
            self.REVIEWED_STATUSES,
            self.REVIEW_STATUS_OPTIONS,
            self.DEFAULT_REVIEW_STATUS,
        )

    def _update_review_row_highlight(self, row_index: int, preview_row: dict[str, str] | None = None) -> None:
        review_status_col: int | None = None
        technical_notes_col: int | None = None
        business_notes_col: int | None = None
        for col_idx, field_name in enumerate(self.user_preview_visible_columns):
            if field_name == "REVIEW_STATUS":
                review_status_col = col_idx
            elif field_name in {"TECH_REVIEW_NOTES", "REVIEW_NOTES"}:
                technical_notes_col = col_idx
            elif field_name == "BUS_REVIEW_NOTES":
                business_notes_col = col_idx

        if preview_row is not None:
            review_status_text = str(preview_row.get("REVIEW_STATUS", "")).strip()
            findings_text = str(preview_row.get("FINDINGS_DESCRIPTION", "")).strip()
            technical_notes_text = str(preview_row.get("TECH_REVIEW_NOTES", "")).strip()
            business_notes_text = str(preview_row.get("BUS_REVIEW_NOTES", "")).strip()
        else:
            review_status_text = ""
            if review_status_col is not None:
                combo = self.user_preview_table.cellWidget(row_index, review_status_col)
                if isinstance(combo, QComboBox):
                    review_status_text = self.format_rtl_text(combo.currentText())
                else:
                    status_item = self.user_preview_table.item(row_index, review_status_col)
                    if status_item is not None:
                        review_status_text = status_item.text().strip()

            findings_text = ""
            try:
                findings_col = self.user_preview_visible_columns.index("FINDINGS_DESCRIPTION")
            except ValueError:
                findings_col = -1
            if findings_col >= 0:
                findings_item = self.user_preview_table.item(row_index, findings_col)
                if findings_item is not None:
                    findings_text = findings_item.text().strip()

            technical_notes_text = ""
            if technical_notes_col is not None:
                notes_item = self.user_preview_table.item(row_index, technical_notes_col)
                if notes_item is not None:
                    technical_notes_text = notes_item.text().strip()

            business_notes_text = ""
            if business_notes_col is not None:
                business_notes_item = self.user_preview_table.item(row_index, business_notes_col)
                if business_notes_item is not None:
                    business_notes_text = business_notes_item.text().strip()

        is_not_reviewed = review_status_text == "טרם נבדק"
        is_review_complete = self._is_user_review_complete(
            review_status_text,
            findings_text,
            technical_notes_text,
            business_notes_text,
        )
        needs_warning = (not is_not_reviewed) and (not is_review_complete)

        unreviewed_color = QColor("#d6e8ff")
        warning_color = QColor("#fff0c2")
        clear_color = QColor(0, 0, 0, 0)

        for col_idx, field_name in enumerate(self.user_preview_visible_columns):
            combo = self.user_preview_table.cellWidget(row_index, col_idx)
            item = self.user_preview_table.item(row_index, col_idx)

            if is_not_reviewed:
                if isinstance(combo, QComboBox):
                    combo.setStyleSheet("background-color: #d6e8ff;")
                elif item is not None:
                    item.setBackground(unreviewed_color)
                continue

            if needs_warning and field_name in {"REVIEW_STATUS", "TECH_REVIEW_NOTES", "BUS_REVIEW_NOTES", "REVIEW_NOTES"}:
                if isinstance(combo, QComboBox):
                    combo.setStyleSheet("background-color: #fff0c2;")
                elif item is not None:
                    item.setBackground(warning_color)
                continue

            if isinstance(combo, QComboBox):
                combo.setStyleSheet("")
            elif item is not None:
                if field_name == "STATUS":
                    status_value = item.text()
                    if status_value == "פעיל":
                        item.setBackground(QColor("#eaf7ea"))
                    elif "אי-התאמה" in status_value:
                        item.setBackground(QColor("#fff4cc"))
                    elif status_value == "נעול":
                        item.setBackground(QColor("#fdecec"))
                    else:
                        item.setBackground(clear_color)
                else:
                    item.setBackground(clear_color)

    def _get_user_preview_row_review_status(self, row_index: int) -> str:
        try:
            review_status_col = self.user_preview_visible_columns.index("REVIEW_STATUS")
        except ValueError:
            return self.DEFAULT_REVIEW_STATUS

        combo = self.user_preview_table.cellWidget(row_index, review_status_col)
        if isinstance(combo, QComboBox):
            return self._normalize_reviewer_status(combo.currentText())

        item = self.user_preview_table.item(row_index, review_status_col)
        if item is not None:
            return self._normalize_reviewer_status(item.text())
        return self.DEFAULT_REVIEW_STATUS

    def _update_user_review_progress_summary(self, total: int, reviewed: int, unreviewed: int) -> None:
        total = max(0, int(total))
        reviewed = max(0, int(reviewed))
        unreviewed = max(0, int(unreviewed))
        if reviewed + unreviewed != total:
            unreviewed = max(0, total - reviewed)

        percent_complete = int(round((reviewed / total) * 100)) if total > 0 else 0

        self.user_review_total_label.setText(self.format_ui_rtl_text(f"סה\"כ משתמשים בדוח: {total}"))
        self.user_review_reviewed_label.setText(self.format_ui_rtl_text(f"משתמשים שנבדקו: {reviewed}"))
        self.user_review_unreviewed_label.setText(self.format_ui_rtl_text(f"משתמשים שטרם נבדקו: {unreviewed}"))
        self.user_review_progress_percent_label.setText(
            self.format_ui_rtl_text(f"התקדמות השלמת סקירה: {percent_complete}%")
        )

        self.user_review_progress_bar.setMaximum(max(total, 1))
        self.user_review_progress_bar.setValue(min(reviewed, max(total, 1)))
        self.user_review_progress_bar.setFormat(f"{percent_complete}%")

    def _refresh_user_review_progress_summary_from_table(self) -> None:
        preview_rows, reviewed_rows, incomplete_rows = self._get_user_review_completion_snapshot()
        total_rows = len(preview_rows)
        self._update_user_review_progress_summary(total_rows, reviewed_rows, len(incomplete_rows))

    def _update_slot_path_label(self, slot_key: str, file_paths: list[str] | None = None) -> None:
        widget_data = self.slot_widgets.get(slot_key, {})
        label = widget_data.get("path_label")
        if not isinstance(label, QLabel):
            return

        # Apply stylesheet indicator first: Qt's setStyleSheet resets QLabel
        # alignment to the parent's RTL default, so we must call it before
        # setting alignment and text.
        self._update_slot_ipe_indicator(slot_key)

        paths = file_paths if file_paths is not None else list(widget_data.get("selected_paths", []))
        if not paths:
            label.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
            label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            label.setText(self.format_ui_rtl_text("טרם נבחר קובץ"))
        elif len(paths) == 1:
            label.setLayoutDirection(Qt.LayoutDirection.LeftToRight)
            label.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            label.setText(self.format_rtl_text(paths[0]))
        else:
            label.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
            label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            label.setText(self.format_ui_rtl_text(self._format_selected_files(paths)))

    def _remember_slot_load(self, slot_key: str) -> None:
        self.load_history = [item for item in self.load_history if item != slot_key]
        if list(self.slot_widgets.get(slot_key, {}).get("selected_paths", [])):
            self.load_history.append(slot_key)

    def clear_slot_selection(self, slot_key: str) -> None:
        if slot_key not in self.slot_widgets:
            return

        self.slot_widgets[slot_key]["selected_paths"] = []
        self._update_slot_path_label(slot_key, [])
        self.load_history = [item for item in self.load_history if item != slot_key]

        if self.selected_slot_key == slot_key:
            self.selected_slot_key = self.load_history[-1] if self.load_history else None
            if self.selected_slot_key:
                self.required_columns_edit.setText(self._suggest_required_columns(self.selected_slot_key))
            else:
                self.required_columns_edit.setText("")

        if slot_key in self.USER_PREVIEW_SLOTS:
            self.refresh_user_preview()
        self._apply_system_settings_availability()
        self._update_slot_ipe_indicator(slot_key)

    def clear_last_loaded_slot(self) -> None:
        while self.load_history:
            last_slot_key = self.load_history[-1]
            if list(self.slot_widgets.get(last_slot_key, {}).get("selected_paths", [])):
                self.clear_slot_selection(last_slot_key)
                return
            self.load_history.pop()

    # ------------------------------------------------------------------
    # IPE Evidence helpers
    # ------------------------------------------------------------------

    def _add_ipe_evidence(self, slot_key: str) -> None:
        """Open a file picker for images and let the user tag controls."""
        image_filter = "Images (*.png *.jpg *.jpeg *.bmp *.tiff *.tif *.gif);;All files (*.*)"
        file_paths, _ = QFileDialog.getOpenFileNames(
            self, f"בחירת תמונות ראיה עבור {slot_key}", "", image_filter
        )
        if not file_paths:
            return

        for path_str in file_paths:
            source = Path(path_str)
            control_ids = list(self._slot_control_mapping.get(slot_key, []))
            entry = self.ipe_repository.add_image(slot_key, source, control_ids, self.ipe_evidence_data)
            self._append_thumbnail(slot_key, entry)
            self._update_slot_ipe_indicator(slot_key)

    def _remove_ipe_evidence(self, slot_key: str, image_id: str) -> None:
        """Remove a single IPE image from the slot."""
        self.ipe_repository.remove_image(slot_key, image_id, self.ipe_evidence_data)
        self._refresh_slot_thumbnails(slot_key)
        self._update_slot_ipe_indicator(slot_key)

    def _append_thumbnail(self, slot_key: str, entry: dict[str, Any]) -> None:
        """Add a single thumbnail widget to the slot's thumbnail strip."""
        widget_data = self.slot_widgets.get(slot_key)
        if widget_data is None:
            return
        thumb_layout: QHBoxLayout = widget_data["thumb_layout"]
        thumb_scroll: QScrollArea = widget_data["thumb_scroll"]

        thumb_widget = self._build_thumbnail_widget(slot_key, entry)
        # Insert before the trailing stretch (last item)
        count = thumb_layout.count()
        thumb_layout.insertWidget(count - 1, thumb_widget)

        thumb_scroll.setVisible(True)

    def _build_thumbnail_widget(self, slot_key: str, entry: dict[str, Any]) -> QWidget:
        """Return a small widget with an image preview and a remove button."""
        image_id: str = entry.get("id", "")
        stored_path = entry.get("stored_path", "")
        filename = entry.get("original_filename", stored_path)
        control_ids: list[str] = entry.get("control_ids", [])

        container = QWidget()
        container.setFixedSize(90, 82)
        container.setToolTip(
            f"{filename}\nבקרות: {', '.join(control_ids) if control_ids else '—'}"
        )
        layout = QVBoxLayout(container)
        layout.setContentsMargins(2, 2, 2, 2)
        layout.setSpacing(2)

        thumb_label = QLabel()
        thumb_label.setFixedSize(82, 62)
        thumb_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        thumb_label.setStyleSheet("border: 1px solid #90caf9; background: #ffffff;")
        pixmap = QPixmap(stored_path)
        if pixmap.isNull():
            thumb_label.setText("🖼")
        else:
            thumb_label.setPixmap(pixmap.scaled(82, 62, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
        layout.addWidget(thumb_label)

        remove_btn = QPushButton("✕")
        remove_btn.setFixedHeight(16)
        remove_btn.setStyleSheet(
            "QPushButton { font-size: 10px; color: #c62828; border: none; background: transparent; }"
            "QPushButton:hover { color: #b71c1c; text-decoration: underline; }"
        )
        remove_btn.setToolTip("הסר ראיה")
        remove_btn.clicked.connect(lambda _checked=False, sk=slot_key, iid=image_id: self._remove_ipe_evidence(sk, iid))
        layout.addWidget(remove_btn, 0, Qt.AlignmentFlag.AlignHCenter)

        return container

    def _refresh_slot_thumbnails(self, slot_key: str) -> None:
        """Rebuild the thumbnail strip for *slot_key* from current evidence data."""
        widget_data = self.slot_widgets.get(slot_key)
        if widget_data is None:
            return
        thumb_layout: QHBoxLayout = widget_data["thumb_layout"]
        thumb_scroll: QScrollArea = widget_data["thumb_scroll"]

        # Remove all existing thumbnail widgets (keep the trailing stretch)
        while thumb_layout.count() > 1:
            item = thumb_layout.takeAt(0)
            if item is not None:
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()

        entries = self.ipe_evidence_data.get(slot_key, [])
        for entry in entries:
            thumb_widget = self._build_thumbnail_widget(slot_key, entry)
            thumb_layout.insertWidget(thumb_layout.count() - 1, thumb_widget)

        thumb_scroll.setVisible(bool(entries))

    def _populate_all_slot_thumbnails(self) -> None:
        """Called once after _build_ui to restore thumbnails from persisted evidence."""
        for slot_key in self.slot_widgets:
            if self.ipe_evidence_data.get(slot_key):
                self._refresh_slot_thumbnails(slot_key)
            self._update_slot_ipe_indicator(slot_key)

    def choose_file(self, slot_key: str) -> None:
        initial_directory = self._get_last_file_dialog_directory()
        if slot_key in self.MULTI_FILE_SLOTS:
            file_paths, _ = QFileDialog.getOpenFileNames(
                self,
                f"בחירת קבצים עבור {slot_key}",
                initial_directory,
                "Supported files (*.txt *.csv *.xlsx *.xlsm);;All files (*.*)",
            )
        else:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                f"בחירת קובץ עבור {slot_key}",
                initial_directory,
                "Supported files (*.txt *.csv *.xlsx *.xlsm);;All files (*.*)",
            )
            file_paths = [file_path] if file_path else []

        if file_paths:
            self._save_last_file_dialog_directory(Path(file_paths[0]).parent)
            self.selected_slot_key = slot_key
            self.slot_widgets[slot_key]["selected_paths"] = file_paths
            self._remember_slot_load(slot_key)
            self._update_slot_path_label(slot_key, file_paths)
            self.required_columns_edit.setText(self._suggest_required_columns(slot_key))
            if slot_key in self.USER_PREVIEW_SLOTS:
                self.refresh_user_preview()
            self._apply_system_settings_availability()

    def _parse_required_columns(self, raw_value: str | None = None) -> list[str]:
        value = self.required_columns_edit.text() if raw_value is None else raw_value
        normalized_value = value.replace(";", ",").replace("\n", ",")
        return [item.strip() for item in normalized_value.split(",") if item.strip()]

    def _required_columns_for_slot(self, slot_key: str) -> list[str]:
        if self.selected_slot_key == slot_key and self.required_columns_edit.text().strip():
            return self._parse_required_columns()
        return self._parse_required_columns(self._suggest_required_columns(slot_key))

    def _get_category_slots(self, sub_category: str) -> list[str]:
        return [
            slot_key
            for slot_key, metadata in self.SLOT_DEFINITIONS.items()
            if metadata.get("sub_category") == sub_category
        ]

    def _get_domain_slots(self, domain: str) -> list[str]:
        return [
            slot_key
            for slot_key, metadata in self.SLOT_DEFINITIONS.items()
            if metadata.get("domain") == domain
        ]

    def _current_file_paths(self) -> list[str]:
        if not self.selected_slot_key:
            return []
        return list(self.slot_widgets[self.selected_slot_key].get("selected_paths", []))

    def _build_input_files_dict(
        self,
        preferred_slot_key: str | None = None,
        fallback_file_paths: Sequence[str | Path] | None = None,
    ) -> dict[str, list[str | Path]]:
        """Collect selected input files from all slot widgets for pipeline execution."""
        input_files: dict[str, list[str | Path]] = {}

        if preferred_slot_key and fallback_file_paths:
            normalized_fallback: list[str | Path] = [
                str(path).strip()
                for path in fallback_file_paths
                if str(path).strip()
            ]
            if normalized_fallback:
                input_files[preferred_slot_key] = normalized_fallback

        for slot_key, widget_data in self.slot_widgets.items():
            selected_paths: list[str | Path] = []
            raw_paths = widget_data.get("selected_paths", [])

            if isinstance(raw_paths, str):
                if raw_paths.strip():
                    selected_paths = [raw_paths.strip()]
            elif isinstance(raw_paths, (list, tuple, set)):
                selected_paths = [str(path).strip() for path in raw_paths if str(path).strip()]

            if selected_paths:
                input_files[slot_key] = selected_paths

        return input_files

    def _format_selected_files(self, file_paths: list[str]) -> str:
        if len(file_paths) == 1:
            return self.format_rtl_text(file_paths[0])

        preview_names = [Path(path).name for path in file_paths[:3]]
        suffix = "" if len(file_paths) <= 3 else " ..."
        return self.format_rtl_text(
            f"נבחרו {len(file_paths)} קבצים: {', '.join(preview_names)}{suffix}"
        )

    def _suggest_required_columns(self, slot_key: str) -> str:
        suggestions = {
            "USR02": "BNAME,UFLAG,TRDAT,LTIME",
            "ADR6_USR21": "",
            "AGR_USERS": "AGR_NAME,UNAME",
            "AGR_1251": "AGR_NAME,OBJECT,FIELD",
            "AGR_1252": "AGR_NAME,LOW",
            "AGR_DEFINE": "AGR_NAME,PARENT_AGR",
            "UST04": "BNAME,PROFILE",
            "E070": "TRKORR,AS4USER,TRFUNCTION",
            "T000": "MANDT,CCCATEGORY",
            "STMS": "TRKORR",
            "RSPARAM": "PARAMETER,VALUE",
            "TPFET": "PARAMETER,VALUE",
        }
        return suggestions.get(slot_key, "")

    def _file_dialog_state_path(self) -> Path:
        return self.ui_state_repository.file_dialog_state_path()

    def _load_last_file_dialog_directory(self) -> Path:
        return self.ui_state_repository.load_last_file_dialog_directory(self._allow_user_preview_persistence)

    def _save_last_file_dialog_directory(self, directory_path: object) -> None:
        saved_directory = self.ui_state_repository.save_last_file_dialog_directory(
            directory_path,
            self._allow_user_preview_persistence,
        )
        if saved_directory is not None:
            self.last_file_dialog_directory = saved_directory

    def _get_last_file_dialog_directory(self) -> str:
        candidate_directory = getattr(self, "last_file_dialog_directory", self.config.input_dir)
        if not isinstance(candidate_directory, Path) or not candidate_directory.exists() or not candidate_directory.is_dir():
            candidate_directory = self.config.input_dir
        return str(candidate_directory)

    def _user_preview_settings_path(self) -> Path:
        return self.ui_state_repository.user_preview_settings_path()

    def _user_reviewer_state_path(self) -> Path:
        return self.ui_state_repository.user_reviewer_state_path()

    @staticmethod
    def _user_reviewer_state_key(mandt: object, bname: object) -> str:
        return reviewer_state_key(mandt, bname)

    @classmethod
    def _normalize_reviewer_status(cls, value: object) -> str:
        return normalize_reviewer_status(value, cls.REVIEW_STATUS_OPTIONS, cls.DEFAULT_REVIEW_STATUS)

    @classmethod
    def _default_reviewer_values(cls) -> dict[str, str]:
        return default_reviewer_values(cls.DEFAULT_REVIEW_STATUS)

    def _load_user_reviewer_state(self) -> dict[str, dict[str, str]]:
        return self.ui_state_repository.load_user_reviewer_state(
            self._allow_user_preview_persistence,
            self._normalize_reviewer_status,
        )

    def _save_user_reviewer_state(self) -> None:
        self.ui_state_repository.save_user_reviewer_state(
            self._allow_user_preview_persistence,
            self.user_reviewer_state,
        )

    def _get_reviewer_values(self, mandt: object, bname: object) -> dict[str, str]:
        review_key = self._user_reviewer_state_key(mandt, bname)
        stored_values = self.user_reviewer_state.get(review_key)
        if not isinstance(stored_values, dict):
            return self._default_reviewer_values().copy()
        legacy_notes = str(stored_values.get("REVIEW_NOTES", "")).strip()
        return {
            "REVIEW_STATUS": self._normalize_reviewer_status(stored_values.get("REVIEW_STATUS")),
            "TECH_REVIEW_NOTES": str(stored_values.get("TECH_REVIEW_NOTES", "")).strip() or legacy_notes,
            "BUS_REVIEW_NOTES": str(stored_values.get("BUS_REVIEW_NOTES", "")).strip(),
            "LAST_IMPORT_DATE": str(stored_values.get("LAST_IMPORT_DATE", "")).strip(),
        }

    def _update_reviewer_value(self, review_key: str, field_name: str, value: object) -> None:
        normalized_field = normalize_review_field(field_name)
        if not review_key or normalized_field not in {"REVIEW_STATUS", "TECH_REVIEW_NOTES", "BUS_REVIEW_NOTES"}:
            return

        current_values = self.user_reviewer_state.setdefault(review_key, self._default_reviewer_values().copy())
        if normalized_field == "REVIEW_STATUS":
            current_values[normalized_field] = self._normalize_reviewer_status(value)
        else:
            current_values[normalized_field] = "" if value is None else str(value).strip()
        self._save_user_reviewer_state()

    def _normalize_user_preview_columns(self, selected_columns: list[str] | None) -> list[str]:
        allowed_fields = [column["field"] for column in self.USER_PREVIEW_COLUMN_DEFINITIONS]
        if not selected_columns:
            return list(self.DEFAULT_USER_PREVIEW_COLUMNS)

        normalized_input = ["TECH_REVIEW_NOTES" if field == "REVIEW_NOTES" else field for field in selected_columns]
        normalized = [field for field in allowed_fields if field in normalized_input]
        return normalized or list(self.DEFAULT_USER_PREVIEW_COLUMNS)

    def _load_user_preview_column_selection(self) -> list[str]:
        return self.ui_state_repository.load_user_preview_column_selection(
            self._allow_user_preview_persistence,
            list(self.DEFAULT_USER_PREVIEW_COLUMNS),
            self.CURRENT_USER_PREVIEW_SETTINGS_VERSION,
            self.USER_PREVIEW_SETTINGS_MIGRATIONS,
            self._normalize_user_preview_columns,
        )

    def _save_user_preview_column_selection(self) -> None:
        self.ui_state_repository.save_user_preview_column_selection(
            self._allow_user_preview_persistence,
            self.CURRENT_USER_PREVIEW_SETTINGS_VERSION,
            self.user_preview_visible_columns,
        )

    def _load_user_preview_column_filters(self) -> dict[str, set[str]]:
        return self.ui_state_repository.load_user_preview_column_filters(
            self._allow_user_preview_persistence,
        )

    def _persist_user_preview_column_filters(self) -> None:
        self.ui_state_repository.save_user_preview_column_filters(
            self._allow_user_preview_persistence,
            self.user_preview_column_filters,
        )

    def _clear_all_user_preview_column_filters(self) -> None:
        """Remove all active per-column filters and refresh the user preview table."""
        self.user_preview_column_filters.clear()
        self._persist_user_preview_column_filters()
        self._update_header_filter_icons()
        self.refresh_user_preview()

    def _get_user_preview_column_definition(self, field_name: str) -> dict[str, Any]:
        for column in self.USER_PREVIEW_COLUMN_DEFINITIONS:
            if column["field"] == field_name:
                return column
        return {"field": field_name, "formal": field_name, "technical": field_name, "source": "לא ידוע", "width": 120}

    def _handle_user_preview_item_changed(self, item: QTableWidgetItem) -> None:
        if self._refreshing_user_preview or item is None:
            return

        field_name = item.data(Qt.ItemDataRole.UserRole + 1)
        review_key = item.data(Qt.ItemDataRole.UserRole)
        if field_name not in {"TECH_REVIEW_NOTES", "BUS_REVIEW_NOTES", "REVIEW_NOTES"} or not review_key:
            return

        normalized_text = self.format_rtl_text(item.text())
        item.setToolTip(normalized_text)
        self._update_reviewer_value(str(review_key), str(field_name), normalized_text)
        self._update_review_row_highlight(item.row())

    def _get_user_preview_cell_text(self, row_index: int, column_index: int) -> str:
        cell_widget = self.user_preview_table.cellWidget(row_index, column_index)
        if isinstance(cell_widget, QComboBox):
            return self.format_rtl_text(cell_widget.currentText())

        item = self.user_preview_table.item(row_index, column_index)
        if item is None:
            return ""
        return self.format_rtl_text(item.text())

    def _configure_user_preview_table(self) -> None:
        self.user_preview_visible_columns = self._normalize_user_preview_columns(self.user_preview_visible_columns)
        self.user_preview_table.setColumnCount(len(self.user_preview_visible_columns))

        # Install the filterable header once (on first configure call)
        if not isinstance(self.user_preview_table.horizontalHeader(), _FilterableHeaderView):
            filterable_header = _FilterableHeaderView(Qt.Orientation.Horizontal, self.user_preview_table)
            self.user_preview_table.setHorizontalHeader(filterable_header)
            filterable_header.filterRequested.connect(self._on_user_preview_header_filter_requested)

        self.user_preview_table.setHorizontalHeaderLabels([
            str(self._get_user_preview_column_definition(field_name).get("formal", field_name))
            for field_name in self.user_preview_visible_columns
        ])
        header = self.user_preview_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        header.setSectionsMovable(False)
        header.setSectionsClickable(True)
        header.setMinimumSectionSize(70)
        self.user_preview_table.verticalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        for column_index, field_name in enumerate(self.user_preview_visible_columns):
            header.setSectionResizeMode(column_index, QHeaderView.ResizeMode.Interactive)
            default_width = int(self._get_user_preview_column_definition(field_name).get("width", 120))
            self.user_preview_table.setColumnWidth(column_index, default_width)
        self.user_preview_table.setSortingEnabled(True)
        self._update_header_filter_icons()

    def _update_header_filter_icons(self) -> None:
        """Refresh the ▼ icon state (active/inactive) on the filterable header."""
        header = self.user_preview_table.horizontalHeader()
        if not isinstance(header, _FilterableHeaderView):
            return
        active_columns: set[int] = set()
        for column_index, field_name in enumerate(self.user_preview_visible_columns):
            if field_name in self.user_preview_column_filters and self.user_preview_column_filters[field_name]:
                active_columns.add(column_index)
        header.set_active_filter_sections(active_columns)

    def _create_user_preview_columns_dialog(self) -> tuple[QDialog, QTableWidget]:
        dialog = QDialog(self)
        dialog.setWindowTitle(self.format_rtl_text("בחירת עמודות לסקירת משתמשים"))
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(720, 460)

        layout = QVBoxLayout(dialog)
        hint_label = QLabel(
            self.format_ui_rtl_text("סמן את העמודות שברצונך להציג. לחיצה על OK תרענן את הטבלה, ו-Cancel תשאיר את המצב הקיים.")
        )
        hint_label.setWordWrap(True)
        hint_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)
        layout.addWidget(hint_label)

        selection_table = QTableWidget(len(self.USER_PREVIEW_COLUMN_DEFINITIONS), 3)
        selection_table.setHorizontalHeaderLabels(["שם פורמלי", "שם טכני", "הצג"])
        selection_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        selection_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        selection_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        selection_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        selection_table.verticalHeader().setVisible(False)
        selection_table.setAlternatingRowColors(True)
        selection_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

        for row_index, column in enumerate(self.USER_PREVIEW_COLUMN_DEFINITIONS):
            formal_item = QTableWidgetItem(self.format_rtl_text(str(column["formal"])))
            formal_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            formal_item.setToolTip(self.format_ui_rtl_text(f"מקור נתון: {column['source']}"))
            technical_item = QTableWidgetItem(str(column["technical"]))
            technical_item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            technical_item.setToolTip(self.format_ui_rtl_text(f"מקור נתון: {column['source']}"))
            checkbox_item = QTableWidgetItem("")
            checkbox_item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsUserCheckable)
            checkbox_item.setCheckState(Qt.CheckState.Checked if column["field"] in self.user_preview_visible_columns else Qt.CheckState.Unchecked)
            checkbox_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            checkbox_item.setToolTip(self.format_ui_rtl_text(f"מקור נתון: {column['source']}"))
            selection_table.setItem(row_index, 0, formal_item)
            selection_table.setItem(row_index, 1, technical_item)
            selection_table.setItem(row_index, 2, checkbox_item)

        layout.addWidget(selection_table)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        return dialog, selection_table

    def _get_selected_user_preview_columns(self, selection_table: QTableWidget) -> list[str]:
        selected_columns: list[str] = []
        for row_index, column in enumerate(self.USER_PREVIEW_COLUMN_DEFINITIONS):
            checkbox_item = selection_table.item(row_index, 2)
            if checkbox_item is not None and checkbox_item.checkState() == Qt.CheckState.Checked:
                selected_columns.append(str(column["field"]))
        return selected_columns

    def _apply_user_preview_columns(self, selected_columns: list[str]) -> None:
        normalized_columns = self._normalize_user_preview_columns(selected_columns)
        if not normalized_columns:
            QMessageBox.warning(self, "בחירת עמודות", "יש לבחור לפחות עמודה אחת להצגה בטבלת הסקירה.")
            return

        self.user_preview_visible_columns = normalized_columns
        self._save_user_preview_column_selection()

        # Prune column filters for columns no longer visible
        pruned = {k: v for k, v in self.user_preview_column_filters.items() if k in normalized_columns}
        if pruned != self.user_preview_column_filters:
            self.user_preview_column_filters = pruned
            self._persist_user_preview_column_filters()

        self._configure_user_preview_table()
        self.refresh_user_preview()

    def show_user_preview_column_dialog(self) -> None:
        dialog, selection_table = self._create_user_preview_columns_dialog()
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        selected_columns = self._get_selected_user_preview_columns(selection_table)
        if not selected_columns:
            QMessageBox.warning(self, "בחירת עמודות", "יש לבחור לפחות עמודה אחת להצגה בטבלת הסקירה.")
            return

        self._apply_user_preview_columns(selected_columns)

    @staticmethod
    def _get_row_value(row: dict[str, Any], *candidates: str) -> str:
        normalized_row = {
            str(key).strip().upper(): value
            for key, value in row.items()
            if not str(key).startswith("__")
        }
        for candidate in candidates:
            for alias in get_column_aliases(candidate):
                if alias in normalized_row:
                    value = normalized_row[alias]
                    if value is None:
                        continue
                    return str(value).strip()
        return ""

    def _get_authorized_stms_users(self) -> list[str]:
        table = self.system_settings_widgets.get("authorized_stms_users")
        if not isinstance(table, QTableWidget):
            table = self.system_settings_widgets.get("super_users")
        if not isinstance(table, QTableWidget):
            return []

        users: list[str] = []
        for row_index in range(table.rowCount()):
            bname_item = table.item(row_index, 1)
            if isinstance(bname_item, QTableWidgetItem):
                bname = bname_item.text().strip().upper()
                if bname and bname not in users:
                    users.append(bname)
        return users

    def _get_critical_roles(self) -> list[str]:
        """Return the user-configured "פרופילים משתמשיי על" (strong profiles)
        from the system settings panel.

        An empty list is returned when the user cleared the configuration -
        downstream validators will then skip the strong-profile check entirely
        and the working paper will note that the check was not performed.
        """
        settings = self._current_system_settings() if hasattr(self, "_current_system_settings") else {}
        raw = settings.get("critical_roles", []) if isinstance(settings, dict) else []
        if isinstance(raw, str):
            raw = [raw]
        cleaned: list[str] = []
        for value in raw or []:
            text = str(value).strip()
            if text and text not in cleaned:
                cleaned.append(text)
        return cleaned

    def _load_preview_rows(self, slot_key: str) -> list[dict[str, Any]]:
        file_paths = list(self.slot_widgets.get(slot_key, {}).get("selected_paths", []))
        if not file_paths:
            return []

        preview_rows: list[dict[str, Any]] = []
        text_reader = TextFileReader()
        excel_reader = ExcelFileReader()

        for raw_path in file_paths:
            try:
                path = Path(raw_path)
                suffix = path.suffix.lower()
                if suffix in {".txt", ".csv"}:
                    rows = text_reader.read(path)
                elif suffix in {".xlsx", ".xlsm"}:
                    rows = excel_reader.read(path)
                else:
                    continue

                preview_rows.extend(
                    {
                        **row,
                        "__source_file": path.name,
                    }
                    for row in rows
                )
            except Exception:
                continue

        return preview_rows

    @staticmethod
    def _format_user_status(flag_value: object) -> str:
        normalized_value = "" if flag_value is None else str(flag_value).strip()
        if normalized_value in {"", "0", "00"}:
            return "פעיל"
        if normalized_value in {"64", "128", "129"}:
            return "נעול"
        return normalized_value

    @staticmethod
    def _parse_user_preview_date(raw_value: object) -> datetime | None:
        return parse_user_preview_date(raw_value)

    @classmethod
    def _format_user_preview_value_for_display(cls, field_name: str, value: object) -> str:
        _ = cls
        return format_user_preview_value_for_display(field_name, value)

    @classmethod
    def _get_user_preview_sort_value(cls, field_name: str, value: object) -> str:
        return get_user_preview_sort_value(field_name, value, cls.USER_PREVIEW_DATE_FIELDS)

    def _get_user_preview_filter_mode(self) -> str:
        filter_widget = getattr(self, "user_preview_status_filter", None)
        if isinstance(filter_widget, QComboBox):
            selected_value = filter_widget.currentData()
            if selected_value:
                return str(selected_value)
        return "all"

    def _filter_user_preview_rows(self, preview_rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], str]:
        filter_mode = self._get_user_preview_filter_mode()
        start_text = self.audit_period_from_edit.text().strip() if hasattr(self, "audit_period_from_edit") else ""
        end_text = self.audit_period_to_edit.text().strip() if hasattr(self, "audit_period_to_edit") else ""
        return filter_user_preview_rows(preview_rows, filter_mode, start_text, end_text)

    def _build_user_preview_rows(
        self,
        usr02_rows: list[dict[str, Any]],
        combined_rows: list[dict[str, Any]],
    ) -> list[dict[str, str]]:
        settings = self._current_system_settings()
        ai_cfg = settings.get("ai_settings", {}) if isinstance(settings, dict) else {}
        ai_enabled = bool(ai_cfg.get("enabled", False))
        features = ai_cfg.get("features", {}) if isinstance(ai_cfg, dict) else {}

        # --- Anomaly detection callback ---
        get_anomaly_data: Any = None
        if ai_enabled and bool(features.get("anomaly_detection", True)):
            try:
                from src.services.user_anomaly_detector import (
                    UserAnomalyDetector,
                    anomaly_score,
                    anomaly_codes,
                )
                detector = UserAnomalyDetector(settings)
                cohort = detector.build_cohort_stats(usr02_rows)

                def _anomaly_cb(row: dict[str, Any]) -> tuple[str, str]:
                    findings = detector.score_user(row, cohort)
                    return str(anomaly_score(findings)), anomaly_codes(findings)

                get_anomaly_data = _anomaly_cb
            except Exception:
                get_anomaly_data = None

        # --- AI narration callback ---
        get_ai_narration: Any = None
        if ai_enabled and bool(features.get("findings_narration", True)):
            try:
                from src.services.findings_narrator import narrate_user_finding
                _client = OllamaClient(ai_cfg)
                _work_env = self._current_work_environment_label()

                def _narration_cb(row: dict[str, Any], raw_findings: str) -> str:
                    return narrate_user_finding(row, raw_findings, _client, _work_env)

                get_ai_narration = _narration_cb
            except Exception:
                get_ai_narration = None

        return build_user_preview_rows(
            usr02_rows,
            combined_rows,
            self._get_row_value,
            self._format_user_status,
            self._get_slot_extraction_date("USR02"),
            self._current_work_environment_label(),
            self._get_reviewer_values,
            self._build_user_findings_description,
            self.DEFAULT_REVIEW_STATUS,
            get_anomaly_data=get_anomaly_data,
            get_ai_narration=get_ai_narration,
        )

    def _apply_column_filters_to_rows(self, rows: list[dict[str, str]]) -> list[dict[str, str]]:
        """Filter rows by active per-column value sets. Empty/absent filters pass all rows."""
        active = {f: v for f, v in self.user_preview_column_filters.items() if v}
        if not active:
            return rows
        result = []
        for row in rows:
            match = True
            for field_name, allowed_values in active.items():
                raw_value = row.get(field_name, "") or ""
                display_value = self._format_user_preview_value_for_display(field_name, raw_value)
                cell_text = display_value.strip()
                if not cell_text:
                    cell_text = "(ריקים)"
                if cell_text not in allowed_values:
                    match = False
                    break
            if match:
                result.append(row)
        return result

    def _on_user_preview_header_filter_requested(self, logical_index: int) -> None:
        """Called when the ▼ filter icon in a column header is clicked."""
        if logical_index < 0 or logical_index >= len(self.user_preview_visible_columns):
            return
        self._show_user_preview_column_filter_popup(logical_index)

    def _show_user_preview_column_filter_popup(self, column_index: int) -> None:
        """Open an Excel-style filter popup for the given column."""
        if column_index >= len(self.user_preview_visible_columns):
            return

        field_name = self.user_preview_visible_columns[column_index]
        col_def = self._get_user_preview_column_definition(field_name)
        formal_name = str(col_def.get("formal", field_name))

        # Build all unique display values across the current base-filtered population (no column filters)
        usr02_rows = self._load_preview_rows("USR02")
        combined_rows = self._load_preview_rows("ADR6_USR21")
        preview_rows = self._build_user_preview_rows(usr02_rows, combined_rows)
        base_filtered, _ = self._filter_user_preview_rows(preview_rows)

        # Collect unique display values for THIS column from base-filtered rows
        all_display_values: set[str] = set()
        for row in base_filtered:
            raw = row.get(field_name, "") or ""
            display = self._format_user_preview_value_for_display(field_name, raw).strip()
            all_display_values.add(display if display else "(ריקים)")

        sorted_values = sorted(all_display_values, key=lambda x: ("" if x == "(ריקים)" else x))

        # Current selection for this field (None means "show all")
        current_selected: set[str] | None = self.user_preview_column_filters.get(field_name)

        # Build dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(self.format_rtl_text(f"סינון לפי: {formal_name}"))
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(320, 440)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(6)

        # Search box
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("חיפוש...")
        search_edit.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        search_edit.setAlignment(Qt.AlignmentFlag.AlignRight)
        layout.addWidget(search_edit)

        # "Select All" checkbox
        from PySide6.QtWidgets import QCheckBox, QListWidget, QListWidgetItem  # noqa: PLC0415
        select_all_cb = QCheckBox(self.format_rtl_text("(בחר הכל)"))
        select_all_cb.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        layout.addWidget(select_all_cb)

        # Values list
        values_list = QListWidget()
        values_list.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        values_list.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)

        def _populate_list(filter_text: str = "") -> None:
            values_list.blockSignals(True)
            values_list.clear()
            for val in sorted_values:
                if filter_text and filter_text.lower() not in val.lower():
                    continue
                item = QListWidgetItem(self.format_rtl_text(val))
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                checked = (
                    Qt.CheckState.Checked
                    if (current_selected is None or val in current_selected)
                    else Qt.CheckState.Unchecked
                )
                item.setCheckState(checked)
                values_list.addItem(item)
            values_list.blockSignals(False)
            _sync_select_all()

        def _sync_select_all() -> None:
            total = values_list.count()
            checked_count = sum(
                1 for i in range(total)
                if values_list.item(i) and values_list.item(i).checkState() == Qt.CheckState.Checked
            )
            select_all_cb.blockSignals(True)
            if checked_count == 0:
                select_all_cb.setCheckState(Qt.CheckState.Unchecked)
            elif checked_count == total:
                select_all_cb.setCheckState(Qt.CheckState.Checked)
            else:
                select_all_cb.setCheckState(Qt.CheckState.PartiallyChecked)
            select_all_cb.blockSignals(False)

        def _on_select_all(state: int) -> None:
            new_state = Qt.CheckState.Checked if state == Qt.CheckState.Checked.value else Qt.CheckState.Unchecked
            values_list.blockSignals(True)
            for i in range(values_list.count()):
                item = values_list.item(i)
                if item:
                    item.setCheckState(new_state)
            values_list.blockSignals(False)

        def _on_item_changed(_item: Any) -> None:
            _sync_select_all()

        select_all_cb.setTristate(True)
        select_all_cb.stateChanged.connect(_on_select_all)
        values_list.itemChanged.connect(_on_item_changed)
        search_edit.textChanged.connect(_populate_list)

        _populate_list()
        layout.addWidget(values_list)

        # Buttons
        btn_row = QHBoxLayout()
        ok_btn = QPushButton(self.format_rtl_text("אישור"))
        cancel_btn = QPushButton(self.format_rtl_text("ביטול"))
        clear_btn = QPushButton(self.format_rtl_text("נקה סינון"))
        clear_btn.setToolTip(self.format_rtl_text("הסר את סינון העמודה הזו"))
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(clear_btn)
        layout.addLayout(btn_row)

        def _on_ok() -> None:
            selected_vals: set[str] = set()
            for i in range(values_list.count()):
                item = values_list.item(i)
                if item and item.checkState() == Qt.CheckState.Checked:
                    selected_vals.add(item.text().strip())
            # If all values selected → no filter active (same as "show all")
            if selected_vals >= all_display_values:
                self.user_preview_column_filters.pop(field_name, None)
            else:
                self.user_preview_column_filters[field_name] = selected_vals
            self._persist_user_preview_column_filters()
            self._update_header_filter_icons()
            self.refresh_user_preview()
            dialog.accept()

        def _on_clear() -> None:
            self.user_preview_column_filters.pop(field_name, None)
            self._persist_user_preview_column_filters()
            self._update_header_filter_icons()
            self.refresh_user_preview()
            dialog.accept()

        ok_btn.clicked.connect(_on_ok)
        cancel_btn.clicked.connect(dialog.reject)
        clear_btn.clicked.connect(_on_clear)

        dialog.exec()

    def refresh_user_preview(self) -> None:
        # Prevent nested refresh calls from transient Qt events while the table is rebuilding.
        if self._refreshing_user_preview:
            return
        self._configure_user_preview_table()
        self._refreshing_user_preview = True
        self.user_preview_table.blockSignals(True)
        self.user_preview_table.setSortingEnabled(False)

        try:
            self.user_preview_table.setRowCount(0)

            usr02_rows = self._load_preview_rows("USR02")
            combined_rows = self._load_preview_rows("ADR6_USR21")
            preview_rows = self._build_user_preview_rows(usr02_rows, combined_rows)

            if not preview_rows:
                self.user_preview_hint.setText(
                    self.format_ui_rtl_text(
                        "לא זוהו עדיין משתמשים להצגה. יש לטעון קבצי USR02 ו-ADR6 / USER_ADDR."
                    )
                )
                self._update_user_review_progress_summary(0, 0, 0)
                return

            filtered_rows, filter_note = self._filter_user_preview_rows(preview_rows)
            filter_mode = self._get_user_preview_filter_mode()

            if filter_note:
                self.user_preview_hint.setText(
                    self.format_ui_rtl_text(
                        f"{filter_note} הטבלה מציגה כעת את כל {len(preview_rows)} המשתמשים שנטענו לכלי."
                    )
                )
                rows_to_display = preview_rows
            else:
                rows_to_display = filtered_rows
                if filter_mode == "all":
                    self.user_preview_hint.setText(
                        self.format_ui_rtl_text(f"הטבלה מציגה כעת {len(rows_to_display)} משתמשים שנטענו לכלי.")
                    )
                else:
                    self.user_preview_hint.setText(
                        self.format_ui_rtl_text(
                            f"הטבלה מציגה כעת {len(rows_to_display)} משתמשים מתוך {len(preview_rows)} בהתאם לטווח התאריכים שנבחר."
                        )
                    )

            # Apply per-column filters (display-only; does not affect export or progress summary)
            col_filtered = self._apply_column_filters_to_rows(rows_to_display)
            if len(col_filtered) < len(rows_to_display):
                active_count = sum(1 for f in self.user_preview_column_filters.values() if f)
                self.user_preview_hint.setText(
                    self.format_ui_rtl_text(
                        f"הטבלה מציגה כעת {len(col_filtered)} משתמשים לאחר סינון עמודות ({active_count} סינון/ים פעיל/ים)."
                    )
                )
            rows_to_display = col_filtered

            if not rows_to_display:
                self._update_user_review_progress_summary(0, 0, 0)
                return

            rows_to_display = sorted(rows_to_display, key=self._export_sort_key)

            for preview_row in rows_to_display:
                row_index = self.user_preview_table.rowCount()
                self.user_preview_table.insertRow(row_index)
                review_key = self._user_reviewer_state_key(preview_row.get("MANDT", ""), preview_row.get("BNAME", ""))
                for column, field_name in enumerate(self.user_preview_visible_columns):
                    value = preview_row.get(field_name, "") or ""

                    if field_name == "REVIEW_STATUS":
                        display_status = self._normalize_reviewer_status(value)
                        status_item = SortableTableWidgetItem(self.format_rtl_text(display_status))
                        status_item.setData(SortableTableWidgetItem.SORT_ROLE, display_status)
                        status_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        status_item.setToolTip(self.format_rtl_text(display_status))
                        status_item.setFlags(status_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        self.user_preview_table.setItem(row_index, column, status_item)
                        continue

                    display_value = self._format_user_preview_value_for_display(field_name, value)
                    item = SortableTableWidgetItem(self.format_rtl_text(display_value))
                    item.setData(SortableTableWidgetItem.SORT_ROLE, self._get_user_preview_sort_value(field_name, value))
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    item.setToolTip(self.format_rtl_text(display_value))
                    item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)

                    if field_name in {"TECH_REVIEW_NOTES", "BUS_REVIEW_NOTES", "REVIEW_NOTES"}:
                        pass  # read-only for audit tool users

                    self.user_preview_table.setItem(row_index, column, item)

                self._update_review_row_highlight(row_index, preview_row)

            self.user_preview_table.resizeColumnsToContents()
            for column_index, field_name in enumerate(self.user_preview_visible_columns):
                default_width = int(self._get_user_preview_column_definition(field_name).get("width", 120))
                self.user_preview_table.setColumnWidth(
                    column_index,
                    max(self.user_preview_table.columnWidth(column_index), default_width),
                )
        finally:
            self.user_preview_table.blockSignals(False)
            self._refreshing_user_preview = False
            self.user_preview_table.setSortingEnabled(True)
            self._refresh_user_review_progress_summary_from_table()
            self._refresh_audit_summary_table()

    def run_validation(self) -> None:
        file_paths = self._current_file_paths()
        if not file_paths or not self.selected_slot_key:
            QMessageBox.warning(self, "חסר קובץ", "יש לבחור קובץ מתוך אחד ממשבצות הקלט לפני הרצת הבדיקה.")
            self.tabs.setCurrentIndex(0)
            return

        if self.validation_thread is not None:
            QMessageBox.information(self, "בדיקה פעילה", "בדיקה כבר רצה ברקע. יש להמתין לסיום.")
            return

        analysis_tab_index = self.tabs.indexOf(self.analysis_tab)
        if analysis_tab_index >= 0:
            self.tabs.setCurrentIndex(analysis_tab_index)
        self._start_slot_validation_async(self.selected_slot_key, file_paths)

    def _set_validation_running_state(self, is_running: bool, slot_key: str | None = None) -> None:
        self.audit_run_button.setEnabled(not is_running)
        self.audit_export_button.setEnabled(not is_running)

        # Disable / re-enable the top-bar action buttons
        self.clear_button.setEnabled(not is_running)
        self.export_log_button.setEnabled(not is_running)
        self.output_button.setEnabled(not is_running)

        # Disable / re-enable every slot's file-chooser and clear button
        for slot_entry in self.slot_widgets.values():
            btn = slot_entry.get("button")
            clr = slot_entry.get("clear_button")
            if isinstance(btn, QPushButton):
                btn.setEnabled(not is_running)
            if isinstance(clr, QPushButton):
                clr.setEnabled(not is_running)

        if is_running:
            slot_text = slot_key or ""
            self.analysis_progress_label.setText(self.format_ui_rtl_text(f"מעבד כעת את המשבצת {slot_text}..."))
            self.analysis_progress_bar.setRange(0, 0)
            self.analysis_progress_container.show()

            # Show a window-modal loading dialog so the user knows to wait
            if self._loading_dialog is not None:
                try:
                    self._loading_dialog.close()
                except Exception:
                    pass
            dialog = _LoadingProgressDialog(self, slot_text)
            self._loading_dialog = dialog
            # Centre over the main window
            dialog.adjustSize()
            dialog.move(
                self.geometry().center().x() - dialog.width() // 2,
                self.geometry().center().y() - dialog.height() // 2,
            )
            dialog.show()
            dialog.raise_()
            dialog.activateWindow()
            QApplication.processEvents()  # force the dialog to render before the thread starts
        else:
            self.analysis_progress_container.hide()

            # Close the loading dialog
            if self._loading_dialog is not None:
                try:
                    # Temporarily allow close so our programmatic close works
                    self._loading_dialog.closeEvent = lambda e: e.accept()  # type: ignore[method-assign]
                    self._loading_dialog.close()
                except Exception:
                    pass
                self._loading_dialog = None

    def _start_slot_validation_async(self, slot_key: str, file_paths: list[str]) -> None:
        input_files_dict: dict[str, list[str | Path]] = {
            slot_key: [str(path) for path in file_paths]
        }
        required_columns = self._required_columns_for_slot(slot_key)
        authorized_users = self._get_authorized_stms_users()
        strong_profiles = self._get_critical_roles()

        self.validation_thread = QThread(self)
        self.validation_worker = SlotValidationWorker(
            slot_key=slot_key,
            file_paths=list(file_paths),
            input_files_dict=input_files_dict,
            required_columns=required_columns,
            output_dir=self.config.output_dir,
            authorized_users=authorized_users,
            strong_profiles=strong_profiles,
        )
        self.validation_worker.moveToThread(self.validation_thread)

        self.validation_thread.started.connect(self.validation_worker.run)
        self.validation_worker.succeeded.connect(self._on_slot_validation_worker_succeeded)
        self.validation_worker.failed.connect(self._on_slot_validation_worker_failed)
        self.validation_worker.finished.connect(self.validation_thread.quit)
        self.validation_worker.finished.connect(self.validation_worker.deleteLater)
        self.validation_thread.finished.connect(self.validation_thread.deleteLater)
        self.validation_thread.finished.connect(self._on_slot_validation_worker_finished)

        self._set_validation_running_state(True, slot_key)
        self.validation_thread.start()

    @Slot(str, list, object)
    def _on_slot_validation_worker_succeeded(self, slot_key: str, file_paths: list[str], result: object) -> None:
        self._handle_slot_validation_success(slot_key, file_paths, result, show_feedback=True)

    @Slot(str, list, str)
    def _on_slot_validation_worker_failed(self, slot_key: str, file_paths: list[str], error_text: str) -> None:
        self._handle_slot_validation_error(slot_key, file_paths, error_text, show_feedback=True)

    @Slot()
    def _on_slot_validation_worker_finished(self) -> None:
        self.validation_worker = None
        self.validation_thread = None
        self._set_validation_running_state(False)

    # ------------------------------------------------------------------
    # IPE pre-run prerequisites check & visual indicators
    # ------------------------------------------------------------------

    def _check_ipe_prerequisites(self, slot_keys: list[str]) -> list[dict[str, str]]:
        """Return a list of blocking issues for *slot_keys*.

        Each issue dict has keys:
          ``slot``        — slot key (e.g. "USR02")
          ``display_name``— label shown in SLOT_DEFINITIONS (or slot key)
          ``type``        — ``"missing_file"`` | ``"missing_ipe"``

        Only required slots are checked.  A slot without an IPE image is
        blocking even if a file was loaded — the screenshot is mandatory.
        """
        issues: list[dict[str, str]] = []
        for slot_key in slot_keys:
            metadata = self.SLOT_DEFINITIONS.get(slot_key, {})
            if not metadata.get("required", False):
                continue
            display_name = str(metadata.get("label", slot_key))
            file_paths = list(self.slot_widgets.get(slot_key, {}).get("selected_paths", []))
            if not file_paths:
                issues.append({"slot": slot_key, "display_name": display_name, "type": "missing_file"})
            elif not self.ipe_evidence_data.get(slot_key):
                issues.append({"slot": slot_key, "display_name": display_name, "type": "missing_ipe"})
        return issues

    def _show_prerequisites_error(self, scope_label: str, issues: list[dict[str, str]]) -> None:
        """Display a blocking QMessageBox.critical listing all *issues*."""
        lines = [f"לא ניתן להריץ את {scope_label} — חסרים תנאים מקדימים:", ""]
        for issue in issues:
            name = issue["display_name"]
            if issue["type"] == "missing_file":
                lines.append(f"  • {name} *  —  קובץ לא נטען")
            else:
                lines.append(f"  • {name} *  —  חסרה ראיה IPE (צילום מסך)")
        lines += [
            "",
            "יש לטעון את כל הקבצים המנדטוריים (מסומנים ב-*) ולצרף לכל אחד",
            "לפחות תמונת ראיה אחת לפני הרצת הבדיקה.",
        ]
        QMessageBox.critical(self, "חסרים תנאים מקדימים להרצה", "\n".join(lines))

    def _update_slot_ipe_indicator(self, slot_key: str) -> None:
        """Set the visual style of the slot's path_label based on IPE state.

        - required + file loaded + IPE present  → green border
        - required + file loaded + IPE missing  → orange border / background
        - anything else (no file / not required) → default style
        """
        widget_data = self.slot_widgets.get(slot_key)
        if widget_data is None:
            return
        label = widget_data.get("path_label")
        if not isinstance(label, QLabel):
            return

        metadata = self.SLOT_DEFINITIONS.get(slot_key, {})
        is_required = bool(metadata.get("required", False))
        has_file = bool(widget_data.get("selected_paths"))
        has_ipe = bool(self.ipe_evidence_data.get(slot_key))

        if is_required and has_file and has_ipe:
            label.setStyleSheet(
                "padding: 6px; background: #e8f5e9; border: 1px solid #388e3c;"
            )
        elif is_required and has_file and not has_ipe:
            label.setStyleSheet(
                "padding: 6px; background: #fff3e0; border: 1px solid #ff8f00;"
            )
        else:
            label.setStyleSheet(
                "padding: 6px; background: #ffffff; border: 1px solid #cfd6e4;"
            )

    def run_domain_validation(self, domain: str) -> None:
        if bool(self.DOMAIN_DEFINITIONS.get(domain, {}).get("in_development", False)):
            QMessageBox.information(
                self,
                "תחום בפיתוח",
                f"תחום '{domain}' נמצא בפיתוח ואינו כולל בדיקות אוטומטיות עדיין.\n\nבדיקות לתחום זה יתווספו בגרסאות הבאות.",
            )
            return

        domain_slots = self._get_domain_slots(domain)

        # IPE prerequisite gate — blocks execution if any required slot has no
        # file loaded or no IPE screenshot attached.
        ipe_issues = self._check_ipe_prerequisites(domain_slots)
        if ipe_issues:
            self._show_prerequisites_error(f"תחום {domain}", ipe_issues)
            return

        selected_slots: list[tuple[str, list[str]]] = []
        for slot_key in domain_slots:
            file_paths = list(self.slot_widgets[slot_key].get("selected_paths", []))
            if file_paths:
                selected_slots.append((slot_key, file_paths))

        if not selected_slots:
            QMessageBox.warning(
                self,
                "לא נבחרו קבצים",
                f"לא נבחרו קבצים עבור תחום {domain}. יש לבחור לפחות קובץ אחד לפני הרצת הבדיקה.",
            )
            return

        processed_slots = 0
        processed_files = 0
        total_rows = 0
        total_invalid_rows = 0
        invalid_slots = 0
        failed_slots: list[str] = []

        # Keep the intake summary hidden while multiple slots are still processing.
        self.summary_group.hide()
        self.results_group.hide()

        for slot_key, file_paths in selected_slots:
            slot_summary = self._run_slot_validation(
                slot_key,
                file_paths,
                show_feedback=False,
                update_summary_ui=False,
            )
            processed_slots += 1
            processed_files += int(slot_summary["file_count"])
            total_rows += int(slot_summary["total_rows"])
            total_invalid_rows += int(slot_summary["invalid_rows"])

            if slot_summary["status"] == "error":
                failed_slots.append(slot_key)
            elif not bool(slot_summary["is_valid"]):
                invalid_slots += 1

        summary_lines = [
            f"בדיקת תחום {domain} הושלמה.",
            f"משבצות שנבדקו: {processed_slots}",
            f"קבצים שנבדקו: {processed_files}",
            f"שורות שנבדקו: {total_rows}",
        ]

        if invalid_slots:
            summary_lines.append(f"משבצות עם ממצאים: {invalid_slots}")
        if failed_slots:
            summary_lines.append(f"משבצות שנכשלו בעיבוד: {', '.join(failed_slots)}")
        summary_lines.append("ניתן לבצע לחיצה כפולה על הרשומה בלוג לצפייה בפירוט.")

        total_valid_rows = max(total_rows - total_invalid_rows, 0)
        self.summary_labels["total"].setText(str(total_rows))
        self.summary_labels["valid"].setText(str(total_valid_rows))
        self.summary_labels["invalid"].setText(str(total_invalid_rows))
        if failed_slots:
            self.summary_labels["status"].setText("הושלם עם שגיאות עיבוד")
        elif total_invalid_rows > 0:
            self.summary_labels["status"].setText("הושלם עם שגיאות קליטה")
        else:
            self.summary_labels["status"].setText("תקין")

        self.issues_table.setRowCount(0)
        self.issues_table.insertRow(0)
        for column, value in enumerate(["-", "-", "הסיכום מתייחס לכל המשבצות שנבדקו בריצה הנוכחית"]):
            item = QTableWidgetItem(self.format_rtl_text(value))
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.issues_table.setItem(0, column, item)
        self.issues_table.resizeColumnsToContents()

        self.summary_group.show()
        if total_invalid_rows > 0 or bool(failed_slots):
            self.results_group.show()
        else:
            self.results_group.hide()

        # Auto-export the intake log to Excel at the end of a domain run so the
        # auditor has an immediate record of what was processed.
        try:
            self.export_run_log_to_excel(open_after_export=False)
        except Exception:
            # Auto-export must never block the run summary dialog.
            pass

        if invalid_slots or failed_slots:
            QMessageBox.warning(self, "בדיקת תחום הושלמה עם ממצאים", "\n".join(summary_lines))
        else:
            QMessageBox.information(self, "בדיקת תחום הושלמה", "\n".join(summary_lines))

    def run_category_validation(self, category: str) -> None:
        category_slots = self._get_category_slots(category)

        # IPE prerequisite gate — blocks execution if any required slot in this
        # category has no file loaded or no IPE screenshot attached.
        ipe_issues = self._check_ipe_prerequisites(category_slots)
        if ipe_issues:
            self._show_prerequisites_error(f"קטגוריה {category}", ipe_issues)
            return

        selected_slots: list[tuple[str, list[str]]] = []
        for slot_key in category_slots:
            file_paths = list(self.slot_widgets[slot_key].get("selected_paths", []))
            if file_paths:
                selected_slots.append((slot_key, file_paths))

        if not selected_slots:
            QMessageBox.warning(
                self,
                "לא נבחרו קבצים",
                f"לא נבחרו קבצים עבור הקבוצה {category}. יש לבחור לפחות קובץ אחד לפני הרצת הבדיקה.",
            )
            return

        processed_slots = 0
        processed_files = 0
        total_rows = 0
        total_invalid_rows = 0
        invalid_slots = 0
        failed_slots: list[str] = []

        self._enter_intake_stage_b()
        # Keep the intake summary hidden while multiple slots are still processing.
        self.summary_group.hide()
        self.results_group.hide()

        for slot_key, file_paths in selected_slots:
            slot_summary = self._run_slot_validation(
                slot_key,
                file_paths,
                show_feedback=False,
                update_summary_ui=False,
            )
            processed_slots += 1
            processed_files += int(slot_summary["file_count"])
            total_rows += int(slot_summary["total_rows"])
            total_invalid_rows += int(slot_summary["invalid_rows"])

            if slot_summary["status"] == "error":
                failed_slots.append(slot_key)
            elif not bool(slot_summary["is_valid"]):
                invalid_slots += 1

        summary_lines = [
            f"בדיקת הקבוצה {category} הושלמה.",
            f"משבצות שנבדקו: {processed_slots}",
            f"קבצים שנבדקו: {processed_files}",
            f"שורות שנבדקו: {total_rows}",
        ]

        if invalid_slots:
            summary_lines.append(f"משבצות עם ממצאים: {invalid_slots}")
        if failed_slots:
            summary_lines.append(f"משבצות שנכשלו בעיבוד: {', '.join(failed_slots)}")
        summary_lines.append("ניתן לבצע לחיצה כפולה על הרשומה בלוג לצפייה בפירוט.")

        total_valid_rows = max(total_rows - total_invalid_rows, 0)
        self.summary_labels["total"].setText(str(total_rows))
        self.summary_labels["valid"].setText(str(total_valid_rows))
        self.summary_labels["invalid"].setText(str(total_invalid_rows))
        if failed_slots:
            self.summary_labels["status"].setText("הושלם עם שגיאות עיבוד")
        elif total_invalid_rows > 0:
            self.summary_labels["status"].setText("הושלם עם שגיאות קליטה")
        else:
            self.summary_labels["status"].setText("תקין")

        self.issues_table.setRowCount(0)
        self.issues_table.insertRow(0)
        for column, value in enumerate(["-", "-", "הסיכום מתייחס לכל המשבצות שנבדקו בריצה הנוכחית"]):
            item = QTableWidgetItem(self.format_rtl_text(value))
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.issues_table.setItem(0, column, item)
        self.issues_table.resizeColumnsToContents()

        self.summary_group.show()
        if total_invalid_rows > 0 or bool(failed_slots):
            self.results_group.show()
        else:
            self.results_group.hide()

        # Auto-export the intake log to Excel at the end of a category run.
        try:
            self.export_run_log_to_excel(open_after_export=False)
        except Exception:
            pass

        if invalid_slots or failed_slots:
            QMessageBox.warning(self, "בדיקת קבוצה הושלמה עם ממצאים", "\n".join(summary_lines))
        else:
            QMessageBox.information(self, "בדיקת קבוצה הושלמה", "\n".join(summary_lines))

    def _run_slot_validation(
        self,
        slot_key: str,
        file_paths: list[str],
        show_feedback: bool = True,
        update_summary_ui: bool = True,
    ) -> dict[str, Any]:
        file_names = ", ".join(Path(p).name for p in file_paths)
        self._log_to_console(f"קולט: {slot_key}  ▸  {file_names}", "title")
        try:
            result = self._process_slot_validation(slot_key, file_paths)
        except Exception as error:
            self._log_to_console(f"שגיאה בעיבוד {slot_key}: {error}", "error")
            return self._handle_slot_validation_error(
                slot_key,
                file_paths,
                str(error),
                show_feedback,
                update_summary_ui,
            )

        return self._handle_slot_validation_success(
            slot_key,
            file_paths,
            result,
            show_feedback,
            update_summary_ui,
        )

    def _process_slot_validation(self, slot_key: str, file_paths: list[str]) -> Any:
        if slot_key == "AGR_1251":
            self.summary_labels["status"].setText("מעבד קובצי הרשאות גדולים במנות...")
            self._log_to_console("AGR_1251: קובצי הרשאות גדולים — מעבד במנות, אנא המתן...", "warn")
            QApplication.processEvents()

        input_files_dict: dict[str, list[str | Path]] = {
            slot_key: [str(path) for path in file_paths]
        }

        return process_file(
            input_files=input_files_dict,
            required_columns=self._required_columns_for_slot(slot_key),
            output_dir=self.config.output_dir,
            source_name_override=slot_key,
            authorized_users=self._get_authorized_stms_users(),
            strong_profiles=self._get_critical_roles(),
        )

    def _handle_slot_validation_error(
        self,
        slot_key: str,
        file_paths: list[str],
        error_text: str,
        show_feedback: bool,
        update_summary_ui: bool = True,
    ) -> dict[str, Any]:
        if update_summary_ui:
            self.summary_labels["status"].setText(f"שגיאה בעיבוד {slot_key}")
            self.issues_table.setRowCount(0)
            error_message = f"אירעה שגיאה במהלך העיבוד של המשבצת {slot_key}: {error_text}"
            self.issues_table.insertRow(0)
            for column, value in enumerate(["מבנה", "SYSTEM", error_message]):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.issues_table.setItem(0, column, item)
            self.issues_table.resizeColumnsToContents()
        self._append_error_log_entries(slot_key, file_paths, str(error_text))
        if show_feedback:
            QMessageBox.critical(self, "שגיאה", f"אירעה שגיאה במהלך העיבוד של המשבצת {slot_key}:\n{error_text}")
        return {
            "slot_key": slot_key,
            "status": "error",
            "file_count": len(file_paths),
            "total_rows": 0,
            "invalid_rows": 0,
            "is_valid": False,
        }

    def _handle_slot_validation_success(
        self,
        slot_key: str,
        file_paths: list[str],
        result: Any,
        show_feedback: bool,
        update_summary_ui: bool = True,
    ) -> dict[str, Any]:

        intake_issues = [iss for iss in result.issues if self._is_intake_issue(iss)]
        valid_rows, invalid_rows = self._compute_intake_summary(result.summary.total_rows, intake_issues)

        if update_summary_ui:
            self._enter_intake_stage_b()
            self.summary_group.show()
            if intake_issues:
                self.results_group.show()
            else:
                self.results_group.hide()
            self.summary_labels["total"].setText(str(result.summary.total_rows))
            self.summary_labels["valid"].setText(str(valid_rows))
            self.summary_labels["invalid"].setText(str(invalid_rows))

            # Only intake-level issues (structural / missing required) surface in this tab.
            # Audit/analysis findings (e.g. RSPARAM policy) are deferred to the analysis tab.
            status_text = "תקין" if not intake_issues else f"שגיאות קליטה - {slot_key}"
            self.summary_labels["status"].setText(status_text)

            self.issues_table.setRowCount(0)
            if intake_issues:
                for issue in intake_issues:
                    row_index = self.issues_table.rowCount()
                    self.issues_table.insertRow(row_index)
                    values = [
                        str(issue.row_number if issue.row_number > 0 else "מבנה"),
                        self.format_rtl_text(issue.column_name),
                        self.format_rtl_text(issue.message),
                    ]
                    for column, value in enumerate(values):
                        item = QTableWidgetItem(value)
                        item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                        self.issues_table.setItem(row_index, column, item)
            else:
                self.issues_table.insertRow(0)
                for column, value in enumerate(["-", "-", "לא נמצאו שגיאות קליטה"]):
                    item = QTableWidgetItem(value)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                    self.issues_table.setItem(0, column, item)
            self.issues_table.resizeColumnsToContents()

        audit_issues = [iss for iss in result.issues if not self._is_intake_issue(iss)]
        self._upsert_audit_control_data(
            slot_key=slot_key,
            result=result,
            audit_issues=audit_issues,
            extraction_date=self._get_slot_extraction_date(slot_key),
        )

        # Capture raw rows per control for working-paper export
        try:
            detected_profile = str(getattr(result, "detected_profile", "") or slot_key).upper()
            slot_rows_raw = list(getattr(result, "rows", []) or [])
            # Tag each row with __profile so downstream report code (and the
            # MA3-3 working paper specifically) can render a per-row
            # "טבלת מקור" when multiple slots are merged into one control.
            slot_rows: list[dict[str, Any]] = []
            for r in slot_rows_raw:
                if isinstance(r, dict):
                    if not r.get("__profile"):
                        r = {**r, "__profile": detected_profile}
                    slot_rows.append(r)
                else:
                    slot_rows.append(r)
            self.slot_to_row_count[slot_key] = (
                getattr(result, "total_rows_override", None) or len(slot_rows)
            )
            expected_controls = get_profile_audit_controls(getattr(result, "detected_profile", slot_key))
            control_ids = sorted(set([iss.control_id for iss in audit_issues if iss.control_id] + expected_controls))
            for cid in control_ids:
                # Aggregate rows across all slots that map to this control
                # (e.g. MA3-3_AYALON_14 receives both UST04 and USH04 rows).
                existing = self.control_to_slot_rows.get(cid)
                if existing is None:
                    self.control_to_slot_rows[cid] = list(slot_rows)
                else:
                    # Replace rows that came from THIS slot key (re-run support),
                    # while preserving rows from other slots.
                    other_slot_rows = [
                        r for r in existing
                        if not (isinstance(r, dict) and str(r.get("__profile", "")).upper() == detected_profile)
                    ]
                    self.control_to_slot_rows[cid] = other_slot_rows + list(slot_rows)
                # Keep the *first* slot key encountered for a control — IPE
                # enrichment iterates all slots anyway via control_ids filter.
                self.control_to_slot_key.setdefault(cid, slot_key)
        except Exception:
            pass

        self._refresh_audit_summary_table()
        self._upsert_permissions_control_data(
            slot_key=slot_key,
            result=result,
            audit_issues=audit_issues,
            extraction_date=self._get_slot_extraction_date(slot_key),
        )
        self._refresh_permissions_summary_table()

        # Cache AGR_1251 / AGR_USERS rows for cross-join user-management permission check
        detected_profile = str(getattr(result, "detected_profile", "") or "").upper()
        if detected_profile == "AGR_1251":
            self.agr_1251_cached_rows = list(getattr(result, "rows", []))
            self._compute_user_mgmt_permissions()
            self._compute_auth_mgmt_permissions()
            self._compute_rscdok99_permissions()
            self._compute_data_mgmt_permissions()
            self._compute_transport_permissions()
            self._compute_debug_permissions()
            self._compute_job_mgmt_permissions()
        elif detected_profile == "AGR_USERS":
            self.agr_users_cached_rows = list(getattr(result, "rows", []))
            # Compute unique-user population per MANDT for summary table
            _pop_sets: dict[str, set[str]] = {}
            for _row in self.agr_users_cached_rows:
                _mandt_val = self._resolve_row_value_by_priority(_row, "MANDT")
                if _mandt_val is not None and str(_mandt_val).strip():
                    _mandt = str(_mandt_val).strip()
                else:
                    _src = str(_row.get("__source_file", ""))
                    _dm = re.search(r"\d{3}", Path(_src).name)
                    _mandt = _dm.group(0) if _dm else "-"
                _uname_val = self._resolve_row_value_by_priority(_row, "UNAME")
                if _uname_val is not None and str(_uname_val).strip():
                    _pop_sets.setdefault(_mandt, set()).add(str(_uname_val).strip().upper())
            self.agr_users_population_by_mandt = {m: len(s) for m, s in _pop_sets.items()}
            self._compute_user_mgmt_permissions()
            self._compute_auth_mgmt_permissions()
            self._compute_rscdok99_permissions()
            self._compute_data_mgmt_permissions()
            self._compute_transport_permissions()
            self._compute_debug_permissions()
            self._compute_job_mgmt_permissions()

        self._sync_permissions_findings_into_analysis_summary()
        self._refresh_audit_summary_table()

        self._append_run_log_entries(slot_key, file_paths, result)
        if result.report_path is not None:
            self.report_path = result.report_path
        self.report_button.setEnabled(self.report_path is not None)
        file_count = len(result.source_files) if result.source_files else len(file_paths)

        if show_feedback:
            if not intake_issues:
                QMessageBox.information(
                    self,
                    "הבדיקה הושלמה",
                    f"בדיקת המשבצת {slot_key} הסתיימה ללא שגיאות קליטה. נקלטו {file_count} קבצים.",
                )
            else:
                ordered_messages: list[str] = []
                structure_messages = [iss.message for iss in intake_issues if "אינו תואם למבנה" in iss.message]
                other_messages = [iss.message for iss in intake_issues if "אינו תואם למבנה" not in iss.message]
                for message in structure_messages + other_messages:
                    if message not in ordered_messages:
                        ordered_messages.append(message)
                    if len(ordered_messages) == 3:
                        break
                summary_text = "\n".join(f"• {message}" for message in ordered_messages)
                QMessageBox.warning(
                    self,
                    "נמצאו שגיאות קליטה",
                    f"בדיקת המשבצת {slot_key} הסתיימה עם שגיאות קליטה.\n\n{summary_text}\n\nממצאי ביקורת יוצגו בטאב 'ביצוע ניתוח לביקורת'.",
                )

        self._log_to_console(
            f"הושלם: {slot_key}  ◂  {result.summary.total_rows} שורות"
            + (f"  |  {len(intake_issues)} שגיאות קליטה" if intake_issues else "  |  תקין"),
            "warn" if intake_issues else "info",
        )
        return {
            "slot_key": slot_key,
            "status": "ok",
            "file_count": file_count,
            "total_rows": result.summary.total_rows,
            "invalid_rows": invalid_rows,
            "is_valid": len(intake_issues) == 0,
        }

    def _append_run_log_entries(self, slot_key: str, file_paths: list[str], result) -> None:
        issues_by_file: dict[str, list] = {Path(path).name: [] for path in file_paths}
        row_counts_by_file = dict(getattr(result, "file_row_counts", {}))
        if not row_counts_by_file:
            for row in getattr(result, "rows", []):
                source_file = Path(str(row.get("__source_file", ""))).name
                if source_file:
                    row_counts_by_file[source_file] = row_counts_by_file.get(source_file, 0) + 1
        display_slot_name = self._get_slot_display_name(slot_key)
        report_group = self._get_slot_category(slot_key)
        extraction_date = self._get_slot_extraction_date(slot_key)

        for issue in result.issues:
            if issue.source_file:
                issue_name = Path(issue.source_file).name
                issues_by_file.setdefault(issue_name, []).append(issue)
            else:
                for issue_list in issues_by_file.values():
                    issue_list.append(issue)

        for path in file_paths:
            file_name = Path(path).name
            file_issues = issues_by_file.get(file_name, [])
            # Status and preview in the intake log are based only on intake-level issues.
            # All issues (including audit findings) are stored for drill-down details.
            intake_file_issues = [iss for iss in file_issues if self._is_intake_issue(iss)]
            status_text = "שגוי" if intake_file_issues else "תקין"
            checked_at = datetime.now()
            row_count = row_counts_by_file.get(file_name, 0)
            # Track per-file counts so the working-paper IPE sheet can show the
            # rows for the specific file referenced by a screenshot (stem match)
            # instead of summing across all files in the slot.
            file_stem_upper = Path(file_name).stem.upper()
            per_file_map = self._slot_file_row_counts.setdefault(slot_key, {})
            per_file_map[file_stem_upper] = row_count
            record = {
                "slot_key": display_slot_name,
                "report_group": report_group,
                "file_name": file_name,
                "extraction_date": extraction_date,
                "row_count": row_count,
                "status": status_text,
                "error_count": len(intake_file_issues),
                "error_preview": self._build_issue_preview(intake_file_issues),
                "date": checked_at.strftime("%Y-%m-%d"),
                "time": checked_at.strftime("%H:%M:%S"),
                "issues": list(file_issues),
            }
            self.run_log_records.append(record)

            row_index = self.run_log_table.rowCount()
            self.run_log_table.insertRow(row_index)
            values = [
                display_slot_name,
                report_group,
                file_name,
                extraction_date,
                str(row_count),
                status_text,
                str(len(intake_file_issues)),
                str(record["error_preview"]),
                str(record["date"]),
                str(record["time"]),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                item.setToolTip(self.format_rtl_text(value))
                if column == 5:
                    item.setBackground(QColor("#fdecec") if status_text in {"שגוי", "שגיאה"} else QColor("#eaf7ea"))
                self.run_log_table.setItem(row_index, column, item)
        self.run_log_table.resizeColumnsToContents()

    def _append_error_log_entries(self, slot_key: str, file_paths: list[str], error_text: str) -> None:
        checked_at = datetime.now()
        display_slot_name = self._get_slot_display_name(slot_key)
        report_group = self._get_slot_category(slot_key)
        extraction_date = self._get_slot_extraction_date(slot_key)

        for path in file_paths:
            file_name = Path(path).name
            issue = ValidationIssue(
                row_number=0,
                column_name="SYSTEM",
                message=f"אירעה שגיאה במהלך העיבוד: {error_text}",
                source_file=file_name,
            )
            record = {
                "slot_key": display_slot_name,
                "report_group": report_group,
                "file_name": file_name,
                "extraction_date": extraction_date,
                "row_count": 0,
                "status": "שגיאה",
                "error_count": 1,
                "error_preview": issue.message,
                "date": checked_at.strftime("%Y-%m-%d"),
                "time": checked_at.strftime("%H:%M:%S"),
                "issues": [issue],
            }
            self.run_log_records.append(record)

            row_index = self.run_log_table.rowCount()
            self.run_log_table.insertRow(row_index)
            values = [
                display_slot_name,
                report_group,
                file_name,
                extraction_date,
                "0",
                "שגיאה",
                "1",
                issue.message,
                str(record["date"]),
                str(record["time"]),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                item.setToolTip(self.format_rtl_text(value))
                if column == 5:
                    item.setBackground(QColor("#fdecec"))
                self.run_log_table.setItem(row_index, column, item)
        self.run_log_table.resizeColumnsToContents()

    @staticmethod
    def _count_stms_control_records(rows: list[dict[str, Any]]) -> int:
        count = 0
        for row in rows:
            normalized_row = {
                str(key).strip().upper(): value
                for key, value in row.items()
                if not str(key).startswith("__")
            }
            trkorr = ""
            for candidate in get_column_aliases("TRKORR"):
                if candidate in normalized_row and str(normalized_row.get(candidate, "")).strip():
                    trkorr = str(normalized_row.get(candidate, "")).strip()
                    break
            import_user = ""
            for candidate in get_column_aliases("IMPORT_USER"):
                if candidate in normalized_row and str(normalized_row.get(candidate, "")).strip():
                    import_user = str(normalized_row.get(candidate, "")).strip()
                    break
            if trkorr and import_user:
                count += 1
        return count

    @staticmethod
    def _find_row_column_by_alias(row: dict[str, Any], candidate: str) -> str | None:
        normalized_map = {str(column).strip().upper(): str(column) for column in row.keys()}
        for alias in get_column_aliases(candidate):
            if alias in normalized_map:
                return normalized_map[alias]
        return None

    @classmethod
    def _resolve_row_value_by_priority(cls, row: dict[str, Any], candidate: str) -> object | None:
        normalized_map = {str(column).strip().upper(): str(column) for column in row.keys()}
        fallback_value: object | None = None
        for alias in get_column_aliases(candidate):
            if alias not in normalized_map:
                continue
            value = row.get(normalized_map[alias])
            if fallback_value is None:
                fallback_value = value
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return value
        return fallback_value

    @classmethod
    def _build_password_control_snapshots(cls, rows: list[dict[str, Any]]) -> dict[str, list[dict[str, str]]]:
        param_map: dict[str, object] = {}
        for row in rows:
            parameter_column = cls._find_row_column_by_alias(row, "PARAMETER")
            if not parameter_column:
                continue

            parameter_name = str(row.get(parameter_column, "")).strip().casefold()
            if not parameter_name:
                continue

            value = cls._resolve_row_value_by_priority(row, "VALUE")
            if value is None:
                continue

            param_map[parameter_name] = value

        snapshots: dict[str, list[dict[str, str]]] = {}
        for control_id, param_name, expected, rule_type, message in SAP_APP_RSPARAM_RULES:
            if param_name not in param_map:
                entry: dict[str, str] = {
                    "check_type": param_name,
                    "description": message,
                    "actual_value": "לא נמצא",
                    "expected_value": str(expected),
                    "status": "עם ממצא",
                    "full_description": f"הפרמטר {param_name} לא נמצא בדוח RSPARAM. הערך המצופה הוא {expected}.",
                }
                snapshots.setdefault(control_id, []).append(entry)
                continue

            actual = param_map[param_name]
            try:
                actual_float = float(str(actual).replace(",", "").strip())
                if rule_type == "minimum":
                    passes = actual_float >= float(expected)
                elif rule_type == "maximum":
                    passes = actual_float <= float(expected)
                else:
                    passes = True
            except (ValueError, TypeError):
                passes = False

            status = "תקין" if passes else "עם ממצא"
            if passes:
                full_desc = f"הערך בפועל עבור {param_name} הוא {actual}, בעוד שהערך המצופה הוא {expected}. ההגדרה תקינה לפי דרישת הבקרה."
            else:
                full_desc = f"הערך בפועל עבור {param_name} הוא {actual}, בעוד שהערך המצופה הוא {expected}. ההגדרה אינה עומדת בדרישת הבקרה."

            entry = {
                "check_type": param_name,
                "description": message,
                "actual_value": str(actual),
                "expected_value": str(expected),
                "status": status,
                "full_description": full_desc,
            }
            snapshots.setdefault(control_id, []).append(entry)

        return snapshots

    def _build_strong_profiles_permissions_section(self, parent_layout: QVBoxLayout) -> None:
        self.permissions_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי הרשאות - משתמשים חזקים"))
        self.permissions_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        permissions_summary_layout = QVBoxLayout(self.permissions_summary_group)
        permissions_summary_layout.setContentsMargins(8, 14, 8, 8)

        self.permissions_summary_table = QTableWidget(0, 6)
        self.permissions_summary_table.setItemDelegate(_RightAlignDelegate(self.permissions_summary_table))
        self.permissions_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("מזהה בקרה"),
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("ממצא"),
            self.format_rtl_text("כמות משתמשים"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("סטטוס"),
        ])
        _perm_summary_hdr = self.permissions_summary_table.horizontalHeader()
        _perm_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _perm_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _perm_summary_hdr.setStretchLastSection(False)
        self.permissions_summary_table.setColumnWidth(2, 220)  # ממצא
        self.permissions_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.permissions_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.permissions_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.permissions_summary_table.setAlternatingRowColors(True)
        self.permissions_summary_table.setMinimumHeight(160)
        self.permissions_summary_table.itemSelectionChanged.connect(self._refresh_selected_permissions_users)
        permissions_summary_layout.addWidget(self.permissions_summary_table)
        parent_layout.addWidget(self.permissions_summary_group)

        self.permissions_users_group = QGroupBox(self.format_ui_rtl_text("משתמשים הכלולים בממצא"))
        self.permissions_users_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        permissions_users_layout = QVBoxLayout(self.permissions_users_group)
        permissions_users_layout.setContentsMargins(8, 14, 8, 8)

        self.permissions_users_table = QTableWidget(0, 2)
        self.permissions_users_table.setItemDelegate(_RightAlignDelegate(self.permissions_users_table))
        self.permissions_users_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
        ])
        _perm_users_hdr = self.permissions_users_table.horizontalHeader()
        _perm_users_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _perm_users_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _perm_users_hdr.setStretchLastSection(True)  # משתמש
        self.permissions_users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.permissions_users_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.permissions_users_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.permissions_users_table.setAlternatingRowColors(True)
        self.permissions_users_table.setMinimumHeight(180)
        self.permissions_users_table.setToolTip(self.format_ui_rtl_text("לחיצה כפולה על משתמש תציג את הפרופילים החזקים שלו"))
        self.permissions_users_table.cellDoubleClicked.connect(self.show_permissions_user_profiles_dialog)
        permissions_users_layout.addWidget(self.permissions_users_table)
        parent_layout.addWidget(self.permissions_users_group)

        self._refresh_permissions_summary_table()

    def _build_user_mgmt_permissions_section(self, parent_layout: QVBoxLayout) -> None:
        self.user_mgmt_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי הרשאות - ניהול משתמשים"))
        self.user_mgmt_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        user_mgmt_summary_layout = QVBoxLayout(self.user_mgmt_summary_group)
        user_mgmt_summary_layout.setContentsMargins(8, 14, 8, 8)

        self.user_mgmt_summary_table = QTableWidget(0, 5)
        self.user_mgmt_summary_table.setItemDelegate(_RightAlignDelegate(self.user_mgmt_summary_table))
        self.user_mgmt_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("ממצא"),
            self.format_rtl_text("כמות משתמשים"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("סטטוס"),
        ])
        _usrmgmt_summary_hdr = self.user_mgmt_summary_table.horizontalHeader()
        _usrmgmt_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _usrmgmt_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _usrmgmt_summary_hdr.setStretchLastSection(False)
        self.user_mgmt_summary_table.setColumnWidth(1, 280)  # ממצא
        self.user_mgmt_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.user_mgmt_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.user_mgmt_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.user_mgmt_summary_table.setAlternatingRowColors(True)
        self.user_mgmt_summary_table.setMinimumHeight(160)
        self.user_mgmt_summary_table.setToolTip(
            self.format_ui_rtl_text("לחיצה על שורה תציג את המשתמשים בעלי הרשאת ניהול משתמשים")
        )
        self.user_mgmt_summary_table.itemSelectionChanged.connect(self._refresh_selected_user_mgmt_users)
        user_mgmt_summary_layout.addWidget(self.user_mgmt_summary_table)
        parent_layout.addWidget(self.user_mgmt_summary_group)

        self.user_mgmt_users_group = QGroupBox(self.format_ui_rtl_text("משתמשים בעלי הרשאת ניהול משתמשים"))
        self.user_mgmt_users_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        user_mgmt_users_layout = QVBoxLayout(self.user_mgmt_users_group)
        user_mgmt_users_layout.setContentsMargins(8, 14, 8, 8)

        self.user_mgmt_users_table = QTableWidget(0, 2)
        self.user_mgmt_users_table.setItemDelegate(_RightAlignDelegate(self.user_mgmt_users_table))
        self.user_mgmt_users_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
        ])
        _usrmgmt_users_hdr = self.user_mgmt_users_table.horizontalHeader()
        _usrmgmt_users_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _usrmgmt_users_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _usrmgmt_users_hdr.setStretchLastSection(True)
        self.user_mgmt_users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.user_mgmt_users_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.user_mgmt_users_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.user_mgmt_users_table.setAlternatingRowColors(True)
        self.user_mgmt_users_table.setMinimumHeight(180)
        self.user_mgmt_users_table.setToolTip(
            self.format_ui_rtl_text("לחיצה כפולה על משתמש תציג את הרולים המעניקים לו הרשאת ניהול משתמשים")
        )
        self.user_mgmt_users_table.cellDoubleClicked.connect(self.show_user_mgmt_user_dialog)
        user_mgmt_users_layout.addWidget(self.user_mgmt_users_table)
        parent_layout.addWidget(self.user_mgmt_users_group)

        self._refresh_user_mgmt_summary_table()

    def _upsert_permissions_control_data(
        self,
        slot_key: str,
        result: Any,
        audit_issues: list[ValidationIssue],
        extraction_date: str,
    ) -> None:
        del slot_key
        del extraction_date
        detected_profile = str(getattr(result, "detected_profile", "") or "").upper()
        control_id = "MA3-3_AYALON_14"
        if detected_profile not in ("UST04", "USH04"):
            return
        if control_id not in get_profile_audit_controls(detected_profile):
            return

        strong_issues = [issue for issue in audit_issues if issue.control_id == control_id]
        slot_users_by_client: dict[str, dict[str, set[str]]] = {}
        rows = list(getattr(result, "rows", []))

        for issue in strong_issues:
            client_name = "-"
            user_name = ""
            profile_names = {
                profile.strip().upper()
                for profile in str(issue.actual_value or "").split(",")
                if profile.strip()
            }

            if issue.row_number > 0 and issue.row_number <= len(rows):
                row = rows[issue.row_number - 1]
                row_client = self._resolve_row_value_by_priority(row, "MANDT")
                if row_client is not None and str(row_client).strip():
                    client_name = str(row_client).strip()
                row_user = self._resolve_row_value_by_priority(row, "BNAME")
                if row_user is not None:
                    user_name = str(row_user).strip().upper()
                if not profile_names:
                    row_profile = self._resolve_row_value_by_priority(row, "PROFILE")
                    if row_profile is not None:
                        profile_names.add(str(row_profile).strip().upper())
                    row_profiles = self._resolve_row_value_by_priority(row, "PROFS")
                    if row_profiles is not None:
                        profile_names.update(
                            profile.strip().upper()
                            for profile in str(row_profiles).split()
                            if profile.strip()
                        )
                    row_profiles_modbe = self._resolve_row_value_by_priority(row, "MODBE")
                    if row_profiles_modbe is not None:
                        profile_names.update(
                            profile.strip().upper()
                            for profile in str(row_profiles_modbe).split()
                            if profile.strip()
                        )

            if not user_name:
                continue

            client_users = slot_users_by_client.setdefault(client_name, {})
            client_users.setdefault(user_name, set())
            client_users[user_name].update(profile_names)

        # Store this source-profile's slice (replaces only data from THIS profile,
        # preserving the OTHER profile's findings - fixes the bug where loading USH04
        # wiped previously-loaded UST04 findings and vice versa).
        self._strong_profile_data[detected_profile] = slot_users_by_client

        # Merge UST04 + USH04 slices into a unified per-client view.
        merged_users_by_client: dict[str, dict[str, set[str]]] = {}
        for prof_bucket in self._strong_profile_data.values():
            for client_name, users in prof_bucket.items():
                client_users = merged_users_by_client.setdefault(client_name, {})
                for user_name, profiles in users.items():
                    client_users.setdefault(user_name, set()).update(profiles)

        keys_to_delete = [key for key in self.permissions_summary_records if str(key).startswith(f"{control_id}|")]
        for key in keys_to_delete:
            self.permissions_summary_records.pop(key, None)
            self.permissions_users_by_control.pop(key, None)

        control_meta = get_audit_control_definition(control_id)
        if not merged_users_by_client:
            record_key = f"{control_id}|-"
            self.permissions_summary_records[record_key] = {
                "record_key": record_key,
                "control_id": control_id,
                "client": "-",
                "finding_text": "נמצאו 0 משתמשים בעלי פרופילים חזקים",
                "users_count": 0,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "תקין",
            }
            self.permissions_users_by_control[record_key] = []
            return

        for client_name, client_users in sorted(merged_users_by_client.items(), key=lambda item: item[0]):
            users_count = len(client_users)
            record_key = f"{control_id}|{client_name}"
            self.permissions_summary_records[record_key] = {
                "record_key": record_key,
                "control_id": control_id,
                "client": client_name,
                "finding_text": f"נמצאו {users_count} משתמשים בעלי פרופילים חזקים",
                "users_count": users_count,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "עם ממצא" if users_count > 0 else "תקין",
            }
            self.permissions_users_by_control[record_key] = [
                {
                    "client": client_name,
                    "user_name": user_name,
                    "profiles": sorted(profiles),
                }
                for user_name, profiles in sorted(client_users.items())
            ]

    @staticmethod
    def _extract_control_id_from_record_key(record_key: object, fallback_control_id: str) -> str:
        key_text = str(record_key or "").strip()
        if "|" in key_text:
            parsed_control_id = key_text.split("|", 1)[0].strip()
            if parsed_control_id:
                return parsed_control_id
        return fallback_control_id

    @staticmethod
    def _to_int(value: object) -> int:
        try:
            return int(str(value).strip())
        except (TypeError, ValueError):
            return 0

    def _permission_control_slots(self, control_id: str) -> list[str]:
        control_slots: dict[str, list[str]] = {
            "MA3-3_AYALON_14": ["UST04", "USH04"],
            "MA1-1_AYALON_10": ["AGR_1251", "AGR_USERS"],
            "MA1-1_AYALON_11": ["AGR_1251", "AGR_USERS"],
            "MA1-1_AYALON_12": ["AGR_1251", "AGR_USERS"],
            "MA1-1_AYALON_16": ["AGR_1251", "AGR_USERS"],
            "MA1-1_AYALON_43": ["AGR_1251", "AGR_USERS"],
            "MA1-1_AYALON_45": ["AGR_1251", "AGR_USERS"],
            "MA1-1_AYALON_67": ["AGR_1251", "AGR_USERS"],
        }
        return control_slots.get(control_id, ["AGR_1251", "AGR_USERS"])

    def _permission_source_file_label(self, control_id: str) -> str:
        labels: list[str] = []
        for slot_key in self._permission_control_slots(control_id):
            label = self._get_slot_display_name(slot_key)
            if label not in labels:
                labels.append(label)
        return ", ".join(labels) if labels else "-"

    def _permission_extraction_date_label(self, control_id: str) -> str:
        extraction_dates: list[str] = []
        for slot_key in self._permission_control_slots(control_id):
            extraction_date = self._get_slot_extraction_date(slot_key)
            if extraction_date and extraction_date not in extraction_dates:
                extraction_dates.append(extraction_date)
        return ", ".join(extraction_dates) if extraction_dates else "-"

    def _permission_summary_sources(self) -> list[tuple[str, dict[str, dict[str, Any]]]]:
        return [
            ("MA3-3_AYALON_14", self.permissions_summary_records),
            ("MA1-1_AYALON_10", self.user_mgmt_summary_records),
            ("MA1-1_AYALON_11", self.auth_mgmt_summary_records),
            ("MA1-1_AYALON_12", self.rscdok99_summary_records),
            ("MA1-1_AYALON_16", self.data_mgmt_summary_records),
            ("MA1-1_AYALON_43", self.transport_summary_records),
            ("MA1-1_AYALON_45", self.debug_summary_records),
            ("MA1-1_AYALON_67", self.job_mgmt_summary_records),
        ]

    def _permission_user_sources(self) -> dict[str, dict[str, list[dict[str, Any]]]]:
        return {
            "MA3-3_AYALON_14": self.permissions_users_by_control,
            "MA1-1_AYALON_10": self.user_mgmt_users_by_control,
            "MA1-1_AYALON_11": self.auth_mgmt_users_by_control,
            "MA1-1_AYALON_12": self.rscdok99_users_by_control,
            "MA1-1_AYALON_16": self.data_mgmt_users_by_control,
            "MA1-1_AYALON_43": self.transport_users_by_control,
            "MA1-1_AYALON_45": self.debug_users_by_control,
            "MA1-1_AYALON_67": self.job_mgmt_users_by_control,
        }

    def _sync_permissions_findings_into_analysis_summary(self) -> None:
        permission_user_sources = self._permission_user_sources()
        permission_control_ids = [
            control_id
            for control_id, _records_map in self._permission_summary_sources()
        ]

        for control_id in permission_control_ids:
            self.audit_summary_records.pop(control_id, None)
            self.audit_details_by_control.pop(control_id, None)

        for fallback_control_id, records_map in self._permission_summary_sources():
            grouped_records: dict[str, list[dict[str, Any]]] = {}
            for row_data in records_map.values():
                control_id = str(row_data.get("control_id", "")).strip()
                if not control_id:
                    control_id = self._extract_control_id_from_record_key(
                        row_data.get("record_key", ""),
                        fallback_control_id,
                    )
                grouped_records.setdefault(control_id, []).append(row_data)

            for control_id, rows in grouped_records.items():
                if not rows:
                    continue

                control_meta = get_audit_control_definition(control_id)
                sorted_rows = sorted(rows, key=lambda item: str(item.get("client", "-")))
                user_map = permission_user_sources.get(control_id, {})
                all_users: set[tuple[str, str]] = set()
                finding_users: set[tuple[str, str]] = set()

                for row in sorted_rows:
                    record_key = str(row.get("record_key", "") or "")
                    users = user_map.get(record_key, [])
                    row_has_finding = (
                        self._to_int(row.get("users_count", 0)) > 0
                        or str(row.get("status", "")).strip() == "עם ממצא"
                    )
                    for user_data in users:
                        client_name = str(user_data.get("client", "-") or "-")
                        user_name = str(user_data.get("user_name", "-") or "-").strip()
                        if not user_name or user_name == "-":
                            continue
                        user_key = (client_name, user_name.upper())
                        all_users.add(user_key)
                        if row_has_finding:
                            finding_users.add(user_key)

                # For AGR-based controls: population = all unique users in AGR_USERS.
                # MA3-3_AYALON_14 uses UST04/USH04 data, so it keeps the current logic.
                is_agr_control = (
                    control_id != "MA3-3_AYALON_14"
                    and bool(self.agr_users_population_by_mandt)
                )
                if is_agr_control:
                    total_records = sum(self.agr_users_population_by_mandt.values())
                    finding_records = len(all_users)
                elif all_users:
                    total_records = len(all_users)
                    finding_records = len(finding_users)
                else:
                    total_records = sum(max(self._to_int(row.get("users_count", 0)), 0) for row in sorted_rows)
                    finding_records = sum(
                        max(self._to_int(row.get("users_count", 0)), 0)
                        for row in sorted_rows
                        if str(row.get("status", "")).strip() == "עם ממצא"
                        or self._to_int(row.get("users_count", 0)) > 0
                    )
                valid_records = max(total_records - finding_records, 0)

                self.audit_summary_records[control_id] = {
                    "control_id": control_id,
                    "check_type": control_meta.get("check_type", "סקירת הרשאות"),
                    "source_file": self._permission_source_file_label(control_id),
                    "extraction_date": self._permission_extraction_date_label(control_id),
                    "work_environment": self._current_work_environment_label(),
                    "risk_level": control_meta.get("risk_level", "-"),
                    "description": control_meta.get("description", "-"),
                    "valid_records": valid_records,
                    "finding_records": finding_records,
                    "total_records": total_records,
                }
                source_file_label = self._permission_source_file_label(control_id)
                extraction_date_label = self._permission_extraction_date_label(control_id)
                detail_rows: list[dict[str, Any]] = []

                for row in sorted_rows:
                    record_key = str(row.get("record_key", "") or "")
                    users = user_map.get(record_key, [])
                    sorted_users = sorted(
                        users,
                        key=lambda item: (
                            str(item.get("client", "-")),
                            str(item.get("user_name", "-")),
                        ),
                    )
                    for user_data in sorted_users:
                        client_name = str(user_data.get("client", "-") or "-")
                        user_name = str(user_data.get("user_name", "-") or "-")
                        profiles_list = list(user_data.get("profiles") or [])
                        roles_list = list(user_data.get("roles") or [])
                        if profiles_list:
                            profiles_block = "\n".join(f"- {name}" for name in profiles_list)
                            full_desc_text = (
                                f"קליינט: {client_name}\n"
                                f"משתמש: {user_name}\n\n"
                                f"פרופילים חזקים:\n{profiles_block}"
                            )
                            actual_value_text = ", ".join(profiles_list)
                            expected_value_text = ""
                        elif roles_list:
                            role_lines: list[str] = []
                            role_names: list[str] = []
                            for role_entry in roles_list:
                                if isinstance(role_entry, dict):
                                    agr_name = str(role_entry.get("agr_name", "-") or "-")
                                    objects_seq = role_entry.get("objects") or []
                                else:
                                    agr_name = str(role_entry or "-")
                                    objects_seq = []
                                role_names.append(agr_name)
                                role_lines.append(f"- {agr_name}")
                                for obj_item in objects_seq:
                                    if isinstance(obj_item, (list, tuple)) and len(obj_item) >= 3:
                                        role_lines.append(
                                            f"    {obj_item[0]} | {obj_item[1]} | {obj_item[2]}"
                                        )
                                    elif isinstance(obj_item, dict):
                                        role_lines.append(
                                            "    "
                                            + " | ".join(
                                                str(obj_item.get(k, "-"))
                                                for k in ("object", "field", "low")
                                            )
                                        )
                                    else:
                                        role_lines.append(f"    {obj_item}")
                            roles_block = "\n".join(role_lines)
                            full_desc_text = (
                                f"קליינט: {client_name}\n"
                                f"משתמש: {user_name}\n\n"
                                f"רולים ואובייקטי הרשאה:\n{roles_block}"
                            )
                            actual_value_text = ", ".join(role_names) if role_names else "-"
                            expected_value_text = ""
                        else:
                            full_desc_text = (
                                f"משתמש: {user_name}. קליינט: {client_name}. "
                                f"{row.get('finding_text', '-')}."
                            )
                            actual_value_text = str(row.get("finding_text", "-") or "-")
                            expected_value_text = ""
                        # Extract unique auth objects from roles for working-paper "אובייקט הרשאה" column
                        _auth_objects: set[str] = set()
                        for _role in roles_list:
                            if isinstance(_role, dict):
                                for _obj_item in (_role.get("objects") or []):
                                    if isinstance(_obj_item, (list, tuple)) and len(_obj_item) > 0:
                                        _auth_objects.add(str(_obj_item[0]))
                                    elif isinstance(_obj_item, dict):
                                        _v = _obj_item.get("object") or _obj_item.get("OBJECT", "")
                                        if _v:
                                            _auth_objects.add(str(_v))
                        _auth_object_value = ", ".join(sorted(_auth_objects)) if _auth_objects else "-"
                        detail_rows.append(
                            {
                                "control_id": control_id,
                                "source_file": source_file_label,
                                "extraction_date": extraction_date_label,
                                "work_environment": self._current_work_environment_label(),
                                "category": control_meta.get("category", "-"),
                                "risk_level": row.get("risk_level", control_meta.get("risk_level", "-")),
                                "description": control_meta.get("description", "-"),
                                "check_type": control_meta.get("check_type", "סקירת הרשאות"),
                                "client": client_name,
                                "user_name": user_name,
                                "actual_value": actual_value_text,
                                "expected_value": expected_value_text,
                                "auth_object": _auth_object_value,
                                "status": "עם ממצא",
                                "full_description": full_desc_text,
                            }
                        )

                if not detail_rows:
                    detail_rows = [
                        {
                            "control_id": control_id,
                            "source_file": source_file_label,
                            "extraction_date": extraction_date_label,
                            "work_environment": self._current_work_environment_label(),
                            "category": control_meta.get("category", "-"),
                            "risk_level": control_meta.get("risk_level", "-"),
                            "description": control_meta.get("description", "-"),
                            "check_type": control_meta.get("check_type", "סקירת הרשאות"),
                            "client": "-",
                            "user_name": "-",
                            "actual_value": "-",
                            "expected_value": "לא נמצאו משתמשים",
                            "status": "תקין",
                            "full_description": "לא נמצאו משתמשים עם ממצא עבור בקרה זו.",
                        }
                    ]

                self.audit_details_by_control[control_id] = detail_rows

    def _refresh_permissions_summary_table(self) -> None:
        self.permissions_summary_table.setRowCount(0)
        self.permissions_users_table.setRowCount(0)
        if not self.permissions_summary_records:
            self.permissions_users_table.insertRow(0)
            item = QTableWidgetItem(self.format_rtl_text("אין משתמשים להצגה"))
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.permissions_users_table.setItem(0, 1, item)
            return

        for row_data in sorted(
            self.permissions_summary_records.values(),
            key=lambda item: (str(item.get("control_id", "")), str(item.get("client", ""))),
        ):
            row_index = self.permissions_summary_table.rowCount()
            self.permissions_summary_table.insertRow(row_index)
            values = [
                str(row_data.get("control_id", "-")),
                str(row_data.get("client", "-")),
                str(row_data.get("finding_text", "-")),
                str(row_data.get("users_count", 0)),
                str(row_data.get("risk_level", "-")),
                str(row_data.get("status", "-")),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("record_key", ""))
                self.permissions_summary_table.setItem(row_index, column, item)

        if self.permissions_summary_table.rowCount() > 0:
            self.permissions_summary_table.selectRow(0)
            self._refresh_selected_permissions_users()
        self.permissions_summary_table.resizeColumnsToContents()

    def _refresh_selected_permissions_users(self) -> None:
        selected_items = self.permissions_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.permissions_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.permissions_users_by_control.get(record_key, [])
        self.permissions_users_table.setRowCount(0)

        if not user_rows:
            self.permissions_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("לא נמצאו משתמשים להצגה"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.permissions_users_table.setItem(0, 1, empty_item)
            return

        for user_data in user_rows:
            row_index = self.permissions_users_table.rowCount()
            self.permissions_users_table.insertRow(row_index)
            client_name = str(user_data.get("client", "-") or "-")
            user_name = str(user_data.get("user_name", "-") or "-")
            client_item = QTableWidgetItem(self.format_rtl_text(client_name))
            client_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item = QTableWidgetItem(self.format_rtl_text(user_name))
            user_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item.setData(Qt.ItemDataRole.UserRole, {"client": client_name, "user": user_name})
            self.permissions_users_table.setItem(row_index, 0, client_item)
            self.permissions_users_table.setItem(row_index, 1, user_item)
        self.permissions_users_table.resizeColumnsToContents()

    def show_permissions_user_profiles_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.permissions_users_table.rowCount():
            return

        user_item = self.permissions_users_table.item(row_index, 1)
        if user_item is None:
            return

        user_payload = user_item.data(Qt.ItemDataRole.UserRole)
        user_name = ""
        client_name = "-"
        if isinstance(user_payload, dict):
            user_name = str(user_payload.get("user", "") or "").strip()
            client_name = str(user_payload.get("client", "-") or "-").strip()
        if not user_name:
            user_name = str(user_item.text() or "").strip()
        if not user_name or user_name.startswith("לא נמצאו") or user_name.startswith("אין "):
            return

        selected_items = self.permissions_summary_table.selectedItems()
        if not selected_items:
            return

        control_item = self.permissions_summary_table.item(selected_items[0].row(), 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.permissions_users_by_control.get(record_key, [])
        profiles: list[str] = []
        for user_data in user_rows:
            if (
                str(user_data.get("user_name", "")).strip().upper() == user_name.upper()
                and str(user_data.get("client", "-") or "-").strip() == client_name
            ):
                profiles = [str(profile) for profile in user_data.get("profiles", []) if str(profile).strip()]
                break

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט פרופילים חזקים למשתמש")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(640, 360)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        lines = [f"קליינט: {client_name}", f"משתמש: {user_name}", "", "פרופילים חזקים:"]
        if profiles:
            lines.extend(f"- {profile}" for profile in profiles)
        else:
            lines.append("- לא נמצאו פרופילים להצגה")
        details_box.setPlainText(self.format_rtl_text("\n".join(lines)))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    # ------------------------------------------------------------------
    # User-Management Permissions (MA1-1_AYALON_10)
    # Cross-join: AGR_1251 (permission objects) × AGR_USERS (role assignments)
    # ------------------------------------------------------------------

    def _compute_user_mgmt_permissions(self) -> None:
        """Recompute user-management permission findings from cached AGR_1251 + AGR_USERS rows."""
        if not self.agr_1251_cached_rows or not self.agr_users_cached_rows:
            return

        control_id = "MA1-1_AYALON_10"
        control_meta = get_audit_control_definition(control_id)

        # Step A: find AGR_NAMEs that carry user-management permission objects.
        # qualifying_values: (OBJECT_upper, FIELD_upper) -> set of qualifying LOW/HIGH values
        qualifying_map: dict[tuple[str, str], set[str]] = {
            (obj.upper(), fld.upper()): {v.upper() for v in vals}
            for (obj, fld), vals in USER_MGMT_PERMISSION_CRITERIA.items()
        }

        # agr_name_objects: AGR_NAME -> set of (OBJECT, FIELD, LOW_display) tuples that qualified
        agr_name_objects: dict[str, set[tuple[str, str, str]]] = {}
        for row in self.agr_1251_cached_rows:
            obj_val = self._resolve_row_value_by_priority(row, "OBJECT")
            fld_val = self._resolve_row_value_by_priority(row, "FIELD")
            low_val = self._resolve_row_value_by_priority(row, "LOW")
            high_val = self._resolve_row_value_by_priority(row, "HIGH")
            if obj_val is None or fld_val is None:
                continue
            obj_upper = str(obj_val).strip().upper()
            fld_upper = str(fld_val).strip().upper()
            key = (obj_upper, fld_upper)
            if key not in qualifying_map:
                continue
            low_str = str(low_val).strip().upper() if low_val is not None else ""
            high_str = str(high_val).strip().upper() if high_val is not None else ""
            # wildcard or value in qualifying set
            if low_str == "*" or high_str == "*":
                qualifies = True
            else:
                qualifies = bool(low_str and low_str in qualifying_map[key]) or bool(
                    high_str and high_str in qualifying_map[key]
                )
            if not qualifies:
                continue
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None or not str(agr_name_val).strip():
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            low_display = low_str if low_str else "-"
            agr_name_objects.setdefault(agr_name_upper, set()).add((obj_upper, fld_upper, low_display))

        matching_agr_names: set[str] = set(agr_name_objects.keys())

        # Step B: scan AGR_USERS; for each row whose AGR_NAME is in matching set,
        #         group by MANDT → UNAME → set of AGR_NAMEs.
        users_by_client: dict[str, dict[str, set[str]]] = {}
        for row in self.agr_users_cached_rows:
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None:
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            if agr_name_upper not in matching_agr_names:
                continue

            # Resolve MANDT with fallback to filename
            mandt_val = self._resolve_row_value_by_priority(row, "MANDT")
            if mandt_val is not None and str(mandt_val).strip():
                mandt = str(mandt_val).strip()
            else:
                source_file = str(row.get("__source_file", ""))
                digits_match = re.search(r"\d{3}", Path(source_file).name)
                mandt = digits_match.group(0) if digits_match else "-"

            uname_val = self._resolve_row_value_by_priority(row, "UNAME")
            if uname_val is None or not str(uname_val).strip():
                continue
            uname = str(uname_val).strip().upper()

            client_users = users_by_client.setdefault(mandt, {})
            client_users.setdefault(uname, set()).add(agr_name_upper)

        # Step C: rebuild summary/users dicts
        self.user_mgmt_summary_records.clear()
        self.user_mgmt_users_by_control.clear()

        if not users_by_client:
            record_key = f"{control_id}|-"
            self.user_mgmt_summary_records[record_key] = {
                "record_key": record_key,
                "client": "-",
                "finding_text": "לא נמצאו משתמשים בעלי הרשאות ניהול משתמשים",
                "users_count": 0,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "תקין",
            }
            self.user_mgmt_users_by_control[record_key] = []
        else:
            for mandt, client_users in sorted(users_by_client.items()):
                users_count = len(client_users)
                record_key = f"{control_id}|{mandt}"
                self.user_mgmt_summary_records[record_key] = {
                    "record_key": record_key,
                    "client": mandt,
                    "finding_text": f"נמצאו {users_count} משתמשים בעלי הרשאות ניהול משתמשים",
                    "users_count": users_count,
                    "risk_level": control_meta.get("risk_level", "-"),
                    "status": "עם ממצא" if users_count > 0 else "תקין",
                }
                self.user_mgmt_users_by_control[record_key] = [
                    {
                        "client": mandt,
                        "user_name": uname,
                        "roles": [
                            {
                                "agr_name": r,
                                "objects": sorted(agr_name_objects.get(r, set())),
                            }
                            for r in sorted(roles)
                        ],
                    }
                    for uname, roles in sorted(client_users.items())
                ]

        self._refresh_user_mgmt_summary_table()

    def _refresh_user_mgmt_summary_table(self) -> None:
        self.user_mgmt_summary_table.setRowCount(0)
        self.user_mgmt_users_table.setRowCount(0)
        if not self.user_mgmt_summary_records:
            self.user_mgmt_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("יש לטעון קבצי AGR_1251 ו-AGR_USERS"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.user_mgmt_users_table.setItem(0, 1, empty_item)
            return

        for row_data in sorted(
            self.user_mgmt_summary_records.values(),
            key=lambda item: str(item.get("client", "")),
        ):
            row_index = self.user_mgmt_summary_table.rowCount()
            self.user_mgmt_summary_table.insertRow(row_index)
            values = [
                str(row_data.get("client", "-")),
                str(row_data.get("finding_text", "-")),
                str(row_data.get("users_count", 0)),
                str(row_data.get("risk_level", "-")),
                str(row_data.get("status", "-")),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("record_key", ""))
                self.user_mgmt_summary_table.setItem(row_index, column, item)

        if self.user_mgmt_summary_table.rowCount() > 0:
            self.user_mgmt_summary_table.selectRow(0)
            self._refresh_selected_user_mgmt_users()
        self.user_mgmt_summary_table.resizeColumnsToContents()

    def _refresh_selected_user_mgmt_users(self) -> None:
        selected_items = self.user_mgmt_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.user_mgmt_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.user_mgmt_users_by_control.get(record_key, [])
        self.user_mgmt_users_table.setRowCount(0)

        if not user_rows:
            self.user_mgmt_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("לא נמצאו משתמשים להצגה"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.user_mgmt_users_table.setItem(0, 1, empty_item)
            return

        for user_data in user_rows:
            row_index = self.user_mgmt_users_table.rowCount()
            self.user_mgmt_users_table.insertRow(row_index)
            client_name = str(user_data.get("client", "-") or "-")
            user_name = str(user_data.get("user_name", "-") or "-")
            client_item = QTableWidgetItem(self.format_rtl_text(client_name))
            client_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item = QTableWidgetItem(self.format_rtl_text(user_name))
            user_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item.setData(Qt.ItemDataRole.UserRole, {"client": client_name, "user": user_name})
            self.user_mgmt_users_table.setItem(row_index, 0, client_item)
            self.user_mgmt_users_table.setItem(row_index, 1, user_item)
        self.user_mgmt_users_table.resizeColumnsToContents()

    def show_user_mgmt_user_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.user_mgmt_users_table.rowCount():
            return

        user_item = self.user_mgmt_users_table.item(row_index, 1)
        if user_item is None:
            return

        user_payload = user_item.data(Qt.ItemDataRole.UserRole)
        user_name = ""
        client_name = "-"
        if isinstance(user_payload, dict):
            user_name = str(user_payload.get("user", "") or "").strip()
            client_name = str(user_payload.get("client", "-") or "-").strip()
        if not user_name:
            user_name = str(user_item.text() or "").strip()
        if not user_name or user_name.startswith("לא נמצאו") or user_name.startswith("אין "):
            return

        selected_items = self.user_mgmt_summary_table.selectedItems()
        if not selected_items:
            return

        summary_item = self.user_mgmt_summary_table.item(selected_items[0].row(), 0)
        if summary_item is None:
            return

        record_key = str(summary_item.data(Qt.ItemDataRole.UserRole) or summary_item.text())
        user_rows = self.user_mgmt_users_by_control.get(record_key, [])
        roles: list[dict] = []
        for user_data in user_rows:
            if (
                str(user_data.get("user_name", "")).strip().upper() == user_name.upper()
                and str(user_data.get("client", "-") or "-").strip() == client_name
            ):
                roles = list(user_data.get("roles", []))
                break

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט הרשאות ניהול משתמשים")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(640, 480)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        lines = [
            f"קליינט: {client_name}",
            f"משתמש: {user_name}",
            "",
            "רולים ואובייקטי הרשאה:",
        ]
        if roles:
            for role_entry in roles:
                lines.append(f"- {role_entry.get('agr_name', '')}")
                for obj, fld, low in role_entry.get("objects", []):
                    lines.append(f"    {obj} | {fld} | {low}")
        else:
            lines.append("- לא נמצאו רולים להצגה")
        details_box.setPlainText(self.format_rtl_text("\n".join(lines)))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    # ------------------------------------------------------------------
    # Authorization-Management Permissions (MA-AUTHMGMT-01)
    # Cross-join: AGR_1251 (permission objects) × AGR_USERS (role assignments)
    # ------------------------------------------------------------------

    def _build_auth_mgmt_permissions_section(self, parent_layout: QVBoxLayout) -> None:
        self.auth_mgmt_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי הרשאות - ניהול הרשאות"))
        self.auth_mgmt_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        auth_mgmt_summary_layout = QVBoxLayout(self.auth_mgmt_summary_group)
        auth_mgmt_summary_layout.setContentsMargins(8, 14, 8, 8)

        self.auth_mgmt_summary_table = QTableWidget(0, 5)
        self.auth_mgmt_summary_table.setItemDelegate(_RightAlignDelegate(self.auth_mgmt_summary_table))
        self.auth_mgmt_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("ממצא"),
            self.format_rtl_text("כמות משתמשים"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("סטטוס"),
        ])
        _authmgmt_summary_hdr = self.auth_mgmt_summary_table.horizontalHeader()
        _authmgmt_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _authmgmt_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _authmgmt_summary_hdr.setStretchLastSection(False)
        self.auth_mgmt_summary_table.setColumnWidth(1, 280)  # ממצא
        self.auth_mgmt_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.auth_mgmt_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.auth_mgmt_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.auth_mgmt_summary_table.setAlternatingRowColors(True)
        self.auth_mgmt_summary_table.setMinimumHeight(160)
        self.auth_mgmt_summary_table.setToolTip(
            self.format_ui_rtl_text("לחיצה על שורה תציג את המשתמשים בעלי הרשאת ניהול הרשאות")
        )
        self.auth_mgmt_summary_table.itemSelectionChanged.connect(self._refresh_selected_auth_mgmt_users)
        auth_mgmt_summary_layout.addWidget(self.auth_mgmt_summary_table)
        parent_layout.addWidget(self.auth_mgmt_summary_group)

        self.auth_mgmt_users_group = QGroupBox(self.format_ui_rtl_text("משתמשים בעלי הרשאת ניהול הרשאות"))
        self.auth_mgmt_users_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        auth_mgmt_users_layout = QVBoxLayout(self.auth_mgmt_users_group)
        auth_mgmt_users_layout.setContentsMargins(8, 14, 8, 8)

        self.auth_mgmt_users_table = QTableWidget(0, 2)
        self.auth_mgmt_users_table.setItemDelegate(_RightAlignDelegate(self.auth_mgmt_users_table))
        self.auth_mgmt_users_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
        ])
        _authmgmt_users_hdr = self.auth_mgmt_users_table.horizontalHeader()
        _authmgmt_users_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _authmgmt_users_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _authmgmt_users_hdr.setStretchLastSection(True)
        self.auth_mgmt_users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.auth_mgmt_users_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.auth_mgmt_users_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.auth_mgmt_users_table.setAlternatingRowColors(True)
        self.auth_mgmt_users_table.setMinimumHeight(180)
        self.auth_mgmt_users_table.setToolTip(
            self.format_ui_rtl_text("לחיצה כפולה על משתמש תציג את הרולים המעניקים לו הרשאת ניהול הרשאות")
        )
        self.auth_mgmt_users_table.cellDoubleClicked.connect(self.show_auth_mgmt_user_dialog)
        auth_mgmt_users_layout.addWidget(self.auth_mgmt_users_table)
        parent_layout.addWidget(self.auth_mgmt_users_group)

        self._refresh_auth_mgmt_summary_table()

    def _compute_auth_mgmt_permissions(self) -> None:
        """Recompute authorization-management permission findings from cached AGR_1251 + AGR_USERS rows."""
        if not self.agr_1251_cached_rows or not self.agr_users_cached_rows:
            return

        control_id = "MA1-1_AYALON_11"
        control_meta = get_audit_control_definition(control_id)

        qualifying_map: dict[tuple[str, str], set[str]] = {
            (obj.upper(), fld.upper()): {v.upper() for v in vals}
            for (obj, fld), vals in AUTH_MGMT_PERMISSION_CRITERIA.items()
        }

        # agr_name_objects: AGR_NAME -> set of (OBJECT, FIELD, LOW_display) tuples that qualified
        agr_name_objects: dict[str, set[tuple[str, str, str]]] = {}
        for row in self.agr_1251_cached_rows:
            obj_val = self._resolve_row_value_by_priority(row, "OBJECT")
            fld_val = self._resolve_row_value_by_priority(row, "FIELD")
            low_val = self._resolve_row_value_by_priority(row, "LOW")
            high_val = self._resolve_row_value_by_priority(row, "HIGH")
            if obj_val is None or fld_val is None:
                continue
            obj_upper = str(obj_val).strip().upper()
            fld_upper = str(fld_val).strip().upper()
            key = (obj_upper, fld_upper)
            if key not in qualifying_map:
                continue
            low_str = str(low_val).strip().upper() if low_val is not None else ""
            high_str = str(high_val).strip().upper() if high_val is not None else ""
            if low_str == "*" or high_str == "*":
                qualifies = True
            else:
                qualifies = bool(low_str and low_str in qualifying_map[key]) or bool(
                    high_str and high_str in qualifying_map[key]
                )
            if not qualifies:
                continue
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None or not str(agr_name_val).strip():
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            low_display = low_str if low_str else "-"
            agr_name_objects.setdefault(agr_name_upper, set()).add((obj_upper, fld_upper, low_display))

        matching_agr_names: set[str] = set(agr_name_objects.keys())

        users_by_client: dict[str, dict[str, set[str]]] = {}
        for row in self.agr_users_cached_rows:
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None:
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            if agr_name_upper not in matching_agr_names:
                continue

            mandt_val = self._resolve_row_value_by_priority(row, "MANDT")
            if mandt_val is not None and str(mandt_val).strip():
                mandt = str(mandt_val).strip()
            else:
                source_file = str(row.get("__source_file", ""))
                digits_match = re.search(r"\d{3}", Path(source_file).name)
                mandt = digits_match.group(0) if digits_match else "-"

            uname_val = self._resolve_row_value_by_priority(row, "UNAME")
            if uname_val is None or not str(uname_val).strip():
                continue
            uname = str(uname_val).strip().upper()

            client_users = users_by_client.setdefault(mandt, {})
            client_users.setdefault(uname, set()).add(agr_name_upper)

        self.auth_mgmt_summary_records.clear()
        self.auth_mgmt_users_by_control.clear()

        if not users_by_client:
            record_key = f"{control_id}|-"
            self.auth_mgmt_summary_records[record_key] = {
                "record_key": record_key,
                "client": "-",
                "finding_text": "לא נמצאו משתמשים בעלי הרשאות ניהול הרשאות",
                "users_count": 0,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "תקין",
            }
            self.auth_mgmt_users_by_control[record_key] = []
        else:
            for mandt, client_users in sorted(users_by_client.items()):
                users_count = len(client_users)
                record_key = f"{control_id}|{mandt}"
                self.auth_mgmt_summary_records[record_key] = {
                    "record_key": record_key,
                    "client": mandt,
                    "finding_text": f"נמצאו {users_count} משתמשים בעלי הרשאות ניהול הרשאות",
                    "users_count": users_count,
                    "risk_level": control_meta.get("risk_level", "-"),
                    "status": "עם ממצא" if users_count > 0 else "תקין",
                }
                self.auth_mgmt_users_by_control[record_key] = [
                    {
                        "client": mandt,
                        "user_name": uname,
                        "roles": [
                            {
                                "agr_name": r,
                                "objects": sorted(agr_name_objects.get(r, set())),
                            }
                            for r in sorted(roles)
                        ],
                    }
                    for uname, roles in sorted(client_users.items())
                ]

        self._refresh_auth_mgmt_summary_table()

    def _refresh_auth_mgmt_summary_table(self) -> None:
        self.auth_mgmt_summary_table.setRowCount(0)
        self.auth_mgmt_users_table.setRowCount(0)
        if not self.auth_mgmt_summary_records:
            self.auth_mgmt_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("יש לטעון קבצי AGR_1251 ו-AGR_USERS"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.auth_mgmt_users_table.setItem(0, 1, empty_item)
            return

        for row_data in sorted(
            self.auth_mgmt_summary_records.values(),
            key=lambda item: str(item.get("client", "")),
        ):
            row_index = self.auth_mgmt_summary_table.rowCount()
            self.auth_mgmt_summary_table.insertRow(row_index)
            values = [
                str(row_data.get("client", "-")),
                str(row_data.get("finding_text", "-")),
                str(row_data.get("users_count", 0)),
                str(row_data.get("risk_level", "-")),
                str(row_data.get("status", "-")),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("record_key", ""))
                self.auth_mgmt_summary_table.setItem(row_index, column, item)

        if self.auth_mgmt_summary_table.rowCount() > 0:
            self.auth_mgmt_summary_table.selectRow(0)
            self._refresh_selected_auth_mgmt_users()
        self.auth_mgmt_summary_table.resizeColumnsToContents()

    def _refresh_selected_auth_mgmt_users(self) -> None:
        selected_items = self.auth_mgmt_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.auth_mgmt_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.auth_mgmt_users_by_control.get(record_key, [])
        self.auth_mgmt_users_table.setRowCount(0)

        if not user_rows:
            self.auth_mgmt_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("לא נמצאו משתמשים להצגה"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.auth_mgmt_users_table.setItem(0, 1, empty_item)
            return

        for user_data in user_rows:
            row_index = self.auth_mgmt_users_table.rowCount()
            self.auth_mgmt_users_table.insertRow(row_index)
            client_name = str(user_data.get("client", "-") or "-")
            user_name = str(user_data.get("user_name", "-") or "-")
            client_item = QTableWidgetItem(self.format_rtl_text(client_name))
            client_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item = QTableWidgetItem(self.format_rtl_text(user_name))
            user_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item.setData(Qt.ItemDataRole.UserRole, {"client": client_name, "user": user_name})
            self.auth_mgmt_users_table.setItem(row_index, 0, client_item)
            self.auth_mgmt_users_table.setItem(row_index, 1, user_item)
        self.auth_mgmt_users_table.resizeColumnsToContents()

    def show_auth_mgmt_user_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.auth_mgmt_users_table.rowCount():
            return

        user_item = self.auth_mgmt_users_table.item(row_index, 1)
        if user_item is None:
            return

        user_payload = user_item.data(Qt.ItemDataRole.UserRole)
        user_name = ""
        client_name = "-"
        if isinstance(user_payload, dict):
            user_name = str(user_payload.get("user", "") or "").strip()
            client_name = str(user_payload.get("client", "-") or "-").strip()
        if not user_name:
            user_name = str(user_item.text() or "").strip()
        if not user_name or user_name.startswith("לא נמצאו") or user_name.startswith("אין "):
            return

        selected_items = self.auth_mgmt_summary_table.selectedItems()
        if not selected_items:
            return

        summary_item = self.auth_mgmt_summary_table.item(selected_items[0].row(), 0)
        if summary_item is None:
            return

        record_key = str(summary_item.data(Qt.ItemDataRole.UserRole) or summary_item.text())
        user_rows = self.auth_mgmt_users_by_control.get(record_key, [])
        roles: list[dict] = []
        for user_data in user_rows:
            if (
                str(user_data.get("user_name", "")).strip().upper() == user_name.upper()
                and str(user_data.get("client", "-") or "-").strip() == client_name
            ):
                roles = list(user_data.get("roles", []))
                break

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט הרשאות ניהול הרשאות")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(640, 480)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        lines = [
            f"קליינט: {client_name}",
            f"משתמש: {user_name}",
            "",
            "רולים ואובייקטי הרשאה:",
        ]
        if roles:
            for role_entry in roles:
                lines.append(f"- {role_entry.get('agr_name', '')}")
                for obj, fld, low in role_entry.get("objects", []):
                    lines.append(f"    {obj} | {fld} | {low}")
        else:
            lines.append("- לא נמצאו רולים להצגה")
        details_box.setPlainText(self.format_rtl_text("\n".join(lines)))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    # ------------------------------------------------------------------
    # RSCDOK99 Program Permissions (MA1-1_AYALON_12)
    # Cross-join: AGR_1251 (permission objects) × AGR_USERS (role assignments)
    # AND logic: an AGR_NAME qualifies only when it satisfies ALL criteria
    # (S_PROGRAM/P_GROUP=RSCDOK99 AND S_PROGRAM/P_ACTION=SUB).
    # ------------------------------------------------------------------

    def _build_rscdok99_permissions_section(self, parent_layout: QVBoxLayout) -> None:
        self.rscdok99_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי הרשאות - תוכנית RSCDOK99"))
        self.rscdok99_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        rscdok99_summary_layout = QVBoxLayout(self.rscdok99_summary_group)
        rscdok99_summary_layout.setContentsMargins(8, 14, 8, 8)

        self.rscdok99_summary_table = QTableWidget(0, 5)
        self.rscdok99_summary_table.setItemDelegate(_RightAlignDelegate(self.rscdok99_summary_table))
        self.rscdok99_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("ממצא"),
            self.format_rtl_text("כמות משתמשים"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("סטטוס"),
        ])
        _rscdok99_summary_hdr = self.rscdok99_summary_table.horizontalHeader()
        _rscdok99_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _rscdok99_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _rscdok99_summary_hdr.setStretchLastSection(False)
        self.rscdok99_summary_table.setColumnWidth(1, 300)  # ממצא
        self.rscdok99_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.rscdok99_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.rscdok99_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.rscdok99_summary_table.setAlternatingRowColors(True)
        self.rscdok99_summary_table.setMinimumHeight(160)
        self.rscdok99_summary_table.setToolTip(
            self.format_ui_rtl_text("לחיצה על שורה תציג את המשתמשים בעלי הרשאה לתוכנית RSCDOK99")
        )
        self.rscdok99_summary_table.itemSelectionChanged.connect(self._refresh_selected_rscdok99_users)
        rscdok99_summary_layout.addWidget(self.rscdok99_summary_table)
        parent_layout.addWidget(self.rscdok99_summary_group)

        self.rscdok99_users_group = QGroupBox(self.format_ui_rtl_text("משתמשים בעלי הרשאה לתוכנית RSCDOK99"))
        self.rscdok99_users_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        rscdok99_users_layout = QVBoxLayout(self.rscdok99_users_group)
        rscdok99_users_layout.setContentsMargins(8, 14, 8, 8)

        self.rscdok99_users_table = QTableWidget(0, 2)
        self.rscdok99_users_table.setItemDelegate(_RightAlignDelegate(self.rscdok99_users_table))
        self.rscdok99_users_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
        ])
        _rscdok99_users_hdr = self.rscdok99_users_table.horizontalHeader()
        _rscdok99_users_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _rscdok99_users_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _rscdok99_users_hdr.setStretchLastSection(True)
        self.rscdok99_users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.rscdok99_users_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.rscdok99_users_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.rscdok99_users_table.setAlternatingRowColors(True)
        self.rscdok99_users_table.setMinimumHeight(180)
        self.rscdok99_users_table.setToolTip(
            self.format_ui_rtl_text("לחיצה כפולה על משתמש תציג את הרולים המעניקים לו הרשאה לתוכנית RSCDOK99")
        )
        self.rscdok99_users_table.cellDoubleClicked.connect(self.show_rscdok99_user_dialog)
        rscdok99_users_layout.addWidget(self.rscdok99_users_table)
        parent_layout.addWidget(self.rscdok99_users_group)

        self._refresh_rscdok99_summary_table()

    def _compute_rscdok99_permissions(self) -> None:
        """Recompute RSCDOK99 permission findings from cached AGR_1251 + AGR_USERS rows.

        An AGR_NAME qualifies only when ALL criteria in RSCDOK99_PERMISSION_CRITERIA
        are satisfied (AND logic): both S_PROGRAM/P_GROUP=RSCDOK99 and
        S_PROGRAM/P_ACTION=SUB must appear in the role's permission rows.
        """
        if not self.agr_1251_cached_rows or not self.agr_users_cached_rows:
            return

        control_id = "MA1-1_AYALON_12"
        control_meta = get_audit_control_definition(control_id)

        # Build per-AGR_NAME satisfied-criteria tracking.
        # criteria_count[agr_name] = set of criterion indexes satisfied
        # agr_name_objects[agr_name] = set of (OBJECT, FIELD, LOW_display) tuples that qualified
        criteria_count: dict[str, set[int]] = {}
        agr_name_objects: dict[str, set[tuple[str, str, str]]] = {}
        total_criteria = len(RSCDOK99_PERMISSION_CRITERIA)

        for row in self.agr_1251_cached_rows:
            obj_val = self._resolve_row_value_by_priority(row, "OBJECT")
            fld_val = self._resolve_row_value_by_priority(row, "FIELD")
            if obj_val is None or fld_val is None:
                continue
            obj_upper = str(obj_val).strip().upper()
            fld_upper = str(fld_val).strip().upper()
            low_val = self._resolve_row_value_by_priority(row, "LOW")
            high_val = self._resolve_row_value_by_priority(row, "HIGH")
            low_str = str(low_val).strip().upper() if low_val is not None else ""
            high_str = str(high_val).strip().upper() if high_val is not None else ""

            for idx, (crit_obj, crit_fld, crit_values) in enumerate(RSCDOK99_PERMISSION_CRITERIA):
                if obj_upper != crit_obj.upper() or fld_upper != crit_fld.upper():
                    continue
                crit_upper = {v.upper() for v in crit_values}
                if low_str == "*" or high_str == "*":
                    qualifies = True
                else:
                    qualifies = bool(low_str and low_str in crit_upper) or bool(
                        high_str and high_str in crit_upper
                    )
                if not qualifies:
                    continue
                agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
                if agr_name_val is None or not str(agr_name_val).strip():
                    continue
                agr_name_upper = str(agr_name_val).strip().upper()
                criteria_count.setdefault(agr_name_upper, set()).add(idx)
                low_display = low_str if low_str else "-"
                agr_name_objects.setdefault(agr_name_upper, set()).add((obj_upper, fld_upper, low_display))

        # Only keep roles that satisfy ALL criteria
        matching_agr_names: set[str] = {
            agr for agr, satisfied in criteria_count.items()
            if len(satisfied) >= total_criteria
        }

        users_by_client: dict[str, dict[str, set[str]]] = {}
        for row in self.agr_users_cached_rows:
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None:
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            if agr_name_upper not in matching_agr_names:
                continue

            mandt_val = self._resolve_row_value_by_priority(row, "MANDT")
            if mandt_val is not None and str(mandt_val).strip():
                mandt = str(mandt_val).strip()
            else:
                source_file = str(row.get("__source_file", ""))
                digits_match = re.search(r"\d{3}", Path(source_file).name)
                mandt = digits_match.group(0) if digits_match else "-"

            uname_val = self._resolve_row_value_by_priority(row, "UNAME")
            if uname_val is None or not str(uname_val).strip():
                continue
            uname = str(uname_val).strip().upper()

            client_users = users_by_client.setdefault(mandt, {})
            client_users.setdefault(uname, set()).add(agr_name_upper)

        self.rscdok99_summary_records.clear()
        self.rscdok99_users_by_control.clear()

        if not users_by_client:
            record_key = f"{control_id}|-"
            self.rscdok99_summary_records[record_key] = {
                "record_key": record_key,
                "client": "-",
                "finding_text": "לא נמצאו משתמשים בעלי הרשאה לתוכנית RSCDOK99",
                "users_count": 0,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "תקין",
            }
            self.rscdok99_users_by_control[record_key] = []
        else:
            for mandt, client_users in sorted(users_by_client.items()):
                users_count = len(client_users)
                record_key = f"{control_id}|{mandt}"
                self.rscdok99_summary_records[record_key] = {
                    "record_key": record_key,
                    "client": mandt,
                    "finding_text": f"נמצאו {users_count} משתמשים בעלי הרשאה לתוכנית RSCDOK99",
                    "users_count": users_count,
                    "risk_level": control_meta.get("risk_level", "-"),
                    "status": "עם ממצא" if users_count > 0 else "תקין",
                }
                self.rscdok99_users_by_control[record_key] = [
                    {
                        "client": mandt,
                        "user_name": uname,
                        "roles": [
                            {
                                "agr_name": r,
                                "objects": sorted(agr_name_objects.get(r, set())),
                            }
                            for r in sorted(roles)
                        ],
                    }
                    for uname, roles in sorted(client_users.items())
                ]

        self._refresh_rscdok99_summary_table()

    def _refresh_rscdok99_summary_table(self) -> None:
        self.rscdok99_summary_table.setRowCount(0)
        self.rscdok99_users_table.setRowCount(0)
        if not self.rscdok99_summary_records:
            self.rscdok99_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("יש לטעון קבצי AGR_1251 ו-AGR_USERS"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.rscdok99_users_table.setItem(0, 1, empty_item)
            return

        for row_data in sorted(
            self.rscdok99_summary_records.values(),
            key=lambda item: str(item.get("client", "")),
        ):
            row_index = self.rscdok99_summary_table.rowCount()
            self.rscdok99_summary_table.insertRow(row_index)
            values = [
                str(row_data.get("client", "-")),
                str(row_data.get("finding_text", "-")),
                str(row_data.get("users_count", 0)),
                str(row_data.get("risk_level", "-")),
                str(row_data.get("status", "-")),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("record_key", ""))
                self.rscdok99_summary_table.setItem(row_index, column, item)

        if self.rscdok99_summary_table.rowCount() > 0:
            self.rscdok99_summary_table.selectRow(0)
            self._refresh_selected_rscdok99_users()
        self.rscdok99_summary_table.resizeColumnsToContents()

    def _refresh_selected_rscdok99_users(self) -> None:
        selected_items = self.rscdok99_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.rscdok99_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.rscdok99_users_by_control.get(record_key, [])
        self.rscdok99_users_table.setRowCount(0)

        if not user_rows:
            self.rscdok99_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("לא נמצאו משתמשים להצגה"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.rscdok99_users_table.setItem(0, 1, empty_item)
            return

        for user_data in user_rows:
            row_index = self.rscdok99_users_table.rowCount()
            self.rscdok99_users_table.insertRow(row_index)
            client_name = str(user_data.get("client", "-") or "-")
            user_name = str(user_data.get("user_name", "-") or "-")
            client_item = QTableWidgetItem(self.format_rtl_text(client_name))
            client_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item = QTableWidgetItem(self.format_rtl_text(user_name))
            user_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item.setData(Qt.ItemDataRole.UserRole, {"client": client_name, "user": user_name})
            self.rscdok99_users_table.setItem(row_index, 0, client_item)
            self.rscdok99_users_table.setItem(row_index, 1, user_item)
        self.rscdok99_users_table.resizeColumnsToContents()

    def show_rscdok99_user_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.rscdok99_users_table.rowCount():
            return

        user_item = self.rscdok99_users_table.item(row_index, 1)
        if user_item is None:
            return

        user_payload = user_item.data(Qt.ItemDataRole.UserRole)
        user_name = ""
        client_name = "-"
        if isinstance(user_payload, dict):
            user_name = str(user_payload.get("user", "") or "").strip()
            client_name = str(user_payload.get("client", "-") or "-").strip()
        if not user_name:
            user_name = str(user_item.text() or "").strip()
        if not user_name or user_name.startswith("לא נמצאו") or user_name.startswith("אין "):
            return

        selected_items = self.rscdok99_summary_table.selectedItems()
        if not selected_items:
            return

        summary_item = self.rscdok99_summary_table.item(selected_items[0].row(), 0)
        if summary_item is None:
            return

        record_key = str(summary_item.data(Qt.ItemDataRole.UserRole) or summary_item.text())
        user_rows = self.rscdok99_users_by_control.get(record_key, [])
        roles: list[dict] = []
        for user_data in user_rows:
            if (
                str(user_data.get("user_name", "")).strip().upper() == user_name.upper()
                and str(user_data.get("client", "-") or "-").strip() == client_name
            ):
                roles = list(user_data.get("roles", []))
                break

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט הרשאה לתוכנית RSCDOK99")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(640, 480)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        lines = [
            f"קליינט: {client_name}",
            f"משתמש: {user_name}",
            "",
            "רולים ואובייקטי הרשאה:",
        ]
        if roles:
            for role_entry in roles:
                lines.append(f"- {role_entry.get('agr_name', '')}")
                for obj, fld, low in role_entry.get("objects", []):
                    lines.append(f"    {obj} | {fld} | {low}")
        else:
            lines.append("- לא נמצאו רולים להצגה")
        details_box.setPlainText(self.format_rtl_text("\n".join(lines)))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    # ------------------------------------------------------------------
    # Data Management Permissions (MA-DATAMGMT-01)
    # Cross-join: AGR_1251 (permission objects) × AGR_USERS (role assignments)
    # OR logic: any single matching AGR_1251 row qualifies the AGR_NAME.
    # Qualifying objects: S_TCODE/TCD (SE16/SM30/SM31/SE16N/SE17/SM38/SE37),
    #   S_TABU_DIS/ACTVT (01/02/06), S_TABU_NAM/ACTVT (01/02/06),
    #   S_TABU_NAM/TABLE (any value), S_TABU_CLI/CLIIDMAINT (X),
    #   S_DATASET/ACTVT (06/34).
    # ------------------------------------------------------------------

    def _build_data_mgmt_permissions_section(self, parent_layout: QVBoxLayout) -> None:
        self.data_mgmt_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי הרשאות - ניהול נתונים"))
        self.data_mgmt_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        data_mgmt_summary_layout = QVBoxLayout(self.data_mgmt_summary_group)
        data_mgmt_summary_layout.setContentsMargins(8, 14, 8, 8)

        self.data_mgmt_summary_table = QTableWidget(0, 5)
        self.data_mgmt_summary_table.setItemDelegate(_RightAlignDelegate(self.data_mgmt_summary_table))
        self.data_mgmt_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("ממצא"),
            self.format_rtl_text("כמות משתמשים"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("סטטוס"),
        ])
        _datamgmt_summary_hdr = self.data_mgmt_summary_table.horizontalHeader()
        _datamgmt_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _datamgmt_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _datamgmt_summary_hdr.setStretchLastSection(False)
        self.data_mgmt_summary_table.setColumnWidth(1, 300)  # ממצא
        self.data_mgmt_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.data_mgmt_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.data_mgmt_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.data_mgmt_summary_table.setAlternatingRowColors(True)
        self.data_mgmt_summary_table.setMinimumHeight(160)
        self.data_mgmt_summary_table.setToolTip(
            self.format_ui_rtl_text("לחיצה על שורה תציג את המשתמשים בעלי הרשאת ניהול נתונים")
        )
        self.data_mgmt_summary_table.itemSelectionChanged.connect(self._refresh_selected_data_mgmt_users)
        data_mgmt_summary_layout.addWidget(self.data_mgmt_summary_table)
        parent_layout.addWidget(self.data_mgmt_summary_group)

        self.data_mgmt_users_group = QGroupBox(self.format_ui_rtl_text("משתמשים בעלי הרשאת ניהול נתונים"))
        self.data_mgmt_users_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        data_mgmt_users_layout = QVBoxLayout(self.data_mgmt_users_group)
        data_mgmt_users_layout.setContentsMargins(8, 14, 8, 8)

        self.data_mgmt_users_table = QTableWidget(0, 2)
        self.data_mgmt_users_table.setItemDelegate(_RightAlignDelegate(self.data_mgmt_users_table))
        self.data_mgmt_users_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
        ])
        _datamgmt_users_hdr = self.data_mgmt_users_table.horizontalHeader()
        _datamgmt_users_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _datamgmt_users_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _datamgmt_users_hdr.setStretchLastSection(True)
        self.data_mgmt_users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.data_mgmt_users_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.data_mgmt_users_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.data_mgmt_users_table.setAlternatingRowColors(True)
        self.data_mgmt_users_table.setMinimumHeight(180)
        self.data_mgmt_users_table.setToolTip(
            self.format_ui_rtl_text("לחיצה כפולה על משתמש תציג את הרולים המעניקים לו הרשאת ניהול נתונים")
        )
        self.data_mgmt_users_table.cellDoubleClicked.connect(self.show_data_mgmt_user_dialog)
        data_mgmt_users_layout.addWidget(self.data_mgmt_users_table)
        parent_layout.addWidget(self.data_mgmt_users_group)

        self._refresh_data_mgmt_summary_table()

    def _compute_data_mgmt_permissions(self) -> None:
        """Recompute data-management permission findings from cached AGR_1251 + AGR_USERS rows.

        OR logic: any single qualifying row in AGR_1251 makes the AGR_NAME a match.
        Special wildcard criterion: if the qualifying set for a (OBJECT, FIELD) pair
        contains "*", then any non-empty LOW or HIGH value qualifies.
        """
        if not self.agr_1251_cached_rows or not self.agr_users_cached_rows:
            return

        control_id = "MA1-1_AYALON_16"
        control_meta = get_audit_control_definition(control_id)

        # Step A: find AGR_NAMEs that carry data-management permission objects.
        qualifying_map: dict[tuple[str, str], set[str]] = {
            (obj.upper(), fld.upper()): {v.upper() for v in vals}
            for (obj, fld), vals in DATA_MGMT_PERMISSION_CRITERIA.items()
        }

        # agr_name_objects: AGR_NAME -> set of (OBJECT, FIELD, LOW_display) tuples that qualified
        agr_name_objects: dict[str, set[tuple[str, str, str]]] = {}
        for row in self.agr_1251_cached_rows:
            obj_val = self._resolve_row_value_by_priority(row, "OBJECT")
            fld_val = self._resolve_row_value_by_priority(row, "FIELD")
            low_val = self._resolve_row_value_by_priority(row, "LOW")
            high_val = self._resolve_row_value_by_priority(row, "HIGH")
            if obj_val is None or fld_val is None:
                continue
            obj_upper = str(obj_val).strip().upper()
            fld_upper = str(fld_val).strip().upper()
            key = (obj_upper, fld_upper)
            if key not in qualifying_map:
                continue
            low_str = str(low_val).strip().upper() if low_val is not None else ""
            high_str = str(high_val).strip().upper() if high_val is not None else ""
            # SAP data has a wildcard → always qualifies
            if low_str == "*" or high_str == "*":
                qualifies = True
            elif "*" in qualifying_map[key]:
                # Criterion allows any value → any non-empty LOW or HIGH qualifies
                qualifies = bool(low_str) or bool(high_str)
            else:
                qualifies = bool(low_str and low_str in qualifying_map[key]) or bool(
                    high_str and high_str in qualifying_map[key]
                )
            if not qualifies:
                continue
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None or not str(agr_name_val).strip():
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            low_display = low_str if low_str else "-"
            agr_name_objects.setdefault(agr_name_upper, set()).add((obj_upper, fld_upper, low_display))

        matching_agr_names: set[str] = set(agr_name_objects.keys())

        # Step B: cross-join with AGR_USERS to collect users per client.
        users_by_client: dict[str, dict[str, set[str]]] = {}
        for row in self.agr_users_cached_rows:
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None:
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            if agr_name_upper not in matching_agr_names:
                continue

            mandt_val = self._resolve_row_value_by_priority(row, "MANDT")
            if mandt_val is not None and str(mandt_val).strip():
                mandt = str(mandt_val).strip()
            else:
                source_file = str(row.get("__source_file", ""))
                digits_match = re.search(r"\d{3}", Path(source_file).name)
                mandt = digits_match.group(0) if digits_match else "-"

            uname_val = self._resolve_row_value_by_priority(row, "UNAME")
            if uname_val is None or not str(uname_val).strip():
                continue
            uname = str(uname_val).strip().upper()

            client_users = users_by_client.setdefault(mandt, {})
            client_users.setdefault(uname, set()).add(agr_name_upper)

        # Step C+D: store results.
        self.data_mgmt_summary_records.clear()
        self.data_mgmt_users_by_control.clear()

        if not users_by_client:
            record_key = f"{control_id}|-"
            self.data_mgmt_summary_records[record_key] = {
                "record_key": record_key,
                "client": "-",
                "finding_text": "לא נמצאו משתמשים בעלי הרשאות ניהול נתונים",
                "users_count": 0,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "תקין",
            }
            self.data_mgmt_users_by_control[record_key] = []
        else:
            for mandt, client_users in sorted(users_by_client.items()):
                users_count = len(client_users)
                record_key = f"{control_id}|{mandt}"
                self.data_mgmt_summary_records[record_key] = {
                    "record_key": record_key,
                    "client": mandt,
                    "finding_text": f"נמצאו {users_count} משתמשים בעלי הרשאות ניהול נתונים",
                    "users_count": users_count,
                    "risk_level": control_meta.get("risk_level", "-"),
                    "status": "עם ממצא" if users_count > 0 else "תקין",
                }
                self.data_mgmt_users_by_control[record_key] = [
                    {
                        "client": mandt,
                        "user_name": uname,
                        "roles": [
                            {
                                "agr_name": r,
                                "objects": sorted(agr_name_objects.get(r, set())),
                            }
                            for r in sorted(roles)
                        ],
                    }
                    for uname, roles in sorted(client_users.items())
                ]

        self._refresh_data_mgmt_summary_table()

    def _refresh_data_mgmt_summary_table(self) -> None:
        self.data_mgmt_summary_table.setRowCount(0)
        self.data_mgmt_users_table.setRowCount(0)
        if not self.data_mgmt_summary_records:
            self.data_mgmt_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("יש לטעון קבצי AGR_1251 ו-AGR_USERS"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.data_mgmt_users_table.setItem(0, 1, empty_item)
            return

        for row_data in sorted(
            self.data_mgmt_summary_records.values(),
            key=lambda item: str(item.get("client", "")),
        ):
            row_index = self.data_mgmt_summary_table.rowCount()
            self.data_mgmt_summary_table.insertRow(row_index)
            values = [
                str(row_data.get("client", "-")),
                str(row_data.get("finding_text", "-")),
                str(row_data.get("users_count", 0)),
                str(row_data.get("risk_level", "-")),
                str(row_data.get("status", "-")),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("record_key", ""))
                self.data_mgmt_summary_table.setItem(row_index, column, item)

        if self.data_mgmt_summary_table.rowCount() > 0:
            self.data_mgmt_summary_table.selectRow(0)
            self._refresh_selected_data_mgmt_users()
        self.data_mgmt_summary_table.resizeColumnsToContents()

    def _refresh_selected_data_mgmt_users(self) -> None:
        selected_items = self.data_mgmt_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.data_mgmt_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.data_mgmt_users_by_control.get(record_key, [])
        self.data_mgmt_users_table.setRowCount(0)

        if not user_rows:
            self.data_mgmt_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("לא נמצאו משתמשים להצגה"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.data_mgmt_users_table.setItem(0, 1, empty_item)
            return

        for user_data in user_rows:
            row_index = self.data_mgmt_users_table.rowCount()
            self.data_mgmt_users_table.insertRow(row_index)
            client_name = str(user_data.get("client", "-") or "-")
            user_name = str(user_data.get("user_name", "-") or "-")
            client_item = QTableWidgetItem(self.format_rtl_text(client_name))
            client_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item = QTableWidgetItem(self.format_rtl_text(user_name))
            user_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item.setData(Qt.ItemDataRole.UserRole, {"client": client_name, "user": user_name})
            self.data_mgmt_users_table.setItem(row_index, 0, client_item)
            self.data_mgmt_users_table.setItem(row_index, 1, user_item)
        self.data_mgmt_users_table.resizeColumnsToContents()

    def show_data_mgmt_user_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.data_mgmt_users_table.rowCount():
            return

        user_item = self.data_mgmt_users_table.item(row_index, 1)
        if user_item is None:
            return

        user_payload = user_item.data(Qt.ItemDataRole.UserRole)
        user_name = ""
        client_name = "-"
        if isinstance(user_payload, dict):
            user_name = str(user_payload.get("user", "") or "").strip()
            client_name = str(user_payload.get("client", "-") or "-").strip()
        if not user_name:
            user_name = str(user_item.text() or "").strip()
        if not user_name or user_name.startswith("לא נמצאו") or user_name.startswith("אין "):
            return

        selected_items = self.data_mgmt_summary_table.selectedItems()
        if not selected_items:
            return

        summary_item = self.data_mgmt_summary_table.item(selected_items[0].row(), 0)
        if summary_item is None:
            return

        record_key = str(summary_item.data(Qt.ItemDataRole.UserRole) or summary_item.text())
        user_rows = self.data_mgmt_users_by_control.get(record_key, [])
        roles: list[dict] = []
        for user_data in user_rows:
            if (
                str(user_data.get("user_name", "")).strip().upper() == user_name.upper()
                and str(user_data.get("client", "-") or "-").strip() == client_name
            ):
                roles = list(user_data.get("roles", []))
                break

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט הרשאות ניהול נתונים")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(640, 480)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        lines = [
            f"קליינט: {client_name}",
            f"משתמש: {user_name}",
            "",
            "רולים ואובייקטי הרשאה:",
        ]
        if roles:
            for role_entry in roles:
                lines.append(f"- {role_entry.get('agr_name', '')}")
                for obj, fld, low in role_entry.get("objects", []):
                    lines.append(f"    {obj} | {fld} | {low}")
        else:
            lines.append("- לא נמצאו רולים להצגה")
        details_box.setPlainText(self.format_rtl_text("\n".join(lines)))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    # ------------------------------------------------------------------
    # Transport / Change-Management Permissions (MA-TRANSPORT-01)
    # Cross-join: AGR_1251 (permission objects) × AGR_USERS (role assignments)
    # OR logic: any single qualifying AGR_1251 row makes the AGR_NAME a match.
    # Qualifying objects: S_TCODE/TCD (STMS/STMS_IMPORT/SCC4),
    #   S_TABU_DIS/DICBERCLS=SS, S_TABU_DIS/ACTVT=02,
    #   S_TRANSPORT/ACTVT (01/02/50/60/06/43),
    #   S_CTS_ADMI/CTS_ADMFCT (IMPT/IMPA/IMP*).
    # ------------------------------------------------------------------

    def _build_transport_permissions_section(self, parent_layout: QVBoxLayout) -> None:
        self.transport_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי הרשאות - העברת שינויים"))
        self.transport_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        transport_summary_layout = QVBoxLayout(self.transport_summary_group)
        transport_summary_layout.setContentsMargins(8, 14, 8, 8)

        self.transport_summary_table = QTableWidget(0, 5)
        self.transport_summary_table.setItemDelegate(_RightAlignDelegate(self.transport_summary_table))
        self.transport_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("ממצא"),
            self.format_rtl_text("כמות משתמשים"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("סטטוס"),
        ])
        _transport_summary_hdr = self.transport_summary_table.horizontalHeader()
        _transport_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _transport_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _transport_summary_hdr.setStretchLastSection(False)
        self.transport_summary_table.setColumnWidth(1, 300)  # ממצא
        self.transport_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.transport_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.transport_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.transport_summary_table.setAlternatingRowColors(True)
        self.transport_summary_table.setMinimumHeight(160)
        self.transport_summary_table.setToolTip(
            self.format_ui_rtl_text("לחיצה על שורה תציג את המשתמשים בעלי הרשאת העברת שינויים")
        )
        self.transport_summary_table.itemSelectionChanged.connect(self._refresh_selected_transport_users)
        transport_summary_layout.addWidget(self.transport_summary_table)
        parent_layout.addWidget(self.transport_summary_group)

        self.transport_users_group = QGroupBox(self.format_ui_rtl_text("משתמשים בעלי הרשאת העברת שינויים"))
        self.transport_users_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        transport_users_layout = QVBoxLayout(self.transport_users_group)
        transport_users_layout.setContentsMargins(8, 14, 8, 8)

        self.transport_users_table = QTableWidget(0, 2)
        self.transport_users_table.setItemDelegate(_RightAlignDelegate(self.transport_users_table))
        self.transport_users_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
        ])
        _transport_users_hdr = self.transport_users_table.horizontalHeader()
        _transport_users_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _transport_users_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _transport_users_hdr.setStretchLastSection(True)
        self.transport_users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.transport_users_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.transport_users_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.transport_users_table.setAlternatingRowColors(True)
        self.transport_users_table.setMinimumHeight(180)
        self.transport_users_table.setToolTip(
            self.format_ui_rtl_text("לחיצה כפולה על משתמש תציג את הרולים המעניקים לו הרשאת העברת שינויים")
        )
        self.transport_users_table.cellDoubleClicked.connect(self.show_transport_user_dialog)
        transport_users_layout.addWidget(self.transport_users_table)
        parent_layout.addWidget(self.transport_users_group)

        self._refresh_transport_summary_table()

    def _compute_transport_permissions(self) -> None:
        """Recompute transport/change-management permission findings from cached AGR_1251 + AGR_USERS rows.

        OR logic: any single qualifying row in AGR_1251 makes the AGR_NAME a match.
        """
        if not self.agr_1251_cached_rows or not self.agr_users_cached_rows:
            return

        control_id = "MA1-1_AYALON_43"
        control_meta = get_audit_control_definition(control_id)

        # Step A: find AGR_NAMEs that carry transport permission objects.
        qualifying_map: dict[tuple[str, str], set[str]] = {
            (obj.upper(), fld.upper()): {v.upper() for v in vals}
            for (obj, fld), vals in TRANSPORT_PERMISSION_CRITERIA.items()
        }

        # agr_name_objects: AGR_NAME -> set of (OBJECT, FIELD, LOW_display) tuples that qualified
        agr_name_objects: dict[str, set[tuple[str, str, str]]] = {}
        for row in self.agr_1251_cached_rows:
            obj_val = self._resolve_row_value_by_priority(row, "OBJECT")
            fld_val = self._resolve_row_value_by_priority(row, "FIELD")
            low_val = self._resolve_row_value_by_priority(row, "LOW")
            high_val = self._resolve_row_value_by_priority(row, "HIGH")
            if obj_val is None or fld_val is None:
                continue
            obj_upper = str(obj_val).strip().upper()
            fld_upper = str(fld_val).strip().upper()
            key = (obj_upper, fld_upper)
            if key not in qualifying_map:
                continue
            low_str = str(low_val).strip().upper() if low_val is not None else ""
            high_str = str(high_val).strip().upper() if high_val is not None else ""
            if low_str == "*" or high_str == "*":
                qualifies = True
            else:
                qualifies = bool(low_str and low_str in qualifying_map[key]) or bool(
                    high_str and high_str in qualifying_map[key]
                )
            if not qualifies:
                continue
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None or not str(agr_name_val).strip():
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            low_display = low_str if low_str else "-"
            agr_name_objects.setdefault(agr_name_upper, set()).add((obj_upper, fld_upper, low_display))

        matching_agr_names: set[str] = set(agr_name_objects.keys())

        # Step B: cross-join with AGR_USERS to collect users per client.
        users_by_client: dict[str, dict[str, set[str]]] = {}
        for row in self.agr_users_cached_rows:
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None:
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            if agr_name_upper not in matching_agr_names:
                continue

            mandt_val = self._resolve_row_value_by_priority(row, "MANDT")
            if mandt_val is not None and str(mandt_val).strip():
                mandt = str(mandt_val).strip()
            else:
                source_file = str(row.get("__source_file", ""))
                digits_match = re.search(r"\d{3}", Path(source_file).name)
                mandt = digits_match.group(0) if digits_match else "-"

            uname_val = self._resolve_row_value_by_priority(row, "UNAME")
            if uname_val is None or not str(uname_val).strip():
                continue
            uname = str(uname_val).strip().upper()

            client_users = users_by_client.setdefault(mandt, {})
            client_users.setdefault(uname, set()).add(agr_name_upper)

        # Step C+D: store results.
        self.transport_summary_records.clear()
        self.transport_users_by_control.clear()

        if not users_by_client:
            record_key = f"{control_id}|-"
            self.transport_summary_records[record_key] = {
                "record_key": record_key,
                "client": "-",
                "finding_text": "לא נמצאו משתמשים בעלי הרשאת העברת שינויים",
                "users_count": 0,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "תקין",
            }
            self.transport_users_by_control[record_key] = []
        else:
            for mandt, client_users in sorted(users_by_client.items()):
                users_count = len(client_users)
                record_key = f"{control_id}|{mandt}"
                self.transport_summary_records[record_key] = {
                    "record_key": record_key,
                    "client": mandt,
                    "finding_text": f"נמצאו {users_count} משתמשים בעלי הרשאת העברת שינויים",
                    "users_count": users_count,
                    "risk_level": control_meta.get("risk_level", "-"),
                    "status": "עם ממצא" if users_count > 0 else "תקין",
                }
                self.transport_users_by_control[record_key] = [
                    {
                        "client": mandt,
                        "user_name": uname,
                        "roles": [
                            {
                                "agr_name": r,
                                "objects": sorted(agr_name_objects.get(r, set())),
                            }
                            for r in sorted(roles)
                        ],
                    }
                    for uname, roles in sorted(client_users.items())
                ]

        self._refresh_transport_summary_table()

    def _refresh_transport_summary_table(self) -> None:
        self.transport_summary_table.setRowCount(0)
        self.transport_users_table.setRowCount(0)
        if not self.transport_summary_records:
            self.transport_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("יש לטעון קבצי AGR_1251 ו-AGR_USERS"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.transport_users_table.setItem(0, 1, empty_item)
            return

        for row_data in sorted(
            self.transport_summary_records.values(),
            key=lambda item: str(item.get("client", "")),
        ):
            row_index = self.transport_summary_table.rowCount()
            self.transport_summary_table.insertRow(row_index)
            values = [
                str(row_data.get("client", "-")),
                str(row_data.get("finding_text", "-")),
                str(row_data.get("users_count", 0)),
                str(row_data.get("risk_level", "-")),
                str(row_data.get("status", "-")),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("record_key", ""))
                self.transport_summary_table.setItem(row_index, column, item)

        if self.transport_summary_table.rowCount() > 0:
            self.transport_summary_table.selectRow(0)
            self._refresh_selected_transport_users()
        self.transport_summary_table.resizeColumnsToContents()

    def _refresh_selected_transport_users(self) -> None:
        selected_items = self.transport_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.transport_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.transport_users_by_control.get(record_key, [])
        self.transport_users_table.setRowCount(0)

        if not user_rows:
            self.transport_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("לא נמצאו משתמשים להצגה"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.transport_users_table.setItem(0, 1, empty_item)
            return

        for user_data in user_rows:
            row_index = self.transport_users_table.rowCount()
            self.transport_users_table.insertRow(row_index)
            client_name = str(user_data.get("client", "-") or "-")
            user_name = str(user_data.get("user_name", "-") or "-")
            client_item = QTableWidgetItem(self.format_rtl_text(client_name))
            client_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item = QTableWidgetItem(self.format_rtl_text(user_name))
            user_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item.setData(Qt.ItemDataRole.UserRole, {"client": client_name, "user": user_name})
            self.transport_users_table.setItem(row_index, 0, client_item)
            self.transport_users_table.setItem(row_index, 1, user_item)
        self.transport_users_table.resizeColumnsToContents()

    def show_transport_user_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.transport_users_table.rowCount():
            return

        user_item = self.transport_users_table.item(row_index, 1)
        if user_item is None:
            return

        user_payload = user_item.data(Qt.ItemDataRole.UserRole)
        user_name = ""
        client_name = "-"
        if isinstance(user_payload, dict):
            user_name = str(user_payload.get("user", "") or "").strip()
            client_name = str(user_payload.get("client", "-") or "-").strip()
        if not user_name:
            user_name = str(user_item.text() or "").strip()
        if not user_name or user_name.startswith("לא נמצאו") or user_name.startswith("אין "):
            return

        selected_items = self.transport_summary_table.selectedItems()
        if not selected_items:
            return

        summary_item = self.transport_summary_table.item(selected_items[0].row(), 0)
        if summary_item is None:
            return

        record_key = str(summary_item.data(Qt.ItemDataRole.UserRole) or summary_item.text())
        user_rows = self.transport_users_by_control.get(record_key, [])
        roles: list[dict] = []
        for user_data in user_rows:
            if (
                str(user_data.get("user_name", "")).strip().upper() == user_name.upper()
                and str(user_data.get("client", "-") or "-").strip() == client_name
            ):
                roles = list(user_data.get("roles", []))
                break

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט הרשאת העברת שינויים")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(640, 480)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        lines = [
            f"קליינט: {client_name}",
            f"משתמש: {user_name}",
            "",
            "רולים ואובייקטי הרשאה:",
        ]
        if roles:
            for role_entry in roles:
                lines.append(f"- {role_entry.get('agr_name', '')}")
                for obj, fld, low in role_entry.get("objects", []):
                    lines.append(f"    {obj} | {fld} | {low}")
        else:
            lines.append("- לא נמצאו רולים להצגה")
        details_box.setPlainText(self.format_rtl_text("\n".join(lines)))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    # ------------------------------------------------------------------
    # DEBUG Permissions (MA-DEBUG-01)
    # Cross-join: AGR_1251 (permission objects) × AGR_USERS (role assignments)
    # OR logic: any single qualifying AGR_1251 row makes the AGR_NAME a match.
    # Qualifying objects: S_TCODE/TCD (SE38/SA38/SE80/ST05),
    #   S_DEVELOP/OBJTYPE=DEBUG, S_DEVELOP/ACTVT (01/02),
    #   S_PROGRAM/P_ACTION=SUB, S_PROGRAM/P_GROUP (any value),
    #   S_ADMI_FCD/S_ADMI_FCD=PADM.
    # ------------------------------------------------------------------

    def _build_debug_permissions_section(self, parent_layout: QVBoxLayout) -> None:
        self.debug_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי הרשאות - שימוש ב-DEBUG"))
        self.debug_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        debug_summary_layout = QVBoxLayout(self.debug_summary_group)
        debug_summary_layout.setContentsMargins(8, 14, 8, 8)

        self.debug_summary_table = QTableWidget(0, 5)
        self.debug_summary_table.setItemDelegate(_RightAlignDelegate(self.debug_summary_table))
        self.debug_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("ממצא"),
            self.format_rtl_text("כמות משתמשים"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("סטטוס"),
        ])
        _debug_summary_hdr = self.debug_summary_table.horizontalHeader()
        _debug_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _debug_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _debug_summary_hdr.setStretchLastSection(False)
        self.debug_summary_table.setColumnWidth(1, 300)  # ממצא
        self.debug_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.debug_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.debug_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.debug_summary_table.setAlternatingRowColors(True)
        self.debug_summary_table.setMinimumHeight(160)
        self.debug_summary_table.setToolTip(
            self.format_ui_rtl_text("לחיצה על שורה תציג את המשתמשים בעלי הרשאות DEBUG")
        )
        self.debug_summary_table.itemSelectionChanged.connect(self._refresh_selected_debug_users)
        debug_summary_layout.addWidget(self.debug_summary_table)
        parent_layout.addWidget(self.debug_summary_group)

        self.debug_users_group = QGroupBox(self.format_ui_rtl_text("משתמשים בעלי הרשאות DEBUG"))
        self.debug_users_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        debug_users_layout = QVBoxLayout(self.debug_users_group)
        debug_users_layout.setContentsMargins(8, 14, 8, 8)

        self.debug_users_table = QTableWidget(0, 2)
        self.debug_users_table.setItemDelegate(_RightAlignDelegate(self.debug_users_table))
        self.debug_users_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
        ])
        _debug_users_hdr = self.debug_users_table.horizontalHeader()
        _debug_users_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _debug_users_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _debug_users_hdr.setStretchLastSection(True)
        self.debug_users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.debug_users_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.debug_users_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.debug_users_table.setAlternatingRowColors(True)
        self.debug_users_table.setMinimumHeight(180)
        self.debug_users_table.setToolTip(
            self.format_ui_rtl_text("לחיצה כפולה על משתמש תציג את הרולים המעניקים לו הרשאות DEBUG")
        )
        self.debug_users_table.cellDoubleClicked.connect(self.show_debug_user_dialog)
        debug_users_layout.addWidget(self.debug_users_table)
        parent_layout.addWidget(self.debug_users_group)

        self._refresh_debug_summary_table()

    def _compute_debug_permissions(self) -> None:
        """Recompute DEBUG permission findings from cached AGR_1251 + AGR_USERS rows.

        OR logic: any single qualifying row in AGR_1251 makes the AGR_NAME a match.
        Special wildcard criterion: if the qualifying set for a (OBJECT, FIELD) pair
        contains "*", then any non-empty LOW or HIGH value qualifies.
        """
        if not self.agr_1251_cached_rows or not self.agr_users_cached_rows:
            return

        control_id = "MA1-1_AYALON_45"
        control_meta = get_audit_control_definition(control_id)

        # Step A: find AGR_NAMEs that carry DEBUG permission objects.
        qualifying_map: dict[tuple[str, str], set[str]] = {
            (obj.upper(), fld.upper()): {v.upper() for v in vals}
            for (obj, fld), vals in DEBUG_PERMISSION_CRITERIA.items()
        }

        # agr_name_objects: AGR_NAME -> set of (OBJECT, FIELD, LOW_display) tuples that qualified
        agr_name_objects: dict[str, set[tuple[str, str, str]]] = {}
        for row in self.agr_1251_cached_rows:
            obj_val = self._resolve_row_value_by_priority(row, "OBJECT")
            fld_val = self._resolve_row_value_by_priority(row, "FIELD")
            low_val = self._resolve_row_value_by_priority(row, "LOW")
            high_val = self._resolve_row_value_by_priority(row, "HIGH")
            if obj_val is None or fld_val is None:
                continue
            obj_upper = str(obj_val).strip().upper()
            fld_upper = str(fld_val).strip().upper()
            key = (obj_upper, fld_upper)
            if key not in qualifying_map:
                continue
            low_str = str(low_val).strip().upper() if low_val is not None else ""
            high_str = str(high_val).strip().upper() if high_val is not None else ""
            if low_str == "*" or high_str == "*":
                qualifies = True
            elif "*" in qualifying_map[key]:
                # Criterion allows any value → any non-empty LOW or HIGH qualifies
                qualifies = bool(low_str) or bool(high_str)
            else:
                qualifies = bool(low_str and low_str in qualifying_map[key]) or bool(
                    high_str and high_str in qualifying_map[key]
                )
            if not qualifies:
                continue
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None or not str(agr_name_val).strip():
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            low_display = low_str if low_str else "-"
            agr_name_objects.setdefault(agr_name_upper, set()).add((obj_upper, fld_upper, low_display))

        matching_agr_names: set[str] = set(agr_name_objects.keys())

        # Step B: cross-join with AGR_USERS to collect users per client.
        users_by_client: dict[str, dict[str, set[str]]] = {}
        for row in self.agr_users_cached_rows:
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None:
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            if agr_name_upper not in matching_agr_names:
                continue

            mandt_val = self._resolve_row_value_by_priority(row, "MANDT")
            if mandt_val is not None and str(mandt_val).strip():
                mandt = str(mandt_val).strip()
            else:
                source_file = str(row.get("__source_file", ""))
                digits_match = re.search(r"\d{3}", Path(source_file).name)
                mandt = digits_match.group(0) if digits_match else "-"

            uname_val = self._resolve_row_value_by_priority(row, "UNAME")
            if uname_val is None or not str(uname_val).strip():
                continue
            uname = str(uname_val).strip().upper()

            client_users = users_by_client.setdefault(mandt, {})
            client_users.setdefault(uname, set()).add(agr_name_upper)

        # Step C+D: store results.
        self.debug_summary_records.clear()
        self.debug_users_by_control.clear()

        if not users_by_client:
            record_key = f"{control_id}|-"
            self.debug_summary_records[record_key] = {
                "record_key": record_key,
                "client": "-",
                "finding_text": "לא נמצאו משתמשים בעלי הרשאות DEBUG",
                "users_count": 0,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "תקין",
            }
            self.debug_users_by_control[record_key] = []
        else:
            for mandt, client_users in sorted(users_by_client.items()):
                users_count = len(client_users)
                record_key = f"{control_id}|{mandt}"
                self.debug_summary_records[record_key] = {
                    "record_key": record_key,
                    "client": mandt,
                    "finding_text": f"נמצאו {users_count} משתמשים בעלי הרשאות DEBUG",
                    "users_count": users_count,
                    "risk_level": control_meta.get("risk_level", "-"),
                    "status": "עם ממצא" if users_count > 0 else "תקין",
                }
                self.debug_users_by_control[record_key] = [
                    {
                        "client": mandt,
                        "user_name": uname,
                        "roles": [
                            {
                                "agr_name": r,
                                "objects": sorted(agr_name_objects.get(r, set())),
                            }
                            for r in sorted(roles)
                        ],
                    }
                    for uname, roles in sorted(client_users.items())
                ]

        self._refresh_debug_summary_table()

    def _refresh_debug_summary_table(self) -> None:
        self.debug_summary_table.setRowCount(0)
        self.debug_users_table.setRowCount(0)
        if not self.debug_summary_records:
            self.debug_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("יש לטעון קבצי AGR_1251 ו-AGR_USERS"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.debug_users_table.setItem(0, 1, empty_item)
            return

        for row_data in sorted(
            self.debug_summary_records.values(),
            key=lambda item: str(item.get("client", "")),
        ):
            row_index = self.debug_summary_table.rowCount()
            self.debug_summary_table.insertRow(row_index)
            values = [
                str(row_data.get("client", "-")),
                str(row_data.get("finding_text", "-")),
                str(row_data.get("users_count", 0)),
                str(row_data.get("risk_level", "-")),
                str(row_data.get("status", "-")),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("record_key", ""))
                self.debug_summary_table.setItem(row_index, column, item)

        if self.debug_summary_table.rowCount() > 0:
            self.debug_summary_table.selectRow(0)
            self._refresh_selected_debug_users()
        self.debug_summary_table.resizeColumnsToContents()

    def _refresh_selected_debug_users(self) -> None:
        selected_items = self.debug_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.debug_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.debug_users_by_control.get(record_key, [])
        self.debug_users_table.setRowCount(0)

        if not user_rows:
            self.debug_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("לא נמצאו משתמשים להצגה"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.debug_users_table.setItem(0, 1, empty_item)
            return

        for user_data in user_rows:
            row_index = self.debug_users_table.rowCount()
            self.debug_users_table.insertRow(row_index)
            client_name = str(user_data.get("client", "-") or "-")
            user_name = str(user_data.get("user_name", "-") or "-")
            client_item = QTableWidgetItem(self.format_rtl_text(client_name))
            client_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item = QTableWidgetItem(self.format_rtl_text(user_name))
            user_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item.setData(Qt.ItemDataRole.UserRole, {"client": client_name, "user": user_name})
            self.debug_users_table.setItem(row_index, 0, client_item)
            self.debug_users_table.setItem(row_index, 1, user_item)
        self.debug_users_table.resizeColumnsToContents()

    def show_debug_user_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.debug_users_table.rowCount():
            return

        user_item = self.debug_users_table.item(row_index, 1)
        if user_item is None:
            return

        user_payload = user_item.data(Qt.ItemDataRole.UserRole)
        user_name = ""
        client_name = "-"
        if isinstance(user_payload, dict):
            user_name = str(user_payload.get("user", "") or "").strip()
            client_name = str(user_payload.get("client", "-") or "-").strip()
        if not user_name:
            user_name = str(user_item.text() or "").strip()
        if not user_name or user_name.startswith("לא נמצאו") or user_name.startswith("אין "):
            return

        selected_items = self.debug_summary_table.selectedItems()
        if not selected_items:
            return

        summary_item = self.debug_summary_table.item(selected_items[0].row(), 0)
        if summary_item is None:
            return

        record_key = str(summary_item.data(Qt.ItemDataRole.UserRole) or summary_item.text())
        user_rows = self.debug_users_by_control.get(record_key, [])
        roles: list[dict] = []
        for user_data in user_rows:
            if (
                str(user_data.get("user_name", "")).strip().upper() == user_name.upper()
                and str(user_data.get("client", "-") or "-").strip() == client_name
            ):
                roles = list(user_data.get("roles", []))
                break

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט הרשאות DEBUG")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(640, 480)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        lines = [
            f"קליינט: {client_name}",
            f"משתמש: {user_name}",
            "",
            "רולים ואובייקטי הרשאה:",
        ]
        if roles:
            for role_entry in roles:
                lines.append(f"- {role_entry.get('agr_name', '')}")
                for obj, fld, low in role_entry.get("objects", []):
                    lines.append(f"    {obj} | {fld} | {low}")
        else:
            lines.append("- לא נמצאו רולים להצגה")
        details_box.setPlainText(self.format_rtl_text("\n".join(lines)))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    # ------------------------------------------------------------------
    # Job-Management Permissions (MA-JOBMGMT-01)
    # Cross-join: AGR_1251 (permission objects) × AGR_USERS (role assignments)
    # OR logic: any single qualifying AGR_1251 row makes the AGR_NAME a match.
    # Qualifying objects: S_TCODE/TCD (SE37/SE36/SM36/SE35/SE30/SE34/SHDB/SM36WIZ),
    #   S_BTCH_ADM/BTCADMIN=Y, S_BTCH_JOB/JOBACTION (DELE/RELE/PROT),
    #   S_BTCH_NAM/BTCUNAME (any value), S_BTCH_MONI/BSCAKTI (DELE/RELE).
    # ------------------------------------------------------------------

    def _build_job_mgmt_permissions_section(self, parent_layout: QVBoxLayout) -> None:
        self.job_mgmt_summary_group = QGroupBox(self.format_ui_rtl_text("ממצאי הרשאות - עידכון ג'ובים"))
        self.job_mgmt_summary_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        job_mgmt_summary_layout = QVBoxLayout(self.job_mgmt_summary_group)
        job_mgmt_summary_layout.setContentsMargins(8, 14, 8, 8)

        self.job_mgmt_summary_table = QTableWidget(0, 5)
        self.job_mgmt_summary_table.setItemDelegate(_RightAlignDelegate(self.job_mgmt_summary_table))
        self.job_mgmt_summary_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("ממצא"),
            self.format_rtl_text("כמות משתמשים"),
            self.format_rtl_text("רמת סיכון"),
            self.format_rtl_text("סטטוס"),
        ])
        _job_mgmt_summary_hdr = self.job_mgmt_summary_table.horizontalHeader()
        _job_mgmt_summary_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _job_mgmt_summary_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _job_mgmt_summary_hdr.setStretchLastSection(False)
        self.job_mgmt_summary_table.setColumnWidth(1, 300)  # ממצא
        self.job_mgmt_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.job_mgmt_summary_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.job_mgmt_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.job_mgmt_summary_table.setAlternatingRowColors(True)
        self.job_mgmt_summary_table.setMinimumHeight(160)
        self.job_mgmt_summary_table.setToolTip(
            self.format_ui_rtl_text("לחיצה על שורה תציג את המשתמשים בעלי הרשאות עידכון ג'ובים")
        )
        self.job_mgmt_summary_table.itemSelectionChanged.connect(self._refresh_selected_job_mgmt_users)
        job_mgmt_summary_layout.addWidget(self.job_mgmt_summary_table)
        parent_layout.addWidget(self.job_mgmt_summary_group)

        self.job_mgmt_users_group = QGroupBox(self.format_ui_rtl_text("משתמשים בעלי הרשאות עידכון ג'ובים"))
        self.job_mgmt_users_group.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        job_mgmt_users_layout = QVBoxLayout(self.job_mgmt_users_group)
        job_mgmt_users_layout.setContentsMargins(8, 14, 8, 8)

        self.job_mgmt_users_table = QTableWidget(0, 2)
        self.job_mgmt_users_table.setItemDelegate(_RightAlignDelegate(self.job_mgmt_users_table))
        self.job_mgmt_users_table.setHorizontalHeaderLabels([
            self.format_rtl_text("קליינט"),
            self.format_rtl_text("משתמש"),
        ])
        _job_mgmt_users_hdr = self.job_mgmt_users_table.horizontalHeader()
        _job_mgmt_users_hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
        _job_mgmt_users_hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        _job_mgmt_users_hdr.setStretchLastSection(True)
        self.job_mgmt_users_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.job_mgmt_users_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.job_mgmt_users_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.job_mgmt_users_table.setAlternatingRowColors(True)
        self.job_mgmt_users_table.setMinimumHeight(180)
        self.job_mgmt_users_table.setToolTip(
            self.format_ui_rtl_text("לחיצה כפולה על משתמש תציג את הרולים המעניקים לו הרשאות ג'ובים")
        )
        self.job_mgmt_users_table.cellDoubleClicked.connect(self.show_job_mgmt_user_dialog)
        job_mgmt_users_layout.addWidget(self.job_mgmt_users_table)
        parent_layout.addWidget(self.job_mgmt_users_group)

        self._refresh_job_mgmt_summary_table()

    def _compute_job_mgmt_permissions(self) -> None:
        """Recompute job-management permission findings from cached AGR_1251 + AGR_USERS rows.

        OR logic: any single qualifying row in AGR_1251 makes the AGR_NAME a match.
        Special wildcard criterion: if the qualifying set for a (OBJECT, FIELD) pair
        contains "*", then any non-empty LOW or HIGH value qualifies.
        """
        if not self.agr_1251_cached_rows or not self.agr_users_cached_rows:
            return

        control_id = "MA1-1_AYALON_67"
        control_meta = get_audit_control_definition(control_id)

        qualifying_map: dict[tuple[str, str], set[str]] = {
            (obj.upper(), fld.upper()): {v.upper() for v in vals}
            for (obj, fld), vals in JOB_MGMT_PERMISSION_CRITERIA.items()
        }

        # agr_name_objects: AGR_NAME -> set of (OBJECT, FIELD, LOW_display) tuples that qualified
        agr_name_objects: dict[str, set[tuple[str, str, str]]] = {}
        for row in self.agr_1251_cached_rows:
            obj_val = self._resolve_row_value_by_priority(row, "OBJECT")
            fld_val = self._resolve_row_value_by_priority(row, "FIELD")
            low_val = self._resolve_row_value_by_priority(row, "LOW")
            high_val = self._resolve_row_value_by_priority(row, "HIGH")
            if obj_val is None or fld_val is None:
                continue
            obj_upper = str(obj_val).strip().upper()
            fld_upper = str(fld_val).strip().upper()
            key = (obj_upper, fld_upper)
            if key not in qualifying_map:
                continue
            low_str = str(low_val).strip().upper() if low_val is not None else ""
            high_str = str(high_val).strip().upper() if high_val is not None else ""
            if low_str == "*" or high_str == "*":
                qualifies = True
            elif "*" in qualifying_map[key]:
                # Criterion allows any value → any non-empty LOW or HIGH qualifies
                qualifies = bool(low_str) or bool(high_str)
            else:
                qualifies = bool(low_str and low_str in qualifying_map[key]) or bool(
                    high_str and high_str in qualifying_map[key]
                )
            if not qualifies:
                continue
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None or not str(agr_name_val).strip():
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            low_display = low_str if low_str else "-"
            agr_name_objects.setdefault(agr_name_upper, set()).add((obj_upper, fld_upper, low_display))

        matching_agr_names: set[str] = set(agr_name_objects.keys())

        # Cross-join with AGR_USERS to collect users per client.
        users_by_client: dict[str, dict[str, set[str]]] = {}
        for row in self.agr_users_cached_rows:
            agr_name_val = self._resolve_row_value_by_priority(row, "AGR_NAME")
            if agr_name_val is None:
                continue
            agr_name_upper = str(agr_name_val).strip().upper()
            if agr_name_upper not in matching_agr_names:
                continue

            mandt_val = self._resolve_row_value_by_priority(row, "MANDT")
            if mandt_val is not None and str(mandt_val).strip():
                mandt = str(mandt_val).strip()
            else:
                source_file = str(row.get("__source_file", ""))
                digits_match = re.search(r"\d{3}", Path(source_file).name)
                mandt = digits_match.group(0) if digits_match else "-"

            uname_val = self._resolve_row_value_by_priority(row, "UNAME")
            if uname_val is None or not str(uname_val).strip():
                continue
            uname = str(uname_val).strip().upper()

            client_users = users_by_client.setdefault(mandt, {})
            client_users.setdefault(uname, set()).add(agr_name_upper)

        self.job_mgmt_summary_records.clear()
        self.job_mgmt_users_by_control.clear()

        if not users_by_client:
            record_key = f"{control_id}|-"
            self.job_mgmt_summary_records[record_key] = {
                "record_key": record_key,
                "client": "-",
                "finding_text": "לא נמצאו משתמשים בעלי הרשאות עידכון ג'ובים",
                "users_count": 0,
                "risk_level": control_meta.get("risk_level", "-"),
                "status": "תקין",
            }
            self.job_mgmt_users_by_control[record_key] = []
        else:
            for mandt, client_users in sorted(users_by_client.items()):
                users_count = len(client_users)
                record_key = f"{control_id}|{mandt}"
                self.job_mgmt_summary_records[record_key] = {
                    "record_key": record_key,
                    "client": mandt,
                    "finding_text": f"נמצאו {users_count} משתמשים בעלי הרשאות עידכון ג'ובים",
                    "users_count": users_count,
                    "risk_level": control_meta.get("risk_level", "-"),
                    "status": "עם ממצא" if users_count > 0 else "תקין",
                }
                self.job_mgmt_users_by_control[record_key] = [
                    {
                        "client": mandt,
                        "user_name": uname,
                        "roles": [
                            {
                                "agr_name": r,
                                "objects": sorted(agr_name_objects.get(r, set())),
                            }
                            for r in sorted(roles)
                        ],
                    }
                    for uname, roles in sorted(client_users.items())
                ]

        self._refresh_job_mgmt_summary_table()

    def _refresh_job_mgmt_summary_table(self) -> None:
        self.job_mgmt_summary_table.setRowCount(0)
        self.job_mgmt_users_table.setRowCount(0)
        if not self.job_mgmt_summary_records:
            self.job_mgmt_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("יש לטעון קבצי AGR_1251 ו-AGR_USERS"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.job_mgmt_users_table.setItem(0, 1, empty_item)
            return

        for row_data in sorted(
            self.job_mgmt_summary_records.values(),
            key=lambda item: str(item.get("client", "")),
        ):
            row_index = self.job_mgmt_summary_table.rowCount()
            self.job_mgmt_summary_table.insertRow(row_index)
            values = [
                str(row_data.get("client", "-")),
                str(row_data.get("finding_text", "-")),
                str(row_data.get("users_count", 0)),
                str(row_data.get("risk_level", "-")),
                str(row_data.get("status", "-")),
            ]
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if column == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("record_key", ""))
                self.job_mgmt_summary_table.setItem(row_index, column, item)

        if self.job_mgmt_summary_table.rowCount() > 0:
            self.job_mgmt_summary_table.selectRow(0)
            self._refresh_selected_job_mgmt_users()
        self.job_mgmt_summary_table.resizeColumnsToContents()

    def _refresh_selected_job_mgmt_users(self) -> None:
        selected_items = self.job_mgmt_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.job_mgmt_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        record_key = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        user_rows = self.job_mgmt_users_by_control.get(record_key, [])
        self.job_mgmt_users_table.setRowCount(0)

        if not user_rows:
            self.job_mgmt_users_table.insertRow(0)
            empty_item = QTableWidgetItem(self.format_rtl_text("לא נמצאו משתמשים להצגה"))
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.job_mgmt_users_table.setItem(0, 1, empty_item)
            return

        for user_data in user_rows:
            row_index = self.job_mgmt_users_table.rowCount()
            self.job_mgmt_users_table.insertRow(row_index)
            client_name = str(user_data.get("client", "-") or "-")
            user_name = str(user_data.get("user_name", "-") or "-")
            client_item = QTableWidgetItem(self.format_rtl_text(client_name))
            client_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item = QTableWidgetItem(self.format_rtl_text(user_name))
            user_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            user_item.setData(Qt.ItemDataRole.UserRole, {"client": client_name, "user": user_name})
            self.job_mgmt_users_table.setItem(row_index, 0, client_item)
            self.job_mgmt_users_table.setItem(row_index, 1, user_item)
        self.job_mgmt_users_table.resizeColumnsToContents()

    def show_job_mgmt_user_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.job_mgmt_users_table.rowCount():
            return

        user_item = self.job_mgmt_users_table.item(row_index, 1)
        if user_item is None:
            return

        user_payload = user_item.data(Qt.ItemDataRole.UserRole)
        user_name = ""
        client_name = "-"
        if isinstance(user_payload, dict):
            user_name = str(user_payload.get("user", "") or "").strip()
            client_name = str(user_payload.get("client", "-") or "-").strip()
        if not user_name:
            user_name = str(user_item.text() or "").strip()
        if not user_name or user_name.startswith("לא נמצאו") or user_name.startswith("אין "):
            return

        selected_items = self.job_mgmt_summary_table.selectedItems()
        if not selected_items:
            return

        summary_item = self.job_mgmt_summary_table.item(selected_items[0].row(), 0)
        if summary_item is None:
            return

        record_key = str(summary_item.data(Qt.ItemDataRole.UserRole) or summary_item.text())
        user_rows = self.job_mgmt_users_by_control.get(record_key, [])
        roles: list[dict] = []
        for user_data in user_rows:
            if (
                str(user_data.get("user_name", "")).strip().upper() == user_name.upper()
                and str(user_data.get("client", "-") or "-").strip() == client_name
            ):
                roles = list(user_data.get("roles", []))
                break

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט הרשאות עידכון ג'ובים")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(640, 480)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        lines = [
            f"קליינט: {client_name}",
            f"משתמש: {user_name}",
            "",
            "רולים ואובייקטי הרשאה:",
        ]
        if roles:
            for role_entry in roles:
                lines.append(f"- {role_entry.get('agr_name', '')}")
                for obj, fld, low in role_entry.get("objects", []):
                    lines.append(f"    {obj} | {fld} | {low}")
        else:
            lines.append("- לא נמצאו רולים להצגה")
        details_box.setPlainText(self.format_rtl_text("\n".join(lines)))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    def _build_audit_detail_row(
        self,
        issue: ValidationIssue | None,
        control_id: str,
        source_file: str,
        extraction_date: str,
        control_meta: dict[str, str],
        control_snapshot: dict[str, str] | None = None,
    ) -> dict[str, Any]:
        return build_audit_detail_row(
            issue,
            control_id,
            source_file,
            extraction_date,
            control_meta,
            self._current_work_environment_label(),
            control_snapshot,
        )

    def _upsert_audit_control_data(
        self,
        slot_key: str,
        result: Any,
        audit_issues: list[ValidationIssue],
        extraction_date: str,
    ) -> None:
        upsert_audit_control_data(
            self.audit_summary_records,
            self.audit_details_by_control,
            slot_key,
            result,
            audit_issues,
            extraction_date,
            self._current_work_environment_label(),
            self._get_slot_display_name(slot_key),
            get_audit_control_definition,
            get_profile_audit_controls,
            self._count_stms_control_records,
            self._build_password_control_snapshots,
        )

    def _sync_user_review_completion_finding(self) -> None:
        control_id = self.REVIEW_COMPLETION_CONTROL_ID
        preview_rows, reviewed_rows, incomplete_rows = self._get_user_review_completion_snapshot()
        sync_user_review_completion_finding(
            self.audit_summary_records,
            self.audit_details_by_control,
            control_id,
            get_audit_control_definition(control_id),
            self._get_slot_display_name("USR02"),
            self._get_slot_extraction_date("USR02") or "-",
            self._current_work_environment_label(),
            reviewed_rows,
            preview_rows,
            incomplete_rows,
            self._build_user_review_incomplete_reason,
        )

    def _sync_developer_sod_finding(self) -> None:
        control_id = "MC5-23_AYALON_48"
        self.audit_summary_records.pop(control_id, None)
        self.audit_details_by_control.pop(control_id, None)

        if self._current_work_environment_code() != "FPP":
            return

        settings = self._current_system_settings()
        dev_list = settings.get("authorized_developers", [])
        if not dev_list:
            return

        usr02_rows = self._load_preview_rows("USR02")
        if not usr02_rows:
            return

        control_meta = get_audit_control_definition(control_id)
        dev_lookup = {(str(d.get("MANDT", "")).strip(), str(d.get("BNAME", "")).strip().upper()) for d in dev_list}
        
        findings: list[dict[str, Any]] = []
        active_developers_count = 0
        
        for row in usr02_rows:
            mandt = self._get_row_value(row, "MANDT")
            bname = self._get_row_value(row, "BNAME").upper()
            uflag = self._get_row_value(row, "UFLAG")
            
            if (mandt, bname) in dev_lookup and not self._is_user_locked(uflag):
                active_developers_count += 1
                findings.append({
                    "control_id": control_id,
                    "source_file": self._get_slot_display_name("USR02"),
                    "extraction_date": self._get_slot_extraction_date("USR02"),
                    "work_environment": self._current_work_environment_label(),
                    "category": control_meta.get("category", "-"),
                    "risk_level": control_meta.get("risk_level", "-"),
                    "description": control_meta.get("description", "-"),
                    "check_type": control_meta.get("check_type", "-"),
                    "actual_value": bname,
                    "expected_value": "מפתח ללא גישת ייצור",
                    "status": "עם ממצא",
                    "full_description": f"המשתמש {bname} (MANDT {mandt}) מוגדר כמפתח וזוהה כפעיל בסביבת הייצור.",
                })

        if findings:
            self.audit_summary_records[control_id] = {
                "control_id": control_id,
                "check_type": control_meta.get("check_type"),
                "source_file": self._get_slot_display_name("USR02"),
                "extraction_date": self._get_slot_extraction_date("USR02"),
                "work_environment": self._current_work_environment_label(),
                "risk_level": control_meta.get("risk_level"),
                "description": control_meta.get("description"),
                "valid_records": len(dev_list) - active_developers_count,
                "finding_records": active_developers_count,
                "total_records": len(dev_list),
            }
            self.audit_details_by_control[control_id] = findings
        else:
            self.audit_summary_records[control_id] = {
                "control_id": control_id,
                "check_type": control_meta.get("check_type"),
                "source_file": self._get_slot_display_name("USR02"),
                "extraction_date": self._get_slot_extraction_date("USR02"),
                "work_environment": self._current_work_environment_label(),
                "risk_level": control_meta.get("risk_level"),
                "description": control_meta.get("description"),
                "valid_records": len(dev_list),
                "finding_records": 0,
                "total_records": len(dev_list),
            }

    def _refresh_audit_summary_table(self) -> None:
        self._sync_user_review_completion_finding()
        self._sync_developer_sod_finding()
        self.audit_summary_table.setRowCount(0)
        if not self.audit_summary_records:
            self.audit_detail_table.setRowCount(0)
            self.audit_detail_table.insertRow(0)
            for column, value in enumerate(["-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "אין ממצאים להצגה"]):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.audit_detail_table.setItem(0, column, item)
            return

        for row_data in sorted_audit_summary_rows(self.audit_summary_records):
            row_index = self.audit_summary_table.rowCount()
            self.audit_summary_table.insertRow(row_index)
            values = build_audit_summary_values(row_data)
            # Insert values: cols 0..5 directly, then skip col 6 (button), shift cols 6..9 to 7..10
            for value_index, value in enumerate(values):
                target_col = value_index if value_index < 6 else value_index + 1
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                if value_index == 0:
                    item.setData(Qt.ItemDataRole.UserRole, row_data.get("control_id", ""))
                self.audit_summary_table.setItem(row_index, target_col, item)

            # Add working-paper export button in column 6 (between סהכ רשומות and סביבת עבודה)
            control_id_value = str(row_data.get("control_id", ""))
            wp_button = QPushButton("📄")
            wp_button.setToolTip(self.format_rtl_text("ייצוא נייר עבודה"))
            wp_button.setCursor(Qt.CursorShape.PointingHandCursor)
            wp_button.clicked.connect(
                lambda _checked=False, cid=control_id_value: self._export_control_working_paper(cid)
            )
            self.audit_summary_table.setCellWidget(row_index, 6, wp_button)

        if self.audit_summary_table.rowCount() > 0:
            self.audit_summary_table.selectRow(0)
            self._refresh_selected_audit_detail()
        self.audit_summary_table.resizeColumnsToContents()

    def _refresh_selected_audit_detail(self) -> None:
        selected_items = self.audit_summary_table.selectedItems()
        if not selected_items:
            return

        selected_row = selected_items[0].row()
        control_item = self.audit_summary_table.item(selected_row, 0)
        if control_item is None:
            return

        control_id = str(control_item.data(Qt.ItemDataRole.UserRole) or control_item.text())
        detail_rows = self.audit_details_by_control.get(control_id, [])
        self.audit_detail_table.setRowCount(0)

        for detail in detail_rows:
            row_index = self.audit_detail_table.rowCount()
            self.audit_detail_table.insertRow(row_index)
            values = build_audit_detail_values(detail)
            for column, value in enumerate(values):
                item = QTableWidgetItem(self.format_rtl_text(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.audit_detail_table.setItem(row_index, column, item)
        self.audit_detail_table.resizeColumnsToContents()

    def export_audit_findings_to_excel(self, open_after_export: bool = False) -> Path | None:
        if not self.audit_summary_records:
            QMessageBox.warning(self, "אין נתונים לייצוא", "לא קיימים ממצאי ביקורת לייצוא.")
            return None

        summary_rows = sorted(self.audit_summary_records.values(), key=lambda item: str(item.get("control_id", "")))
        detail_rows: list[dict[str, Any]] = []
        for control_id in sorted(self.audit_details_by_control.keys()):
            detail_rows.extend(self.audit_details_by_control.get(control_id, []))

        export_path = self.config.output_dir / f"audit_findings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ExcelReportWriter.write_audit_findings_report(summary_rows, detail_rows, export_path)
        self.audit_findings_export_path = export_path

        if open_after_export:
            QMessageBox.information(self, "הייצוא הושלם", f"קובץ הממצאים נשמר בהצלחה:\n{export_path}")

        return export_path

    def _export_control_working_paper(self, control_id: str) -> None:
        """Export a per-control working paper Excel file."""
        if not control_id:
            return
        summary_record = self.audit_summary_records.get(control_id)
        if not summary_record:
            QMessageBox.warning(
                self,
                "אין נתונים לייצוא",
                f"לא נמצאו נתוני בקרה עבור {control_id}.",
            )
            return

        detail_rows = list(self.audit_details_by_control.get(control_id, []))
        raw_population_rows = list(self.control_to_slot_rows.get(control_id, []))

        # For cross-join controls (AGR_1251 × AGR_USERS), the raw population is
        # never stored in control_to_slot_rows. Build it on-demand from cached rows.
        _AGR_CROSS_JOIN_CONTROLS = frozenset({
            "MA1-1_AYALON_10", "MA1-1_AYALON_11", "MA1-1_AYALON_12",
            "MA1-1_AYALON_16", "MA1-1_AYALON_43", "MA1-1_AYALON_45",
            "MA1-1_AYALON_67", "MA5.1-13_AYALON_24", "MA7-17_AYALON_30",
        })
        raw_population_note: str | None = None
        if not raw_population_rows and control_id in _AGR_CROSS_JOIN_CONTROLS:
            # Build AGR_1251 lookup by role (O(m))
            _agr1251_by_role: dict[str, list[dict[str, Any]]] = {}
            for _r in self.agr_1251_cached_rows:
                _k = str(_r.get("AGR_NAME") or "").strip().upper()
                if _k:
                    _agr1251_by_role.setdefault(_k, []).append(_r)

            # Count cross-join total in O(n)
            _cross_total = sum(
                len(_agr1251_by_role.get(str(r.get("AGR_NAME") or "").strip().upper(), []))
                for r in self.agr_users_cached_rows
            )

            if _cross_total > 100_000:
                # Large population: show only findings, add note with total count
                raw_population_rows = [
                    r for r in detail_rows
                    if str(r.get("status", "")).strip() == "עם ממצא"
                ]
                raw_population_note = (
                    f"לא כל האוכלוסייה נשלפה מפאת היקף גדול של נתונים"
                    f" (סה\u05f4כ אוכלוסייה: {_cross_total:,} רשומות)"
                )
            else:
                # Small population: build full cross-join
                raw_population_rows = []
                for _u_row in self.agr_users_cached_rows:
                    _agr_name = str(_u_row.get("AGR_NAME") or "").strip().upper()
                    for _a_row in _agr1251_by_role.get(_agr_name, []):
                        raw_population_rows.append({
                            "MANDT":    _u_row.get("MANDT", "-"),
                            "UNAME":    _u_row.get("UNAME", "-"),
                            "AGR_NAME": _u_row.get("AGR_NAME", "-"),
                            "OBJECT":   _a_row.get("OBJECT", "-"),
                            "FIELD":    _a_row.get("FIELD", "-"),
                            "LOW":      _a_row.get("LOW", "-"),
                            "HIGH":     _a_row.get("HIGH", "-"),
                        })

        # For the user-review completion control, the raw population is the full review report
        if not raw_population_rows and control_id == self.REVIEW_COMPLETION_CONTROL_ID:
            _review_preview_rows, _, _ = self._get_user_review_completion_snapshot()
            raw_population_rows = [
                {field: row.get(field, "-") for field in self.EXPORT_REVIEW_FIELDS}
                for row in _review_preview_rows
            ]

        # Filter IPE entries by matching slot key, enriching each with extraction_date + population_count
        slot_key = self.control_to_slot_key.get(control_id, "")
        ipe_entries: list[dict[str, Any]] = []

        def _enrich(entry: dict[str, Any], source_slot: str) -> dict[str, Any]:
            enriched = dict(entry)
            enriched["extraction_date"] = self._get_slot_extraction_date(source_slot) or "-"
            # Match the screenshot's stem to the data-file row count so each
            # IPE entry reflects the population_count of *its own* source file.
            per_file_map = self._slot_file_row_counts.get(source_slot, {}) or {}
            screenshot_stem = Path(
                str(entry.get("original_filename") or entry.get("stored_filename") or "")
            ).stem

            def _norm(s: Any) -> str:
                return re.sub(r"[_\-.]+", " ", str(s).strip().lower()).strip()

            screenshot_norm = _norm(screenshot_stem)
            per_file_norm = {_norm(k): v for k, v in per_file_map.items() if k}

            matched_count: Any = None
            if screenshot_norm and per_file_norm:
                # Exact normalized match first
                if screenshot_norm in per_file_norm:
                    matched_count = per_file_norm[screenshot_norm]
                else:
                    # Fall back to substring match on normalized stems.
                    for data_norm, count in per_file_norm.items():
                        if data_norm and (data_norm in screenshot_norm or screenshot_norm in data_norm):
                            matched_count = count
                            break
                if matched_count is None:
                    # Numeric-token fallback: match when both stems share the same
                    # set of digit-sequences, e.g. screenshot "AGR_1251_100" →
                    # {1251, 100} uniquely matches data file "UGR_1251_100".
                    _screenshot_nums = frozenset(re.findall(r"\d+", screenshot_norm))
                    if _screenshot_nums:
                        _num_candidates = [
                            cnt for dn, cnt in per_file_norm.items()
                            if _screenshot_nums == frozenset(re.findall(r"\d+", dn))
                        ]
                        if len(_num_candidates) == 1:
                            matched_count = _num_candidates[0]
            if matched_count is None:
                matched_count = self.slot_to_row_count.get(source_slot, "-")
            enriched["population_count"] = matched_count
            return enriched

        if slot_key:
            for entry in (self.ipe_evidence_data.get(slot_key, []) or []):
                ipe_entries.append(_enrich(entry, slot_key))
        # Also include entries explicitly assigned to this control_id across other slots
        for s_key, entries in (self.ipe_evidence_data or {}).items():
            if s_key == slot_key:
                continue
            for entry in entries or []:
                if control_id in (entry.get("control_ids") or []):
                    ipe_entries.append(_enrich(entry, s_key))

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_id = re.sub(r"[\\/*?:\[\]&]", "_", control_id)
        default_name = f"{safe_id}_working_paper_{timestamp}.xlsx"
        default_path = str(self.config.output_dir / default_name)

        chosen_path, _ = QFileDialog.getSaveFileName(
            self,
            "שמירת נייר עבודה",
            default_path,
            "Excel Files (*.xlsx)",
        )
        if not chosen_path:
            return
        if not chosen_path.lower().endswith(".xlsx"):
            chosen_path += ".xlsx"

        try:
            notes: list[str] = []
            if control_id == "MA3-3_AYALON_14" and not self._get_critical_roles():
                notes.append(
                    "לא הוגדרו פרופילי משתמשיי על בהגדרות המערכת - "
                    "בדיקת פרופילים חזקים לא בוצעה."
                )
            write_control_working_paper(
                control_id=control_id,
                summary_record=summary_record,
                detail_rows=detail_rows,
                raw_population_rows=raw_population_rows,
                ipe_entries=ipe_entries,
                work_environment_label=self._current_work_environment_label(),
                output_path=Path(chosen_path),
                notes=notes,
                critical_roles=self._get_critical_roles(),
                raw_population_note=raw_population_note,
            )
        except Exception as exc:
            QMessageBox.critical(
                self,
                "שגיאה בייצוא",
                f"שגיאה ביצירת נייר העבודה:\n{exc}",
            )
            return

        QMessageBox.information(
            self,
            "הייצוא הושלם",
            f"נייר העבודה נשמר בהצלחה:\n{chosen_path}",
        )

    def _build_audit_detail_dialog_text(self, row_index: int) -> str:
        if row_index < 0 or row_index >= self.audit_detail_table.rowCount():
            return self.format_rtl_text("לא נמצא פירוט עבור הרשומה שנבחרה.")

        field_labels = [
            "קובץ מקור",
            "תאריך הפקה",
            "סביבת עבודה",
            "קטגוריה",
            "רמת סיכון",
            "תיאור",
            "סוג בדיקה",
            "קליינט",
            "משתמש",
            "ערך בפועל",
            "ערך מצופה",
            "סטטוס",
            "תיאור מלא",
        ]
        lines = ["פירוט ממצא ביקורת:", ""]
        for column, field_label in enumerate(field_labels):
            item = self.audit_detail_table.item(row_index, column)
            field_value = item.text() if item is not None else "-"
            lines.append(f"{field_label}: {field_value}")
        return self.format_rtl_text("\n".join(lines))

    def show_audit_detail_dialog(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= self.audit_detail_table.rowCount():
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט ממצא ביקורת")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(760, 420)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        details_box.setPlainText(self._build_audit_detail_dialog_text(row_index))
        layout.addWidget(details_box)

        # Compensating controls button — only shown when AI is enabled
        settings = self._current_system_settings()
        ai_cfg = settings.get("ai_settings", {}) if isinstance(settings, dict) else {}
        ai_enabled = bool(ai_cfg.get("enabled", False))
        features = ai_cfg.get("features", {}) if isinstance(ai_cfg, dict) else {}
        compensating_enabled = ai_enabled and bool(features.get("compensating_controls", True))

        if compensating_enabled:
            comp_btn_row = QWidget()
            comp_btn_row.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
            comp_btn_layout = QHBoxLayout(comp_btn_row)
            comp_btn_layout.setContentsMargins(0, 0, 0, 0)
            comp_btn = QPushButton(self.format_ui_rtl_text("המלצות בקרות מפצות (AI)"))
            comp_btn.setMaximumWidth(240)
            comp_result_label = QLabel("")
            comp_result_label.setWordWrap(True)
            comp_result_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop)

            def _on_compensating_controls() -> None:
                comp_btn.setEnabled(False)
                comp_btn.setText(self.format_ui_rtl_text("מחשב המלצות..."))
                try:
                    from src.services.compensating_advisor import CompensatingAdvisor
                    # Extract fields from dialog text
                    _item_risk = self.audit_detail_table.item(row_index, 4)
                    _item_desc = self.audit_detail_table.item(row_index, 5)
                    _item_actual = self.audit_detail_table.item(row_index, 7)
                    _item_expected = self.audit_detail_table.item(row_index, 8)
                    _selected = self.audit_summary_table.selectedItems()
                    _ctrl_item = self.audit_summary_table.item(_selected[0].row(), 0) if _selected else None
                    control_id = str(_ctrl_item.data(Qt.ItemDataRole.UserRole) or (_ctrl_item.text() if _ctrl_item else "")) if _ctrl_item else ""
                    control_desc = _item_desc.text() if _item_desc else ""
                    risk_level = _item_risk.text().strip().lower() if _item_risk else "high"
                    risk_map = {"גבוה": "high", "בינוני": "medium", "נמוך": "low", "high": "high", "medium": "medium", "low": "low"}
                    risk_level = risk_map.get(risk_level, "high")

                    client = OllamaClient(ai_cfg)
                    advisor = CompensatingAdvisor(client, work_environment=self._current_work_environment_label())
                    recs = advisor.recommend(control_id, control_desc, risk_level=risk_level)

                    if recs:
                        lines = ["המלצות בקרות מפצות:", ""]
                        for rec in recs:
                            lines.append(f"#{rec.get('rank','')}. {rec.get('title','')}")
                            lines.append(f"   נימוק: {rec.get('rationale','')}")
                            lines.append(f"   עדות: {rec.get('evidence_needed','')}")
                            lines.append("")
                        comp_result_label.setText(self.format_rtl_text("\n".join(lines)))
                    else:
                        comp_result_label.setText(self.format_rtl_text("לא נמצאו המלצות רלוונטיות לבקרה זו."))
                except Exception as exc:
                    comp_result_label.setText(self.format_rtl_text(f"שגיאה בקבלת המלצות: {exc}"))
                finally:
                    comp_btn.setEnabled(True)
                    comp_btn.setText(self.format_ui_rtl_text("המלצות בקרות מפצות (AI)"))

            comp_btn.clicked.connect(_on_compensating_controls)
            comp_btn_layout.addStretch(1)
            comp_btn_layout.addWidget(comp_btn)
            layout.addWidget(comp_btn_row)
            layout.addWidget(comp_result_label)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    def _build_log_details(self, row_index: int) -> str:
        if row_index < 0 or row_index >= len(self.run_log_records):
            return "לא נמצא פירוט עבור הרשומה שנבחרה."

        record = self.run_log_records[row_index]
        lines = [
            f"משבצת: {record['slot_key']}",
            f"קבוצת דוחות: {record['report_group']}",
            f"קובץ: {record['file_name']}",
            f"תאריך הפקה: {record['extraction_date']}",
            f"מספר רשומות שנקלטו: {record['row_count']}",
            f"סטטוס: {record['status']}",
            f"מספר שגיאות: {record['error_count']}",
            f"תיאור קצר: {record['error_preview']}",
            f"תאריך בדיקה: {record['date']}",
            f"שעת בדיקה: {record['time']}",
            "",
            "פירוט:",
        ]

        issues = record.get("issues", [])
        intake = [iss for iss in issues if self._is_intake_issue(iss)]
        audit = [iss for iss in issues if not self._is_intake_issue(iss)]

        if not issues:
            lines.append("לא נמצאו שגיאות קליטה וממצאי ביקורת.")
        else:
            if intake:
                lines.append("--- שגיאות קליטה ---")
                for issue in intake:
                    row_label = issue.row_number if issue.row_number > 0 else "מבנה"
                    lines.append(f"- שורה {row_label} / {issue.column_name}: {issue.message}")
            else:
                lines.append("--- שגיאות קליטה: אין ---")

            if audit:
                lines.append("")
                lines.append("--- ממצאי ביקורת (לפירוט ראה טאב 'ביצוע ניתוח לביקורת') ---")
                for issue in audit:
                    row_label = issue.row_number if issue.row_number > 0 else "מבנה"
                    lines.append(f"- שורה {row_label} / {issue.column_name}: {issue.message}")

        return self.format_rtl_text("\n".join(lines))

    def show_log_details(self, row_index: int, _column: int) -> None:
        if row_index < 0 or row_index >= len(self.run_log_records):
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("פירוט קובץ שנבדק")
        dialog.setLayoutDirection(Qt.LayoutDirection.RightToLeft)
        dialog.resize(760, 420)

        layout = QVBoxLayout(dialog)
        details_box = QTextEdit()
        details_box.setReadOnly(True)
        details_box.setPlainText(self._build_log_details(row_index))
        layout.addWidget(details_box)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok)
        buttons.accepted.connect(dialog.accept)
        layout.addWidget(buttons)
        dialog.exec()

    def clear_results(self) -> None:
        self.selected_slot_key = None
        self.load_history = []
        for slot_key, widget_data in self.slot_widgets.items():
            widget_data["selected_paths"] = []
            self._update_slot_path_label(slot_key, [])
            date_edit = widget_data.get("extraction_date_edit")
            if isinstance(date_edit, QLineEdit):
                date_edit.setText(self._default_extraction_date())
        self.required_columns_edit.setText("")
        self.summary_labels["total"].setText("0")
        self.summary_labels["valid"].setText("0")
        self.summary_labels["invalid"].setText("0")
        self.summary_labels["status"].setText("ממתין להרצה")
        self.summary_group.hide()
        self.results_group.hide()
        self.run_log_group.hide()
        if hasattr(self, "console_output"):
            self.console_output.clear()
            self._log_to_console("מוכן לקליטה...", "dim")
        self.report_path = None
        self.log_export_path = None
        self.audit_findings_export_path = None
        self.report_button.setEnabled(False)
        self.issues_table.setRowCount(0)
        self.audit_summary_records = {}
        self.audit_details_by_control = {}
        self.permissions_summary_records = {}
        self.permissions_users_by_control = {}
        self._strong_profile_data = {}
        self.slot_to_row_count = {}
        self._slot_file_row_counts = {}
        self.control_to_slot_rows = {}
        self.control_to_slot_key = {}
        self.audit_summary_table.setRowCount(0)
        self.audit_detail_table.setRowCount(0)
        self.permissions_summary_table.setRowCount(0)
        self.permissions_users_table.setRowCount(0)
        self.run_log_records = []
        self.run_log_table.setRowCount(0)
        self.refresh_user_preview()
        self.tabs.setCurrentIndex(0)

    def export_run_log_to_excel(self, open_after_export: bool = False) -> Path | None:
        if not self.run_log_records:
            QMessageBox.warning(self, "אין נתונים לייצוא", "טרם תועדו קבצים שנבדקו לייצוא לאקסל.")
            return None

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = self.format_rtl_text("קבצים שנבדקו")
        headers = [
            "משבצת",
            "קבוצת דוחות",
            "קובץ",
            "תאריך הפקה",
            "רשומות שנקלטו",
            "סטטוס",
            "מספר שגיאות",
            "תיאור שגיאה",
            "תאריך בדיקה",
            "שעת בדיקה",
        ]
        sheet.append(headers)

        for record in self.run_log_records:
            sheet.append([
                record["slot_key"],
                record["report_group"],
                record["file_name"],
                record["extraction_date"],
                record["row_count"],
                record["status"],
                record["error_count"],
                record["error_preview"],
                record["date"],
                record["time"],
            ])

        export_path = self.config.output_dir / f"intake_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(export_path)
        self.log_export_path = export_path

        if open_after_export:
            QMessageBox.information(self, "הייצוא הושלם", f"קובץ התיעוד נשמר בהצלחה:\n{export_path}")

        return export_path

    def export_user_preview_to_excel(self, open_after_export: bool = False) -> Path | None:
        usr02_rows = self._load_preview_rows("USR02")
        combined_rows = self._load_preview_rows("ADR6_USR21")
        if not usr02_rows and not combined_rows:
            QMessageBox.warning(self, "אין נתונים לייצוא", "טרם נטענו משתמשים לסקירה לצורך ייצוא לאקסל.")
            return None

        all_preview_rows = self._build_user_preview_rows(usr02_rows, combined_rows)
        if not all_preview_rows:
            QMessageBox.warning(self, "אין נתונים לייצוא", "טרם נטענו משתמשים לסקירה לצורך ייצוא לאקסל.")
            return None

        field_to_col_def = {col["field"]: col for col in self.USER_PREVIEW_COLUMN_DEFINITIONS}
        export_field_names = [f for f in self.EXPORT_REVIEW_FIELDS if f in field_to_col_def]
        export_formal_names = [str(field_to_col_def[f]["formal"]) for f in export_field_names]

        sorted_rows = sorted(all_preview_rows, key=self._export_sort_key)

        workbook = Workbook()
        sheet = workbook.active
        assert sheet is not None
        sheet.title = self.format_rtl_text("סקירת משתמשים")
        sheet.append(export_formal_names)

        review_status_col_index: int | None = None
        technical_notes_col_index: int | None = None
        for idx, field in enumerate(export_field_names):
            if field == "REVIEW_STATUS":
                review_status_col_index = idx + 1  # 1-based Excel column
            elif field in {"TECH_REVIEW_NOTES", "REVIEW_NOTES"}:
                technical_notes_col_index = idx + 1  # 1-based Excel column

        total_data_rows = len(sorted_rows)
        for preview_row in sorted_rows:
            sheet.append([
                preview_row.get(field, "") or ""
                for field in export_field_names
            ])

        if review_status_col_index is not None and total_data_rows > 0:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(review_status_col_index)
            dv_formula = '"' + ",".join(self.REVIEW_STATUS_OPTIONS) + '"'
            dv = DataValidation(
                type="list",
                formula1=dv_formula,
                allow_blank=True,
                showDropDown=False,
            )
            dv.sqref = f"{col_letter}2:{col_letter}{total_data_rows + 1}"
            sheet.add_data_validation(dv)

        if total_data_rows > 0 and (review_status_col_index is not None or technical_notes_col_index is not None):
            from openpyxl.utils import get_column_letter  # noqa: F811
            warning_fill = PatternFill("solid", fgColor="FFF0C2")
            for excel_row_idx, preview_row in enumerate(sorted_rows, start=2):
                is_not_ok = preview_row.get("REVIEW_STATUS", "") == "נבדק - לא תקין"
                has_notes = bool((preview_row.get("TECH_REVIEW_NOTES", "") or "").strip())
                if is_not_ok and not has_notes:
                    if review_status_col_index is not None:
                        sheet.cell(row=excel_row_idx, column=review_status_col_index).fill = warning_fill
                    if technical_notes_col_index is not None:
                        sheet.cell(row=excel_row_idx, column=technical_notes_col_index).fill = warning_fill

        export_path = self.config.output_dir / f"users_review_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(export_path)
        self.user_preview_export_path = export_path

        if open_after_export:
            QMessageBox.information(self, "הייצוא הושלם", f"קובץ הסקירה נשמר בהצלחה:\n{export_path}")

        return export_path

    def import_user_review_from_excel(self) -> None:
        initial_directory = self._get_last_file_dialog_directory()
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "ייבוא סקירת משתמשים מאקסל",
            initial_directory,
            "Excel files (*.xlsx *.xlsm);;All files (*.*)",
        )
        if not file_path:
            return

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            QMessageBox.warning(self, "שגיאת ייבוא", f"לא ניתן לפתוח את קובץ האקסל:\n{exc}")
            return

        sheet = workbook.active
        if sheet is None:
            QMessageBox.warning(self, "שגיאת ייבוא", "הגיליון הפעיל בקובץ ריק או לא נמצא.")
            workbook.close()
            return

        rows_iter = iter(sheet.iter_rows(values_only=True))
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            QMessageBox.warning(self, "שגיאת ייבוא", "הקובץ ריק - לא נמצאו כותרות.")
            workbook.close()
            return

        headers = [str(h).strip() if h is not None else "" for h in raw_headers]

        FORMAL_TO_FIELD = {
            str(col["formal"]).strip(): str(col["field"])
            for col in self.USER_PREVIEW_COLUMN_DEFINITIONS
        }
        TECHNICAL_FIELDS = {str(col["field"]) for col in self.USER_PREVIEW_COLUMN_DEFINITIONS}

        def _resolve(header: str) -> str | None:
            if header in TECHNICAL_FIELDS:
                return header
            return FORMAL_TO_FIELD.get(header)

        col_map: dict[str, int] = {}
        for col_idx, header in enumerate(headers):
            field = _resolve(header)
            if field and field not in col_map:
                col_map[field] = col_idx

        expected_import_fields = list(self.EXPORT_REVIEW_FIELDS)
        missing = [field_name for field_name in expected_import_fields if field_name not in col_map]
        non_empty_headers = [header for header in headers if header]
        if missing or len(non_empty_headers) < len(expected_import_fields):
            missing_labels = [
                str(self._get_user_preview_column_definition(field_name).get("formal", field_name))
                for field_name in missing
            ]
            QMessageBox.warning(
                self,
                "שגיאת ייבוא",
                "קובץ הסקירה אינו תואם לתבנית המעודכנת של המערכת.\n"
                f"מספר עמודות מזוהות: {len(non_empty_headers)} מתוך {len(expected_import_fields)} נדרשות.\n"
                f"עמודות חסרות: {', '.join(missing_labels) if missing_labels else '-'}\n"
                "יש לייבא קובץ אקסל שיוצא מהכלי לאחר השינוי האחרון במבנה דוח הסקירה.",
            )
            workbook.close()
            return

        mandt_col = col_map.get("MANDT")
        bname_col = col_map["BNAME"]
        status_col = col_map["REVIEW_STATUS"]
        tech_notes_col = col_map.get("TECH_REVIEW_NOTES")
        if tech_notes_col is None:
            tech_notes_col = col_map.get("REVIEW_NOTES")
        business_notes_col = col_map.get("BUS_REVIEW_NOTES")

        # ── Phase 1: Parse all rows from the file ──────────────────────────
        imported_data: dict[str, dict[str, str]] = {}
        for row_values in rows_iter:
            bname = str(row_values[bname_col]).strip() if bname_col < len(row_values) and row_values[bname_col] is not None else ""
            if not bname:
                continue
            mandt = str(row_values[mandt_col]).strip() if mandt_col is not None and mandt_col < len(row_values) and row_values[mandt_col] is not None else ""
            review_key = self._user_reviewer_state_key(mandt, bname)

            raw_status = row_values[status_col] if status_col < len(row_values) else None
            status_value = self._normalize_reviewer_status(str(raw_status).strip() if raw_status is not None else "")

            raw_tech = row_values[tech_notes_col] if tech_notes_col is not None and tech_notes_col < len(row_values) else None
            tech_notes_value = str(raw_tech).strip() if raw_tech is not None else ""
            raw_bus = row_values[business_notes_col] if business_notes_col is not None and business_notes_col < len(row_values) else None
            bus_notes_value = str(raw_bus).strip() if raw_bus is not None else ""

            imported_data[review_key] = {
                "BNAME": bname,
                "REVIEW_STATUS": status_value,
                "TECH_REVIEW_NOTES": tech_notes_value,
                "BUS_REVIEW_NOTES": bus_notes_value,
            }

        workbook.close()

        # ── Phase 2: Integrity analysis ────────────────────────────────────
        default_status = self.DEFAULT_REVIEW_STATUS
        imported_keys = set(imported_data.keys())

        # Records with non-default reviewer data absent from the imported file
        missing_with_data: list[str] = []
        for key, state in self.user_reviewer_state.items():
            if key in imported_keys or not isinstance(state, dict):
                continue
            has_non_default = (
                state.get("REVIEW_STATUS", default_status) != default_status
                or bool(str(state.get("TECH_REVIEW_NOTES", "")).strip())
                or bool(str(state.get("BUS_REVIEW_NOTES", "")).strip())
            )
            if has_non_default:
                bname_part = key.split("|", 1)[-1] if "|" in key else key
                missing_with_data.append(bname_part)

        # Notes that were populated and will be erased by the import
        notes_cleared: list[tuple[str, list[str]]] = []
        for key, new_vals in imported_data.items():
            old_state = self.user_reviewer_state.get(key)
            if not isinstance(old_state, dict):
                continue
            cleared_fields: list[str] = []
            for note_field in ("TECH_REVIEW_NOTES", "BUS_REVIEW_NOTES"):
                old_val = str(old_state.get(note_field, "")).strip()
                new_val = new_vals.get(note_field, "")
                if old_val and not new_val:
                    formal = str(self._get_user_preview_column_definition(note_field).get("formal", note_field))
                    cleared_fields.append(formal)
            if cleared_fields:
                notes_cleared.append((new_vals["BNAME"], cleared_fields))

        # ── Phase 3: Confirmation dialog ───────────────────────────────────
        confirm_dialog = _ImportReviewConfirmDialog(
            self,
            total_in_file=len(imported_data),
            missing_with_data=sorted(missing_with_data),
            notes_cleared=notes_cleared,
        )
        if confirm_dialog.exec() != QDialog.DialogCode.Accepted:
            return

        selected_mode = confirm_dialog.selected_mode
        import_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")

        # ── Phase 4: Apply ─────────────────────────────────────────────────
        imported_count = 0
        for key, new_vals in imported_data.items():
            current = self.user_reviewer_state.setdefault(key, self._default_reviewer_values().copy())
            current["REVIEW_STATUS"] = new_vals["REVIEW_STATUS"]
            if selected_mode == _ImportReviewConfirmDialog.MODE_PRESERVE_NOTES:
                for note_field in ("TECH_REVIEW_NOTES", "BUS_REVIEW_NOTES"):
                    new_val = new_vals.get(note_field, "")
                    if new_val or not str(current.get(note_field, "")).strip():
                        current[note_field] = new_val
            else:
                current["TECH_REVIEW_NOTES"] = new_vals["TECH_REVIEW_NOTES"]
                current["BUS_REVIEW_NOTES"] = new_vals["BUS_REVIEW_NOTES"]
            current["LAST_IMPORT_DATE"] = import_timestamp
            imported_count += 1

        # ── Phase 5: Persist, log and refresh ─────────────────────────────
        self._save_user_reviewer_state()

        mode_label = "שמור הערות קיימות" if selected_mode == _ImportReviewConfirmDialog.MODE_PRESERVE_NOTES else "כל השינויים"
        summary_parts = [f"עודכנו: {imported_count}"]
        if missing_with_data:
            summary_parts.append(f"לא בקובץ: {len(missing_with_data)}")
        if notes_cleared:
            summary_parts.append(f"הערות שנמחקו: {len(notes_cleared)}")
        summary_parts.append(f"מצב: {mode_label}")
        error_preview = " | ".join(summary_parts)

        import_now = datetime.now()
        record: dict[str, Any] = {
            "slot_key": "ייבוא סקירה",
            "report_group": "סקירת משתמשים",
            "file_name": Path(file_path).name,
            "extraction_date": "",
            "row_count": imported_count,
            "status": "יובא",
            "error_count": 0,
            "error_preview": error_preview,
            "date": import_now.strftime("%Y-%m-%d"),
            "time": import_now.strftime("%H:%M:%S"),
            "issues": [],
        }
        self.run_log_records.append(record)

        row_index = self.run_log_table.rowCount()
        self.run_log_table.insertRow(row_index)
        log_values = [
            record["slot_key"],
            record["report_group"],
            record["file_name"],
            record["extraction_date"],
            str(record["row_count"]),
            record["status"],
            str(record["error_count"]),
            record["error_preview"],
            record["date"],
            record["time"],
        ]
        for column, value in enumerate(log_values):
            item = QTableWidgetItem(self.format_rtl_text(value))
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            item.setToolTip(self.format_rtl_text(value))
            if column == 5:
                item.setBackground(QColor("#e8f4fd"))
            self.run_log_table.setItem(row_index, column, item)
        self.run_log_table.resizeColumnsToContents()

        self.refresh_user_preview()

        notes_warn = (
            f"\n⚠ {len(notes_cleared)} הערות שאוכלסו נמחקו."
            if notes_cleared and selected_mode != _ImportReviewConfirmDialog.MODE_PRESERVE_NOTES
            else ""
        )
        missing_warn = (
            f"\n⚠ {len(missing_with_data)} רשומות עם נתונים קיימים לא נמצאו בקובץ (נשמרו ללא שינוי)."
            if missing_with_data
            else ""
        )
        QMessageBox.information(
            self,
            "הייבוא הושלם",
            f"יובאו בהצלחה {imported_count} שורות מקובץ הסקירה.{notes_warn}{missing_warn}",
        )

    # ------------------------------------------------------------------
    # Batch AI narration for user findings (Phase 1 batch flow)
    # ------------------------------------------------------------------
    def generate_ai_narration_for_all_findings(self) -> None:
        """Run AI narration on every user with findings, using a background QThread.

        Pre-checks: AI must be enabled in system settings, the findings_narration feature
        must be on, and Ollama must be reachable.  Results are cached so the next
        ``refresh_user_preview()`` picks them up automatically into the
        ``FINDINGS_DESCRIPTION_AI`` column.
        """
        if self.batch_narration_thread is not None:
            QMessageBox.information(self, "פעולה כבר רצה", "כבר רץ תהליך הפקת נרטיב ברקע. נא להמתין לסיומו.")
            return

        settings = self._current_system_settings()
        ai_cfg = settings.get("ai_settings", {}) if isinstance(settings, dict) else {}
        if not bool(ai_cfg.get("enabled", False)):
            QMessageBox.warning(
                self,
                "AI לא מופעל",
                "כדי להפיק נרטיב ממצאים יש להפעיל את ה-AI בהגדרות המערכת (טאב הגדרות → הגדרות AI).",
            )
            return
        features = ai_cfg.get("features", {}) if isinstance(ai_cfg, dict) else {}
        if not bool(features.get("findings_narration", True)):
            QMessageBox.warning(
                self,
                "תכונת AI כבויה",
                "תכונת 'נרטיב ממצאים AI' אינה מסומנת בהגדרות. נא להפעילה ולנסות שוב.",
            )
            return

        try:
            client_check = OllamaClient(ai_cfg)
            if not client_check.is_available():
                QMessageBox.warning(
                    self,
                    "Ollama אינו זמין",
                    f"לא ניתן להתחבר לשרת Ollama בכתובת {ai_cfg.get('ollama_host', '')}\n"
                    "ודא ש-Ollama פועל מקומית ושהמודל המוגדר מותקן.",
                )
                return
        except Exception as error:
            QMessageBox.warning(self, "שגיאת חיבור", f"בדיקת זמינות Ollama נכשלה:\n{error}")
            return

        # Build list of items {row, raw_findings} from preview data
        try:
            usr02_rows = self._load_preview_rows("USR02")
            combined_rows = self._load_preview_rows("ADR6_USR21")
        except Exception as error:
            QMessageBox.warning(self, "אין נתונים", f"טעינת נתוני המשתמשים נכשלה:\n{error}")
            return

        if not usr02_rows:
            QMessageBox.warning(self, "אין נתונים", "טרם נטענו משתמשים. נא לבחור ולעבד את קובץ USR02 בלשונית קליטת קבצים.")
            return

        # Index ADR6 rows by BNAME for quick merge
        combined_by_bname: dict[str, dict[str, Any]] = {}
        for entry in combined_rows:
            key = str(entry.get("BNAME", "")).strip().upper()
            if key:
                combined_by_bname[key] = entry

        extraction_date = self._get_slot_extraction_date("USR02")
        items: list[dict[str, Any]] = []
        for usr in usr02_rows:
            merged = dict(usr)
            bname_key = str(usr.get("BNAME", "")).strip().upper()
            if bname_key and bname_key in combined_by_bname:
                for k, v in combined_by_bname[bname_key].items():
                    merged.setdefault(k, v)
            raw_findings = self._build_user_findings_description(merged, extraction_date)
            if raw_findings and raw_findings.strip():
                items.append({"row": merged, "raw_findings": raw_findings})

        if not items:
            QMessageBox.information(self, "אין ממצאים", "לא נמצאו משתמשים עם ממצאים שדורשים נרטיב AI.")
            return

        total = len(items)
        progress = QProgressDialog(
            self.format_rtl_text(f"מפיק נרטיב AI לכל הממצאים... (0 / {total})"),
            "ביטול",
            0,
            total,
            self,
        )
        progress.setWindowTitle("הפקת נרטיב AI")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setAutoClose(False)
        progress.setAutoReset(False)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        self.batch_narration_progress_dialog = progress

        self.batch_narration_thread = QThread(self)
        self.batch_narration_worker = BatchNarrationWorker(
            items=items,
            ai_settings=dict(ai_cfg),
            work_environment=self._current_work_environment_label(),
        )
        self.batch_narration_worker.moveToThread(self.batch_narration_thread)

        self.batch_narration_thread.started.connect(self.batch_narration_worker.run)
        self.batch_narration_worker.progress.connect(self._on_batch_narration_progress)
        self.batch_narration_worker.finished.connect(self._on_batch_narration_finished)
        self.batch_narration_worker.finished.connect(self.batch_narration_thread.quit)
        self.batch_narration_worker.finished.connect(self.batch_narration_worker.deleteLater)
        self.batch_narration_thread.finished.connect(self.batch_narration_thread.deleteLater)

        progress.canceled.connect(self._on_batch_narration_cancel)

        self.user_preview_ai_narrate_button.setEnabled(False)
        self.batch_narration_thread.start()

    @Slot(int, int, str)
    def _on_batch_narration_progress(self, done: int, total: int, bname: str) -> None:
        dialog = self.batch_narration_progress_dialog
        if dialog is None:
            return
        try:
            dialog.setValue(done)
            label = f"מפיק נרטיב AI לכל הממצאים... ({done} / {total})"
            if bname:
                label += f"  -  {bname}"
            dialog.setLabelText(self.format_rtl_text(label))
        except Exception:
            pass

    @Slot()
    def _on_batch_narration_cancel(self) -> None:
        worker = self.batch_narration_worker
        if worker is not None:
            try:
                worker.request_cancel()
            except Exception:
                pass

    @Slot(int, int, int)
    def _on_batch_narration_finished(self, processed: int, skipped: int, failed: int) -> None:
        dialog = self.batch_narration_progress_dialog
        if dialog is not None:
            try:
                dialog.close()
            except Exception:
                pass
        self.batch_narration_progress_dialog = None
        self.batch_narration_worker = None
        self.batch_narration_thread = None
        try:
            self.user_preview_ai_narrate_button.setEnabled(True)
        except Exception:
            pass

        # Refresh preview so cached narrations populate FINDINGS_DESCRIPTION_AI
        try:
            self.refresh_user_preview()
        except Exception:
            pass

        QMessageBox.information(
            self,
            "הפקת נרטיב AI הסתיימה",
            f"הופקו: {processed}\nדולגו: {skipped}\nנכשלו: {failed}\n\n"
            "התוצאות נשמרות במטמון מקומי וניתן לראותן בעמודה 'תיאור ממצאים (AI)'.",
        )

    def _email_from_settings(self, settings_key: str) -> str:
        settings = self._current_system_settings()
        return str(settings.get(settings_key, "")).strip()

    @staticmethod
    def _validate_email_address(email_value: str) -> bool:
        normalized = email_value.strip()
        return bool(normalized and "@" in normalized and "." in normalized.split("@")[-1])

    def _create_outlook_review_draft(self, recipient_email: str, role_label: str) -> None:
        if not self._validate_email_address(recipient_email):
            QMessageBox.warning(self, "מייל לא מוגדר", f"לא הוגדרה כתובת מייל תקינה עבור {role_label} במסך ההגדרות.")
            return

        export_path = self.export_user_preview_to_excel(open_after_export=False)
        if export_path is None:
            return

        if not sys.platform.startswith("win"):
            QMessageBox.warning(self, "מערכת לא נתמכת", "יצירת טיוטת מייל נתמכת כרגע ב-Windows בלבד.")
            return

        try:
            import win32com.client  # type: ignore[import-not-found]
        except ModuleNotFoundError:
            install_command = f'"{sys.executable}" -m pip install pywin32'
            QMessageBox.warning(
                self,
                "רכיב חסר ל-Outlook",
                "לא ניתן ליצור טיוטת Outlook כי חסרה חבילת pywin32.\n\n"
                f"יש להריץ פעם אחת בסביבת העבודה:\n{install_command}",
            )
            return
        except Exception as error:
            QMessageBox.warning(self, "Outlook לא זמין", f"לא ניתן לטעון Outlook COM ליצירת טיוטה:\n{error}")
            return

        try:
            outlook = win32com.client.Dispatch("Outlook.Application")
            mail_item = outlook.CreateItem(0)
            mail_item.To = recipient_email
            mail_item.Subject = self.format_rtl_text(f"סקירת דוח משתמשים - {datetime.now().strftime('%Y-%m-%d')}")
            mail_item.HTMLBody = (
                "<div dir='rtl' style='text-align:right; font-family:Arial, sans-serif; font-size:12pt;'>"
                "<p>שלום,</p>"
                "<p>מצורף דוח סקירת משתמשים עדכני מתוך המערכת.</p>"
                "<p>נא לעבור על הממצאים ולעדכן סטטוס/הערות בהתאם.</p>"
                "<p>בברכה,<br>מערכת בקרות ITGC</p>"
                "</div>"
            )
            mail_item.Attachments.Add(str(export_path))
            mail_item.Display(False)
        except Exception as error:
            QMessageBox.warning(self, "שגיאת מייל", f"יצירת טיוטת מייל נכשלה:\n{error}")
            return

        QMessageBox.information(
            self,
            "טיוטת מייל נוצרה",
            f"נוצרה טיוטה ל-{role_label} עם קובץ מצורף:\n{export_path}",
        )

    def draft_user_review_email_to_business(self) -> None:
        self._create_outlook_review_draft(
            recipient_email=self._email_from_settings("business_reviewer_email"),
            role_label="גורם מהכספים",
        )

    def draft_user_review_email_to_technical(self) -> None:
        self._create_outlook_review_draft(
            recipient_email=self._email_from_settings("technical_reviewer_email"),
            role_label="גורם טכני",
        )

    def open_output_folder(self) -> None:
        self._open_path(self.config.output_dir)

    def open_report(self) -> None:
        if self.report_path and self.report_path.exists():
            self._open_path(self.report_path)
        else:
            QMessageBox.warning(self, "דוח לא זמין", "טרם נוצר דוח אקסל לפתיחה.")

    @staticmethod
    def _open_path(path: Path) -> None:
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore[attr-defined]
            return
        if sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=False)
            return
        subprocess.run(["xdg-open", str(path)], check=False)


def launch_desktop_app() -> None:
    app = get_qt_app()
    window = ValidationDesktopApp()
    window.show()
    app.exec()
