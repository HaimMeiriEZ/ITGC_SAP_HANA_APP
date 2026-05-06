from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook

from src.models.validation_result import ValidationIssue, ValidationResult


class ExcelReportWriter:
    def write(
        self,
        result: ValidationResult,
        source_file: Path,
        output_dir: Path,
    ) -> Path:
        output_dir.mkdir(parents=True, exist_ok=True)
        generated_at = datetime.now()
        report_path = output_dir / (
            f"{source_file.stem}_שגיאות_קליטה_{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "סיכום"
        self._write_summary(summary_sheet, result, source_file, generated_at)

        issues_sheet = workbook.create_sheet("שגיאות")
        self._write_issues(issues_sheet, result.issues)

        workbook.save(report_path)
        return report_path

    @staticmethod
    def _write_summary(sheet, result: ValidationResult, source_file: Path, generated_at: datetime) -> None:

        sheet.append(["מדד", "ערך"])
        sheet.append(["שורות שנבדקו", result.summary.total_rows])
        sheet.append(["שורות תקינות", result.summary.valid_rows])
        sheet.append(["שורות שגויות", result.summary.invalid_rows])
        sheet.append(["הקובץ תקין", result.summary.is_valid])
        if result.source_files:
            unique_sources: list[str] = []
            for name in result.source_files:
                source_name = str(name).strip()
                if source_name and source_name not in unique_sources:
                    unique_sources.append(source_name)
            source_label = " | ".join(unique_sources) if unique_sources else source_file.name
        else:
            source_label = source_file.name
        sheet.append(["קובץ מקור", source_label])
        sheet.append(["סיבת קליטה שגויה", ExcelReportWriter._intake_failure_reason_text(result.issues)])
        sheet.append(["פרופיל בדיקה", result.detected_profile or "GENERIC"])
        sheet.append(["מספר קבצים שנבחרו", max(len(result.source_files), 1)])
        sheet.append(["תאריך הפקה", generated_at.strftime("%Y-%m-%d")])
        sheet.append(["שעת הפקה", generated_at.strftime("%H:%M:%S")])

    @staticmethod
    def _is_intake_issue(issue: ValidationIssue) -> bool:
        msg = str(issue.message or "")
        if issue.row_number == 0:
            return (
                "עמודת חובה חסרה" in msg
                or "אינו תואם למבנה" in msg
                or "נדרשת לפחות" in msg
            )
        return issue.row_number > 0 and "ערך חובה חסר" in msg

    @staticmethod
    def _intake_failure_reason_text(issues: Iterable[ValidationIssue]) -> str:
        reasons: list[str] = []
        for issue in issues:
            if not ExcelReportWriter._is_intake_issue(issue):
                continue
            text = str(issue.message or "").strip()
            if text and text not in reasons:
                reasons.append(text)
        if not reasons:
            return "-"
        return " | ".join(reasons)

    @staticmethod
    def _write_issues(sheet, issues: Iterable[ValidationIssue]) -> None:
        sheet.append(["מספר שורה", "שם עמודה", "הודעת שגיאה", "קובץ מקור"])
        for issue in issues:
            sheet.append([
                ExcelReportWriter._safe_excel_value(issue.row_number),
                ExcelReportWriter._safe_excel_value(issue.column_name),
                ExcelReportWriter._safe_excel_value(issue.message),
                ExcelReportWriter._safe_excel_value(issue.source_file),
            ])

    @staticmethod
    def _safe_excel_value(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, list):
            return " | ".join(str(item) for item in value)
        if isinstance(value, dict):
            return " | ".join(f"{key}={item}" for key, item in value.items())
        return value

    @staticmethod
    def write_audit_findings_report(
        summary_rows: list[dict[str, Any]],
        detail_rows: list[dict[str, Any]],
        output_path: Path,
    ) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "ריכוז ממצאים"
        summary_sheet.append([
            "מזהה בקרה",
            "סוג בדיקה",
            "קובץ מקור",
            "תאריך הפקה",
            "סביבת עבודה",
            "רמת סיכון",
            "תיאור הבדיקה",
            "רשומות תקינות",
            "רשומות עם ממצא",
            "סהכ רשומות",
        ])
        for row in summary_rows:
            summary_sheet.append([
                ExcelReportWriter._safe_excel_value(row.get("control_id", "")),
                ExcelReportWriter._safe_excel_value(row.get("check_type", "")),
                ExcelReportWriter._safe_excel_value(row.get("source_file", "")),
                ExcelReportWriter._safe_excel_value(row.get("extraction_date", "")),
                ExcelReportWriter._safe_excel_value(row.get("work_environment", "")),
                ExcelReportWriter._safe_excel_value(row.get("risk_level", "")),
                ExcelReportWriter._safe_excel_value(row.get("description", "")),
                ExcelReportWriter._safe_excel_value(row.get("valid_records", 0)),
                ExcelReportWriter._safe_excel_value(row.get("finding_records", 0)),
                ExcelReportWriter._safe_excel_value(row.get("total_records", 0)),
            ])

        details_sheet = workbook.create_sheet("פירוט ממצאים")
        details_sheet.append([
            "מזהה בקרה",
            "קובץ מקור",
            "תאריך הפקה",
            "סביבת עבודה",
            "קטגוריה",
            "רמת סיכון",
            "תיאור",
            "סוג בדיקה",
            "ערך בפועל",
            "ערך מצופה",
            "סטטוס",
            "תיאור מלא",
        ])
        for row in detail_rows:
            details_sheet.append([
                ExcelReportWriter._safe_excel_value(row.get("control_id", "")),
                ExcelReportWriter._safe_excel_value(row.get("source_file", "")),
                ExcelReportWriter._safe_excel_value(row.get("extraction_date", "")),
                ExcelReportWriter._safe_excel_value(row.get("work_environment", "")),
                ExcelReportWriter._safe_excel_value(row.get("category", "")),
                ExcelReportWriter._safe_excel_value(row.get("risk_level", "")),
                ExcelReportWriter._safe_excel_value(row.get("description", "")),
                ExcelReportWriter._safe_excel_value(row.get("check_type", "")),
                ExcelReportWriter._safe_excel_value(row.get("actual_value", "-")),
                ExcelReportWriter._safe_excel_value(row.get("expected_value", "-")),
                ExcelReportWriter._safe_excel_value(row.get("status", "")),
                ExcelReportWriter._safe_excel_value(row.get("full_description", "")),
            ])

        workbook.save(output_path)
        return output_path
