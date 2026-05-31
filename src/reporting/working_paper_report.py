"""Generate per-control 'Working Paper' Excel reports.

Each working paper Excel file contains 3 sheets:
1. {control_id} (sanitized) - general details (key-value layout)
2. IPE - metadata table + embedded screenshot images
3. אוכלוסיה נבחנת - raw population with finding highlights
4. ריכוז ממצאים - filtered (relevant) population
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.validators.spec_rules import (
    AUDIT_CONTROL_DEFINITIONS,
    CONTROL_REQUIRED_TABLES,
    PROFILE_REQUIRED_TABLES,
    build_test_steps_for_control,
)


_HEADER_FILL = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_KEY_FILL = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
_KEY_FONT = Font(bold=True, size=11)
_FINDING_FILL = PatternFill(start_color="FFFFE6E6", end_color="FFFFE6E6", fill_type="solid")
_SECTION_FILL = PatternFill(start_color="FF8EA9DB", end_color="FF8EA9DB", fill_type="solid")
_SECTION_FONT = Font(bold=True, color="FFFFFFFF", size=12)
_THIN = Side(border_style="thin", color="FF808080")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP_RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", name).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:31]


def _set_rtl(sheet) -> None:
    sheet.sheet_view.rightToLeft = True


def _apply_header(cell) -> None:
    cell.fill = _HEADER_FILL
    cell.font = _HEADER_FONT
    cell.alignment = _CENTER
    cell.border = _BORDER


def _apply_value_cell(cell, *, fill: PatternFill | None = None, font: Font | None = None) -> None:
    cell.alignment = _WRAP_RIGHT
    cell.border = _BORDER
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font


def write_control_working_paper(
    *,
    control_id: str,
    summary_record: dict[str, Any],
    detail_rows: list[dict[str, Any]],
    raw_population_rows: list[dict[str, Any]],
    ipe_entries: list[dict[str, Any]],
    work_environment_label: str,
    output_path: Path,
    notes: list[str] | None = None,
    critical_roles: list[str] | None = None,
    raw_population_note: str | None = None,
) -> Path:
    """Build the working-paper workbook and save to *output_path*.

    Parameters
    ----------
    control_id : str
        e.g. "MA2-2_AYALON_6".
    summary_record : dict
        Single row from `audit_summary_records[control_id]`.
    detail_rows : list[dict]
        Rows from `audit_details_by_control[control_id]` (filtered/relevant population).
    raw_population_rows : list[dict]
        Full source rows (e.g. cached AGR_USERS, RSPARAM rows, etc.). May be empty.
    ipe_entries : list[dict]
        IPE evidence records associated with the matching slot, each having
        keys: original_filename, stored_path, control_ids, added_at, id.
    work_environment_label : str
        Current work environment string ("ייצור", "פיתוח", etc.).
    output_path : Path
        Destination .xlsx path.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    overview_sheet = workbook.active
    overview_sheet.title = _sanitize_sheet_name(control_id)
    _write_overview_sheet(
        overview_sheet,
        control_id=control_id,
        summary_record=summary_record,
        detail_rows=detail_rows,
        raw_population_rows=raw_population_rows,
        work_environment_label=work_environment_label,
        notes=notes,
        critical_roles=critical_roles,
    )

    ipe_sheet = workbook.create_sheet("IPE")
    _write_ipe_sheet(ipe_sheet, ipe_entries)

    examined_population_sheet = workbook.create_sheet("אוכלוסיה נבחנת")
    _write_examined_population_sheet(
        examined_population_sheet,
        detail_rows,
        raw_population_rows,
        raw_population_note=raw_population_note,
    )

    findings_sheet = workbook.create_sheet("ריכוז ממצאים")
    _write_findings_sheet(findings_sheet, detail_rows)

    workbook.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Sheet 1: General details (key/value)
# ---------------------------------------------------------------------------

def _write_overview_sheet(
    sheet,
    *,
    control_id: str,
    summary_record: dict[str, Any],
    detail_rows: list[dict[str, Any]],
    raw_population_rows: list[dict[str, Any]],
    work_environment_label: str,
    notes: list[str] | None = None,
    critical_roles: list[str] | None = None,
) -> None:
    _set_rtl(sheet)
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 110

    meta = AUDIT_CONTROL_DEFINITIONS.get(control_id, {}) or {}
    process = meta.get("process", "-") or "-"
    risk_desc = meta.get("risk_description", "-") or "-"
    description = meta.get("description") or summary_record.get("description", "-") or "-"

    # Test steps: prefer CSV override if present, otherwise dynamic template
    test_steps_override = meta.get("test_steps_override", "").strip()
    test_steps = test_steps_override or build_test_steps_for_control(
        control_id, critical_roles=critical_roles
    )

    docs = (
        CONTROL_REQUIRED_TABLES.get(control_id)
        or PROFILE_REQUIRED_TABLES.get(_detect_profile_from_summary(summary_record))
        or "-"
    )

    examined_user_pairs = _distinct_user_pairs(
        raw_population_rows,
        client_columns=("MANDT", "CLIENT"),
        user_columns=("BNAME", "UNAME", "USER", "USER_NAME"),
    )
    finding_user_pairs = _distinct_user_pairs(
        detail_rows,
        client_columns=("client", "MANDT", "CLIENT"),
        user_columns=("user_name", "BNAME", "UNAME", "USER"),
        finding_only=True,
    )

    if examined_user_pairs:
        total = len(examined_user_pairs)
        findings = len(finding_user_pairs)
        pct = (findings / total * 100.0) if total > 0 else 0.0
        summary_text = f"נמצאו {findings} מתוך {total} מקרים שנבדקו ({pct:.1f}%)."
    else:
        total = int(summary_record.get("total_records", 0) or 0)
        findings = int(summary_record.get("finding_records", 0) or 0)
        pct = (findings / total * 100.0) if total > 0 else 0.0
        summary_text = (
            f"נמצאו {findings} ממצאים מתוך {total} רשומות שנבדקו ({pct:.1f}%)."
            if total > 0
            else "לא בוצעה בדיקה / אין רשומות."
        )

    extraction_date = str(summary_record.get("extraction_date", "-") or "-")

    rows: list[tuple[str, str]] = [
        ("מספר בקרה", control_id),
        ("שם מערכת", "SAP HANA Application"),
        ("סביבת מערכת", work_environment_label or "-"),
        ("תאריך הפקה", extraction_date),
        ("תהליך", process),
        ("תיאור הסיכון", risk_desc),
        ("תיאור הבקרה", description),
        ("צעדי טסט", test_steps),
        ("תיעוד נדרש", docs),
        ("סיכום ממצאים", summary_text),
    ]
    cleaned_notes = [str(note).strip() for note in (notes or []) if str(note).strip()]
    if cleaned_notes:
        rows.append(("הערות", "\n".join(cleaned_notes)))

    # Title row
    sheet.cell(row=1, column=1, value="פרטי הבקרה")
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.fill = _SECTION_FILL
    title_cell.font = _SECTION_FONT
    title_cell.alignment = _CENTER
    title_cell.border = _BORDER
    sheet.row_dimensions[1].height = 24

    for idx, (key, value) in enumerate(rows, start=2):
        key_cell = sheet.cell(row=idx, column=1, value=key)
        _apply_value_cell(key_cell, fill=_KEY_FILL, font=_KEY_FONT)
        value_cell = sheet.cell(row=idx, column=2, value=value)
        _apply_value_cell(value_cell)
        if key == "צעדי טסט":
            # Force right-to-left reading order on this specific cell value
            # (openpyxl uses readingOrder=2 for RTL).
            value_cell.alignment = Alignment(
                horizontal="right",
                vertical="top",
                wrap_text=True,
                readingOrder=2,
            )
        # Allow tall multiline cells (esp. test steps)
        line_count = max(1, str(value).count("\n") + 1)
        sheet.row_dimensions[idx].height = max(22, min(line_count * 16, 420))


def _detect_profile_from_summary(summary_record: dict[str, Any]) -> str:
    profile = str(summary_record.get("detected_profile", "") or "").upper()
    if profile:
        return profile
    # Fall back to source_file string hints
    src = str(summary_record.get("source_file", "") or "").upper()
    for key in ("RSPARAM", "TPFET", "USR02", "AGR_1251", "AGR_USERS", "UST04", "USH04", "STMS"):
        if key in src:
            return key
    return ""


# ---------------------------------------------------------------------------
# Sheet 2: IPE
# ---------------------------------------------------------------------------

def _write_ipe_sheet(sheet, ipe_entries: list[dict[str, Any]]) -> None:
    _set_rtl(sheet)
    headers = [
        "#",
        "שם קובץ מקורי",
        "תאריך הפקה",
        "כמות רשומות שנקלטה",
        "תאריך הוספה",
        "תצוגה",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        _apply_header(cell)

    sheet.column_dimensions[get_column_letter(1)].width = 5
    sheet.column_dimensions[get_column_letter(2)].width = 38
    sheet.column_dimensions[get_column_letter(3)].width = 18
    sheet.column_dimensions[get_column_letter(4)].width = 22
    sheet.column_dimensions[get_column_letter(5)].width = 22
    sheet.column_dimensions[get_column_letter(6)].width = 55

    if not ipe_entries:
        cell = sheet.cell(row=2, column=1, value="לא נמצאו תיעודי IPE עבור בקרה זו.")
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        _apply_value_cell(cell)
        return

    current_row = 2
    for idx, entry in enumerate(ipe_entries, start=1):
        original = str(entry.get("original_filename", "-"))
        extraction_date = str(entry.get("extraction_date", "-") or "-")
        population_count = entry.get("population_count", "-")
        if isinstance(population_count, (int, float)):
            population_text = f"{int(population_count):,}"
        else:
            population_text = str(population_count or "-")
        added_at = str(entry.get("added_at", "-"))

        for col_idx, value in enumerate(
            [idx, original, extraction_date, population_text, added_at, ""],
            start=1,
        ):
            cell = sheet.cell(row=current_row, column=col_idx, value=value if col_idx != 6 else "")
            _apply_value_cell(cell)

        stored_path = Path(str(entry.get("stored_path", "")))
        image_row_height = 22
        if stored_path.exists() and stored_path.suffix.lower() in {".png", ".jpg", ".jpeg", ".gif", ".bmp"}:
            try:
                img = XLImage(str(stored_path))
                # Scale: cap width at ~420px and height at ~280px
                max_w, max_h = 420, 280
                scale = min(max_w / max(img.width, 1), max_h / max(img.height, 1), 1.0)
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
                anchor_cell = f"F{current_row}"
                sheet.add_image(img, anchor_cell)
                image_row_height = max(image_row_height, img.height * 0.75 + 8)
            except Exception as exc:
                cell = sheet.cell(
                    row=current_row,
                    column=6,
                    value=f"(נכשלה טעינת תמונה: {original} — {exc})",
                )
                _apply_value_cell(cell)
        else:
            cell = sheet.cell(
                row=current_row,
                column=6,
                value=f"(קובץ לא תמונה / חסר: {stored_path.name})",
            )
            _apply_value_cell(cell)

        sheet.row_dimensions[current_row].height = image_row_height
        current_row += 1


# ---------------------------------------------------------------------------
# Sheet 3: Detail (raw + filtered)
# ---------------------------------------------------------------------------

_INTERNAL_KEYS = {"__source_file", "__row_number", "__profile"}


def _ordered_keys(
    rows: Iterable[dict[str, Any]],
    drop_columns: set[str] | None = None,
) -> list[str]:
    seen: list[str] = []
    drop = drop_columns or set()
    for row in rows:
        for k in row.keys():
            if k in _INTERNAL_KEYS:
                continue
            if k in drop:
                continue
            if k not in seen:
                seen.append(k)
    return seen


def _write_table_block(
    sheet,
    *,
    start_row: int,
    title: str,
    rows: list[dict[str, Any]],
    finding_keys: set[tuple[str, ...]] | None,
    key_columns: list[str] | None,
    use_status_column: bool = False,
    drop_columns: set[str] | None = None,
    row_status_matcher: Callable[[dict[str, Any]], bool] | None = None,
) -> int:
    """Write a titled table to *sheet* starting at *start_row*. Returns next free row."""
    section_cell = sheet.cell(row=start_row, column=1, value=title)
    section_cell.fill = _SECTION_FILL
    section_cell.font = _SECTION_FONT
    section_cell.alignment = _CENTER
    section_cell.border = _BORDER

    if not rows:
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
        empty_cell = sheet.cell(row=start_row + 1, column=1, value="(אין רשומות)")
        sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=4)
        _apply_value_cell(empty_cell)
        return start_row + 3

    columns = _ordered_keys(rows, drop_columns=drop_columns)
    columns_with_status = columns + ["סטטוס"]

    sheet.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=len(columns_with_status),
    )

    header_row = start_row + 1
    for col_idx, header in enumerate(columns_with_status, start=1):
        cell = sheet.cell(row=header_row, column=col_idx, value=header)
        _apply_header(cell)
        sheet.column_dimensions[get_column_letter(col_idx)].width = max(
            14, min(40, len(str(header)) + 4)
        )

    for row_offset, row in enumerate(rows, start=header_row + 1):
        if use_status_column:
            is_finding = str(row.get("status", "")).strip() == "עם ממצא"
        elif row_status_matcher is not None:
            is_finding = row_status_matcher(row)
        else:
            is_finding = _row_is_finding(row, finding_keys, key_columns)
        row_fill = _FINDING_FILL if is_finding else None
        for col_idx, key in enumerate(columns, start=1):
            value = row.get(key, "")
            cell = sheet.cell(row=row_offset, column=col_idx, value=_excel_safe(value))
            _apply_value_cell(cell, fill=row_fill)
        status_cell = sheet.cell(
            row=row_offset,
            column=len(columns_with_status),
            value="עם ממצא" if is_finding else "תקין",
        )
        _apply_value_cell(status_cell, fill=row_fill)

    return header_row + len(rows) + 2


def _row_is_finding(
    row: dict[str, Any],
    finding_keys: set[tuple[str, ...]] | None,
    key_columns: list[str] | None,
) -> bool:
    if not finding_keys or not key_columns:
        return False
    try:
        key = tuple(str(row.get(col, "") or "").strip().upper() for col in key_columns)
    except Exception:
        return False
    return key in finding_keys


def _excel_safe(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(v) for v in value)
    if isinstance(value, dict):
        return ", ".join(f"{k}={v}" for k, v in value.items())
    return value


def _build_finding_keys(
    detail_rows: list[dict[str, Any]],
    key_columns: list[str] | None,
) -> set[tuple[str, ...]]:
    if not key_columns:
        return set()
    keys: set[tuple[str, ...]] = set()
    for row in detail_rows:
        if str(row.get("status", "")).strip() != "עם ממצא":
            continue
        # Generic fallback for controls whose detail rows still carry the source
        # row identifier in actual_value (for example parameter name / BNAME).
        actual = str(row.get("actual_value", "") or "").strip().upper()
        if actual:
            keys.add((actual,) * len(key_columns) if len(key_columns) > 1 else (actual,))
    return keys


def _split_profile_values(value: object) -> set[str]:
    return {
        part.strip().upper()
        for part in re.split(r"[,\s]+", str(value or ""))
        if part.strip()
    }


def _row_value(row: dict[str, Any], *column_names: str) -> str:
    lookup = {str(key).upper(): value for key, value in row.items()}
    for column_name in column_names:
        value = lookup.get(column_name.upper())
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _is_valid_identifier(value: str) -> bool:
    normalized = str(value or "").strip()
    return bool(normalized and normalized.lower() not in {"-", "nan", "none"})


def _distinct_user_pairs(
    rows: Iterable[dict[str, Any]],
    *,
    client_columns: tuple[str, ...],
    user_columns: tuple[str, ...],
    finding_only: bool = False,
) -> set[tuple[str, str]]:
    pairs: set[tuple[str, str]] = set()
    for row in rows:
        if finding_only and str(row.get("status", "") or "").strip() != "עם ממצא":
            continue
        client = _row_value(row, *client_columns)
        user_name = _row_value(row, *user_columns)
        if not (_is_valid_identifier(client) and _is_valid_identifier(user_name)):
            continue
        pairs.add((client.strip().upper(), user_name.strip().upper()))
    return pairs


def _build_strong_profile_finding_lookup(
    detail_rows: list[dict[str, Any]],
) -> dict[tuple[str, str], set[str]]:
    lookup: dict[tuple[str, str], set[str]] = {}
    for row in detail_rows:
        if str(row.get("status", "")).strip() != "עם ממצא":
            continue
        control_id = str(row.get("control_id", "") or "").strip()
        if control_id and control_id != "MA3-3_AYALON_14":
            continue
        client = str(row.get("client", "") or "").strip().upper()
        user_name = str(row.get("user_name", "") or "").strip().upper()
        profiles = _split_profile_values(row.get("actual_value", ""))
        if client and user_name and profiles:
            lookup.setdefault((client, user_name), set()).update(profiles)
    return lookup


def _row_matches_strong_profile_finding(
    row: dict[str, Any],
    row_profile: str,
    finding_lookup: dict[tuple[str, str], set[str]],
) -> bool:
    if row_profile not in {"UST04", "USH04"}:
        return False
    client = _row_value(row, "MANDT", "CLIENT").upper()
    user_name = _row_value(row, "BNAME", "UNAME", "USER").upper()
    if not client or not user_name:
        return False
    finding_profiles = finding_lookup.get((client, user_name), set())
    if not finding_profiles:
        return False
    if row_profile == "UST04":
        row_profiles = {_row_value(row, "PROFILE").upper()}
    else:
        row_profiles = _split_profile_values(_row_value(row, "PROFS"))
        row_profiles.update(_split_profile_values(_row_value(row, "MODBE")))
    row_profiles.discard("")
    return bool(row_profiles & finding_profiles)


def _detect_key_columns(profile: str, available_columns: list[str]) -> list[str]:
    """Heuristic: which columns identify a unique row for finding lookup."""
    upper = {c.upper(): c for c in available_columns}
    if profile in {"RSPARAM", "TPFET"}:
        for cand in ("PARAMETER", "NAME"):
            if cand in upper:
                return [upper[cand]]
    if profile in {"USR02", "AGR_USERS"}:
        if "BNAME" in upper:
            return [upper["BNAME"]]
        if "UNAME" in upper:
            return [upper["UNAME"]]
    if profile in {"UST04", "USH04"}:
        if "BNAME" in upper:
            return [upper["BNAME"]]
    if profile == "AGR_1251" and "AGR_NAME" in upper:
        return [upper["AGR_NAME"]]
    return []


def _write_examined_population_sheet(
    sheet,
    detail_rows: list[dict[str, Any]],
    raw_population_rows: list[dict[str, Any]],
    *,
    raw_population_note: str | None = None,
) -> None:
    _set_rtl(sheet)

    # Detect a fallback profile (used only for key-column heuristic, not for
    # the per-row "טבלת מקור" column — which is now derived per row).
    profile = ""
    if raw_population_rows:
        profile = str(raw_population_rows[0].get("__profile", "") or "").upper()
        if not profile:
            src = str(raw_population_rows[0].get("__source_file", "") or "").upper()
            for key in ("RSPARAM", "TPFET", "USR02", "AGR_1251", "AGR_USERS", "UST04", "USH04", "STMS"):
                if key in src:
                    profile = key
                    break

    raw_columns = _ordered_keys(raw_population_rows) if raw_population_rows else []
    key_columns = _detect_key_columns(profile, raw_columns)
    finding_keys = _build_finding_keys(detail_rows, key_columns)
    strong_profile_lookup = _build_strong_profile_finding_lookup(detail_rows)

    def _row_profile(row: dict[str, Any]) -> str:
        p = str(row.get("__profile", "") or "").upper()
        if p:
            return p
        src = str(row.get("__source_file", "") or "").upper()
        for key in ("RSPARAM", "TPFET", "USR02", "AGR_1251", "AGR_USERS", "UST04", "USH04", "STMS"):
            if key in src:
                return key
        return profile  # last-resort fallback

    def _raw_status_matcher(row: dict[str, Any]) -> bool:
        row_profile = _row_profile(row)
        if strong_profile_lookup and row_profile in {"UST04", "USH04"}:
            return _row_matches_strong_profile_finding(row, row_profile, strong_profile_lookup)
        return _row_is_finding(row, finding_keys, key_columns)

    # Enrich raw rows with per-row source-table name so mixed populations
    # (e.g. UST04 + USH04 merged into one control) display correctly.
    if raw_population_rows:
        enriched_raw: list[dict[str, Any]] = []
        for row in raw_population_rows:
            enriched_raw.append({"טבלת מקור": _row_profile(row), **row})
        raw_for_table = enriched_raw
    else:
        raw_for_table = raw_population_rows

    next_row = _write_table_block(
        sheet,
        start_row=1,
        title="אוכלוסייה גולמית - כל הרשומות שנקלטו",
        rows=raw_for_table,
        finding_keys=finding_keys,
        key_columns=key_columns,
        row_status_matcher=_raw_status_matcher,
    )

    if raw_population_note:
        note_cell = sheet.cell(row=next_row, column=1, value=raw_population_note)
        sheet.merge_cells(
            start_row=next_row, start_column=1, end_row=next_row, end_column=6
        )
        note_cell.font = Font(italic=True, color="FFC00000", size=9)
        note_cell.alignment = Alignment(horizontal="right", vertical="center")
        next_row += 2


def _write_findings_sheet(
    sheet,
    detail_rows: list[dict[str, Any]],
) -> None:
    _set_rtl(sheet)

    # Enrich detail rows with the authorization-object that triggered the
    # finding (if available) so the table explains *why* each user is flagged.
    # We also drop the verbose "description" column from this table per spec.
    enriched_detail: list[dict[str, Any]] = []
    for row in detail_rows:
        copy = dict(row)
        copy["אובייקט הרשאה"] = row.get("auth_object", "-") or "-"
        copy.pop("auth_object", None)
        enriched_detail.append(copy)

    _write_table_block(
        sheet,
        start_row=1,
        title="אוכלוסייה רלוונטית / רשומות לפי כללי הבקרה",
        rows=enriched_detail,
        finding_keys=None,
        key_columns=None,
        use_status_column=True,
        drop_columns={"description"},
    )
